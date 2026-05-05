"""
fill_excel_from_db.py
======================
DB → Excel 채우기 스크립트
- Task #2: container_info.size_type='40' UPDATE
- Task #3: IN / UNSOLD 시트에 신규 행 추가

작성: Ruby (2026-05-02)
대상: SMQ 입,출고 재고관리 파일.xlsx
소스: backup/sqm_backup_20260429_231958.db (실데이터 40 LOT)
"""

import sqlite3
import shutil
from datetime import datetime, date
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent
DB_PATH = BASE / "backup" / "sqm_backup_20260429_231958.db"
EXCEL_SRC = BASE / "SMQ 입,출고 재고관리 파일.xlsx"
EXCEL_OUT = BASE / "SMQ 입,출고 재고관리 파일_v866.xlsx"

# ──────────────────────────────────────────────
# 헬퍼 매핑
# ──────────────────────────────────────────────
CARRIER_MAP = {
    "MAERSK": "Maersk",
    "HAPAG":  "Hapag-Lloyd",
    "MSC":    "MSC",
    "ONE":    "ONE",
}

PORT_MAP = {
    "KRKAN": "GWCT",
    "GWANGYANG": "GWCT",
}


def derive_carrier(bl_no: str, carrier_id: str) -> str:
    """BL 번호 접두사 또는 carrier_id로 운송사명 반환."""
    if carrier_id and carrier_id.upper() in CARRIER_MAP:
        return CARRIER_MAP[carrier_id.upper()]
    if not bl_no:
        return ""
    prefix = bl_no[:4].upper()
    mapping = {
        "MAEU": "Maersk",
        "HLCU": "Hapag-Lloyd",
        "MSCU": "MSC",
        "MEDU": "MSC",
        "ONEY": "ONE",
        "ONEU": "ONE",
    }
    return mapping.get(prefix, "")


def derive_terminal(bl_pod: str, warehouse_name: str) -> str:
    """양하부두 코드 반환: GWCT/KIT/PNIT."""
    for key, val in PORT_MAP.items():
        if key in (bl_pod or "").upper():
            return val
    # warehouse_name 내 키워드
    if "서부" in (warehouse_name or ""):
        return "GWCT"
    if "KIT" in (warehouse_name or "").upper():
        return "KIT"
    return "GWCT"  # 광양항 기본값


def derive_product_code(product_code: str) -> str:
    """'MIC9000.00/500 KG' → 'MIC9000'"""
    if not product_code:
        return ""
    return product_code.split(".")[0].strip()


def safe_int(v) -> int | None:
    """숫자 변환 가능하면 int, 아니면 None."""
    try:
        return int(str(v).strip())
    except (ValueError, TypeError):
        return None


def to_date(v) -> date | None:
    """'2026-04-10' 문자열 → date 객체."""
    if not v:
        return None
    if isinstance(v, (date, datetime)):
        return v if isinstance(v, date) else v.date()
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


# ──────────────────────────────────────────────
# Task #2: container_info.size_type UPDATE
# ──────────────────────────────────────────────
def task2_update_size_type():
    print("=" * 50)
    print("Task #2: container_info.size_type='40' UPDATE")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    cur.execute("UPDATE container_info SET size_type='40' WHERE size_type='' OR size_type IS NULL")
    updated = cur.rowcount
    conn.commit()
    conn.close()
    print(f"  ✅ {updated}건 UPDATE 완료 (size_type='40')")


# ──────────────────────────────────────────────
# Task #3: Excel 신규 행 추가
# ──────────────────────────────────────────────
def fetch_lot_data() -> list[dict]:
    """DB에서 실데이터 LOT 40건 쿼리."""
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    cur.execute("""
        SELECT
            i.lot_no,
            i.arrival_date,
            i.product_code,
            i.sap_no,
            i.bl_no,
            i.container_no,
            i.net_weight,
            i.gross_weight,
            i.mxbg_pallet,
            i.tonbag_count,
            i.salar_invoice_no,
            i.location,
            i.status,
            i.remarks,
            i.sale_ref,
            COALESCE(db.carrier_id, '') as carrier_id,
            COALESCE(db.port_of_discharge, '') as bl_pod,
            COALESCE(dd.mrn, '') as mrn,
            COALESCE(dd.msn, '') as msn,
            COALESCE(dd.warehouse_name, '') as warehouse_name,
            COALESCE(pl.code, '') as pl_code
        FROM inventory i
        LEFT JOIN document_bl db ON db.lot_no = i.lot_no
        LEFT JOIN document_do dd ON dd.lot_no = i.lot_no
        LEFT JOIN document_pl pl ON pl.lot_no = i.lot_no
        WHERE i.bl_no NOT LIKE 'TEST%'
        ORDER BY i.bl_no, i.lot_no
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()
    print(f"  ✅ DB에서 {len(rows)}건 조회 완료")
    return rows


def build_row(lot: dict, seq: int, sheet: str) -> tuple:
    """
    Excel 1행 tuple 생성 (25컬럼, 0-indexed).
    sheet: 'IN' | 'UNSOLD'
    """
    prod_code = derive_product_code(lot.get("pl_code") or lot.get("product_code") or "")
    carrier   = derive_carrier(lot.get("bl_no", ""), lot.get("carrier_id", ""))
    terminal  = derive_terminal(lot.get("bl_pod", ""), lot.get("warehouse_name", ""))

    # 화물관리번호: mrn + '-' + msn
    mrn = lot.get("mrn", "")
    msn = lot.get("msn", "")
    cargo_mgmt_no = f"{mrn}-{msn}" if mrn and msn else (mrn or None)

    # Salar Invoice no. → int if numeric
    sal_inv = lot.get("salar_invoice_no", "")
    sal_inv_val = safe_int(sal_inv) if sal_inv else None

    # LOT NO → int
    lot_no_int = safe_int(lot.get("lot_no"))

    # SAP NO → int
    sap_int = safe_int(lot.get("sap_no"))

    # 포장갯수 (MT 단위)
    net_mt  = round(float(lot.get("net_weight") or 0) / 1000, 4) or None
    gw_mt   = round(float(lot.get("gross_weight") or 0) / 1000, 5) or None
    plt_cnt = safe_int(lot.get("mxbg_pallet"))

    # 내품
    inner_product = "탄산리튬"  # MIC9000/CRY9000 모두 탄산리튬

    # 총중량
    total_kg = float(lot.get("gross_weight") or 0) or None

    # col16 (index 15): IN=Salar Invoice no.#2 / UNSOLD=SALE REF
    col16_val = lot.get("sale_ref") if sheet == "UNSOLD" else None

    row = (
        seq,                            # 0: 순번
        to_date(lot.get("arrival_date")), # 1: 입고일
        None,                           # 2: 출고일
        prod_code,                      # 3: 품명
        sap_int,                        # 4: SAP NO
        lot.get("bl_no"),               # 5: M B/L
        cargo_mgmt_no,                  # 6: 화물관리번호
        lot.get("container_no"),        # 7: Cont's NO
        lot_no_int,                     # 8: LOT NO
        None,                           # 9: 출고갯수 NET
        None,                           # 10: 출고갯수 GW
        None,                           # 11: 출고갯수 PLT
        sal_inv_val,                    # 12: Salar Invoice no.
        lot.get("location"),            # 13: 위치
        carrier,                        # 14: 운송사
        col16_val,                      # 15: SALE REF / Salar Invoice no.#2
        net_mt,                         # 16: 포장갯수 NET (MT)
        gw_mt,                          # 17: 포장갯수 GW (MT)
        plt_cnt,                        # 18: 포장갯수 PLT
        lot.get("remarks"),             # 19: REMARK
        inner_product,                  # 20: 내품
        None,                           # 21: 양하일 (미수집)
        terminal,                       # 22: 양하부두
        40,                             # 23: size
        total_kg,                       # 24: 총중량 (KG)
    )
    return row


def task3_fill_excel(lots: list[dict]):
    print("=" * 50)
    print("Task #3: Excel 신규 행 추가")

    # 출력 파일 = 기존 파일 복사본
    shutil.copy2(str(EXCEL_SRC), str(EXCEL_OUT))
    wb = openpyxl.load_workbook(str(EXCEL_OUT))

    for sheet_name in ["IN", "UNSOLD"]:
        ws = wb[sheet_name]
        next_row = ws.max_row + 1

        # seq 시작값: 현재 시트의 마지막 seq + 1
        # (같은 BL 그룹 내에서 seq가 restart되는 패턴이지만 새 배치는 1부터)
        # 배치 구조: BL별 그룹 → 각 그룹 내 seq 1..n
        added = 0
        current_bl = None
        seq_in_group = 0

        for lot in lots:
            bl = lot.get("bl_no", "")
            if bl != current_bl:
                current_bl = bl
                seq_in_group = 0

            seq_in_group += 1
            row_data = build_row(lot, seq_in_group, sheet_name)
            ws.append(row_data)
            added += 1

        print(f"  ✅ {sheet_name}: {added}행 추가 (시작 row={next_row})")

    wb.save(str(EXCEL_OUT))
    print(f"\n  📁 저장 완료: {EXCEL_OUT.name}")


# ──────────────────────────────────────────────
# 검증
# ──────────────────────────────────────────────
def task4_verify():
    print("=" * 50)
    print("Task #4: 결과 검증")
    wb = openpyxl.load_workbook(str(EXCEL_OUT))

    for sheet_name in ["IN", "UNSOLD"]:
        ws = wb[sheet_name]
        # 마지막 5행 샘플
        print(f"\n  [{sheet_name}] 마지막 5행:")
        rows = list(ws.iter_rows(min_row=ws.max_row - 4, max_row=ws.max_row, values_only=True))
        for r in rows:
            # 핵심 컬럼만: 순번, 입고일, 품명, BL, LOT, 운송사, 양하부두, size, 총중량
            print(f"    seq={r[0]}, 입고일={r[1]}, 품명={r[3]}, BL={r[5]}, LOT={r[8]}, "
                  f"운송사={r[14]}, 부두={r[22]}, size={r[23]}, kg={r[24]}")

    print("\n  ✅ 검증 완료")


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n🚀 SQM Excel 채우기 스크립트 시작 ({datetime.now():%Y-%m-%d %H:%M:%S})")
    print(f"   DB: {DB_PATH}")
    print(f"   Excel 입력: {EXCEL_SRC}")
    print(f"   Excel 출력: {EXCEL_OUT}\n")

    task2_update_size_type()
    lots = fetch_lot_data()
    task3_fill_excel(lots)
    task4_verify()

    print("\n🏁 완료!")
