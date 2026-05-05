"""
fill_excel_v2.py
=================
Excel 보완 스크립트 v2
- Task A: 기존 40행 UPDATE — 화물관리번호(col7) + 양하일(col22)
          Maersk 20 LOT: 26MAEUK071I-1020 / 입항 2026-04-10
          Hapag  20 LOT: 26HLCU9401I-6006 / 입항 2026-04-11
- Task B: ONE 배치 신규 20행 추가 (IN + UNSOLD)
          BL=ONEYSCLG01825300, SAP=2200034590
          MRN/MSN=26HDMUK026I-5019 / 입항 2026-04-11 / 양하부두=KIT

입력: SMQ 입,출고 재고관리 파일_v866.xlsx   (Task #3 산출물)
출력: SMQ 입,출고 재고관리 파일_v867.xlsx

작성: Ruby (2026-05-03)
"""

import shutil
from datetime import date
from pathlib import Path
import openpyxl

BASE      = Path(__file__).parent.parent
EXCEL_IN  = BASE / "SMQ 입,출고 재고관리 파일_v866.xlsx"
EXCEL_OUT = BASE / "SMQ 입,출고 재고관리 파일_v867.xlsx"

# ──────────────────────────────────────────────
# BL별 보완 데이터
# ──────────────────────────────────────────────
BL_META = {
    "MAEU265083673":   {"cargo_mgmt": "26MAEUK071I-1020", "arrival": date(2026, 4, 10)},
    "HLCUSCL260148627":{"cargo_mgmt": "26HLCU9401I-6006", "arrival": date(2026, 4, 11)},
    "ONEYSCLG01825300":{"cargo_mgmt": "26HDMUK026I-5019", "arrival": date(2026, 4, 11)},
}

# ──────────────────────────────────────────────
# ONE 배치 20 LOT — PL/DO/BL/FA에서 추출
# (container, lot_no, lot_sqm) 순서 = PL seq 1..20
# ──────────────────────────────────────────────
ONE_LOTS = [
    # container      lot_no         lot_sqm
    ("TCLU6404353", "1126012309",  "1019217"),
    ("TCLU6404353", "1126012310",  "1019218"),
    ("TCLU6404353", "1126012311",  "1019219"),
    ("TCLU6404353", "1126012437",  "1019220"),
    ("TGCU5305383", "1126012433",  "1019221"),
    ("TGCU5305383", "1126012434",  "1019222"),
    ("TGCU5305383", "1126012435",  "1019223"),
    ("TGCU5305383", "1126012436",  "1019224"),
    ("FDCU0445465", "1126012404",  "1019226"),
    ("FDCU0445465", "1126012405",  "1019227"),
    ("FDCU0445465", "1126012410",  "1019228"),
    ("FDCU0445465", "1126012411",  "1019229"),
    ("NYKU4915247", "1126012412",  "1019230"),
    ("NYKU4915247", "1126012413",  "1019231"),
    ("NYKU4915247", "1126012416",  "1019232"),
    ("NYKU4915247", "1126012424",  "1019233"),
    ("TGBU4681720", "1126012426",  "1019234"),
    ("TGBU4681720", "1126012428",  "1019235"),
    ("TGBU4681720", "1126012429",  "1019236"),
    ("TGBU4681720", "1126012430",  "1019237"),
]

# ONE 공통 필드
ONE_BL         = "ONEYSCLG01825300"
ONE_SAP        = 2200034590          # int
ONE_CARGO_MGMT = "26HDMUK026I-5019"
ONE_ARRIVAL    = date(2026, 4, 11)
ONE_PRODUCT    = "MIC9000"
ONE_CARRIER    = "ONE"
ONE_TERMINAL   = "KIT"
ONE_NET_MT     = round(5001.0 / 1000, 4)   # 5.001
ONE_GW_MT      = round(5131.25 / 1000, 5)  # 5.13125
ONE_PLT        = 10
ONE_SAL_INV    = 17653               # int (Factura No.17653)
ONE_INNER      = "탄산리튬"
ONE_SIZE       = 40
ONE_TOTAL_KG   = 5131.25            # GW KG


# ──────────────────────────────────────────────
# Task A: 기존 행 UPDATE
# col6(F)=M B/L, col7(G)=화물관리번호, col22(V)=양하일
# ──────────────────────────────────────────────
def task_a_update_existing(ws):
    """BL 기준으로 화물관리번호(col7) + 양하일(col22) 채우기."""
    updated = 0
    for row in ws.iter_rows(min_row=4):
        bl_cell  = row[5]   # col6 (0-indexed=5) = M B/L
        cgn_cell = row[6]   # col7 = 화물관리번호
        arr_cell = row[21]  # col22 = 양하일

        bl_val = bl_cell.value
        if not bl_val or bl_val not in BL_META:
            continue

        meta = BL_META[bl_val]
        changed = False

        if not cgn_cell.value:
            cgn_cell.value = meta["cargo_mgmt"]
            changed = True

        if not arr_cell.value:
            arr_cell.value = meta["arrival"]
            changed = True

        if changed:
            updated += 1

    return updated


# ──────────────────────────────────────────────
# Task B: ONE 배치 신규 행 추가
# ──────────────────────────────────────────────
def build_one_row(seq: int, container: str, lot_no: str, sheet: str) -> tuple:
    """
    25컬럼 tuple (0-indexed 0..24) — 기존 fill_excel_from_db.py 레이아웃과 동일
    sheet: 'IN' | 'UNSOLD'
    UNSOLD col16(index 15) = SALE REF (None for new stock)
    """
    col16_val = None  # UNSOLD SALE REF: 미판매 → 비워 둠

    return (
        seq,            # 0: 순번
        ONE_ARRIVAL,    # 1: 입고일
        None,           # 2: 출고일
        ONE_PRODUCT,    # 3: 품명
        ONE_SAP,        # 4: SAP NO
        ONE_BL,         # 5: M B/L
        ONE_CARGO_MGMT, # 6: 화물관리번호
        container,      # 7: Cont's NO
        int(lot_no),    # 8: LOT NO
        None,           # 9: 출고갯수 NET
        None,           # 10: 출고갯수 GW
        None,           # 11: 출고갯수 PLT
        ONE_SAL_INV,    # 12: Salar Invoice no.
        None,           # 13: 위치 (미배정)
        ONE_CARRIER,    # 14: 운송사
        col16_val,      # 15: SALE REF / Salar Invoice no.#2
        ONE_NET_MT,     # 16: 포장갯수 NET (MT)
        ONE_GW_MT,      # 17: 포장갯수 GW (MT)
        ONE_PLT,        # 18: 포장갯수 PLT
        None,           # 19: REMARK
        ONE_INNER,      # 20: 내품
        ONE_ARRIVAL,    # 21: 양하일
        ONE_TERMINAL,   # 22: 양하부두
        ONE_SIZE,       # 23: size
        ONE_TOTAL_KG,   # 24: 총중량 (KG)
    )


def task_b_add_one_rows(ws, sheet_name: str):
    """ONE 배치 20행 추가."""
    next_row = ws.max_row + 1
    for seq, (container, lot_no, _) in enumerate(ONE_LOTS, start=1):
        row_data = build_one_row(seq, container, lot_no, sheet_name)
        ws.append(row_data)
    return len(ONE_LOTS)


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def main():
    from datetime import datetime
    print(f"\n🚀 fill_excel_v2.py 시작 ({datetime.now():%Y-%m-%d %H:%M:%S})")
    print(f"   입력: {EXCEL_IN.name}")
    print(f"   출력: {EXCEL_OUT.name}\n")

    if not EXCEL_IN.exists():
        print(f"❌ 입력 파일 없음: {EXCEL_IN}")
        return

    shutil.copy2(str(EXCEL_IN), str(EXCEL_OUT))
    wb = openpyxl.load_workbook(str(EXCEL_OUT))

    for sheet_name in ["IN", "UNSOLD"]:
        ws = wb[sheet_name]
        before_rows = ws.max_row
        print(f"{'='*50}")
        print(f"[{sheet_name}] 처리 시작 (현재 {before_rows}행)")

        # Task A
        updated = task_a_update_existing(ws)
        print(f"  ✅ Task A: {updated}행 UPDATE (화물관리번호+양하일)")

        # Task B
        added = task_b_add_one_rows(ws, sheet_name)
        print(f"  ✅ Task B: ONE 배치 {added}행 추가 (row {before_rows+1}~{ws.max_row})")

    wb.save(str(EXCEL_OUT))
    print(f"\n  📁 저장 완료: {EXCEL_OUT.name}")

    # ── 검증 ──
    print(f"\n{'='*50}")
    print("검증")
    wb2 = openpyxl.load_workbook(str(EXCEL_OUT))
    for sheet_name in ["IN", "UNSOLD"]:
        ws2 = wb2[sheet_name]
        print(f"\n  [{sheet_name}] 총 {ws2.max_row}행")

        # 화물관리번호 채워진 행 수
        filled_cgn = sum(
            1 for row in ws2.iter_rows(min_row=4, values_only=True)
            if row[6]  # col7 화물관리번호
        )
        # 양하일 채워진 행 수
        filled_arr = sum(
            1 for row in ws2.iter_rows(min_row=4, values_only=True)
            if row[21]  # col22 양하일
        )
        print(f"    화물관리번호 채움: {filled_cgn}행")
        print(f"    양하일 채움:       {filled_arr}행")

        # 마지막 5행 샘플
        print(f"    마지막 5행 샘플:")
        total = ws2.max_row
        for r in range(max(4, total-4), total+1):
            vals = [ws2.cell(row=r, column=c).value for c in range(1, 26)]
            print(f"      row{r}: seq={vals[0]}, BL={vals[5]}, "
                  f"화물관리번호={vals[6]}, LOT={vals[8]}, "
                  f"운송사={vals[14]}, 양하일={vals[21]}, "
                  f"부두={vals[22]}, size={vals[23]}")

    print("\n  ✅ 검증 완료")
    print("\n🏁 완료!")


if __name__ == "__main__":
    main()
