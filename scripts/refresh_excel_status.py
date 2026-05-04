"""
refresh_excel_status.py — SMQ 입,출고 재고관리 파일_int.xlsx STATUS 컬럼 갱신 유틸

사용법:
    python scripts/refresh_excel_status.py
    python scripts/refresh_excel_status.py --db data/db/sqm_inventory.db
    python scripts/refresh_excel_status.py --excel "SMQ 입,출고 재고관리 파일_int.xlsx"

동작:
    1. DB allocation_plan 에서 RESERVED/CONFIRMED LOT 목록 조회
    2. DB sold_table 또는 inventory 에서 출고완료(SOLD) LOT 확인
    3. _int.xlsx INVENTORY 시트 STATUS 컬럼 갱신
       - RESERVED (allocation_plan 잡힌 LOT)  → 노란색
       - SOLD     (출고완료 LOT)               → 빨간색
       - AVAILABLE (나머지)                    → 초록색

작성일: 2026-05-03
작성자: Ruby (Senior Software Architect)
"""

from __future__ import annotations

import argparse
import logging
import sqlite3
import sys
from pathlib import Path

logger = logging.getLogger(__name__)


# ─── 컬럼 인덱스 (1-based, openpyxl) ───────────────────────────────────────
LOT_COL    = 9   # LOT NO
STATUS_COL = 26  # STATUS
HEADER_ROW = 3
DATA_START  = 4

# ─── 색상 ───────────────────────────────────────────────────────────────────
COLOR = {
    "AVAILABLE": {"fill": "C6EFCE", "font": "006100"},
    "SOLD":      {"fill": "FFCCCC", "font": "9C0006"},
    "RESERVED":  {"fill": "FFEB9C", "font": "7D4F00"},
}


def _get_db_lots(db_path: str) -> tuple[set[str], set[str]]:
    """DB에서 RESERVED·SOLD LOT 집합 반환.

    Returns:
        (reserved_lots, sold_lots)
    """
    reserved_lots: set[str] = set()
    sold_lots: set[str]     = set()

    try:
        conn = sqlite3.connect(db_path)
        cur  = conn.cursor()

        # RESERVED: allocation_plan 에서 미실행 건
        cur.execute("""
            SELECT DISTINCT lot_no FROM allocation_plan
            WHERE status IN ('RESERVED', 'CONFIRMED', 'PENDING')
              AND lot_no IS NOT NULL
        """)
        reserved_lots = {str(r[0]).strip() for r in cur.fetchall()}

        # SOLD: inventory 테이블 status='SOLD' 또는 sold_table
        tables = {r[0] for r in cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        )}

        if "sold_table" in tables:
            cur.execute("SELECT DISTINCT lot_no FROM sold_table WHERE lot_no IS NOT NULL")
            sold_lots = {str(r[0]).strip() for r in cur.fetchall()}
        elif "inventory" in tables:
            cur.execute("""
                SELECT DISTINCT lot_no FROM inventory
                WHERE status = 'SOLD' AND lot_no IS NOT NULL
            """)
            sold_lots = {str(r[0]).strip() for r in cur.fetchall()}

        conn.close()
        logger.info("DB 조회 완료 — RESERVED: %d, SOLD: %d", len(reserved_lots), len(sold_lots))

    except sqlite3.OperationalError as exc:
        logger.warning("DB 조회 오류 (WAL 잠금일 수 있음): %s — /tmp 복사본으로 재시도", exc)
        # WAL 잠금 우회: 임시 복사 후 읽기
        import shutil, tempfile
        with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            shutil.copy2(db_path, tmp_path)
            conn2 = sqlite3.connect(tmp_path)
            cur2  = conn2.cursor()
            cur2.execute("""
                SELECT DISTINCT lot_no FROM allocation_plan
                WHERE status IN ('RESERVED', 'CONFIRMED', 'PENDING')
                  AND lot_no IS NOT NULL
            """)
            reserved_lots = {str(r[0]).strip() for r in cur2.fetchall()}
            tables2 = {r[0] for r in cur2.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )}
            if "sold_table" in tables2:
                cur2.execute("SELECT DISTINCT lot_no FROM sold_table WHERE lot_no IS NOT NULL")
                sold_lots = {str(r[0]).strip() for r in cur2.fetchall()}
            conn2.close()
            logger.info("복사본 DB 조회 완료 — RESERVED: %d, SOLD: %d",
                        len(reserved_lots), len(sold_lots))
        except Exception as exc2:
            logger.error("복사본 DB도 실패: %s", exc2)
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    return reserved_lots, sold_lots


def refresh(excel_path: str, db_path: str, dry_run: bool = False) -> dict:
    """_int.xlsx INVENTORY 시트 STATUS 갱신.

    Returns:
        {"updated": int, "reserved": int, "sold": int, "available": int}
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        logger.error("openpyxl 미설치 — pip install openpyxl")
        sys.exit(1)

    xl = Path(excel_path)
    if not xl.exists():
        logger.error("Excel 파일 없음: %s", xl)
        sys.exit(1)

    reserved_lots, sold_lots = _get_db_lots(db_path)

    wb = load_workbook(str(xl))
    if "INVENTORY" not in wb.sheetnames:
        logger.error("INVENTORY 시트 없음 — 통합본(_int.xlsx)인지 확인 요망")
        sys.exit(1)
    ws = wb["INVENTORY"]

    stats = {"updated": 0, "reserved": 0, "sold": 0, "available": 0}

    for row_idx in range(DATA_START, ws.max_row + 1):
        lot_val = ws.cell(row_idx, LOT_COL).value
        if lot_val is None:
            continue
        lot_str = str(lot_val).strip()

        # 상태 결정 (RESERVED 우선 > SOLD > AVAILABLE)
        if lot_str in reserved_lots:
            new_status = "RESERVED"
            stats["reserved"] += 1
        elif lot_str in sold_lots:
            new_status = "SOLD"
            stats["sold"] += 1
        else:
            new_status = "AVAILABLE"
            stats["available"] += 1

        old_status = ws.cell(row_idx, STATUS_COL).value
        if old_status == new_status:
            continue  # 변경 없으면 스킵

        if not dry_run:
            cfg      = COLOR[new_status]
            row_fill = PatternFill("solid", fgColor=cfg["fill"])
            row_font = Font(name="맑은 고딕", size=9)

            # 행 전체 색상 변경
            for c in range(1, STATUS_COL + 1):
                cell = ws.cell(row_idx, c)
                cell.fill = row_fill
                if c < STATUS_COL:
                    cell.font = row_font

            # STATUS 셀
            sc = ws.cell(row_idx, STATUS_COL)
            sc.value = new_status
            sc.font  = Font(name="맑은 고딕", size=9, bold=True, color=cfg["font"])
            sc.fill  = row_fill

        stats["updated"] += 1
        logger.debug("row %d LOT=%s: %s → %s", row_idx, lot_str, old_status, new_status)

    if not dry_run and stats["updated"] > 0:
        wb.save(str(xl))
        logger.info("저장 완료: %s", xl.name)
    elif stats["updated"] == 0:
        logger.info("변경 없음 — 저장 스킵")

    return stats


def _find_db(base_dir: Path) -> str:
    """DB 경로 자동 탐색."""
    candidates = [
        base_dir / "data" / "db" / "sqm_inventory.db",
        base_dir / "sqm_inventory.db",
    ]
    for c in candidates:
        if c.exists():
            return str(c)
    raise FileNotFoundError(f"sqm_inventory.db 를 찾을 수 없습니다: {base_dir}")


def _find_excel(base_dir: Path) -> str:
    """Excel 경로 자동 탐색 (_int.xlsx 우선)."""
    int_file = base_dir / "SMQ 입,출고 재고관리 파일_int.xlsx"
    if int_file.exists():
        return str(int_file)
    versioned = sorted(base_dir.glob("SMQ 입,출고 재고관리 파일_v*.xlsx"))
    if versioned:
        return str(versioned[-1])
    raise FileNotFoundError("SMQ 입,출고 재고관리 파일_int.xlsx 를 찾을 수 없습니다.")


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(description="Excel STATUS 컬럼 DB 기준 갱신")
    parser.add_argument("--db",    default=None, help="DB 경로 (기본: 자동탐색)")
    parser.add_argument("--excel", default=None, help="Excel 경로 (기본: _int.xlsx 자동탐색)")
    parser.add_argument("--dry-run", action="store_true", help="실제 저장 없이 변경 내역만 출력")
    args = parser.parse_args()

    base = Path(__file__).resolve().parent.parent  # scripts/ 의 상위 = 프로젝트 루트

    db_path    = args.db    or _find_db(base)
    excel_path = args.excel or _find_excel(base)

    logger.info("DB   : %s", db_path)
    logger.info("Excel: %s", excel_path)
    if args.dry_run:
        logger.info("DRY-RUN 모드 — 저장하지 않음")

    stats = refresh(excel_path, db_path, dry_run=args.dry_run)

    print("\n" + "=" * 50)
    print("  STATUS 갱신 결과")
    print("=" * 50)
    print(f"  변경된 행  : {stats['updated']}")
    print(f"  RESERVED   : {stats['reserved']}")
    print(f"  SOLD       : {stats['sold']}")
    print(f"  AVAILABLE  : {stats['available']}")
    print("=" * 50)


if __name__ == "__main__":
    main()
