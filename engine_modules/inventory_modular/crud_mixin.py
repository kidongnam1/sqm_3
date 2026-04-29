# ============================================================
# SQM 재고 계산 원칙 (v8.0.5 확정 — 2026-03-16)
# ------------------------------------------------------------
# 모든 current_weight / picked_weight 변경은
# 상태 변경 후 _recalc_current_weight()로 확정한다.
#
# current_weight = AVAILABLE + RESERVED 일반 tonbag 합 (sample 제외)
# picked_weight  = PICKED 일반 tonbag 합 (sample 제외)
# OUTBOUND       = 계산 제외
#
# Central Truth: _recalc_current_weight(lot_no, reason='...')
# Bulk Helper  : _recalc_current_weight_many(lot_nos, reason='...')
# ============================================================

"""
SQM Inventory Engine - CRUD Mixin
=================================

v2.9.91 - Extracted from inventory.py

CRUD operations and search functions
"""

from engine_modules.constants import STATUS_AVAILABLE
import logging
import sqlite3
from datetime import date, datetime
from typing import Dict

from core.constants import SAMPLE_WEIGHT_KG
from engine_modules.tonbag_weight_rules import calculate_tonbag_weight

logger = logging.getLogger(__name__)


class CRUDMixin:
    """
    CRUD and search mixin
    
    Methods for inventory CRUD operations and LOT search
    """

    def _recalc_current_weight(self, lot_no: str, reason: str = '') -> float:
        """v8.0.0 [P2]: current_weight 중앙 재계산 함수.

        inventory_tonbag의 실제 데이터 기준으로 inventory.current_weight를
        항상 정확하게 재계산. 모든 갱신 경로에서 이 함수를 호출해야 함.

        AVAILABLE + RESERVED 상태 톤백 합계 = current_weight
        PICKED 이상 상태 = picked_weight 반영

        Args:
            lot_no:  LOT 번호
            reason:  갱신 이유 (로그용)

        Returns:
            갱신된 current_weight (kg)
        """
        lot_no = str(lot_no or '').strip()
        if not lot_no:
            return 0.0
        try:
            now = self._now() if hasattr(self, '_now') else __import__('datetime').datetime.now().isoformat()

            # AVAILABLE + RESERVED 합계 → current_weight
            # [BUG-FIX #4-RETURN-RECALC 2026-04-29]
            # RETURN 상태 톤백은 창고 내 실재 재고 (반품 대기 중)
            # verify_lot_integrity도 RETURN을 current에 포함하므로 일치시킴
            row_avail = self.db.fetchone(
                "SELECT COALESCE(SUM(weight), 0) AS s "
                "FROM inventory_tonbag "
                "WHERE lot_no = ? AND status IN ('AVAILABLE','RESERVED','RETURN') "
                "  AND (is_sample IS NULL OR is_sample = 0)",
                (lot_no,)
            )
            avail_kg = float(row_avail.get('s', 0) if isinstance(row_avail, dict) else (row_avail[0] or 0))

            # PICKED 합계 → picked_weight
            row_picked = self.db.fetchone(
                "SELECT COALESCE(SUM(weight), 0) AS s "
                "FROM inventory_tonbag "
                "WHERE lot_no = ? AND status = 'PICKED' "
                "  AND (is_sample IS NULL OR is_sample = 0)",
                (lot_no,)
            )
            picked_kg = float(row_picked.get('s', 0) if isinstance(row_picked, dict) else (row_picked[0] or 0))

            # v8.0.5 [RECALC-LOG]: old 값 조회 후 로그
            _old_row = self.db.fetchone(
                "SELECT current_weight, picked_weight FROM inventory WHERE lot_no = ?",
                (lot_no,)
            )
            _old_cw = float(_old_row.get('current_weight', 0) if isinstance(_old_row, dict)
                           else (_old_row[0] if _old_row else 0))
            _old_pw = float(_old_row.get('picked_weight',  0) if isinstance(_old_row, dict)
                           else (_old_row[1] if _old_row else 0))

            self.db.execute(
                "UPDATE inventory "
                "SET current_weight = ?, picked_weight = ?, updated_at = ? "
                "WHERE lot_no = ?",
                (avail_kg, picked_kg, now, lot_no)
            )

            label = f" [{reason}]" if reason else ""
            # old→new 변화가 있을 때만 INFO, 동일하면 DEBUG
            _changed = abs(_old_cw - avail_kg) > 0.01 or abs(_old_pw - picked_kg) > 0.01
            _log = logger.info if _changed else logger.debug
            _log(
                f"[P2][RECALC]{label} LOT={lot_no} "
                f"CW: {_old_cw:.0f}→{avail_kg:.0f}kg, "
                f"PW: {_old_pw:.0f}→{picked_kg:.0f}kg"
            )
            return avail_kg

        except Exception as e:
            logger.warning(f"[P2][_recalc_current_weight] LOT={lot_no} 오류: {e}")
            return 0.0


    def add_inventory(self, lot_no: str, sap_no: str = None, bl_no: str = None,
                      container_no: str = None, product: str = None,
                      product_code: str = None, mxbg_pallet: int = 20,
                      net_weight: float = 10000, warehouse: str = 'GY',
                      arrival_date=None, stock_date=None, **kwargs) -> Dict:
        """
        Add single LOT inventory (v3.8.7: 18열 전체 지원)
        
        Args:
            lot_no: LOT number (required)
            sap_no: SAP number
            bl_no: B/L number
            container_no: Container number
            product: Product name
            product_code: Product code
            mxbg_pallet: Number of tonbags (default 20)
            net_weight: Total weight in kg (default 10000)
            warehouse: Warehouse code (default 'GY')
            arrival_date: Arrival date
            stock_date: Stock date
            **kwargs: lot_sqm, gross_weight, salar_invoice_no, ship_date, free_time 등
        
        Returns:
            Result dict with success, lot_no, tonbags_created
        """
        try:
            # v3.8.7: kwargs에서 추가 필드 추출
            lot_sqm = kwargs.get('lot_sqm', '')
            folio   = kwargs.get('folio', '')
            vessel  = kwargs.get('vessel', '')
            gross_weight = kwargs.get('gross_weight', net_weight)
            salar_invoice_no = kwargs.get('salar_invoice_no', '') or kwargs.get('invoice_no', '')
            ship_date = kwargs.get('ship_date', '')
            free_time = kwargs.get('free_time', 0)

            # Date handling — arrival_date 미상 시 비움(date.today() 사용 금지)
            if arrival_date is None:
                arrival_date = ''
            if stock_date is None:
                stock_date = date.today()
            if hasattr(arrival_date, 'isoformat'):
                arrival_date = arrival_date.isoformat()
            if hasattr(stock_date, 'isoformat'):
                stock_date = stock_date.isoformat()

            # Check duplicate
            existing = self.db.fetchone(
                "SELECT id FROM inventory WHERE lot_no = ?", (lot_no,)
            )

            if existing:
                return {
                    'success': False,
                    'lot_no': lot_no,
                    'tonbags_created': 0,
                    'message': f'LOT already exists: {lot_no}'
                }

            # Calculate weight per bag (v5.6.0 대원칙: 샘플 1kg 제외 후 균등 분배)
            # LOT 총무게 = (톤백수 × 단가) + 샘플 1kg
            # → 톤백 단가 = (총무게 - SAMPLE_WEIGHT_KG) / 톤백수
            weight_per_bag = calculate_tonbag_weight(net_weight, mxbg_pallet, SAMPLE_WEIGHT_KG) if mxbg_pallet > 0 else 500

            # v8.7.1 [ROOT-CAUSE-FIX]: current_weight는 샘플 제외가 원칙 (crud_mixin.py:7 선언)
            # ──────────────────────────────────────────────────────────────
            # 과거: current_weight = net_weight  (예: 5001, 샘플 포함 → 설계 위반)
            #      → validators.py:345가 매 부팅마다 5001→5000 자동 교정 (데이터 손실)
            # 현재: current_weight = net_weight - SAMPLE_WEIGHT_KG (예: 5000, 샘플 제외)
            #      → _recalc_current_weight() 출력과 일치, validator 개입 불필요
            # initial_weight는 "입고 당시 총량"이므로 net_weight 유지 (샘플 포함).
            # ──────────────────────────────────────────────────────────────
            _current_weight_init = (float(net_weight) - float(SAMPLE_WEIGHT_KG)) if mxbg_pallet > 0 else float(net_weight)

            with self.db.transaction("IMMEDIATE"):
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                # Insert inventory (v3.8.7: 18열 전체, v5.8.8: con_return 추가)
                con_return = kwargs.get('con_return', '')
                self.db.execute("""
                    INSERT INTO inventory (
                        lot_no, sap_no, bl_no, container_no, product, product_code,
                        lot_sqm, folio, vessel, mxbg_pallet, net_weight, gross_weight,
                        current_weight, initial_weight, picked_weight,
                        salar_invoice_no, ship_date, arrival_date, con_return, free_time,
                        warehouse, stock_date, status, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, 'AVAILABLE', ?)
                """, (lot_no, sap_no, bl_no, container_no, product, product_code,
                      lot_sqm, folio, vessel, mxbg_pallet, net_weight, gross_weight,
                      _current_weight_init, net_weight,  # v8.7.1: current_weight = net - sample
                      salar_invoice_no, ship_date, arrival_date, con_return, free_time,
                      warehouse, stock_date, now))

                # P3: DB 독립적 ID 조회 (SQLite: lastrowid, PG: RETURNING)
                if hasattr(self.db, 'insert_returning_id'):
                    inv_id = self.db.insert_returning_id("""
                        SELECT id FROM inventory WHERE lot_no = ?
                    """, (lot_no,))
                else:
                    inv_row = self.db.fetchone("SELECT id FROM inventory WHERE lot_no = ?", (lot_no,))
                    inv_id = inv_row['id'] if inv_row else None

                # v7.7.0: 톤백 N개 개별 INSERT → executemany (N+1 → 1회 처리)
                _tonbag_rows = [
                    (inv_id, lot_no, sub, weight_per_bag, now)
                    for sub in range(1, mxbg_pallet + 1)
                ]
                self.db.executemany("""
                    INSERT INTO inventory_tonbag (
                        inventory_id, lot_no, sub_lt, weight, status,
                        is_sample, created_at
                    ) VALUES (?, ?, ?, ?, 'AVAILABLE', 0, ?)
                """, _tonbag_rows)

                # v3.9.1: 샘플 톤백 자동 생성 (sub_lt=0, SAMPLE_WEIGHT_KG, is_sample=1)
                self.db.execute("""
                    INSERT INTO inventory_tonbag (
                        inventory_id, lot_no, sub_lt, weight, status,
                        is_sample, created_at
                    ) VALUES (?, ?, 0, ?, 'AVAILABLE', 1, ?)
                """, (inv_id, lot_no, SAMPLE_WEIGHT_KG, now))

                logger.info(f"[add_inventory] 샘플 톤백 생성: {lot_no}/0 ({SAMPLE_WEIGHT_KG}kg)")

                # Movement history
                self.db.execute("""
                    INSERT INTO stock_movement (
                        movement_type, lot_no, qty_kg, created_at
                    ) VALUES ('INBOUND', ?, ?, ?)
                """, (lot_no, net_weight, now))

            logger.info(f"[add_inventory] Success: {lot_no}, {mxbg_pallet} tonbags + 1 sample")

            return {
                'success': True,
                'lot_no': lot_no,
                'tonbags_created': mxbg_pallet + 1,  # v3.9.1: 일반 + 샘플
                'sample_created': True,
                'message': 'OK'
            }

        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            logger.error(f"[add_inventory] Error: {e}")
            return {
                'success': False,
                'lot_no': lot_no,
                'tonbags_created': 0,
                'message': str(e)
            }

    def add_inventory_from_dict(self, data: dict) -> Dict:
        """v5.7.3: Excel 입고용 — dict를 받아 add_inventory(**data) 호출, 반환에 tonbags 키 추가 (GUI 호환)"""
        if not isinstance(data, dict):
            return {'success': False, 'lot_no': '', 'tonbags_created': 0, 'tonbags': 0, 'message': 'data must be dict'}
        # add_inventory가 받지 않는 키 제거 (location, remark, status는 INSERT에 없음)
        allowed = {
            'lot_no', 'sap_no', 'bl_no', 'container_no', 'product', 'product_code',
            'mxbg_pallet', 'net_weight', 'gross_weight', 'warehouse', 'arrival_date', 'stock_date',
            'lot_sqm', 'folio', 'vessel', 'salar_invoice_no', 'ship_date', 'con_return', 'free_time', 'initial_weight', 'current_weight',
        }
        kwargs = {k: v for k, v in data.items() if k in allowed}
        result = self.add_inventory(**kwargs)
        if result.get('success'):
            result['tonbags'] = result.get('tonbags_created', 0)
        else:
            result['tonbags'] = 0
        return result

    def delete_inventory(self, lot_no: str, force: bool = False,
                         confirmed: bool = False) -> Dict:
        """
        Delete inventory LOT
        
        Args:
            lot_no: LOT number
            force: Force delete even if not AVAILABLE (requires confirmed=True)
            confirmed: User confirmation for deletion (required for actual deletion)
            
        Returns:
            Result dict
            
        Note:
            데이터 보호 정책에 따라 confirmed=True 없이는 삭제되지 않습니다.
        """
        result = {'success': False, 'error': None}

        # 데이터 보호: confirmed 체크
        if not confirmed:
            result['error'] = "삭제 승인 필요: confirmed=True를 전달해주세요"
            result['requires_confirmation'] = True
            logger.warning(f"[delete_inventory] Blocked - no confirmation: {lot_no}")
            return result

        # force 사용 시에도 confirmed 필수
        if force and not confirmed:
            result['error'] = "force 옵션 사용 시 confirmed=True 필수"
            return result

        try:
            # Check if LOT exists
            lot = self.db.fetchone(
                "SELECT status FROM inventory WHERE lot_no = ?", (lot_no,)
            )

            if not lot:
                result['error'] = f"LOT not found: {lot_no}"
                return result

            # Check status unless forced
            if not force and lot['status'] != STATUS_AVAILABLE:
                result['error'] = f"Cannot delete LOT with status: {lot['status']}"
                return result

            # Check if any tonbags are picked
            if not force:
                picked = self.db.fetchone("""
                    SELECT COUNT(*) as cnt FROM inventory_tonbag
                    WHERE lot_no = ? AND status != 'AVAILABLE'
                """, (lot_no,))

                if picked and picked['cnt'] > 0:
                    result['error'] = f"Cannot delete: {picked['cnt']} tonbags are not AVAILABLE"
                    return result

            with self.db.transaction("IMMEDIATE"):
                # Delete tonbags
                self.db.execute(
                    "DELETE FROM inventory_tonbag WHERE lot_no = ?", (lot_no,)
                )

                # Delete inventory
                self.db.execute(
                    "DELETE FROM inventory WHERE lot_no = ?", (lot_no,)
                )

                # Record movement
                from datetime import datetime
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.db.execute("""
                    INSERT INTO stock_movement (
                        movement_type, lot_no, qty_kg, created_at, remarks
                    ) VALUES ('DELETE', ?, 0, ?, 'Manual deletion')
                """, (lot_no, now))

            result['success'] = True
            # v8.3.0 [Phase 9]: LOT_DELETE audit_log
            try:
                from engine_modules.audit_helper import write_audit, EVT_LOT_DELETE
                write_audit(self.db, EVT_LOT_DELETE, lot_no=lot_no,
                            detail={'force': force})
            except Exception as _ae:
                logger.debug(f"[LOT_DELETE audit] 스킵: {_ae}")
            logger.info(f"[delete_inventory] Deleted: {lot_no}")

        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            result['error'] = str(e)
            logger.error(f"[delete_inventory] Error: {e}")

        return result

    # NOTE: search_lots → QueryMixin으로 이관 완료 (v3.8.4 데드코드 정리)

    def update_inventory(self, lot_no: str, confirmed: bool = False, **updates) -> Dict:
        """
        Update inventory fields
        
        Args:
            lot_no: LOT number
            confirmed: User confirmation for critical field updates
            **updates: Fields to update (e.g., product='NICKEL', warehouse='GY2')
            
        Returns:
            Result dict
            
        Note:
            중요 필드(sap_no, bl_no, net_weight 등) 수정 시 confirmed=True 필요
        """
        result = {'success': False, 'error': None}

        # Allowed fields for update
        allowed_fields = {
            'sap_no', 'bl_no', 'container_no', 'product', 'product_code',
            'warehouse', 'remark', 'condition', 'sold_to', 'sale_ref'
        }

        # 중요 필드 (confirmed 필요)
        critical_fields = {'sap_no', 'bl_no', 'net_weight', 'initial_weight'}

        # Filter valid fields
        valid_updates = {k: v for k, v in updates.items() if k in allowed_fields}

        if not valid_updates:
            result['error'] = "No valid fields to update"
            return result

        # 중요 필드 수정 시 confirmed 체크
        updating_critical = any(f in valid_updates for f in critical_fields)
        if updating_critical and not confirmed:
            result['error'] = f"중요 필드 수정은 승인 필요: {critical_fields & set(valid_updates.keys())}"
            result['requires_confirmation'] = True
            return result

        try:
            # Build update query
            set_clauses = [f"{k} = ?" for k in valid_updates.keys()]
            values = list(valid_updates.values())
            values.append(lot_no)

            query = f"""
                UPDATE inventory 
                SET {', '.join(set_clauses)}, updated_at = CURRENT_TIMESTAMP
                WHERE lot_no = ?
            """

            self.db.execute(query, tuple(values))
            result['success'] = True
            # v8.3.0 [Phase 9]: LOT_UPDATE audit_log
            try:
                from engine_modules.audit_helper import write_audit, EVT_LOT_UPDATE
                write_audit(self.db, EVT_LOT_UPDATE, lot_no=lot_no,
                            detail={'updated_fields': list(valid_updates.keys())})
            except Exception as _ae:
                logger.debug(f"[LOT_UPDATE audit] 스킵: {_ae}")
            logger.info(f"[update_inventory] Updated: {lot_no}")

        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            result['error'] = str(e)
            logger.error(f"[update_inventory] Error: {e}")

        return result

    # NOTE: get_lot_detail → QueryMixin으로 이관 완료 (v3.8.4 데드코드 정리)

    def export_lot_report(self, lot_no: str, filepath: str = None) -> Dict:
        """
        Export LOT detail report to Excel
        
        Args:
            lot_no: LOT number
            filepath: Output file path (optional)
            
        Returns:
            Result dict with filepath
        """
        import os

        from openpyxl import Workbook
        from openpyxl.styles import Border, Font, Side

        result = {'success': False, 'filepath': None, 'error': None}

        # Get LOT detail (QueryMixin 포맷 호환 — v3.8.4)
        detail = self.get_lot_detail(lot_no)
        if detail is None or detail.get('error'):
            result['error'] = detail.get('error', 'LOT not found') if detail else 'LOT not found'
            return result

        # QueryMixin 반환을 CRUDMixin export 포맷으로 래핑
        # QueryMixin: {lot_no, product, ..., tonbags: [...]}
        # export 기대: {success: True, inventory: {...}, tonbags: [...]}
        tonbags = detail.pop('tonbags', [])
        detail_wrapped = {
            'success': True,
            'inventory': detail,
            'tonbags': tonbags if isinstance(tonbags, list) else [],
        }
        detail = detail_wrapped

        # Generate filepath if not provided
        if not filepath:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = f"output/reports/LOT_{lot_no}_{timestamp}.xlsx"
            os.makedirs(os.path.dirname(filepath), exist_ok=True)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"LOT {lot_no}"

            # Styles
            header_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # LOT Info section
            inv = detail['inventory']
            info_rows = [
                ('LOT NO', inv['lot_no']),
                ('SAP NO', inv.get('sap_no', '')),
                ('BL NO', inv.get('bl_no', '')),
                ('Product', inv.get('product', '')),
                ('Warehouse', inv.get('warehouse', '')),
                ('Net Weight', inv.get('net_weight', 0)),
                ('Status', inv.get('status', '')),
            ]

            for idx, (label, value) in enumerate(info_rows, 1):
                ws.cell(row=idx, column=1, value=label).font = header_font
                ws.cell(row=idx, column=2, value=value)

            # Tonbag section
            start_row = len(info_rows) + 2
            tonbag_headers = ['Sub LT', 'Weight', 'Status', 'Location', 'Inbound', 'Outbound', 'Customer']

            for col, h in enumerate(tonbag_headers, 1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border

            for row_idx, tb in enumerate(detail['tonbags'], start_row + 1):
                data = [tb['sub_lt'], tb['weight'], tb['status'], tb.get('location', ''),
                        tb.get('inbound_date', ''), tb.get('outbound_date', ''), tb.get('picked_to', '')]
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val or '')
                    cell.border = thin_border

            wb.save(filepath)
            result['success'] = True
            result['filepath'] = filepath

        except (OSError, ValueError) as e:
            result['error'] = str(e)
            logger.error(f"[export_lot_report] Error: {e}")

        return result

