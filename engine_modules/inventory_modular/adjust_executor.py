# -*- coding: utf-8 -*-
"""
SQM 재고관리 시스템 - 재고조정 실행 모듈
==========================================

v1.0.0 (2026-05-03)

개요:
    재고 불일치 발생 시 DB inventory 테이블과 엑셀 파일을
    동시에 수정하는 재고조정 실행 모듈.

    - DB 수정: inventory.mxbg_pallet / net_weight / gross_weight / current_weight
    - 이력 기록: stock_movement (movement_type='ADJUSTMENT')
    - 엑셀 수정: INVENTORY 단일 시트 (PLT, NET MT, GW MT, STATUS 컬럼) — v868
    - 충돌 검사: RESERVED allocation_plan 합계와 new_count 비교

안전 규칙:
    - DB 수정 전 현재값 백업 -> details_json에 before/after 저장
    - 엑셀 수정 실패해도 DB 수정은 rollback 안 함 (각자 독립)
    - RESERVED/PICKED/OUTBOUND 상태 LOT도 수정 가능 (경고만, 차단 안 함)
    - 1포대당 무게: net=500 kg, gross=513.125 kg

작성자: Ruby (Senior Software Architect)
버전: v1.0.0
"""

import json
import logging
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

logger = logging.getLogger(__name__)

# -------------------------------------------------
# 비즈니스 상수
# -------------------------------------------------
NET_KG_PER_BAG: float = 500.0        # 1포대당 NET 중량 (kg)
GW_KG_PER_BAG: float  = 513.125      # 1포대당 GW 중량 (kg)  <- 5131.25 / 10

# 엑셀 컬럼 인덱스 (0-based, openpyxl row tuple 기준)
COL_LOT_NO: int = 8    # LOT NO
COL_NET_MT: int = 16   # 포장갯수 NET (MT)
COL_GW_MT: int  = 17   # 포장갯수 GW  (MT)
COL_PLT: int    = 18   # 포장갯수 PLT (포대 수)
COL_STATUS: int = 25   # STATUS 컬럼 (v868 INVENTORY 26번째 열, 0-based=25)

EXCEL_DATA_START_ROW: int  = 4              # 데이터 시작 행 (1-based, row 3 = 헤더)
EXCEL_SHEETS: List[str]    = ["INVENTORY"]      # v868 단일시트


# =====================================================================
# 결과 데이터클래스
# =====================================================================

@dataclass
class AdjustResult:
    """재고조정 실행 결과

    Attributes:
        success:  성공한 lot_no 목록
        skipped:  delta=0이라 스킵한 lot_no 목록
        failed:   실패한 lot_no 목록 (형식: "lot_no: 이유")
        log_ids:  stock_movement에 기록된 id 목록
    """
    success: List[str] = field(default_factory=list)
    skipped: List[str] = field(default_factory=list)
    failed:  List[str] = field(default_factory=list)
    log_ids: List[int] = field(default_factory=list)

    @property
    def total(self) -> int:
        return len(self.success) + len(self.skipped) + len(self.failed)

    def summary(self) -> str:
        return (
            f"성공={len(self.success)}, 스킵={len(self.skipped)}, "
            f"실패={len(self.failed)}, 총={self.total}"
        )


# =====================================================================
# 공개 API
# =====================================================================

def execute_adjustment(
    items: Sequence[Dict[str, Any]],
    db,
    excel_path: Optional[str],
    operator: str = "Nam Ki-dong",
) -> AdjustResult:
    """재고조정 메인 실행 함수.

    items의 각 항목을 순회하며 DB와 엑셀을 수정한다.
    DB 수정은 LOT마다 개별 트랜잭션 (한 LOT 실패가 다른 LOT에 영향 없음).
    엑셀 수정은 성공한 LOT 전체에 대해 파일을 한 번만 열고 저장 (성능 최적화).

    Args:
        items:      조정 항목 목록.
                    각 항목 dict 구조:
                    {
                        "lot_no":      "1125122363",   # 필수
                        "new_count":   18,             # 필수 -- 조정 후 포대 수
                        "reason_code": "STOCKTAKE",    # 선택 (기본: "ADJUSTMENT")
                        "reason_text": "실사 결과 조정"  # 선택 (비고)
                    }
        db:         SQMDatabase 인스턴스
        excel_path: 엑셀 파일 경로 (None 이면 엑셀 수정 스킵)
        operator:   작업자 이름 (stock_movement.operator)

    Returns:
        AdjustResult -- 성공/스킵/실패/log_ids 목록
    """
    result = AdjustResult()

    # 1) items 정규화 및 기초 검증
    if not items:
        logger.warning("[조정] items가 비어 있습니다.")
        return result

    normalized: List[Dict[str, Any]] = []
    for raw in items:
        lot_no = str(raw.get("lot_no") or "").strip()
        if not lot_no:
            result.failed.append("(lot_no 없음): lot_no 필드 누락")
            continue
        try:
            new_count = int(float(str(raw.get("new_count", 0)).replace(",", "").strip()))
        except (ValueError, TypeError):
            result.failed.append(f"{lot_no}: new_count 파싱 오류 -- {raw.get('new_count')!r}")
            continue
        if new_count < 0:
            result.failed.append(f"{lot_no}: new_count 음수 불가 ({new_count})")
            continue
        normalized.append({
            "lot_no":      lot_no,
            "new_count":   new_count,
            "reason_code": str(raw.get("reason_code") or "ADJUSTMENT").strip().upper() or "ADJUSTMENT",
            "reason_text": str(raw.get("reason_text") or "").strip(),
        })

    # 2) RESERVED 충돌 사전 경고 (실행 차단 안 함)
    for item in normalized:
        conflict_msg = _check_reserved_conflict(item["lot_no"], item["new_count"], db)
        if conflict_msg:
            logger.warning("[조정][RESERVED 충돌] %s", conflict_msg)

    # 3) 각 LOT DB 수정
    for item in normalized:
        lot_no    = item["lot_no"]
        new_count = item["new_count"]

        # 현재값 조회
        current_row = db.fetchone(
            "SELECT mxbg_pallet FROM inventory WHERE lot_no = ?", (lot_no,)
        )
        if current_row is None:
            result.failed.append(f"{lot_no}: inventory에 존재하지 않는 LOT")
            continue

        old_count = int(current_row.get("mxbg_pallet") or 0)

        # delta=0 이면 스킵
        if old_count == new_count:
            result.skipped.append(lot_no)
            logger.info("[조정] 스킵 (delta=0): LOT=%s, count=%d", lot_no, old_count)
            continue

        # DB 수정 + stock_movement 기록
        try:
            log_id = _update_inventory_db(
                lot_no=lot_no,
                new_count=new_count,
                reason_code=item["reason_code"],
                reason_text=item["reason_text"],
                operator=operator,
                db=db,
            )
            result.success.append(lot_no)
            result.log_ids.append(log_id)
            logger.info(
                "[조정] DB 완료: LOT=%s, %d->%d 포대, stock_movement.id=%d",
                lot_no, old_count, new_count, log_id,
            )
        except Exception as exc:
            err_msg = f"{lot_no}: DB 수정 오류 -- {exc}"
            result.failed.append(err_msg)
            logger.error("[조정] %s", err_msg, exc_info=True)

    # 4) 엑셀 수정 (성공한 LOT만, 파일 1회 열기)
    if excel_path and result.success:
        success_map: Dict[str, int] = {
            item["lot_no"]: item["new_count"]
            for item in normalized
            if item["lot_no"] in result.success
        }
        excel_updated: List[str] = []
        excel_failed:  List[str] = []
        try:
            _update_excel_batch(success_map, excel_path, excel_updated, excel_failed)
        except Exception as exc:
            logger.error("[조정] 엑셀 일괄 수정 예외 (DB는 유지): %s", exc, exc_info=True)
        if excel_updated:
            logger.info("[조정] 엑셀 수정 완료: %s", excel_updated)
        if excel_failed:
            logger.warning("[조정] 엑셀에서 행을 찾지 못한 LOT: %s", excel_failed)
    elif not excel_path:
        logger.info("[조정] excel_path 없음 -- 엑셀 수정 스킵")

    logger.info("[조정] 완료: %s", result.summary())
    return result


# =====================================================================
# 내부 함수
# =====================================================================

def _update_inventory_db(
    lot_no: str,
    new_count: int,
    reason_code: str,
    reason_text: str,
    operator: str,
    db,
) -> int:
    """inventory 테이블 수정 + stock_movement 기록.

    수정 항목:
        mxbg_pallet    = new_count
        net_weight     = new_count * NET_KG_PER_BAG       (500 kg/포대)
        gross_weight   = new_count * GW_KG_PER_BAG        (513.125 kg/포대)
        current_weight = old_current * (new_net / old_net) 비례
                         (old_net=0 이면 new_net 그대로)
        updated_at     = 현재 시각

    stock_movement 기록:
        movement_type = 'ADJUSTMENT'
        qty_kg        = new_net (조정 후 NET 중량)
        reason_code   = 사유 코드
        operator      = 작업자
        details_json  = {"before": {...}, "after": {...}, "reason_text": "..."}
        remarks       = reason_text

    Args:
        lot_no:      LOT 번호
        new_count:   조정 후 포대 수
        reason_code: 사유 코드 (예: STOCKTAKE, MANUAL, SHORTAGE)
        reason_text: 사유 텍스트 (remarks)
        operator:    작업자
        db:          SQMDatabase

    Returns:
        stock_movement.id (int)

    Raises:
        ValueError:  LOT not found
        sqlite3.*:   DB 오류 (상위 execute_adjustment에서 캐치)
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 현재값 전체 조회 (before 스냅샷)
    row = db.fetchone(
        "SELECT mxbg_pallet, net_weight, gross_weight, current_weight, status "
        "FROM inventory WHERE lot_no = ?",
        (lot_no,),
    )
    if row is None:
        raise ValueError(f"LOT not found: {lot_no}")

    old_pallet  = int(row.get("mxbg_pallet")    or 0)
    old_net     = float(row.get("net_weight")    or 0.0)
    old_gross   = float(row.get("gross_weight")  or 0.0)
    old_current = float(row.get("current_weight") or 0.0)
    status      = str(row.get("status") or "")

    # 새 중량 계산
    new_net   = round(new_count * NET_KG_PER_BAG, 3)
    new_gross = round(new_count * GW_KG_PER_BAG, 3)

    # current_weight 비례 업데이트
    if old_net > 0:
        new_current = round(old_current * (new_net / old_net), 3)
    else:
        # old_net=0 이면 데이터가 깨진 상태 -> new_net 그대로
        new_current = new_net

    before_snap = {
        "mxbg_pallet":    old_pallet,
        "net_weight":     old_net,
        "gross_weight":   old_gross,
        "current_weight": old_current,
        "status":         status,
    }
    after_snap = {
        "mxbg_pallet":    new_count,
        "net_weight":     new_net,
        "gross_weight":   new_gross,
        "current_weight": new_current,
    }
    details = json.dumps(
        {"before": before_snap, "after": after_snap, "reason_text": reason_text},
        ensure_ascii=False,
    )

    with db.transaction("IMMEDIATE"):
        # inventory 수정
        db.execute(
            """
            UPDATE inventory
            SET mxbg_pallet    = ?,
                net_weight     = ?,
                gross_weight   = ?,
                current_weight = ?,
                updated_at     = ?
            WHERE lot_no = ?
            """,
            (new_count, new_net, new_gross, new_current, now, lot_no),
        )

        # stock_movement 기록 (movement_type='ADJUSTMENT')
        cursor = db.execute(
            """
            INSERT INTO stock_movement (
                lot_no, movement_type, qty_kg,
                reason_code, operator, details_json,
                remarks, created_at
            ) VALUES (?, 'ADJUSTMENT', ?, ?, ?, ?, ?, ?)
            """,
            (
                lot_no,
                new_net,
                reason_code,
                operator,
                details,
                reason_text or None,
                now,
            ),
        )

    log_id: int = cursor.lastrowid if hasattr(cursor, "lastrowid") and cursor.lastrowid else 0

    logger.debug(
        "[_update_inventory_db] LOT=%s pallet %d->%d "
        "net %.0f->%.0f kg current %.0f->%.0f kg log_id=%d",
        lot_no, old_pallet, new_count,
        old_net, new_net,
        old_current, new_current,
        log_id,
    )
    return log_id


def _update_excel(lot_no: str, new_count: int, excel_path: str) -> bool:
    """엑셀 IN + UNSOLD 시트에서 해당 LOT 행을 찾아 수정.

    수정 컬럼 (0-based tuple index):
        col18 (PLT)    = new_count
        col16 (NET MT) = new_count * 0.5         (500 kg = 0.5 MT)
        col17 (GW MT)  = new_count * 0.513125

    단일 LOT 수정 전용. 여러 LOT 일괄 수정은 _update_excel_batch 사용.

    Args:
        lot_no:     LOT 번호
        new_count:  조정 후 포대 수
        excel_path: 엑셀 파일 경로

    Returns:
        True = 1행 이상 수정됨, False = 수정된 행 없음
    """
    updated: List[str] = []
    failed:  List[str] = []
    _update_excel_batch({lot_no: new_count}, excel_path, updated, failed)
    return lot_no in updated


def _update_excel_batch(
    lot_count_map: Dict[str, int],
    excel_path: str,
    updated_out: List[str],
    failed_out: List[str],
) -> None:
    """여러 LOT을 파일 1회 열기/저장으로 처리 (성능 최적화).

    Args:
        lot_count_map: {lot_no: new_count} 딕셔너리
        excel_path:    엑셀 파일 경로
        updated_out:   수정된 lot_no를 append할 리스트 (out-param)
        failed_out:    못 찾은 lot_no를 append할 리스트 (out-param)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("[엑셀] openpyxl 미설치")
        failed_out.extend(lot_count_map.keys())
        return

    path = Path(excel_path)
    if not path.exists():
        logger.error("[엑셀] 파일 없음: %s", excel_path)
        failed_out.extend(lot_count_map.keys())
        return

    # 찾은 LOT 추적 (시트별 중복 수정 방지: 한 시트에서 찾으면 다른 시트도 수정)
    found: Dict[str, bool] = {lot: False for lot in lot_count_map}

    try:
        wb = load_workbook(str(path))

        for sheet_name in EXCEL_SHEETS:
            if sheet_name not in wb.sheetnames:
                logger.debug("[엑셀] 시트 없음: %s (스킵)", sheet_name)
                continue
            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=EXCEL_DATA_START_ROW):
                if len(row) <= COL_PLT:
                    continue
                lot_cell = row[COL_LOT_NO]
                cell_val = str(lot_cell.value or "").strip()
                if cell_val not in lot_count_map:
                    continue

                new_count  = lot_count_map[cell_val]
                new_net_mt = round(new_count * 0.5, 6)
                new_gw_mt  = round(new_count * 0.513125, 6)

                row[COL_NET_MT].value = new_net_mt
                row[COL_GW_MT].value  = new_gw_mt
                row[COL_PLT].value    = new_count
                # STATUS 컬럼 보존 (v868: AVAILABLE/SOLD/RESERVED 유지)
                if len(row) > COL_STATUS and row[COL_STATUS].value not in (None, "SOLD", "RESERVED"):
                    row[COL_STATUS].value = "AVAILABLE"
                found[cell_val] = True
                logger.debug(
                    "[엑셀][batch] 시트=%s row=%d LOT=%s PLT=%d NET=%.4fMT GW=%.4fMT",
                    sheet_name, lot_cell.row, cell_val, new_count, new_net_mt, new_gw_mt,
                )

        if any(found.values()):
            wb.save(str(path))
            logger.info("[엑셀][batch] 저장 완료: %s", path.name)

    except PermissionError:
        logger.error("[엑셀][batch] 파일 접근 거부 (다른 프로그램이 열고 있을 수 있음): %s", excel_path)
        failed_out.extend(lot_count_map.keys())
        return
    except Exception as exc:
        logger.error("[엑셀][batch] 수정 오류: %s", exc, exc_info=True)
        failed_out.extend(lot_count_map.keys())
        return

    for lot, was_found in found.items():
        if was_found:
            updated_out.append(lot)
        else:
            failed_out.append(lot)
            logger.warning("[엑셀] LOT=%s 행 없음 (INVENTORY 시트 미발견)", lot)


def _check_reserved_conflict(
    lot_no: str, new_count: int, db
) -> Optional[str]:
    """RESERVED 상태 충돌 확인.

    해당 LOT의 allocation_plan에서 status='RESERVED'인 행 수를 집계.
    new_count < reserved_count 이면 경고 메시지 반환.
    충돌 있어도 실행 차단 안 함 (경고만).

    Args:
        lot_no:    LOT 번호
        new_count: 조정 후 포대 수
        db:        SQMDatabase

    Returns:
        None  = 충돌 없음 (또는 테이블 없음)
        str   = 경고 메시지
    """
    try:
        row = db.fetchone(
            """
            SELECT COUNT(*) AS reserved_count
            FROM allocation_plan
            WHERE lot_no = ? AND status = 'RESERVED'
            """,
            (lot_no,),
        )
        if row is None:
            return None
        reserved_count = int(row.get("reserved_count") or 0)
        if reserved_count <= 0:
            return None
        if new_count < reserved_count:
            return (
                f"LOT={lot_no}: 조정 후 포대 수({new_count})가 "
                f"RESERVED 예약 건수({reserved_count})보다 적습니다. "
                f"출고 실행 시 재고 부족 발생 가능."
            )
        return None
    except sqlite3.OperationalError as exc:
        # allocation_plan 테이블이 없는 구버전 DB
        logger.debug("[조정][충돌검사] allocation_plan 조회 스킵: %s", exc)
        return None
    except Exception as exc:
        logger.warning("[조정][충돌검사] LOT=%s 오류: %s", lot_no, exc)
        return None


# =====================================================================
# 편의 함수: 최신 엑셀 파일 자동 탐색
# =====================================================================

def find_latest_excel(base_dir: str) -> Optional[str]:
    """base_dir에서 통합본(_int.xlsx) 우선, 없으면 v*.xlsx 최신 파일 반환."""
    from pathlib import Path
    base = Path(base_dir)

    # 1순위: 통합 단일시트 파일
    int_file = base / "SMQ 입,출고 재고관리 파일_int.xlsx"
    if int_file.exists():
        return str(int_file)

    # 2순위: v* 버전 파일 중 최신
    versioned = sorted(base.glob("SMQ 입,출고 재고관리 파일_v*.xlsx"))
    if versioned:
        return str(versioned[-1])

    # 3순위: 기본 파일
    fallback = base / "SMQ 입,출고 재고관리 파일.xlsx"
    if fallback.exists():
        return str(fallback)

    return None
