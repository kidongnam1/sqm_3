# -*- coding: utf-8 -*-
"""
SQM 재고관리 시스템 - Preflight 검증 모듈 (v2.5.4)

All-or-Nothing 트랜잭션 패턴:
- Import: 검증 오류 1건이라도 있으면 전체 중단 (부분반영 금지)
- Outbound: LOT 누락/재고부족/형식 오류 중 1개라도 있으면 전체 중단 (일부 차감 금지)

이 구조가 적용되면 "엑셀 수기 대조 0회"가 현실적으로 가능합니다.

v2.5.4 개선:
- get_summary() 에러 메시지에 필드명, 문제값, 수정제안 상세 표시
- PreflightIssue.__str__() 개선: 필드명, 문제값 포함
- 사용자 친화적 에러 메시지 제공

Author: Ruby
Version: 2.5.4
Created: 2025-01-13
"""


from engine_modules.constants import STATUS_AVAILABLE
import logging
from datetime import datetime
from typing import List, Dict, Any, Tuple
from dataclasses import dataclass, field
from enum import Enum

logger = logging.getLogger(__name__)


# =============================================================================
# 예외 클래스
# =============================================================================

class PreflightError(Exception):
    """
    Preflight 검증 실패 예외

    이 예외가 발생하면 트랜잭션이 롤백되어야 합니다.
    부분 반영 없이 전체 중단됩니다.
    """

    def __init__(self, message: str, validation_result: 'PreflightResult') -> None:
        super().__init__(message)
        self.validation_result = validation_result

    def get_error_summary(self) -> str:
        """오류 요약 문자열 반환"""
        return self.validation_result.get_summary()


class PreflightAbortError(PreflightError):
    """사용자 요청에 의한 중단 (경고 확인 후 취소)"""


# =============================================================================
# 데이터 클래스
# =============================================================================

class PreflightErrorLevel(str, Enum):
    """검증 오류 심각도"""
    FATAL = "FATAL"      # 치명적 (절대 진행 불가)
    ERROR = "ERROR"      # 오류 (진행 불가)
    WARNING = "WARNING"  # 경고 (진행 가능하나 확인 필요)
    INFO = "INFO"        # 정보 (참고용)


@dataclass
class PreflightIssue:
    """검증 발견 항목"""
    level: PreflightErrorLevel
    row: int                    # Excel 행 번호 (헤더가 1이므로 데이터는 2부터)
    field: str                  # 필드명
    value: Any                  # 문제 값
    message: str                # 오류 메시지
    code: str = ""              # 오류 코드 (예: E001, W002)
    suggestion: str = ""        # 수정 제안
    column: str = ""            # Excel 열 문자 (A, B, C...) - v2.5.4 추가

    # 필드명 → Excel 열 매핑 (입고/출고 공통)
    FIELD_TO_COLUMN = {
        'lot_no': 'A',
        'sub_lt': 'B', 
        'weight': 'C',
        'qty': 'C',
        'container_no': 'D',
        'product': 'E',
        'customer': 'F',
        'lot_no+sub_lt': 'A-B',  # 복합 필드
    }

    def get_cell_location(self) -> str:
        """Excel 셀 위치 반환 (예: A5, B12)"""
        col = self.column or self.FIELD_TO_COLUMN.get(self.field, '')
        if col:
            return f"{col}{self.row}"
        return f"{self.row}행"

    def __str__(self) -> str:
        prefix = "❌" if self.level in (PreflightErrorLevel.FATAL, PreflightErrorLevel.ERROR) else "⚠️"
        
        # ★ 셀 위치 표시 (v2.5.4)
        cell_loc = self.get_cell_location()
        field_info = f"[{self.field}]" if self.field else ""
        
        value_info = f" 값='{self.value}'" if self.value is not None and str(self.value).strip() else ""
        base = f"{prefix} [{cell_loc}] {field_info} {self.message}{value_info}"
        if self.suggestion:
            base += f" 💡 {self.suggestion}"
        return base
    
    def to_popup_str(self) -> str:
        """팝업용 간결한 문자열"""
        cell_loc = self.get_cell_location()
        return f"[{cell_loc}] {self.message}"


@dataclass
class PreflightResult:
    """Preflight 검증 결과"""
    operation: str              # 'INBOUND' 또는 'OUTBOUND'
    total_rows: int             # 총 행 수
    valid_rows: int = 0         # 유효한 행 수
    issues: List[PreflightIssue] = field(default_factory=list)

    # 집계
    fatal_count: int = 0
    error_count: int = 0
    warning_count: int = 0

    # 상태
    is_valid: bool = True       # ERROR 이상이 없으면 True
    can_proceed: bool = True    # FATAL이 없으면 True (경고만 있으면 진행 가능)

    def add_issue(self, issue: PreflightIssue) -> None:
        """검증 이슈 추가"""
        self.issues.append(issue)

        if issue.level == PreflightErrorLevel.FATAL:
            self.fatal_count += 1
            self.is_valid = False
            self.can_proceed = False
        elif issue.level == PreflightErrorLevel.ERROR:
            self.error_count += 1
            self.is_valid = False
        elif issue.level == PreflightErrorLevel.WARNING:
            self.warning_count += 1

    def has_blocking_errors(self) -> bool:
        """진행 불가능한 오류가 있는지 확인"""
        return self.fatal_count > 0 or self.error_count > 0

    def get_summary(self) -> str:
        """검증 결과 요약"""
        lines = [
            f"{'='*60}",
            f"PREFLIGHT 검증 결과: {self.operation}",
            f"{'='*60}",
            f"총 {self.total_rows}행 검증",
            f"  • 유효: {self.valid_rows}행",
            f"  • FATAL: {self.fatal_count}건",
            f"  • ERROR: {self.error_count}건",
            f"  • WARNING: {self.warning_count}건",
            "",
            f"결과: {'✅ 검증 통과' if self.is_valid else '❌ 검증 실패 - 전체 중단'}",
            f"{'='*60}",
        ]

        if self.issues:
            # FATAL + ERROR 먼저
            blocking = [i for i in self.issues if i.level in (PreflightErrorLevel.FATAL, PreflightErrorLevel.ERROR)]
            if blocking:
                lines.append("")
                lines.append("❌ 오류 목록 (전체 중단 사유):")
                lines.append("-" * 40)
                for issue in blocking[:100]:  # 최대 100건
                    # v2.5.4: 필드명, 문제값, 수정제안 상세 표시
                    field_info = f"[{issue.field}]" if issue.field else ""
                    value_info = f" 값='{issue.value}'" if issue.value is not None and str(issue.value).strip() else ""
                    lines.append(f"  [{issue.row}행] [{issue.code}] {field_info} {issue.message}{value_info}")
                    if issue.suggestion:
                        lines.append(f"          💡 수정제안: {issue.suggestion}")
                if len(blocking) > 100:
                    lines.append(f"  ... 외 {len(blocking) - 100}건")

            # WARNING
            warnings = [i for i in self.issues if i.level == PreflightErrorLevel.WARNING]
            if warnings:
                lines.append("")
                lines.append("⚠️ 경고 목록 (확인 권장):")
                lines.append("-" * 40)
                for issue in warnings[:50]:  # 최대 50건
                    field_info = f"[{issue.field}]" if issue.field else ""
                    value_info = f" 값='{issue.value}'" if issue.value is not None and str(issue.value).strip() else ""
                    lines.append(f"  [{issue.row}행] [{issue.code}] {field_info} {issue.message}{value_info}")
                    if issue.suggestion:
                        lines.append(f"          💡 {issue.suggestion}")
                if len(warnings) > 50:
                    lines.append(f"  ... 외 {len(warnings) - 50}건")

        return "\n".join(lines)

    def get_errors_for_gui(self) -> List[Dict]:
        """GUI 표시용 오류 목록"""
        return [
            {
                'row': i.row,
                'level': i.level.value,
                'code': i.code,
                'field': i.field,
                'value': str(i.value)[:50] if i.value else '',
                'message': i.message,
                'suggestion': i.suggestion
            }
            for i in self.issues
            if i.level in (PreflightErrorLevel.FATAL, PreflightErrorLevel.ERROR)
        ]


# =============================================================================
# Preflight 검증기 클래스
# =============================================================================

class PreflightValidator:
    """
    Preflight 전수 검증기

    All-or-Nothing 패턴:
    - 모든 데이터를 먼저 검증
    - 1건이라도 ERROR 이상이면 PreflightError 예외 발생
    - 예외 발생 시 트랜잭션 롤백으로 부분 반영 방지

    Usage:
        validator = PreflightValidator(db)

        # 검증만 수행 (예외 없음)
        result = validator.validate_inbound(data)

        # 검증 + 오류 시 예외 발생
        validator.preflight_inbound(data)  # 실패 시 PreflightError
    """

    # 오류 코드 정의
    class Code:
        # 공통
        EMPTY_DATA = "E000"
        REQUIRED_MISSING = "E001"
        FORMAT_INVALID = "E002"

        # 입고
        DUPLICATE_IN_FILE = "E101"
        DUPLICATE_IN_DB = "E102"
        WEIGHT_INVALID = "E103"
        DATE_INVALID = "E104"

        # 출고
        LOT_NOT_FOUND = "E201"
        ALREADY_OUTBOUND = "E202"
        STOCK_INSUFFICIENT = "E203"
        OUTBOUND_DATE_INVALID = "E204"
        DUPLICATE_OUTBOUND = "E205"

        # 경고
        DATE_SEQUENCE = "W001"
        WEIGHT_ZERO = "W002"
        SAP_NO_MISSING = "W003"

    def __init__(self, db=None) -> None:
        """
        Args:
            db: 데이터베이스 연결 (필수 - DB 체크용)
        """
        self.db = db

    # =========================================================================
    # 입고 검증
    # =========================================================================

    def validate_inbound(self, data: List[Dict], check_db: bool = True) -> PreflightResult:
        """
        입고 데이터 Preflight 검증 (예외 없이 결과만 반환)

        Args:
            data: 검증할 데이터 리스트
            check_db: DB 중복 체크 여부

        Returns:
            PreflightResult: 검증 결과
        """
        result = PreflightResult(operation="INBOUND", total_rows=len(data))

        if not data:
            result.add_issue(PreflightIssue(
                level=PreflightErrorLevel.FATAL,
                row=0, field="", value="",
                message="데이터가 비어있습니다.",
                code=self.Code.EMPTY_DATA
            ))
            return result

        # 내부 중복 체크용
        seen_keys: Dict[Tuple, int] = {}  # (lot_no, sub_lt) -> first_row

        for idx, row in enumerate(data, start=2):  # Excel 헤더가 1행
            row_valid = True

            # 1. 필수값 체크
            for f_name in ['lot_no', 'sub_lt', 'weight']:
                value = row.get(f_name)
                if value is None or (isinstance(value, str) and not value.strip()):
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.ERROR,
                        row=idx, field=f_name, value=value,
                        message=f"필수값 누락: {f_name}",
                        code=self.Code.REQUIRED_MISSING,
                        suggestion=f"'{field}' 값을 입력해주세요."
                    ))
                    row_valid = False

            # v6.9.1 [I2]: product 누락 경고 (제품명 없으면 재고 집계 불가)
            product_val = row.get('product', '')
            if not product_val or (isinstance(product_val, str) and not product_val.strip()):
                result.add_issue(PreflightIssue(
                    level=PreflightErrorLevel.WARNING,
                    row=idx, field='product', value=product_val,
                    message="제품명(product) 누락 — 재고 집계/보고서에 영향",
                    code=self.Code.REQUIRED_MISSING,
                    suggestion="제품명을 입력해주세요 (예: Lithium Carbonate, Nickel Sulfate)"
                ))

            # 2. 중량 형식 체크
            weight = row.get('weight')
            if weight is not None:
                try:
                    w = float(weight)
                    if w <= 0:
                        result.add_issue(PreflightIssue(
                            level=PreflightErrorLevel.WARNING,
                            row=idx, field='weight', value=weight,
                            message=f"중량이 0 이하: {weight}",
                            code=self.Code.WEIGHT_ZERO,
                            suggestion="중량 값을 확인해주세요."
                        ))
                except (ValueError, TypeError):
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.ERROR,
                        row=idx, field='weight', value=weight,
                        message=f"중량 형식 오류: {weight}",
                        code=self.Code.WEIGHT_INVALID,
                        suggestion="숫자만 입력해주세요."
                    ))
                    row_valid = False

            # 3. 내부 중복 체크 (파일 내)
            lot_no = str(row.get('lot_no', '')).strip()
            sub_lt = str(row.get('sub_lt', '')).strip()

            if lot_no and sub_lt:
                key = (lot_no, sub_lt)
                if key in seen_keys:
                    first_row = seen_keys[key]
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.ERROR,
                        row=idx, field='lot_no+sub_lt', value=f"{lot_no}-{sub_lt}",
                        message=f"파일 내 중복: LOT {lot_no}, 톤백 {sub_lt} ({first_row}행과 중복)",
                        code=self.Code.DUPLICATE_IN_FILE,
                        suggestion="중복 행을 제거해주세요."
                    ))
                    row_valid = False
                else:
                    seen_keys[key] = idx

            # 4. DB 중복 체크
            if check_db and self.db and lot_no and sub_lt:
                existing = self.db.fetchone("""
                    SELECT id FROM inventory_tonbag
                    WHERE lot_no = ? AND sub_lt = ?
                """, (lot_no, sub_lt))

                if existing:
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.ERROR,
                        row=idx, field='lot_no+sub_lt', value=f"{lot_no}-{sub_lt}",
                        message=f"DB에 이미 존재: LOT {lot_no}, 톤백 {sub_lt}",
                        code=self.Code.DUPLICATE_IN_DB,
                        suggestion="이미 입고된 항목입니다."
                    ))
                    row_valid = False

            if row_valid:
                result.valid_rows += 1

        return result

    def preflight_inbound(self, data: List[Dict], check_db: bool = True) -> PreflightResult:
        """
        입고 데이터 Preflight 검증 + 오류 시 예외 발생

        Args:
            data: 검증할 데이터 리스트
            check_db: DB 중복 체크 여부

        Returns:
            PreflightResult: 검증 통과 시 결과 반환

        Raises:
            PreflightError: 검증 실패 시 (ERROR 1건 이상)
        """
        result = self.validate_inbound(data, check_db)

        if result.has_blocking_errors():
            logger.warning(f"[PREFLIGHT] 입고 검증 실패: {result.error_count} 오류, {result.fatal_count} 치명적")
            raise PreflightError(
                f"입고 Preflight 검증 실패: {result.error_count + result.fatal_count}건의 오류 발견. 전체 중단됩니다.",
                result
            )

        logger.info(f"[PREFLIGHT] 입고 검증 통과: {result.valid_rows}/{result.total_rows}행")
        return result

    # =========================================================================
    # 출고 검증
    # =========================================================================

    def validate_outbound(self, data: List[Dict], check_db: bool = True) -> PreflightResult:
        """
        출고 데이터 Preflight 검증 (예외 없이 결과만 반환)

        Args:
            data: 검증할 데이터 리스트 (lot_no, sub_lt 또는 qty_mt 포함)
            check_db: DB 재고 체크 여부

        Returns:
            PreflightResult: 검증 결과
        """
        result = PreflightResult(operation="OUTBOUND", total_rows=len(data))

        if not data:
            result.add_issue(PreflightIssue(
                level=PreflightErrorLevel.FATAL,
                row=0, field="", value="",
                message="데이터가 비어있습니다.",
                code=self.Code.EMPTY_DATA
            ))
            return result

        # 내부 중복 체크용 (같은 톤백 2번 출고 방지)
        seen_keys: Dict[Tuple, int] = {}  # (lot_no, sub_lt) -> first_row

        # LOT 단위 수량 집계 (재고 부족 체크용)
        lot_requests: Dict[str, float] = {}  # lot_no -> 총 요청량 (kg)

        for idx, row in enumerate(data, start=2):
            row_valid = True

            # 1. 필수값 체크
            lot_no = str(row.get('lot_no', '')).strip()
            sub_lt = row.get('sub_lt')  # 톤백 번호 (선택적)

            if not lot_no:
                result.add_issue(PreflightIssue(
                    level=PreflightErrorLevel.ERROR,
                    row=idx, field='lot_no', value=lot_no,
                    message="LOT 번호 누락",
                    code=self.Code.REQUIRED_MISSING,
                    suggestion="LOT 번호를 입력해주세요."
                ))
                row_valid = False
                continue

            # 2. 톤백 단위 출고인 경우 (sub_lt 존재)
            if sub_lt:
                sub_lt = str(sub_lt).strip()
                key = (lot_no, sub_lt)

                # 내부 중복 체크
                if key in seen_keys:
                    first_row = seen_keys[key]
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.ERROR,
                        row=idx, field='lot_no+sub_lt', value=f"{lot_no}-{sub_lt}",
                        message=f"파일 내 중복 출고: LOT {lot_no}, 톤백 {sub_lt} ({first_row}행과 중복)",
                        code=self.Code.DUPLICATE_OUTBOUND,
                        suggestion="중복 행을 제거해주세요."
                    ))
                    row_valid = False
                else:
                    seen_keys[key] = idx

                # DB 재고 체크
                if check_db and self.db:
                    tonbag = self.db.fetchone("""
                        SELECT t.id, t.status, t.weight, i.current_weight
                        FROM inventory_tonbag t
                        LEFT JOIN inventory i ON t.lot_no = i.lot_no
                        WHERE t.lot_no = ? AND t.sub_lt = ?
                    """, (lot_no, sub_lt))

                    if not tonbag:
                        result.add_issue(PreflightIssue(
                            level=PreflightErrorLevel.ERROR,
                            row=idx, field='lot_no+sub_lt', value=f"{lot_no}-{sub_lt}",
                            message=f"재고 없음: LOT {lot_no}, 톤백 {sub_lt}",
                            code=self.Code.LOT_NOT_FOUND,
                            suggestion="입고되지 않았거나 LOT/톤백 번호를 확인해주세요."
                        ))
                        row_valid = False
                    elif tonbag['status'] != STATUS_AVAILABLE:
                        result.add_issue(PreflightIssue(
                            level=PreflightErrorLevel.ERROR,
                            row=idx, field='status', value=tonbag['status'],
                            message=f"이미 출고됨: LOT {lot_no}, 톤백 {sub_lt} (상태: {tonbag['status']})",
                            code=self.Code.ALREADY_OUTBOUND,
                            suggestion="이미 출고된 항목입니다."
                        ))
                        row_valid = False

            # 3. LOT 단위 수량 출고인 경우 (qty_mt 존재)
            else:
                qty_mt = row.get('qty_mt', 0)
                try:
                    qty_mt = float(qty_mt) if qty_mt else 0
                    qty_kg = qty_mt * 1000
                except (ValueError, TypeError):
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.ERROR,
                        row=idx, field='qty_mt', value=qty_mt,
                        message=f"수량 형식 오류: {qty_mt}",
                        code=self.Code.FORMAT_INVALID,
                        suggestion="숫자만 입력해주세요."
                    ))
                    row_valid = False
                    continue

                if qty_kg <= 0:
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.WARNING,
                        row=idx, field='qty_mt', value=qty_mt,
                        message=f"출고 수량이 0 이하: {qty_mt}MT",
                        code=self.Code.WEIGHT_ZERO
                    ))

                # 집계
                lot_requests[lot_no] = lot_requests.get(lot_no, 0) + qty_kg

            if row_valid:
                result.valid_rows += 1

        # 4. LOT 단위 재고 부족 체크
        if check_db and self.db and lot_requests:
            for lot_no, total_request_kg in lot_requests.items():
                inv = self.db.fetchone("""
                    SELECT id, current_weight, status
                    FROM inventory
                    WHERE lot_no = ?
                """, (lot_no,))

                if not inv:
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.ERROR,
                        row=0, field='lot_no', value=lot_no,
                        message=f"LOT 없음: {lot_no}",
                        code=self.Code.LOT_NOT_FOUND,
                        suggestion="입고되지 않았거나 LOT 번호를 확인해주세요."
                    ))
                    result.is_valid = False
                elif (inv['current_weight'] or 0) < total_request_kg:
                    current_mt = (inv['current_weight'] or 0) / 1000
                    request_mt = total_request_kg / 1000
                    result.add_issue(PreflightIssue(
                        level=PreflightErrorLevel.ERROR,
                        row=0, field='qty_mt', value=f"{request_mt:.3f}MT",
                        message=f"재고 부족: LOT {lot_no} (잔량: {current_mt:.3f}MT, 요청: {request_mt:.3f}MT)",
                        code=self.Code.STOCK_INSUFFICIENT,
                        suggestion=f"출고량을 {current_mt:.3f}MT 이하로 조정하거나 재고를 확인해주세요."
                    ))
                    result.is_valid = False

        return result

    def preflight_outbound(self, data: List[Dict], check_db: bool = True) -> PreflightResult:
        """
        출고 데이터 Preflight 검증 + 오류 시 예외 발생

        Args:
            data: 검증할 데이터 리스트
            check_db: DB 재고 체크 여부

        Returns:
            PreflightResult: 검증 통과 시 결과 반환

        Raises:
            PreflightError: 검증 실패 시 (ERROR 1건 이상)
        """
        result = self.validate_outbound(data, check_db)

        if result.has_blocking_errors():
            logger.warning(f"[PREFLIGHT] 출고 검증 실패: {result.error_count} 오류, {result.fatal_count} 치명적")
            raise PreflightError(
                f"출고 Preflight 검증 실패: {result.error_count + result.fatal_count}건의 오류 발견. 전체 중단됩니다.",
                result
            )

        logger.info(f"[PREFLIGHT] 출고 검증 통과: {result.valid_rows}/{result.total_rows}행")
        return result

    # =========================================================================
    # 유틸리티
    # =========================================================================

    def validate_and_summarize(self, data: List[Dict], operation: str = 'INBOUND') -> str:
        """
        검증 후 결과 요약 문자열 반환 (GUI 표시용)

        Args:
            data: 검증할 데이터
            operation: 'INBOUND' 또는 'OUTBOUND'

        Returns:
            str: 요약 문자열
        """
        if operation.upper() == 'INBOUND':
            result = self.validate_inbound(data)
        else:
            result = self.validate_outbound(data)

        return result.get_summary()


# =============================================================================
# 헬퍼 함수
# =============================================================================

def run_preflight_check(db, data: List[Dict], operation: str = 'INBOUND') -> PreflightResult:
    """
    Preflight 검증 실행 (간편 함수)

    Args:
        db: 데이터베이스 연결
        data: 검증할 데이터
        operation: 'INBOUND' 또는 'OUTBOUND'

    Returns:
        PreflightResult: 검증 결과

    Raises:
        PreflightError: 검증 실패 시
    """
    validator = PreflightValidator(db)

    if operation.upper() == 'INBOUND':
        return validator.preflight_inbound(data)
    else:
        return validator.preflight_outbound(data)


def format_preflight_errors_for_excel(result: PreflightResult) -> List[Dict]:
    """
    오류 목록을 Excel 내보내기용 형식으로 변환

    Args:
        result: PreflightResult 객체

    Returns:
        List[Dict]: Excel 행으로 변환 가능한 딕셔너리 리스트
    """
    return [
        {
            '행번호': i.row,
            '심각도': i.level.value,
            '오류코드': i.code,
            '필드': i.field,
            '값': str(i.value)[:100] if i.value else '',
            '오류메시지': i.message,
            '수정제안': i.suggestion
        }
        for i in result.issues
    ]


# =============================================================================
# Excel 오류 리포트 생성
# =============================================================================

class PreflightErrorReport:
    """
    Preflight 검증 오류를 Excel로 내보내기

    Usage:
        result = validator.validate_outbound(data)
        if result.has_blocking_errors():
            report = PreflightErrorReport(result)
            output_path = report.save_to_excel()
            logger.debug(f"오류 리포트 저장됨: {output_path}")
    """

    def __init__(self, result: PreflightResult, source_file: str = None) -> None:
        """
        Args:
            result: PreflightResult 객체
            source_file: 원본 파일명 (선택)
        """
        self.result = result
        self.source_file = source_file or "Unknown"

    def save_to_excel(self, output_path: str = None) -> str:
        """
        오류 목록을 Excel로 저장

        Args:
            output_path: 출력 파일 경로 (None이면 자동 생성)

        Returns:
            str: 저장된 파일 경로
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl 모듈 없음 - Excel 저장 불가")
            return self._save_to_csv(output_path)

        # 기본 경로 생성
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"preflight_errors_{self.result.operation}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "검증오류"

        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 요약 정보 (상단)
        summary_data = [
            ["PREFLIGHT 검증 결과", ""],
            ["검증 유형", self.result.operation],
            ["원본 파일", self.source_file],
            ["검증 시간", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["총 행수", self.result.total_rows],
            ["유효 행수", self.result.valid_rows],
            ["FATAL", self.result.fatal_count],
            ["ERROR", self.result.error_count],
            ["WARNING", self.result.warning_count],
            ["결과", "✅ 통과" if self.result.is_valid else "❌ 실패"],
            ["", ""],
        ]

        for row_idx, (key, value) in enumerate(summary_data, start=1):
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)

        # 헤더
        headers = ["행번호", "심각도", "오류코드", "필드", "값", "오류메시지", "수정제안"]
        header_row = len(summary_data) + 1

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 데이터
        for row_idx, issue in enumerate(self.result.issues, start=header_row + 1):
            row_data = [
                issue.row,
                issue.level.value,
                issue.code,
                issue.field,
                str(issue.value)[:100] if issue.value else '',
                issue.message,
                issue.suggestion
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # 심각도별 색상
                if col_idx == 2:  # 심각도 열
                    if value in ('FATAL', 'ERROR'):
                        cell.fill = error_fill
                    elif value == 'WARNING':
                        cell.fill = warning_fill

        # 열 너비 자동 조정
        column_widths = [10, 12, 12, 15, 30, 50, 40]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 저장
        wb.save(output_path)
        logger.info(f"[PREFLIGHT] 오류 리포트 저장: {output_path}")

        return output_path

    def _save_to_csv(self, output_path: str = None) -> str:
        """openpyxl 없을 때 CSV로 저장"""
        import csv

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"preflight_errors_{self.result.operation}_{timestamp}.csv"

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # 요약
            writer.writerow(["PREFLIGHT 검증 결과"])
            writer.writerow(["검증 유형", self.result.operation])
            writer.writerow(["총 행수", self.result.total_rows])
            writer.writerow(["ERROR", self.result.error_count])
            writer.writerow(["WARNING", self.result.warning_count])
            writer.writerow([])

            # 헤더
            writer.writerow(["행번호", "심각도", "오류코드", "필드", "값", "오류메시지", "수정제안"])

            # 데이터
            for issue in self.result.issues:
                writer.writerow([
                    issue.row,
                    issue.level.value,
                    issue.code,
                    issue.field,
                    str(issue.value)[:100] if issue.value else '',
                    issue.message,
                    issue.suggestion
                ])

        logger.info(f"[PREFLIGHT] 오류 리포트 저장 (CSV): {output_path}")
        return output_path


def export_preflight_errors(result: PreflightResult, output_path: str = None,
                           source_file: str = None) -> str:
    """
    Preflight 오류를 Excel로 내보내기 (간편 함수)

    Args:
        result: PreflightResult 객체
        output_path: 출력 파일 경로 (선택)
        source_file: 원본 파일명 (선택)

    Returns:
        str: 저장된 파일 경로
    """
    report = PreflightErrorReport(result, source_file)
    return report.save_to_excel(output_path)


# =============================================================================
# 테스트
# =============================================================================

if __name__ == "__main__":
    # 테스트용 (DB 없이)
    validator = PreflightValidator(db=None)

    # 입고 테스트 데이터
    inbound_data = [
        {'lot_no': 'LOT001', 'sub_lt': 'T001', 'weight': 1000},
        {'lot_no': 'LOT001', 'sub_lt': 'T002', 'weight': 1000},
        {'lot_no': 'LOT001', 'sub_lt': 'T001', 'weight': 1000},  # 중복!
        {'lot_no': 'LOT002', 'sub_lt': '', 'weight': 1000},       # sub_lt 누락!
        {'lot_no': '', 'sub_lt': 'T003', 'weight': 1000},         # lot_no 누락!
    ]

    logger.debug("=" * 60)
    logger.debug("입고 검증 테스트")
    logger.debug("=" * 60)
    result = validator.validate_inbound(inbound_data, check_db=False)
    logger.debug(f"{result.get_summary()}")

    # 출고 테스트 데이터
    outbound_data = [
        {'lot_no': 'LOT001', 'sub_lt': 'T001'},
        {'lot_no': 'LOT001', 'sub_lt': 'T002'},
        {'lot_no': 'LOT001', 'sub_lt': 'T001'},  # 중복 출고!
        {'lot_no': '', 'sub_lt': 'T003'},         # lot_no 누락!
    ]

    logger.debug("\n" + "=" * 60)
    logger.debug("출고 검증 테스트")
    logger.debug("=" * 60)
    result = validator.validate_outbound(outbound_data, check_db=False)
    logger.debug(f"{result.get_summary()}")