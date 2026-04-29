"""
SQM v7.0.1 — 정합성 검증 리포트 생성
=====================================

PDF: 전체 정합성 검사 결과 + 구역별 위치 현황 + 크로스 검증 상세
Excel: 정합성 결과 시트 + LOT별 상세 시트 + 위치 현황 시트

일본 고객 감사 대응용 — "시스템 정합성 검사 리포트" 제출
"""

def _get_malgun_font_paths():
    """플랫폼별 맑은 고딕 폰트 경로 목록 반환."""
    import sys, os
    if sys.platform == 'win32':
        win_fonts = os.environ.get('WINDIR', 'C:/Windows') + '/Fonts'
        return [
            os.path.join(win_fonts, 'malgun.ttf'),
            os.path.join(win_fonts, 'malgunbd.ttf'),
            _get_malgun_font_paths()[0],    # fallback
        ]
    elif sys.platform == 'darwin':
        return [
            '/Library/Fonts/AppleGothic.ttf',
            '/System/Library/Fonts/AppleGothic.ttf',
        ]
    else:  # Linux
        return [
            '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        ]

import logging
import os
from datetime import datetime
from typing import Any, Optional

logger = logging.getLogger(__name__)


def generate_integrity_report_pdf(engine: Any, save_path: str) -> Optional[str]:
    """
    정합성 검증 리포트 PDF 생성

    Args:
        engine: SQMInventoryEngineV3 인스턴스
        save_path: 저장 경로

    Returns:
        저장된 파일 경로 (실패 시 None)
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError:
        logger.error("reportlab 필요: pip install reportlab")
        return None

    try:
        # 한글 폰트
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        font_paths = [
            '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
            _get_malgun_font_paths()[0],
            '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
        ]
        font_name = 'Helvetica'
        for fp in font_paths:
            if os.path.exists(fp):
                pdfmetrics.registerFont(TTFont('KoreanFont', fp))
                font_name = 'KoreanFont'
                break
    except Exception:
        font_name = 'Helvetica'

    now = datetime.now()
    c = canvas.Canvas(save_path, pagesize=A4)
    w, h = A4

    # ── 페이지 1: 요약 ──
    y = h - 40 * mm

    c.setFont(font_name, 18)
    c.drawCentredString(w / 2, y, "SQM 재고관리 시스템 — 정합성 검증 리포트")
    y -= 10 * mm

    c.setFont(font_name, 10)
    c.drawCentredString(w / 2, y, f"생성일시: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 15 * mm

    # 정합성 검사 실행
    from engine_modules.validators import InventoryValidator
    validator = InventoryValidator(db=engine.db)
    result = validator.check_data_integrity()

    # 요약 박스
    c.setFont(font_name, 14)
    status_text = "✅ 정상" if result.is_valid else "🔴 이상 발견"
    c.drawString(20 * mm, y, f"검사 결과: {status_text}")
    y -= 8 * mm

    c.setFont(font_name, 10)
    c.drawString(20 * mm, y, f"오류: {len(result.errors)}건  |  경고: {len(result.warnings)}건")
    y -= 10 * mm

    # 오류 목록
    if result.errors:
        c.setFont(font_name, 12)
        c.setFillColor(colors.red)
        c.drawString(20 * mm, y, "■ 오류 항목")
        c.setFillColor(colors.black)
        y -= 7 * mm

        c.setFont(font_name, 9)
        for err in result.errors[:15]:
            if y < 30 * mm:
                c.showPage()
                y = h - 30 * mm
                c.setFont(font_name, 9)
            # 긴 텍스트 자르기
            display = err[:100] + ('...' if len(err) > 100 else '')
            c.drawString(25 * mm, y, f"• {display}")
            y -= 5 * mm

    # 경고 목록
    if result.warnings:
        y -= 5 * mm
        c.setFont(font_name, 12)
        c.setFillColor(colors.HexColor('#e67e22'))
        c.drawString(20 * mm, y, "■ 경고 항목")
        c.setFillColor(colors.black)
        y -= 7 * mm

        c.setFont(font_name, 9)
        for warn in result.warnings[:10]:
            if y < 30 * mm:
                c.showPage()
                y = h - 30 * mm
                c.setFont(font_name, 9)
            display = warn[:100] + ('...' if len(warn) > 100 else '')
            c.drawString(25 * mm, y, f"• {display}")
            y -= 5 * mm

    # ── 페이지 2: 재고 통계 + 위치 현황 ──
    c.showPage()
    y = h - 30 * mm

    c.setFont(font_name, 14)
    c.drawString(20 * mm, y, "■ 재고 현황 요약")
    y -= 10 * mm

    # 전체 통계
    try:
        stats = engine.db.fetchone("""
            SELECT COUNT(*) as lot_count,
                   SUM(current_weight) as total_weight,
                   SUM(CASE WHEN status='AVAILABLE' THEN 1 ELSE 0 END) as avail_lots,
                   SUM(CASE WHEN status='DEPLETED' THEN 1 ELSE 0 END) as depleted_lots
            FROM inventory
        """)
        tb_stats = engine.db.fetchone("""
            SELECT COUNT(*) as total,
                   SUM(CASE WHEN status='AVAILABLE' THEN 1 ELSE 0 END) as avail,
                   SUM(CASE WHEN status='RESERVED' THEN 1 ELSE 0 END) as reserved,
                   SUM(CASE WHEN status='PICKED' THEN 1 ELSE 0 END) as picked,
                   SUM(CASE WHEN status='SOLD' THEN 1 ELSE 0 END) as sold
            FROM inventory_tonbag WHERE COALESCE(is_sample, 0) = 0
        """)

        c.setFont(font_name, 10)
        items = [
            f"전체 LOT: {stats['lot_count']}개",
            f"총 재고: {(stats['total_weight'] or 0) / 1000:,.1f} MT",
            f"가용 LOT: {stats['avail_lots']}개  /  소진 LOT: {stats['depleted_lots']}개",
            f"톤백 현황: 가용 {tb_stats['avail']}  예약 {tb_stats['reserved']}  "
            f"출고 {tb_stats['picked']}  판매 {tb_stats['sold']}",
        ]
        for item in items:
            c.drawString(25 * mm, y, item)
            y -= 6 * mm
    except Exception as e:
        c.drawString(25 * mm, y, f"통계 조회 실패: {e}")
        y -= 6 * mm

    # 구역별 위치 현황
    y -= 10 * mm
    c.setFont(font_name, 14)
    c.drawString(20 * mm, y, "■ 구역별 위치 현황")
    y -= 10 * mm

    try:
        zone_rows = engine.db.fetchall("""
            SELECT 
                CASE 
                    WHEN location IS NULL OR location = '' THEN '(미지정)'
                    WHEN INSTR(location, '-') > 0 THEN SUBSTR(location, 1, INSTR(location, '-') - 1)
                    ELSE location
                END AS zone,
                COUNT(*) AS count,
                SUM(weight) AS total_weight
            FROM inventory_tonbag
            WHERE status = 'AVAILABLE' AND COALESCE(is_sample, 0) = 0
            GROUP BY zone ORDER BY zone
        """)

        # 테이블 헤더
        c.setFont(font_name, 10)
        c.setFillColor(colors.HexColor('#2c3e50'))
        c.drawString(25 * mm, y, "구역")
        c.drawString(60 * mm, y, "톤백 수")
        c.drawString(95 * mm, y, "중량 (MT)")
        c.setFillColor(colors.black)
        y -= 2 * mm
        c.line(25 * mm, y, 140 * mm, y)
        y -= 5 * mm

        c.setFont(font_name, 9)
        for row in zone_rows:
            zone = row['zone'] or '(미지정)'
            count = row['count'] or 0
            weight_mt = (row['total_weight'] or 0) / 1000
            c.drawString(25 * mm, y, zone)
            c.drawString(60 * mm, y, f"{count:,}")
            c.drawString(95 * mm, y, f"{weight_mt:,.1f}")
            y -= 5 * mm
            if y < 30 * mm:
                c.showPage()
                y = h - 30 * mm
    except Exception as e:
        c.drawString(25 * mm, y, f"위치 조회 실패: {e}")

    # 푸터
    c.setFont(font_name, 8)
    c.drawCentredString(w / 2, 15 * mm, f"SQM Inventory System v7.0.1 — {now.strftime('%Y-%m-%d')}")

    c.save()
    logger.info(f"정합성 리포트 PDF 생성: {save_path}")
    return save_path


def generate_integrity_report_excel(engine: Any, save_path: str) -> Optional[str]:
    """
    정합성 검증 리포트 Excel 생성 (3시트)

    Sheet 1: 검사 요약 + 오류/경고 목록
    Sheet 2: LOT별 정합성 상세
    Sheet 3: 구역별 위치 현황
    """
    try:
        import openpyxl
        from openpyxl.styles import Border, Font, PatternFill, Side
    except ImportError:
        logger.error("openpyxl 필요: pip install openpyxl")
        return None

    try:
        now = datetime.now()
        wb = openpyxl.Workbook()

        # 스타일
        _ = Font(bold=True, size=11)  # header_font: header_font_white 사용
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font_white = Font(bold=True, size=11, color='FFFFFF')
        error_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
        warn_fill = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ── Sheet 1: 검사 요약 ──
        ws1 = wb.active
        ws1.title = "검사 요약"

        from engine_modules.validators import InventoryValidator
        validator = InventoryValidator(db=engine.db)
        result = validator.check_data_integrity()

        ws1.append(["SQM 정합성 검증 리포트"])
        ws1.append([f"생성일시: {now.strftime('%Y-%m-%d %H:%M:%S')}"])
        ws1.append([])
        ws1.append(["검사 결과", "정상" if result.is_valid else "이상 발견"])
        ws1.append(["오류 건수", len(result.errors)])
        ws1.append(["경고 건수", len(result.warnings)])
        ws1.append([])

        # 오류 목록
        if result.errors:
            ws1.append(["■ 오류 항목"])
            for err in result.errors:
                row = ws1.append(["", err])
            ws1.append([])

        # 경고 목록
        if result.warnings:
            ws1.append(["■ 경고 항목"])
            for warn in result.warnings:
                ws1.append(["", warn])

        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 100

        # ── Sheet 2: LOT별 정합성 상세 ──
        ws2 = wb.create_sheet("LOT별 상세")

        headers = ["LOT NO", "initial_weight", "current_weight", "picked_weight",
                    "톤백_가용", "톤백_출고", "톤백_예약", "톤백_수", "샘플_수",
                    "정합성", "오류/경고"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        lots = engine.db.fetchall("SELECT lot_no FROM inventory ORDER BY lot_no")
        row_idx = 2
        for lot in lots:
            lot_no = lot['lot_no']
            try:
                check = engine.verify_lot_integrity(lot_no)
                d = check.get('details', {})

                ws2.cell(row=row_idx, column=1, value=lot_no)
                ws2.cell(row=row_idx, column=2, value=d.get('initial_weight', 0))
                ws2.cell(row=row_idx, column=3, value=d.get('current_weight', 0))
                ws2.cell(row=row_idx, column=4, value=d.get('picked_weight', 0))
                ws2.cell(row=row_idx, column=5, value=d.get('tonbag_available_weight', 0))
                ws2.cell(row=row_idx, column=6, value=d.get('tonbag_picked_weight', 0))
                ws2.cell(row=row_idx, column=7, value=d.get('tonbag_reserved_weight', 0))
                ws2.cell(row=row_idx, column=8, value=d.get('tonbag_count', 0))
                ws2.cell(row=row_idx, column=9, value=d.get('sample_count', 0))
                ws2.cell(row=row_idx, column=10, value="OK" if check['valid'] else "ERROR")
                issues = '; '.join(check['errors'] + check['warnings'])
                ws2.cell(row=row_idx, column=11, value=issues[:200])

                if not check['valid']:
                    for col in range(1, 12):
                        ws2.cell(row=row_idx, column=col).fill = error_fill
                elif check['warnings']:
                    for col in range(1, 12):
                        ws2.cell(row=row_idx, column=col).fill = warn_fill
            except Exception as e:
                ws2.cell(row=row_idx, column=1, value=lot_no)
                ws2.cell(row=row_idx, column=10, value="SKIP")
                ws2.cell(row=row_idx, column=11, value=str(e)[:200])
            row_idx += 1

        for col in range(1, 12):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        ws2.column_dimensions['K'].width = 60

        # ── Sheet 3: 구역별 위치 현황 ──
        ws3 = wb.create_sheet("위치 현황")

        zone_headers = ["구역", "톤백 수", "중량 (kg)", "중량 (MT)"]
        for col, h in enumerate(zone_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        zone_rows = engine.db.fetchall("""
            SELECT 
                CASE 
                    WHEN location IS NULL OR location = '' THEN '(미지정)'
                    WHEN INSTR(location, '-') > 0 THEN SUBSTR(location, 1, INSTR(location, '-') - 1)
                    ELSE location
                END AS zone,
                COUNT(*) AS count,
                SUM(weight) AS total_weight
            FROM inventory_tonbag
            WHERE status = 'AVAILABLE' AND COALESCE(is_sample, 0) = 0
            GROUP BY zone ORDER BY zone
        """)

        for i, row in enumerate(zone_rows, 2):
            ws3.cell(row=i, column=1, value=row['zone'] or '(미지정)')
            ws3.cell(row=i, column=2, value=row['count'] or 0)
            ws3.cell(row=i, column=3, value=row['total_weight'] or 0)
            ws3.cell(row=i, column=4, value=(row['total_weight'] or 0) / 1000)

        for col in range(1, 5):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        wb.save(save_path)
        logger.info(f"정합성 리포트 Excel 생성: {save_path}")
        return save_path

    except Exception as e:
        logger.error(f"정합성 Excel 리포트 생성 실패: {e}")
        return None
