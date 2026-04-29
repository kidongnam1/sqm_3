"""
SQM v864.3 — 엔진+SQL 혼합 액션 엔드포인트
Phase 4-A Group 4: F013, F029, F035, F050, F061

- F013: 🔍 정합성 검증 (inventory vs inventory_tonbag 무결성)
- F029: 💾 백업 생성 (shutil.copy → backup/ 폴더, 최대 5개 유지)
- F035: 📊 LOT 리스트 Excel 내보내기 (openpyxl)
- F050: 🔖 LOT 상세 (inventory + tonbag + snapshot JOIN)
- F061: ℹ️ 시스템 정보 (DB stats + version + 파일 크기)
"""
import os
import shutil
import sqlite3
import logging
import tempfile
import importlib.util
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter
from fastapi.responses import FileResponse
from backend.common.errors import ok_response, err_response

router = APIRouter(prefix="/api/action", tags=["actions"])
logger = logging.getLogger(__name__)

KST = timezone(timedelta(hours=9))
MAX_BACKUPS = 5


# ── 공용 헬퍼 ────────────────────────────────────────────────────
def _project_root() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.dirname(os.path.dirname(here))


def _db_path() -> str:
    return os.path.join(_project_root(), "data", "db", "sqm_inventory.db")


def _db() -> sqlite3.Connection:
    con = sqlite3.connect(_db_path(), timeout=5, check_same_thread=False)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA busy_timeout=3000")
    return con


def _rows_to_list(rows) -> list:
    return [dict(r) for r in rows]


def _table_exists(con: sqlite3.Connection, name: str) -> bool:
    row = con.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1",
        (name,),
    ).fetchone()
    return bool(row)


def _table_columns(con: sqlite3.Connection, name: str) -> set[str]:
    if not _table_exists(con, name):
        return set()
    return {str(r["name"]) for r in con.execute(f"PRAGMA table_info({name})").fetchall()}


def _load_version() -> dict:
    """version.py 동적 로드 — 실패 시 기본값 반환"""
    try:
        root = _project_root()
        spec = importlib.util.spec_from_file_location(
            "version", os.path.join(root, "version.py"))
        ver = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(ver)
        return {
            "app_name":    getattr(ver, "APP_NAME",    "SQM 재고관리 시스템"),
            "version":     getattr(ver, "__version__", "unknown"),
            "release_date":getattr(ver, "RELEASE_DATE","unknown"),
            "build_date":  getattr(ver, "BUILD_DATE",  "unknown"),
        }
    except Exception:
        return {"app_name": "SQM 재고관리 시스템", "version": "8.6.4"}


# ── F013: 정합성 검증 ─────────────────────────────────────────
@router.get("/integrity-check", summary="🔍 정합성 검증 (F013)")
def integrity_check():
    """
    inventory ↔ inventory_tonbag 무결성 검사.
    - LOT에 연결된 톤백 수 vs 선언된 tonbag_count 불일치
    - 고아(orphan) 톤백 (inventory 없는 tonbag)
    - 상태 불일치 (LOT=AVAILABLE, 톤백=SOLD 등 혼재)
    """
    try:
        con = _db()
        issues = []

        # ① tonbag_count 불일치
        mismatch = con.execute("""
            SELECT
                i.id, i.lot_no, i.product,
                i.tonbag_count AS declared,
                COUNT(t.id)   AS actual,
                i.status
            FROM inventory i
            LEFT JOIN inventory_tonbag t ON t.inventory_id = i.id
            GROUP BY i.id
            HAVING declared != actual
        """).fetchall()
        for r in mismatch:
            issues.append({
                "type": "TONBAG_COUNT_MISMATCH",
                "lot_no": r["lot_no"],
                "detail": f"선언={r['declared']} / 실제={r['actual']}"
            })

        # ② 고아 톤백 (inventory 없는 tonbag)
        orphans = con.execute("""
            SELECT t.id, t.lot_no, t.sub_lt
            FROM inventory_tonbag t
            LEFT JOIN inventory i ON i.id = t.inventory_id
            WHERE i.id IS NULL
            LIMIT 50
        """).fetchall()
        for r in orphans:
            issues.append({
                "type": "ORPHAN_TONBAG",
                "lot_no": r["lot_no"],
                "detail": f"tonbag id={r['id']} sub_lt={r['sub_lt']} — inventory 없음"
            })

        # ③ 상태 불일치 (LOT=AVAILABLE 이지만 tonbag 전부 SOLD)
        status_mismatch = con.execute("""
            SELECT
                i.lot_no, i.status AS lot_status,
                GROUP_CONCAT(DISTINCT t.status) AS tonbag_statuses
            FROM inventory i
            JOIN inventory_tonbag t ON t.inventory_id = i.id
            WHERE i.status = 'AVAILABLE'
            GROUP BY i.id
            HAVING tonbag_statuses NOT LIKE '%AVAILABLE%'
               AND tonbag_statuses IS NOT NULL
        """).fetchall()
        for r in status_mismatch:
            issues.append({
                "type": "STATUS_MISMATCH",
                "lot_no": r["lot_no"],
                "detail": f"LOT={r['lot_status']}, 톤백 상태={r['tonbag_statuses']}"
            })

        # 총 통계
        total_lots = con.execute("SELECT COUNT(*) FROM inventory").fetchone()[0]
        total_tbs = con.execute("SELECT COUNT(*) FROM inventory_tonbag").fetchone()[0]
        con.close()

        return ok_response(data={
            "issues": issues,
            "issue_count": len(issues),
            "status": "PASS" if len(issues) == 0 else "FAIL",
            "checked_at": datetime.now(KST).isoformat(timespec="seconds"),
            "stats": {"total_lots": total_lots, "total_tonbags": total_tbs}
        })
    except Exception as e:
        logger.error("integrity-check error: %s", e)
        return err_response(str(e))


def _engine():
    from backend.api import engine, ENGINE_AVAILABLE
    if not ENGINE_AVAILABLE or engine is None:
        raise RuntimeError("엔진 사용 불가")
    return engine


def _build_integrity_v760_rows(result: dict, engine) -> dict:
    """Shape engine.verify_all_integrity() into v864-2 IntegrityV760 dialog rows."""
    error_lots = {
        r.get("lot_no"): r.get("errors", [])
        for r in result.get("error_lots", [])
        if r.get("lot_no")
    }
    warning_lots = {
        r.get("lot_no"): r.get("warnings", [])
        for r in result.get("warning_lots", [])
        if r.get("lot_no")
    }

    try:
        lot_rows = engine.db.fetchall("SELECT lot_no FROM inventory ORDER BY lot_no")
        all_lots = [r["lot_no"] if isinstance(r, dict) else r[0] for r in lot_rows]
    except Exception as e:
        logger.debug("integrity-report lot list fallback: %s", e)
        all_lots = sorted(set(list(error_lots.keys()) + list(warning_lots.keys())))

    rows = []
    summary = {
        "total_lots": result.get("total_lots") or len(all_lots),
        "error_count": 0,
        "warning_count": 0,
        "ok_count": 0,
        "partial_count": 0,
        "allocation_issue_count": 0,
    }

    for lot_no in all_lots:
        errs = error_lots.get(lot_no, []) or []
        warns = warning_lots.get(lot_no, []) or []
        partial = any("부분 출고 잔류" in m for m in warns)
        alloc_err = any("allocation 초과" in m for m in errs)
        alloc_warn = any("allocation 미완결" in m for m in warns)
        sample_err = any("샘플 무게 오류" in m for m in errs)
        sample_warn = any("샘플 잔류 경고" in m for m in warns)

        severity = "ok"
        if errs:
            severity = "error"
            summary["error_count"] += 1
        elif warns:
            severity = "warning"
            summary["warning_count"] += 1
        else:
            summary["ok_count"] += 1
        if partial:
            summary["partial_count"] += 1
        if alloc_err or alloc_warn:
            summary["allocation_issue_count"] += 1

        rows.append({
            "lot_no": lot_no,
            "sample_ok": "❌ 무게오류" if sample_err else ("⚠️ 잔류" if sample_warn else "✅"),
            "partial_out": "⚠️ 잔류" if partial else "✅",
            "alloc_stat": "❌ 초과" if alloc_err else ("⚠️ 미완결" if alloc_warn else "✅"),
            "errors": errs,
            "warnings": warns,
            "error_text": " | ".join(errs[:3]) + ("…" if len(errs) > 3 else ""),
            "warning_text": " | ".join(warns[:3]) + ("…" if len(warns) > 3 else ""),
            "severity": severity,
        })

    return {
        "summary": summary,
        "rows": rows,
        "valid": result.get("valid", True),
        "checked_at": datetime.now(KST).isoformat(timespec="seconds"),
        "details": result.get("details", {}),
    }


@router.get("/integrity-report", summary="📋 정합성 검증 리포트 v7.6.0")
def integrity_report():
    """v864-2 IntegrityV760Dialog read-only report."""
    try:
        eng = _engine()
        if not hasattr(eng, "verify_all_integrity"):
            return err_response("엔진 verify_all_integrity 없음")
        result = eng.verify_all_integrity()
        return ok_response(data=_build_integrity_v760_rows(result, eng))
    except Exception as e:
        logger.error("integrity-report error: %s", e, exc_info=True)
        return err_response(str(e))


@router.post("/fix-integrity", summary="🛠 LOT 상태 정합성 복구")
def fix_integrity():
    """v864-2 fix_lot_status_integrity action."""
    try:
        eng = _engine()
        if not hasattr(eng, "fix_lot_status_integrity"):
            return err_response("엔진 fix_lot_status_integrity 없음")
        result = eng.fix_lot_status_integrity()
        if result.get("success"):
            return ok_response(data=result, message=result.get("message", "정합성 복구 완료"))
        return err_response(result.get("message") or "; ".join(result.get("errors", [])) or "정합성 복구 실패", data=result)
    except Exception as e:
        logger.error("fix-integrity error: %s", e, exc_info=True)
        return err_response(str(e))


# ── F029: 백업 생성 ───────────────────────────────────────────
@router.post("/backup-create", summary="💾 백업 생성 (F029)")
def backup_create():
    """
    sqm_inventory.db → backup/sqm_backup_YYYYMMDD_HHMMSS.db 복사.
    WAL 파일도 함께 복사. 최대 5개 유지 (오래된 것부터 삭제).
    """
    try:
        root = _project_root()
        db_src = _db_path()
        backup_dir = os.path.join(root, "backup")
        os.makedirs(backup_dir, exist_ok=True)

        ts = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
        dst_name = f"sqm_backup_{ts}.db"
        dst_path = os.path.join(backup_dir, dst_name)

        # WAL checkpoint 먼저 수행 (깨끗한 복사 보장)
        con = sqlite3.connect(db_src, timeout=5)
        con.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        con.close()

        shutil.copy2(db_src, dst_path)
        size_kb = round(os.path.getsize(dst_path) / 1024, 1)

        # WAL 파일도 복사 (있으면)
        for ext in ("-wal", "-shm"):
            src_wal = db_src + ext
            if os.path.exists(src_wal):
                shutil.copy2(src_wal, dst_path + ext)

        # 최대 MAX_BACKUPS 개 유지
        all_backups = sorted(
            [f for f in os.listdir(backup_dir) if f.startswith("sqm_backup_") and f.endswith(".db")],
            key=lambda x: os.path.getmtime(os.path.join(backup_dir, x))
        )
        while len(all_backups) > MAX_BACKUPS:
            old = all_backups.pop(0)
            old_path = os.path.join(backup_dir, old)
            os.remove(old_path)
            for ext in ("-wal", "-shm"):
                if os.path.exists(old_path + ext):
                    os.remove(old_path + ext)

        return ok_response(data={
            "filename": dst_name,
            "path": dst_path,
            "size_kb": size_kb,
            "created_at": datetime.now(KST).isoformat(timespec="seconds"),
            "message": f"✅ 백업 생성 완료: {dst_name} ({size_kb} KB)"
        })
    except Exception as e:
        logger.error("backup-create error: %s", e)
        return err_response(str(e))


# ── 백업 복원 ────────────────────────────────────────────────
@router.post("/restore", summary="🔄 백업 복원")
def restore_backup(body: dict = {}):
    """
    backup/ 폴더에서 선택한 .db 파일로 현재 DB를 교체합니다.
    복원 전 현재 DB를 자동 백업합니다.
    """
    try:
        filename = body.get("filename", "")
        if not filename:
            return err_response("복원할 백업 파일명이 필요합니다")

        root = _project_root()
        backup_dir = os.path.join(root, "backup")
        src_path = os.path.join(backup_dir, filename)

        if not os.path.exists(src_path):
            return err_response(f"백업 파일을 찾을 수 없습니다: {filename}")

        # 안전 검증: backup 디렉토리 내부인지 확인
        if not os.path.abspath(src_path).startswith(os.path.abspath(backup_dir)):
            return err_response("잘못된 파일 경로입니다")

        db_dst = _db_path()

        # 복원 전 현재 DB 자동 백업
        ts = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
        pre_restore_name = f"sqm_pre_restore_{ts}.db"
        pre_restore_path = os.path.join(backup_dir, pre_restore_name)
        os.makedirs(backup_dir, exist_ok=True)

        # WAL checkpoint
        try:
            con = sqlite3.connect(db_dst, timeout=5)
            con.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            con.close()
        except Exception:
            pass

        if os.path.exists(db_dst):
            shutil.copy2(db_dst, pre_restore_path)

        # 복원 실행
        shutil.copy2(src_path, db_dst)
        # WAL/SHM 파일도 복원 (있으면)
        for ext in ("-wal", "-shm"):
            src_wal = src_path + ext
            if os.path.exists(src_wal):
                shutil.copy2(src_wal, db_dst + ext)
            else:
                # 기존 WAL/SHM 삭제 (깨끗한 상태)
                dst_wal = db_dst + ext
                if os.path.exists(dst_wal):
                    os.remove(dst_wal)

        size_kb = round(os.path.getsize(db_dst) / 1024, 1)
        return ok_response(
            data={"filename": filename, "size_kb": size_kb, "pre_backup": pre_restore_name},
            message=f"✅ 복원 완료: {filename} ({size_kb} KB) — 이전 DB는 {pre_restore_name}에 백업됨"
        )
    except Exception as e:
        logger.error("restore error: %s", e)
        return err_response(str(e))


# ── F035: LOT 리스트 Excel 내보내기 ──────────────────────────
@router.get("/export-lot-excel", summary="📊 LOT 리스트 Excel 내보내기 (F035)")
def export_lot_excel():
    """
    inventory 전체 → .xlsx 임시 파일 생성 → FileResponse.
    openpyxl 사용, 헤더 색상 적용.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        con = _db()
        rows = con.execute("""
            SELECT
                lot_no, lot_sqm, sap_no, bl_no, product,
                net_weight, current_weight, tonbag_count,
                status, inbound_date, arrival_date,
                warehouse, vessel, do_no, remarks
            FROM inventory
            ORDER BY inbound_date DESC, lot_no
        """).fetchall()
        con.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LOT 재고현황"

        headers = [
            "LOT NO", "LOT SQM", "SAP NO", "BL NO", "제품명",
            "순중량(kg)", "현재중량(kg)", "톤백수",
            "상태", "입고일", "도착일",
            "창고", "선박", "D/O NO", "비고"
        ]

        # 헤더 스타일
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.row_dimensions[1].height = 22
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # 상태 색상 매핑
        status_fills = {
            "AVAILABLE": PatternFill("solid", fgColor="E8F5E9"),
            "RESERVED":  PatternFill("solid", fgColor="E3F2FD"),
            "PICKED":    PatternFill("solid", fgColor="FFF3E0"),
            "OUTBOUND":  PatternFill("solid", fgColor="F3E5F5"),
            "SOLD":      PatternFill("solid", fgColor="ECEFF1"),
            "RETURNED":  PatternFill("solid", fgColor="FFEBEE"),
        }
        body_font = Font(name="맑은 고딕", size=9)

        for r_idx, row in enumerate(rows, 2):
            data = list(row)
            status = data[8] or ""
            row_fill = status_fills.get(status)
            for c_idx, val in enumerate(data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = body_font
                cell.border = border
                if row_fill:
                    cell.fill = row_fill
                if c_idx in (6, 7):  # 중량 숫자 오른쪽 정렬
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = center_align

        # 열 너비 자동 조정
        col_widths = [16, 14, 12, 14, 22, 13, 13, 8, 12, 12, 12, 12, 14, 14, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 합계 행
        last_row = len(rows) + 2
        ws.cell(row=last_row, column=1, value="합계").font = Font(bold=True, name="맑은 고딕", size=9)
        total_net = sum(r[5] or 0 for r in rows)
        total_cur = sum(r[6] or 0 for r in rows)
        ws.cell(row=last_row, column=6, value=round(total_net, 1)).font = Font(bold=True, name="맑은 고딕", size=9)
        ws.cell(row=last_row, column=7, value=round(total_cur, 1)).font = Font(bold=True, name="맑은 고딕", size=9)

        # 임시 파일 저장
        tmp_dir = tempfile.gettempdir()
        ts = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(tmp_dir, f"SQM_LOT_재고현황_{ts}.xlsx")
        wb.save(out_path)

        return FileResponse(
            path=out_path,
            filename=f"SQM_LOT_재고현황_{ts}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error("export-lot-excel error: %s", e)
        return err_response(str(e))


# ── F050: LOT 상세 ────────────────────────────────────────────
@router.get("/lot-detail/{lot_no}", summary="🔖 LOT 상세 (F050)")
def get_lot_detail(lot_no: str):
    """inventory + inventory_tonbag + snapshot → LOT 완전 상세"""
    try:
        con = _db()

        # 메인 LOT 정보
        lot = con.execute("""
            SELECT * FROM inventory WHERE lot_no = ? LIMIT 1
        """, (lot_no,)).fetchone()
        if not lot:
            return err_response(f"LOT {lot_no} 없음")

        # 톤백 목록
        tonbags = con.execute("""
            SELECT id, sub_lt, tonbag_no, tonbag_uid, weight, status,
                   location, picked_to, picked_date, outbound_date,
                   sale_ref, remarks, created_at
            FROM inventory_tonbag
            WHERE lot_no = ?
            ORDER BY sub_lt
        """, (lot_no,)).fetchall()

        # 입출고 이력
        movements = con.execute("""
            SELECT movement_type, qty_kg,
                   COALESCE(movement_date, created_at) AS movement_date,
                   customer, actor, remarks
            FROM stock_movement
            WHERE lot_no = ?
            ORDER BY COALESCE(movement_date, created_at) DESC
            LIMIT 50
        """, (lot_no,)).fetchall()

        # BL 문서
        bl_doc = con.execute("""
            SELECT bl_no, booking_no, vessel, voyage, ship_date,
                   port_of_loading, port_of_discharge, net_weight_kg, carrier_name
            FROM document_bl WHERE lot_no = ? LIMIT 1
        """, (lot_no,)).fetchone()

        # 컨테이너 정보
        containers = con.execute("""
            SELECT container_no, seal_no, size_type, weight_kg, package_count
            FROM container_info WHERE lot_no = ?
        """, (lot_no,)).fetchall()

        # 톤백 상태 집계
        tb_stats = con.execute("""
            SELECT status, COUNT(*) AS cnt, ROUND(SUM(weight)/1000.0,3) AS mt
            FROM inventory_tonbag WHERE lot_no = ? GROUP BY status
        """, (lot_no,)).fetchall()

        con.close()
        return ok_response(data={
            "lot":       dict(lot),
            "tonbags":   _rows_to_list(tonbags),
            "movements": _rows_to_list(movements),
            "bl_doc":    dict(bl_doc) if bl_doc else None,
            "containers":_rows_to_list(containers),
            "tb_stats":  _rows_to_list(tb_stats),
            "summary": {
                "tonbag_count": len(tonbags),
                "movement_count": len(movements),
            }
        })
    except Exception as e:
        logger.error("lot-detail error for %s: %s", lot_no, e)
        return err_response(str(e))


# ── F061: 시스템 정보 ─────────────────────────────────────────
@router.get("/lot-detail-v760/{lot_no}", summary="LOT Detail v7.6.0 3-tab data")
def get_lot_detail_v760(lot_no: str):
    """v864-2 LOT detail dialog port: tonbags, stock movement, allocation audit."""
    try:
        con = _db()
        lot = con.execute(
            "SELECT * FROM inventory WHERE lot_no = ? LIMIT 1",
            (lot_no,),
        ).fetchone()
        if not lot:
            con.close()
            return err_response(f"LOT {lot_no} not found")
        lot_dict = dict(lot)

        tonbags = con.execute("""
            SELECT id, sub_lt, tonbag_no, tonbag_uid, weight, is_sample, status,
                   location, picked_to, picked_date, outbound_date,
                   sale_ref, remarks, created_at
            FROM inventory_tonbag
            WHERE lot_no = ?
            ORDER BY COALESCE(is_sample, 0) DESC, sub_lt
        """, (lot_no,)).fetchall()

        movements = con.execute("""
            SELECT id, movement_type, qty_kg, sub_lt,
                   COALESCE(movement_date, created_at) AS movement_date,
                   from_location, to_location, customer,
                   ref_table, ref_id, source_type, source_file, actor, remarks
            FROM stock_movement
            WHERE lot_no = ?
            ORDER BY COALESCE(movement_date, created_at) DESC
            LIMIT 200
        """, (lot_no,)).fetchall()

        allocation_plans = []
        allocation_summary = []
        if _table_exists(con, "allocation_plan"):
            plan_cols = _table_columns(con, "allocation_plan")
            want_cols = [
                "id", "tonbag_id", "sub_lt", "customer", "sale_ref", "qty_mt",
                "outbound_date", "status", "workflow_status", "created_at",
                "executed_at", "cancelled_at", "source_file",
            ]
            select_cols = [c for c in want_cols if c in plan_cols]
            if select_cols:
                allocation_plans = con.execute(
                    f"""
                    SELECT {', '.join(select_cols)}
                    FROM allocation_plan
                    WHERE lot_no = ?
                    ORDER BY id DESC
                    LIMIT 300
                    """,
                    (lot_no,),
                ).fetchall()
            if {"status", "qty_mt"}.issubset(plan_cols):
                allocation_summary = con.execute("""
                    SELECT status, COUNT(*) AS cnt,
                           ROUND(COALESCE(SUM(qty_mt), 0), 4) AS qty_mt
                    FROM allocation_plan
                    WHERE lot_no = ?
                    GROUP BY status
                    ORDER BY status
                """, (lot_no,)).fetchall()

        audit_events = []
        if _table_exists(con, "audit_log"):
            audit_events = con.execute("""
                SELECT id, event_type, event_data, batch_id, tonbag_id,
                       user_note, created_by, created_at
                FROM audit_log
                WHERE event_data LIKE ? OR batch_id LIKE ? OR user_note LIKE ?
                ORDER BY created_at DESC
                LIMIT 100
            """, (f"%{lot_no}%", f"%{lot_no}%", f"%{lot_no}%")).fetchall()

        tb_stats = con.execute("""
            SELECT status, COUNT(*) AS cnt, ROUND(SUM(weight)/1000.0,3) AS mt
            FROM inventory_tonbag
            WHERE lot_no = ?
            GROUP BY status
            ORDER BY status
        """, (lot_no,)).fetchall()

        tonbag_items = _rows_to_list(tonbags)
        movement_items = _rows_to_list(movements)
        allocation_items = _rows_to_list(allocation_plans)
        audit_items = _rows_to_list(audit_events)
        initial_weight = float(
            lot_dict.get("initial_weight")
            or lot_dict.get("net_weight")
            or lot_dict.get("current_weight")
            or 0
        )
        current_weight = float(lot_dict.get("current_weight") or 0)
        outbound_weight = max(initial_weight - current_weight, 0)
        available_items = [r for r in tonbag_items if str(r.get("status") or "") == "AVAILABLE"]
        picked_items = [
            r for r in tonbag_items
            if str(r.get("status") or "") in {"PICKED", "SOLD", "SHIPPED", "CONFIRMED"}
        ]

        con.close()
        return ok_response(data={
            "lot": lot_dict,
            "tonbags": tonbag_items,
            "movements": movement_items,
            "allocation_plans": allocation_items,
            "allocation_summary": _rows_to_list(allocation_summary),
            "audit_events": audit_items,
            "tb_stats": _rows_to_list(tb_stats),
            "summary": {
                "tonbag_count": len(tonbag_items),
                "movement_count": len(movement_items),
                "allocation_count": len(allocation_items),
                "audit_count": len(audit_items),
                "sample_count": sum(1 for r in tonbag_items if int(r.get("is_sample") or 0)),
                "available_count": len(available_items),
                "available_kg": sum(float(r.get("weight") or 0) for r in available_items),
                "picked_count": len(picked_items),
                "picked_kg": sum(float(r.get("weight") or 0) for r in picked_items),
                "initial_weight": initial_weight,
                "current_weight": current_weight,
                "outbound_weight": outbound_weight,
                "outbound_percent": (outbound_weight / initial_weight * 100) if initial_weight > 0 else 0,
            },
        })
    except Exception as e:
        logger.error("lot-detail-v760 error for %s: %s", lot_no, e, exc_info=True)
        return err_response(str(e))


@router.get("/system-info", summary="ℹ️ 시스템 정보 (F061)")
def get_system_info():
    """DB 통계 + version.py + 파일 크기 + Python 버전"""
    try:
        import sys as _sys
        import platform

        con = _db()
        db_path = _db_path()

        # DB 통계
        stats = {}
        for tbl in ["inventory", "inventory_tonbag", "stock_movement",
                    "audit_log", "allocation_plan", "inventory_snapshot"]:
            cnt = con.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
            stats[tbl] = cnt

        # DB 파일 크기
        db_size_mb = round(os.path.getsize(db_path) / (1024 * 1024), 2)

        # schema 버전
        schema_ver = con.execute(
            "SELECT version_str FROM schema_version ORDER BY id DESC LIMIT 1"
        ).fetchone()
        con.close()

        ver_info = _load_version()
        root = _project_root()

        return ok_response(data={
            "app":  ver_info,
            "python": f"{_sys.version_info.major}.{_sys.version_info.minor}.{_sys.version_info.micro}",
            "platform": platform.system() + " " + platform.release(),
            "db": {
                "path":          db_path,
                "size_mb":       db_size_mb,
                "schema_version": schema_ver[0] if schema_ver else "unknown",
                "row_counts":    stats,
            },
            "paths": {
                "project_root": root,
                "backup_dir":   os.path.join(root, "backup"),
                "data_dir":     os.path.join(root, "data"),
            },
            "checked_at": datetime.now(KST).isoformat(timespec="seconds"),
        })
    except Exception as e:
        logger.error("system-info error: %s", e)
        return err_response(str(e))
