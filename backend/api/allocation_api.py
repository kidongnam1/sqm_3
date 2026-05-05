# -*- coding: utf-8 -*-
"""
SQM v864.3 — Allocation API (Phase 4-B)
POST /api/allocation/bulk-import-excel : Excel 업로드 → reserve_from_allocation
F014 위치 배정 (Allocation 출고 예약) 네이티브 구현
"""
import logging
import os
import tempfile
from typing import Any

from fastapi import APIRouter, HTTPException, UploadFile, File

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/allocation", tags=["allocation"])


# Allocation 행 컬럼 별칭 매핑
_ALLOC_COLUMN_MAP = {
    "lot_no":        ["lot_no", "lot", "lot no", "lot번호", "로트"],
    "sold_to":       ["sold_to", "sold to", "customer", "고객", "고객사"],
    "sale_ref":      ["sale_ref", "sale ref", "sales_ref", "sales ref", "sc_rcvd", "sc rcvd", "판매참조"],
    "qty_mt":        ["qty_mt", "qty", "quantity", "qty mt", "qty(mt)", "kg", "weight", "수량"],
    "outbound_date": ["outbound_date", "outbound date", "ship_date", "ship date", "출고일", "선적일"],
    "sublot_count":  ["sublot_count", "sublot count", "tonbag_count", "tonbag count", "톤백수"],
    "is_sample":     ["is_sample", "sample", "샘플"],
    "export_type":   ["export_type", "export type", "수출종류"],
    "unit":          ["unit", "단위"],
}


def _match_alloc_columns(df_columns) -> dict:
    """Excel 컬럼 → 표준 키 매핑"""
    result = {}
    lowered = {str(c).strip().lower(): c for c in df_columns}
    for std_key, aliases in _ALLOC_COLUMN_MAP.items():
        for alias in aliases:
            a = alias.strip().lower()
            if a in lowered:
                result[std_key] = lowered[a]
                break
    return result


def _clean_value(v: Any) -> Any:
    try:
        import math
        if v is None:
            return None
        if isinstance(v, float) and math.isnan(v):
            return None
    except Exception:
        pass
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


@router.post("/bulk-import-excel", summary="📍 Allocation 입력 — Excel 업로드 (F014)")
async def bulk_import_allocation(file: UploadFile = File(...)):
    """
    Allocation Excel → 톤백 예약 (AVAILABLE → RESERVED).
    - engine.reserve_from_allocation(rows, source_file) 호출
    - All-or-Nothing 트랜잭션
    """
    if not file.filename:
        raise HTTPException(400, "파일명이 없습니다.")
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(400, f"Excel 파일만 지원 (.xlsx/.xls). 받은 파일: {file.filename}")

    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(500, "pandas 미설치")

    try:
        from backend.api import engine, ENGINE_AVAILABLE
    except Exception as e:
        raise HTTPException(500, f"엔진 로드 실패: {e}")
    if not ENGINE_AVAILABLE or engine is None:
        raise HTTPException(500, "엔진 사용 불가")
    if not hasattr(engine, "reserve_from_allocation"):
        raise HTTPException(500, "엔진에 reserve_from_allocation 메서드 없음")

    tmp_path = None
    try:
        content = await file.read()
        if not content:
            raise HTTPException(400, "빈 파일")

        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        logger.info(f"[allocation-import] 수신: {file.filename} ({len(content)} bytes)")

        # header=0/1/2 자동 감지
        df = None
        header_used = None
        for header_row in (0, 1, 2):
            try:
                candidate = pd.read_excel(tmp_path, header=header_row)
                if candidate.empty:
                    continue
                matched = _match_alloc_columns(candidate.columns)
                if "lot_no" in matched and ("sold_to" in matched or "qty_mt" in matched):
                    df = candidate
                    header_used = header_row
                    break
            except Exception as e:
                logger.debug(f"[allocation-import] header={header_row} 실패: {e}")
                continue
        if df is None or df.empty:
            raise HTTPException(400, "Excel 헤더 인식 실패 — lot_no + sold_to/qty_mt 컬럼 필요")

        col_map = _match_alloc_columns(df.columns)
        logger.info(f"[allocation-import] header={header_used}, {len(df)}행, 매핑: {list(col_map.keys())}")

        # row dict 리스트 생성 + 검증 통계
        rows = []
        skip_no_lot   = 0   # lot_no 없어서 건너뛴 행
        warn_no_qty   = []  # qty_mt 0이하 또는 없음
        warn_no_sold  = []  # sold_to/customer 없음
        total_df_rows = len(df)

        for idx, row in df.iterrows():
            r = {}
            for std_key, orig_col in col_map.items():
                r[std_key] = _clean_value(row[orig_col])
            if not r.get("lot_no"):
                skip_no_lot += 1
                continue  # 빈 lot_no 행은 skip
            # customer 별명: sold_to ↔ customer
            if r.get("sold_to") and not r.get("customer"):
                r["customer"] = r["sold_to"]
            # 검증 경고 수집 (skip하지 않고 계속 진행)
            qty = r.get("qty_mt")
            try:
                qty_f = float(qty) if qty not in (None, "") else 0.0
            except (ValueError, TypeError):
                qty_f = 0.0
            if qty_f <= 0:
                warn_no_qty.append(str(r["lot_no"]))
            if not r.get("sold_to") and not r.get("customer"):
                warn_no_sold.append(str(r["lot_no"]))
            rows.append(r)

        validation_summary = {
            "total_rows": total_df_rows,
            "valid_rows": len(rows),
            "skipped_no_lot": skip_no_lot,
            "warn_no_qty": warn_no_qty[:20],
            "warn_no_sold_to": warn_no_sold[:20],
        }

        if not rows:
            raise HTTPException(400, "유효한 데이터 행이 없습니다 (lot_no 전부 비어있음)")

        # 엔진 호출 — 트랜잭션 내부 처리
        result = engine.reserve_from_allocation(rows, source_file=file.filename)

        success = bool(result.get("success"))
        reserved = int(result.get("reserved", 0))
        errors = result.get("errors", [])
        error_details = result.get("error_details", [])
        plan_ids = result.get("plan_ids", [])

        if success:
            logger.info(
                f"[allocation-import] 완료: {reserved}건 예약 / 에러 {len(errors)}건 ({file.filename})"
            )
            return {
                "ok": True,
                "data": {
                    "filename": file.filename,
                    "total_rows": len(rows),
                    "reserved": reserved,
                    "plan_ids": plan_ids[:50],
                    "header_row": header_used,
                    "matched_columns": list(col_map.keys()),
                    "errors": errors[:20],
                    "error_details": error_details[:20],
                    "warnings": result.get("warnings", [])[:20],
                    "validation_summary": validation_summary,
                },
                "message": f"{reserved}건 Allocation 예약 완료 / 경고 {len(errors)}건",
            }
        else:
            return {
                "ok": False,
                "data": {
                    "filename": file.filename,
                    "total_rows": len(rows),
                    "reserved": reserved,
                    "errors": errors[:20],
                    "error_details": error_details[:20],
                    "validation_summary": validation_summary,
                },
                "error": "Allocation 예약 실패",
                "detail": {"code": "ALLOCATION_FAILED", "errors": errors[:10]},
                "message": "Allocation 실패 — 전체 롤백",
            }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[allocation-import] 예기치 않은 에러: {e}")
        raise HTTPException(500, f"Internal error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


# ────────────────────────────────────────────────────────────
# F022 예약 반영 (승인분) — engine.apply_approved_allocation_reservations()
# ────────────────────────────────────────────────────────────
@router.post("/apply-approved", summary="📌 예약 반영 — 승인분 실행 (F022)")
def apply_approved_allocation():
    """
    workflow_status=APPROVED 상태의 allocation_plan 을 RESERVED 로 실제 반영.
    """
    try:
        from backend.api import engine, ENGINE_AVAILABLE
    except Exception as e:
        raise HTTPException(500, f"엔진 로드 실패: {e}")
    if not ENGINE_AVAILABLE or engine is None:
        raise HTTPException(500, "엔진 사용 불가")
    if not hasattr(engine, "apply_approved_allocation_reservations"):
        raise HTTPException(500, "엔진에 apply_approved_allocation_reservations 메서드 없음")

    try:
        result = engine.apply_approved_allocation_reservations()
    except Exception as e:
        logger.exception(f"[apply-approved] 에러: {e}")
        raise HTTPException(500, f"Engine error: {e}")

    if result.get("success"):
        applied = int(result.get("applied", 0))
        return {
            "ok": True,
            "data": {"applied": applied, "errors": result.get("errors", [])[:20]},
            "message": f"{applied}건 승인 예약 반영 완료",
        }
    else:
        return {
            "ok": False,
            "data": {"applied": int(result.get("applied", 0)), "errors": result.get("errors", [])},
            "error": "예약 반영 실패",
            "detail": {"code": "APPLY_APPROVED_FAILED", "errors": result.get("errors", [])[:10]},
            "message": "; ".join(result.get("errors", []))[:200] or "예약 반영 실패",
        }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Stage 2/3: Allocation approve / reject / PATCH inline edit
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
import sqlite3 as _sqlite3
from fastapi import Body, Path as PathParam

def _alloc_db():
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(os.path.dirname(here))
    db_path = os.path.join(root, "data", "db", "sqm_inventory.db")
    con = _sqlite3.connect(db_path, timeout=5, check_same_thread=False)
    con.row_factory = _sqlite3.Row
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA busy_timeout=3000")
    return con


@router.post("/approve", summary="✅ 할당 승인 (Stage 2)")
def approve_allocation(data: dict = Body(...)):
    ids = data.get("ids") or []
    actor = data.get("actor") or "system"
    reason = data.get("reason") or "승인"
    if not ids:
        raise HTTPException(400, "ids 필요")
    try:
        con = _alloc_db()
        updated = 0
        for plan_id in ids:
            cur = con.execute(
                "UPDATE allocation_plan SET approval_status='APPROVED', updated_at=datetime('now') WHERE id=?",
                (plan_id,)
            )
            updated += cur.rowcount
            try:
                con.execute(
                    "INSERT INTO allocation_approval (allocation_plan_id, status, actor, reason, created_at) VALUES (?,?,?,?,datetime('now'))",
                    (plan_id, "APPROVED", actor, reason)
                )
            except Exception:
                pass  # allocation_approval 테이블 없으면 무시
        con.commit(); con.close()
        return {"ok": True, "message": f"{updated}건 승인됨", "data": {"updated": updated}}
    except Exception as e:
        raise HTTPException(500, str(e))


@router.post("/reject", summary="❌ 할당 반려 (Stage 2)")
def reject_allocation(data: dict = Body(...)):
    ids = data.get("ids") or []
    actor = data.get("actor") or "system"
    reason = data.get("reason") or "반려"
    if not ids:
        raise HTTPException(400, "ids 필요")
    try:
        con = _alloc_db()
        updated = 0
        for plan_id in ids:
            cur = con.execute(
                "UPDATE allocation_plan SET approval_status='REJECTED', updated_at=datetime('now') WHERE id=?",
                (plan_id,)
            )
            updated += cur.rowcount
            try:
                con.execute(
                    "INSERT INTO allocation_approval (allocation_plan_id, status, actor, reason, created_at) VALUES (?,?,?,?,datetime('now'))",
                    (plan_id, "REJECTED", actor, reason)
                )
            except Exception:
                pass
        con.commit(); con.close()
        return {"ok": True, "message": f"{updated}건 반려됨", "data": {"updated": updated}}
    except Exception as e:
        raise HTTPException(500, str(e))


ALLOC_EDITABLE_FIELDS = {"qty_mt", "customer", "sale_ref", "outbound_date", "remarks"}

@router.patch("/{lot_no}", summary="인라인 편집 (Stage 3)")
def patch_allocation(lot_no: str = PathParam(...), data: dict = Body(...)):
    fields_to_update = {k: v for k, v in data.items() if k in ALLOC_EDITABLE_FIELDS}
    if not fields_to_update:
        raise HTTPException(400, f"허용된 편집 필드 없음. 허용: {sorted(ALLOC_EDITABLE_FIELDS)}")
    try:
        con = _alloc_db()
        plan_row = con.execute("SELECT id FROM allocation_plan WHERE lot_no=? LIMIT 1", (lot_no,)).fetchone()
        updated = 0
        if plan_row:
            set_clauses = ", ".join(f"{f}=?" for f in fields_to_update)
            vals = list(fields_to_update.values()) + [lot_no]
            cur = con.execute(
                f"UPDATE allocation_plan SET {set_clauses}, updated_at=datetime('now') WHERE lot_no=?", vals
            )
            updated += cur.rowcount
        INV_FIELD_MAP = {"customer": "sold_to", "sale_ref": "sale_ref", "outbound_date": "ship_date", "remarks": "remarks"}
        inv_fields = {INV_FIELD_MAP[f]: v for f, v in fields_to_update.items() if f in INV_FIELD_MAP}
        if inv_fields:
            set_clauses_inv = ", ".join(f"{f}=?" for f in inv_fields)
            con.execute(
                f"UPDATE inventory SET {set_clauses_inv}, updated_at=datetime('now') WHERE lot_no=?",
                list(inv_fields.values()) + [lot_no]
            )
        con.commit(); con.close()
        return {"success": True, "lot_no": lot_no, "updated_fields": list(fields_to_update.keys()), "allocation_rows": updated}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# POST /api/allocation/cancel-by-sale-ref  — SALE REF 일괄 취소
# v864-2 AllocationTabMixin._on_allocation_cancel_by_sale_ref 포팅
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@router.post("/cancel-by-sale-ref", summary="🔖 SALE REF 일괄 취소")
def cancel_by_sale_ref(data: dict = Body(...)):
    """
    sale_ref 기준으로 해당 배정 전체를 CANCELLED 처리하고
    inventory 상태를 AVAILABLE 로 원복.
    """
    sale_ref = (data.get("sale_ref") or "").strip()
    if not sale_ref:
        raise HTTPException(400, "sale_ref 값이 필요합니다")
    try:
        con = _alloc_db()
        rows = con.execute(
            "SELECT DISTINCT lot_no FROM allocation_plan WHERE sale_ref=? AND status NOT IN ('CANCELLED','SOLD')",
            (sale_ref,)
        ).fetchall()
        if not rows:
            con.close()
            return {"ok": False, "message": f"SALE REF '{sale_ref}' 에 해당하는 배정이 없거나 이미 취소됨"}

        lot_list = [r[0] for r in rows]
        cur = con.execute(
            "UPDATE allocation_plan SET status='CANCELLED', cancelled_at=datetime('now') "
            "WHERE sale_ref=? AND status NOT IN ('CANCELLED','SOLD')",
            (sale_ref,)
        )
        cancelled_plans = cur.rowcount
        for lot_no in lot_list:
            con.execute(
                "UPDATE inventory SET status='AVAILABLE', sold_to=NULL, sale_ref=NULL "
                "WHERE lot_no=? AND status NOT IN ('SOLD','OUTBOUND')",
                (lot_no,)
            )
        con.commit(); con.close()
        logger.info(f"[cancel-by-sale-ref] sale_ref={sale_ref}, lots={lot_list}, plans={cancelled_plans}")
        return {
            "ok": True,
            "message": f"SALE REF '{sale_ref}': {len(lot_list)} LOT / {cancelled_plans}건 취소됨",
            "data": {"sale_ref": sale_ref, "lots": lot_list, "cancelled_plans": cancelled_plans},
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[cancel-by-sale-ref] error: {e}")
        raise HTTPException(500, str(e))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# POST /api/allocation/reset-all  — 전체 초기화
# v864-2 AllocationTabMixin._on_allocation_reset_all 포팅
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@router.post("/reset-all", summary="🧹 전체 초기화 — 모든 배정 취소 + AVAILABLE 원복")
def reset_all_allocations():
    """
    RESERVED / PICKED / OUTBOUND 상태의 allocation_plan 행을 전부 CANCELLED 처리하고
    inventory 상태를 AVAILABLE 로 원복. SOLD 는 보호.
    """
    try:
        con = _alloc_db()
        rows = con.execute(
            "SELECT DISTINCT lot_no FROM allocation_plan WHERE status NOT IN ('CANCELLED','SOLD')"
        ).fetchall()
        lot_list = [r[0] for r in rows]
        cur = con.execute(
            "UPDATE allocation_plan SET status='CANCELLED', cancelled_at=datetime('now') "
            "WHERE status NOT IN ('CANCELLED','SOLD')"
        )
        cancelled_plans = cur.rowcount
        for lot_no in lot_list:
            con.execute(
                "UPDATE inventory SET status='AVAILABLE', sold_to=NULL, sale_ref=NULL "
                "WHERE lot_no=? AND status NOT IN ('SOLD','OUTBOUND')",
                (lot_no,)
            )
        con.commit(); con.close()
        logger.info(f"[reset-all] lots={len(lot_list)}, plans={cancelled_plans}")
        return {
            "ok": True,
            "message": f"전체 초기화 완료: {len(lot_list)} LOT / {cancelled_plans}건 배정 취소",
            "data": {"lots_affected": len(lot_list), "cancelled_plans": cancelled_plans},
        }
    except Exception as e:
        logger.exception(f"[reset-all] error: {e}")
        raise HTTPException(500, str(e))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# POST /api/allocation/revert-step  — 단계 되돌리기
# v864-2 AllocationTabMixin._on_revert_step 포팅
# RESERVED → AVAILABLE / PICKED → RESERVED / OUTBOUND → PICKED
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_REVERT_MAP = {
    "RESERVED": ("RESERVED",  "AVAILABLE"),
    "PICKED":   ("PICKED",    "RESERVED"),
    "OUTBOUND": ("OUTBOUND",  "PICKED"),
}

@router.post("/revert-step", summary="↩️ 단계 되돌리기 (RESERVED→AVAILABLE 등)")
def revert_allocation_step(data: dict = Body(...)):
    """
    from_status 에 따라 선택된 lot_no 목록(또는 전체)을 한 단계 되돌린다.
    - RESERVED  → AVAILABLE (allocation_plan CANCELLED + inventory AVAILABLE)
    - PICKED    → RESERVED  (allocation_plan RESERVED + inventory RESERVED)
    - OUTBOUND  → PICKED    (allocation_plan PICKED   + inventory PICKED)
    """
    from_status = (data.get("from_status") or "").upper().strip()
    lot_nos = data.get("lot_nos") or []   # 빈 리스트 = 전체

    if from_status not in _REVERT_MAP:
        raise HTTPException(400, f"from_status 는 {list(_REVERT_MAP.keys())} 중 하나여야 합니다")

    src_status, dst_status = _REVERT_MAP[from_status]
    try:
        con = _alloc_db()
        if lot_nos:
            placeholders = ",".join("?" * len(lot_nos))
            rows = con.execute(
                f"SELECT DISTINCT lot_no FROM allocation_plan "
                f"WHERE status=? AND lot_no IN ({placeholders})",
                [src_status] + lot_nos
            ).fetchall()
        else:
            rows = con.execute(
                "SELECT DISTINCT lot_no FROM allocation_plan WHERE status=?",
                (src_status,)
            ).fetchall()

        affected_lots = [r[0] for r in rows]
        if not affected_lots:
            con.close()
            return {"ok": False, "message": f"{src_status} 상태인 배정 없음"}

        if dst_status == "AVAILABLE":
            cur = con.execute(
                f"UPDATE allocation_plan SET status='CANCELLED', cancelled_at=datetime('now') "
                f"WHERE status='{src_status}' AND lot_no IN ({','.join('?' * len(affected_lots))})",
                affected_lots
            )
            for lot_no in affected_lots:
                con.execute(
                    "UPDATE inventory SET status='AVAILABLE', sold_to=NULL, sale_ref=NULL "
                    "WHERE lot_no=? AND status=?", (lot_no, src_status)
                )
        else:
            cur = con.execute(
                f"UPDATE allocation_plan SET status=?, updated_at=datetime('now') "
                f"WHERE status=? AND lot_no IN ({','.join('?' * len(affected_lots))})",
                [dst_status, src_status] + affected_lots
            )
            for lot_no in affected_lots:
                con.execute(
                    "UPDATE inventory SET status=? WHERE lot_no=? AND status=?",
                    (dst_status, lot_no, src_status)
                )
        changed = cur.rowcount
        con.commit(); con.close()
        logger.info(f"[revert-step] {src_status}→{dst_status}: {len(affected_lots)} lots, {changed} rows")
        return {
            "ok": True,
            "message": f"{src_status} → {dst_status}: {len(affected_lots)} LOT 되돌리기 완료",
            "data": {"from": src_status, "to": dst_status, "lots": affected_lots, "rows_changed": changed},
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[revert-step] error: {e}")
        raise HTTPException(500, str(e))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GET /api/allocation/export-excel  — Excel 내보내기
# v864-2 AllocationTabMixin._on_allocation_export_excel 포팅
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@router.get("/export-excel", summary="📊 Allocation Excel 내보내기")
def export_allocation_excel():
    """
    현재 RESERVED/PICKED 상태 배정 전체를 Excel 파일로 반환.
    """
    import io
    from fastapi.responses import StreamingResponse
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl 미설치")

    try:
        con = _alloc_db()
        rows = con.execute("""
            SELECT ap.lot_no, ap.sub_lt, ap.customer, ap.sale_ref,
                   ap.qty_mt, ap.outbound_date, ap.status, ap.created_at,
                   i.sap_no, i.product, i.warehouse
            FROM allocation_plan ap
            LEFT JOIN inventory i ON ap.lot_no = i.lot_no
            WHERE ap.status NOT IN ('CANCELLED')
            ORDER BY ap.status, ap.lot_no, ap.sub_lt
        """).fetchall()
        con.close()
    except Exception as e:
        raise HTTPException(500, f"DB 오류: {e}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation"

    headers = ["LOT NO", "Sub LOT", "고객사", "SALE REF", "수량(MT)", "출고예정일", "상태", "등록일시", "SAP NO", "제품", "창고"]
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", name="맑은 고딕")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in rows:
        ws.append(list(row))

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 32)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime as _dt
    fname = f"ALLOCATION_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GET /api/allocation/open-picked-excel — Picked 리스트 Excel 저장+열기
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _alloc_exports_dir() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(os.path.dirname(here))
    path = os.path.join(root, "exports")
    os.makedirs(path, exist_ok=True)
    return path


@router.get("/open-picked-excel", summary="📂 Picked 리스트 Excel 저장 후 바로 열기")
def open_picked_excel():
    """
    PICKED 상태 LOT 목록 Excel 생성 → exports/ 폴더 저장 → os.startfile() 열기.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime as _dt
    except ImportError:
        raise HTTPException(500, "openpyxl 미설치")

    try:
        con = _alloc_db()
        rows = con.execute("""
            SELECT i.lot_no, i.lot_sqm, i.sap_no, i.product,
                   i.tonbag_count, i.current_weight,
                   i.status, i.inbound_date,
                   ap.customer, ap.sale_ref, ap.outbound_date
            FROM inventory i
            LEFT JOIN allocation_plan ap ON ap.lot_no = i.lot_no AND ap.status = 'PICKED'
            WHERE i.status = 'PICKED'
            ORDER BY i.lot_no
        """).fetchall()
        con.close()
    except Exception as e:
        raise HTTPException(500, f"DB 오류: {e}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판매화물 결정 LOT"

    headers = ["LOT NO", "LOT SQM", "SAP NO", "제품", "톤백수", "중량(kg)",
               "상태", "입고일", "고객사", "SALE REF", "출고예정일"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="맑은 고딕")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center

    picked_fill = PatternFill("solid", fgColor="FFF9C4")
    for row in rows:
        ws.append(list(row))
        for cell in ws[ws.max_row]:
            cell.fill = picked_fill
            cell.alignment = center

    widths = [16, 14, 12, 22, 8, 12, 10, 12, 20, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    exports = _alloc_exports_dir()
    ts = _dt.now().strftime("%Y%m%d_%H%M%S")
    fname = f"판매화물결정LOT_{ts}.xlsx"
    out_path = os.path.join(exports, fname)
    wb.save(out_path)

    try:
        os.startfile(out_path)
    except Exception as open_err:
        logger.warning("os.startfile 실패: %s", open_err)

    logger.info("Picked LOT Excel 저장+열기: %s (%d LOT)", fname, len(rows))
    from backend.common.errors import ok_response
    return ok_response({"filename": fname, "path": out_path, "rows": len(rows), "opened": True})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GET /api/allocation/lot-overview  — LOT 현황
# v864-2 AllocationTabMixin._on_open_allocation_lot_overview 포팅
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@router.get("/lot-overview", summary="📦 LOT 현황 (배정 요약)")
def get_allocation_lot_overview():
    """
    LOT별 배정 상태 요약: 일반 배정량(MT) / 샘플(1kg) LOT 여부 / 배정 잔여량.
    """
    try:
        con = _alloc_db()
        rows = con.execute("""
            SELECT
                i.lot_no,
                i.sap_no,
                i.product,
                COALESCE(i.net_weight, 0) / 1000.0        AS net_mt,
                COALESCE(i.current_weight, 0) / 1000.0    AS balance_mt,
                COALESCE(SUM(CASE WHEN ap.status NOT IN ('CANCELLED','SOLD')
                               THEN ap.qty_mt ELSE 0 END), 0) AS alloc_mt,
                COUNT(CASE WHEN COALESCE(t.is_sample,0)=1 THEN 1 END) AS sample_bags,
                i.status AS lot_status
            FROM inventory i
            LEFT JOIN allocation_plan ap ON ap.lot_no = i.lot_no
            LEFT JOIN inventory_tonbag t ON t.lot_no = i.lot_no
            WHERE i.status NOT IN ('SOLD','CANCELLED','OUTBOUND','DEPLETED')
            GROUP BY i.lot_no
            ORDER BY i.lot_no
        """).fetchall()
        con.close()
        result = []
        for r in rows:
            net_mt = round(float(r[3] or 0), 4)
            balance_mt = round(float(r[4] or 0), 4)
            alloc_mt = round(float(r[5] or 0), 4)
            result.append({
                "lot_no":      r[0],
                "sap_no":      r[1],
                "product":     r[2],
                "net_mt":      net_mt,
                "balance_mt":  balance_mt,
                "alloc_mt":    alloc_mt,
                "remain_mt":   round(balance_mt - alloc_mt, 4),
                "sample_bags": r[6] or 0,
                "lot_status":  r[7],
            })
        return {"ok": True, "data": result, "count": len(result)}
    except Exception as e:
        logger.exception(f"[lot-overview] error: {e}")
        raise HTTPException(500, str(e))


# ──────────────────────────────────────────────────────────────
# Allocation 양식 가져오기 (v864.2 _mapping_from_excel 포팅)
# POST /api/allocation/template-upload
# ──────────────────────────────────────────────────────────────
import json as _json
import shutil as _shutil
import re as _re
from pathlib import Path as _Path

def _alloc_template_dir() -> _Path:
    """resources/templates/allocation/ 절대경로 반환."""
    root = _Path(__file__).resolve().parents[2]
    d = root / 'resources' / 'templates' / 'allocation'
    d.mkdir(parents=True, exist_ok=True)
    return d

def _safe_template_id(value: str) -> str:
    base = _re.sub(r'[^A-Za-z0-9_.-]+', '_', str(value or '').strip()).strip('._')
    return (base or 'allocation_template')[:80]

def _find_header_row(filepath: _Path):
    """openpyxl로 PRODUCT+LOT 기준 헤더 행 탐색. (sheet, header_row_1based, columns) 반환."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    best = (wb.sheetnames[0], 1, [])
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 12) + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            norm = [str(v).strip() for v in vals if v is not None and str(v).strip()]
            upper = ' '.join(norm).upper()
            if 'PRODUCT' in upper and 'LOT' in upper:
                wb.close()
                return ws.title, r, norm
            if 'LOT' in upper and ('QTY' in upper or 'SAP' in upper):
                best = (ws.title, r, norm)
    wb.close()
    return best

@router.post("/template-upload", summary="📥 Allocation 양식 가져오기 (xlsx → json+xlsx 저장)")
async def template_upload(
    file: UploadFile = File(...),
    label: str = "",
    action: str = "check"
):
    """
    고객사 Allocation xlsx 파일을 업로드하면:
    1. 헤더 행 자동 탐지 (PRODUCT + LOT 기준)
    2. 매핑 JSON 생성 → resources/templates/allocation/<id>.json 저장
    3. xlsx 원본 → resources/templates/allocation/<id>.xlsx 저장
    반환: {ok, id, tab_label, columns, sheet, header_row, overwritten}
    """
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "xlsx 또는 xls 파일만 업로드 가능합니다.")

    # 임시 파일에 저장
    import tempfile
    suffix = _Path(file.filename).suffix
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        contents = await file.read()
        tmp.write(contents)
        tmp.close()
        tmp_path = _Path(tmp.name)

        # 헤더 탐지 + 매핑 생성
        sheet, header_row, columns = _find_header_row(tmp_path)
        stem = label.strip() or _Path(file.filename).stem
        file_id = _safe_template_id(stem)
        mapping = {
            'id': file_id,
            'tab_label': f'📄 {stem}',
            'template_file': f'{file_id}.xlsx',
            'sheet': sheet,
            'header_row': header_row,
            'data_start_row': header_row + 1,
            'description': (
                f'외부 Allocation 양식 │ 시트: {sheet} │ '
                f'헤더: {header_row}행 │ 컬럼: {" / ".join(columns)}'
            ),
            'columns': columns,
            'sample_rule': {
                'field': 'Product',
                'contains': 'sample',
                'qty_mt_max': 0.01,
            },
        }

        # 저장
        base = _alloc_template_dir()
        dst_json = base / f'{file_id}.json'
        dst_xlsx = base / f'{file_id}.xlsx'
        overwritten = dst_json.exists() or dst_xlsx.exists()

        # action=check → 분석만, 저장 안 함
        if action == "check":
            return {
                "ok": True,
                "id": file_id,
                "tab_label": mapping['tab_label'],
                "columns": columns,
                "sheet": sheet,
                "header_row": header_row,
                "duplicate": overwritten,
                "message": "분석 완료 (미저장)"
            }

        # action=keep_both → ID 중복 시 _2, _3 ... 접미사
        if action == "keep_both" and overwritten:
            for i in range(2, 99):
                new_id = f"{file_id}_{i}"
                if not (base / f"{new_id}.json").exists():
                    file_id = new_id
                    mapping['id'] = file_id
                    mapping['tab_label'] = mapping['tab_label'] + f" ({i})"
                    dst_json = base / f"{file_id}.json"
                    dst_xlsx = base / f"{file_id}.xlsx"
                    overwritten = False
                    break

        # action=overwrite 또는 keep_both(새 ID) → 저장
        dst_json.write_text(
            _json.dumps(mapping, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )
        _shutil.copy2(str(tmp_path), str(dst_xlsx))

        logger.info("[template-upload] 저장 완료: %s (헤더=%d행, 컬럼=%d개)", file_id, header_row, len(columns))
        return {
            "ok": True,
            "id": file_id,
            "tab_label": mapping['tab_label'],
            "columns": columns,
            "sheet": sheet,
            "header_row": header_row,
            "overwritten": overwritten,
            "message": f"양식 {'덮어쓰기' if overwritten else '저장'} 완료: {file_id}"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[template-upload] error: %s", e)
        raise HTTPException(500, f"양식 분석 실패: {e}")
    finally:
        try:
            import os as _os
            _os.unlink(tmp.name)
        except Exception:
            pass


# ── Allocation 양식 목록 조회 ──────────────────────────────────
@router.get("/template-list", summary="📋 등록된 Allocation 양식 목록")
async def template_list():
    """resources/templates/allocation/ 의 .json 파일 목록 반환."""
    base = _alloc_template_dir()
    templates = []
    for jf in sorted(base.glob("*.json")):
        try:
            data = _json.loads(jf.read_text(encoding='utf-8'))
            templates.append({
                "id":         data.get("id", jf.stem),
                "tab_label":  data.get("tab_label", jf.stem),
                "columns":    data.get("columns", []),
                "sheet":      data.get("sheet", ""),
                "header_row": data.get("header_row", 1),
            })
        except Exception:
            pass
    return {"ok": True, "templates": templates}


# ── Allocation 양식 삭제 ───────────────────────────────────────
@router.delete("/template/{template_id}", summary="🗑️ Allocation 양식 삭제")
async def template_delete(template_id: str):
    """지정한 ID의 .json + .xlsx 파일 삭제."""
    safe_id = _safe_template_id(template_id)
    base = _alloc_template_dir()
    deleted = []
    for ext in ('.json', '.xlsx'):
        f = base / f"{safe_id}{ext}"
        if f.exists():
            f.unlink()
            deleted.append(f.name)
    if not deleted:
        raise HTTPException(404, f"양식을 찾을 수 없습니다: {safe_id}")
    logger.info("[template-delete] 삭제: %s", safe_id)
    return {"ok": True, "id": safe_id, "deleted": deleted}
