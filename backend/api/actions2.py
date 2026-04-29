"""
SQM v865 — Phase 4-B + MISS-02/03 액션 엔드포인트
actions2.py: inbound-cancel, inventory-move, allocate,
             outbound-confirm, export-tonbag-excel,
             sales-order-upload (MISS-02), swap-report (MISS-03)

모두 DB 직접 조작 (engine 우회) — Rule 4: try/except 의무
"""
import csv
import io
import os
import sqlite3
import logging
import shutil
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, HTTPException, Query as QP, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from backend.common.errors import ok_response, err_response

router = APIRouter(prefix="/api/action2", tags=["actions2"])
logger = logging.getLogger(__name__)


# ── 공통 헬퍼 ────────────────────────────────────────────────────
def _db_path() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(os.path.dirname(here))
    return os.path.join(root, "data", "db", "sqm_inventory.db")


def _db() -> sqlite3.Connection:
    con = sqlite3.connect(_db_path(), timeout=5, check_same_thread=False)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA busy_timeout=3000")
    return con


def _root() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.dirname(os.path.dirname(here))


# ── F003: 입고 취소 ─────────────────────────────────────────────
@router.post("/inbound-cancel", summary="↩️ 입고 취소 (F003-alt)")
def cancel_inbound(payload: dict):
    """
    payload: { lot_no: str, reason: str (optional) }
    입고된 LOT을 CANCELLED 상태로 변경 + audit_log 기록
    """
    lot_no = (payload.get("lot_no") or "").strip()
    reason = (payload.get("reason") or "사용자 취소").strip()
    if not lot_no:
        raise HTTPException(400, "lot_no 필수")

    try:
        con = _db()
        row = con.execute(
            "SELECT id, status FROM inventory WHERE lot_no=?", (lot_no,)
        ).fetchone()

        if not row:
            con.close()
            return err_response(f"LOT '{lot_no}' 을(를) 찾을 수 없습니다")

        old_status = row["status"]
        if old_status in ("OUTBOUND", "CANCELLED"):
            con.close()
            return err_response(f"'{old_status}' 상태는 취소 불가 (이미 출고/취소됨)")

        inv_id = row["id"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        con.execute(
            "UPDATE inventory SET status='CANCELLED', updated_at=? WHERE id=?",
            (ts, inv_id)
        )
        con.execute(
            "UPDATE inventory_tonbag SET status='CANCELLED', updated_at=? WHERE inventory_id=?",
            (ts, inv_id)
        )
        # audit_log 기록
        con.execute("""
            INSERT INTO audit_log (event_type, event_data, user_note, created_by, created_at)
            VALUES ('INBOUND_CANCEL', ?, ?, 'system', ?)
        """, (
            f'{{"lot_no":"{lot_no}","old_status":"{old_status}"}}',
            reason, ts
        ))
        # stock_movement 기록
        con.execute("""
            INSERT INTO stock_movement
                (lot_no, movement_type, qty_kg, source_type, actor, remarks, created_at)
            VALUES (?, 'CANCEL', 0, 'MANUAL', 'user', ?, ?)
        """, (lot_no, f"입고취소: {reason}", ts))

        con.commit()
        con.close()
        return ok_response(data={
            "lot_no": lot_no,
            "old_status": old_status,
            "new_status": "CANCELLED",
            "message": f"{lot_no} 입고 취소 완료",
        })
    except Exception as e:
        logger.error("inbound-cancel error: %s", e)
        return err_response(str(e))


# ── F006: 위치 이동 (톤백 단위) ─────────────────────────────────
@router.post("/inventory-move", summary="🔀 위치 이동 (F006-alt)")
def inventory_move(payload: dict):
    """
    payload: { lot_no: str, from_loc: str (optional), to_loc: str }
    lot_no 산하 모든 톤백 위치 일괄 이동
    """
    lot_no = (payload.get("lot_no") or "").strip()
    to_loc = (payload.get("to_loc") or payload.get("destination") or "").strip()
    from_loc = (payload.get("from_loc") or "").strip() or None

    if not lot_no or not to_loc:
        raise HTTPException(400, "lot_no, to_loc 필수")

    try:
        con = _db()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # LOT 존재 확인
        row = con.execute("SELECT id FROM inventory WHERE lot_no=?", (lot_no,)).fetchone()
        if not row:
            con.close()
            return err_response(f"LOT '{lot_no}' 없음")

        # 톤백 위치 업데이트
        if from_loc:
            updated = con.execute("""
                UPDATE inventory_tonbag
                SET location=?, location_updated_at=?, updated_at=?
                WHERE inventory_id=? AND (location=? OR location IS NULL)
            """, (to_loc, ts, ts, row["id"], from_loc)).rowcount
        else:
            updated = con.execute("""
                UPDATE inventory_tonbag
                SET location=?, location_updated_at=?, updated_at=?
                WHERE inventory_id=?
            """, (to_loc, ts, ts, row["id"])).rowcount

        # inventory.location 업데이트
        con.execute(
            "UPDATE inventory SET location=?, updated_at=? WHERE id=?",
            (to_loc, ts, row["id"])
        )

        # stock_movement 기록
        con.execute("""
            INSERT INTO stock_movement
                (lot_no, movement_type, qty_kg, from_location, to_location,
                 source_type, actor, remarks, created_at)
            VALUES (?, 'MOVE', 0, ?, ?, 'MANUAL', 'user', ?, ?)
        """, (lot_no, from_loc or "", to_loc,
              f"위치이동: {from_loc or '(없음)'} → {to_loc}", ts))

        con.commit()
        con.close()
        return ok_response(data={
            "lot_no": lot_no,
            "from_location": from_loc,
            "to_location": to_loc,
            "tonbags_updated": updated,
            "message": f"{lot_no} → {to_loc} 이동 완료 ({updated}개 톤백)",
        })
    except Exception as e:
        logger.error("inventory-move error: %s", e)
        return err_response(str(e))


# ── 위치 배정 (LOT 단위) ─────────────────────────────────────────
@router.post("/allocate", summary="📍 위치 배정 (F004-alt)")
def allocate_location(payload: dict):
    """
    payload: { lot_no: str, location: str }
    위치 미배정(NULL) LOT에 창고 위치 배정
    """
    lot_no   = (payload.get("lot_no") or "").strip()
    location = (payload.get("location") or "").strip()
    if not lot_no or not location:
        raise HTTPException(400, "lot_no, location 필수")

    try:
        con = _db()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row = con.execute("SELECT id, location FROM inventory WHERE lot_no=?", (lot_no,)).fetchone()
        if not row:
            con.close()
            return err_response(f"LOT '{lot_no}' 없음")

        old_loc = row["location"]
        con.execute(
            "UPDATE inventory SET location=?, updated_at=? WHERE id=?",
            (location, ts, row["id"])
        )
        con.execute("""
            UPDATE inventory_tonbag
            SET location=?, location_updated_at=?, updated_at=?
            WHERE inventory_id=? AND location IS NULL
        """, (location, ts, ts, row["id"]))

        # audit_log
        con.execute("""
            INSERT INTO audit_log (event_type, event_data, user_note, created_by, created_at)
            VALUES ('LOCATION_ASSIGN', ?, ?, 'system', ?)
        """, (
            f'{{"lot_no":"{lot_no}","location":"{location}"}}',
            f"위치배정: {old_loc or '없음'} → {location}", ts
        ))
        con.commit()
        con.close()
        return ok_response(data={
            "lot_no": lot_no,
            "old_location": old_loc,
            "new_location": location,
            "message": f"{lot_no} → '{location}' 배정 완료",
        })
    except Exception as e:
        logger.error("allocate error: %s", e)
        return err_response(str(e))


# ── 출고 확정 (LOT 단위) ─────────────────────────────────────────
@router.post("/outbound-confirm", summary="✅ 출고 확정 (F022-alt)")
def outbound_confirm(payload: dict):
    """
    payload: { lot_no: str, customer: str (optional) }
    PICKED/RESERVED → OUTBOUND 상태 전환 + 이동 기록
    """
    lot_no   = (payload.get("lot_no") or "").strip()
    customer = (payload.get("customer") or "").strip()
    if not lot_no:
        raise HTTPException(400, "lot_no 필수")

    try:
        con = _db()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row = con.execute(
            "SELECT id, status, current_weight, product FROM inventory WHERE lot_no=?",
            (lot_no,)
        ).fetchone()

        if not row:
            con.close()
            return err_response(f"LOT '{lot_no}' 없음")

        if row["status"] == "OUTBOUND":
            con.close()
            return err_response(f"{lot_no} 이미 출고 완료 상태")

        if row["status"] not in ("PICKED", "RESERVED", "AVAILABLE"):
            con.close()
            return err_response(f"'{row['status']}' 상태는 출고 확정 불가")

        weight_kg = row["current_weight"] or 0
        con.execute(
            "UPDATE inventory SET status='OUTBOUND', sold_to=?, updated_at=? WHERE id=?",
            (customer or row["sold_to"], ts, row["id"])
        )
        con.execute("""
            UPDATE inventory_tonbag
            SET status='OUTBOUND', outbound_date=?, updated_at=?
            WHERE inventory_id=? AND status != 'OUTBOUND'
        """, (ts[:10], ts, row["id"]))

        # stock_movement 기록
        con.execute("""
            INSERT INTO stock_movement
                (lot_no, movement_type, qty_kg, customer,
                 movement_date, source_type, actor, remarks, created_at)
            VALUES (?, 'OUTBOUND', ?, ?, ?, 'MANUAL', 'user', ?, ?)
        """, (lot_no, weight_kg, customer, ts[:10],
              f"출고확정: {customer or '고객미지정'}", ts))

        # audit_log
        con.execute("""
            INSERT INTO audit_log (event_type, event_data, user_note, created_by, created_at)
            VALUES ('OUTBOUND_CONFIRM', ?, ?, 'system', ?)
        """, (
            f'{{"lot_no":"{lot_no}","weight_kg":{weight_kg},"customer":"{customer}"}}',
            "출고 확정", ts
        ))
        con.commit()
        con.close()
        return ok_response(data={
            "lot_no": lot_no,
            "status": "OUTBOUND",
            "weight_kg": weight_kg,
            "customer": customer,
            "message": f"{lot_no} 출고 확정 완료",
        })
    except Exception as e:
        logger.error("outbound-confirm error: %s", e)
        return err_response(str(e))


# ── 톤백리스트 Excel 내보내기 ────────────────────────────────────
@router.get("/export-tonbag-excel", summary="🎒 톤백리스트 Excel (F036)")
def export_tonbag_excel(lot_no: Optional[str] = QP(None)):
    """톤백 목록 → Excel FileResponse (openpyxl)"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        con = _db()
        if lot_no:
            rows = con.execute("""
                SELECT t.tonbag_uid, t.lot_no, t.sap_no, t.bl_no,
                       t.sub_lt, t.tonbag_no, t.weight,
                       t.status, t.location, t.inbound_date,
                       t.picked_to, t.sale_ref, t.remarks,
                       i.product, i.warehouse
                FROM inventory_tonbag t
                LEFT JOIN inventory i ON i.id = t.inventory_id
                WHERE t.lot_no = ?
                ORDER BY t.sub_lt, t.tonbag_no
            """, (lot_no,)).fetchall()
        else:
            rows = con.execute("""
                SELECT t.tonbag_uid, t.lot_no, t.sap_no, t.bl_no,
                       t.sub_lt, t.tonbag_no, t.weight,
                       t.status, t.location, t.inbound_date,
                       t.picked_to, t.sale_ref, t.remarks,
                       i.product, i.warehouse
                FROM inventory_tonbag t
                LEFT JOIN inventory i ON i.id = t.inventory_id
                ORDER BY t.lot_no, t.sub_lt, t.tonbag_no
            """).fetchall()
        con.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "톤백리스트"

        # 헤더
        headers = ["톤백 UID","LOT NO","SAP NO","BL NO","Sub LT","톤백 번호",
                   "중량(kg)","상태","위치","입고일","출고대상","Sale Ref","비고","제품","창고"]
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF")
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        # 상태별 색상
        status_colors = {
            "AVAILABLE": "E8F5E9",
            "PICKED":    "FFF9C4",
            "RESERVED":  "E3F2FD",
            "OUTBOUND":  "FAFAFA",
            "CANCELLED": "FFEBEE",
        }
        for r in rows:
            ws.append(list(r))
            status = r[7] or ""
            color  = status_colors.get(status, "FFFFFF")
            fill   = PatternFill("solid", fgColor=color)
            for cell in ws[ws.max_row]:
                cell.fill = fill

        # 열 너비
        widths = [20,16,14,18,8,10,12,12,12,12,14,14,20,20,10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 파일 저장
        import tempfile
        tmp_dir = tempfile.gettempdir()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"tonbag_list_{ts}.xlsx"
        out = os.path.join(tmp_dir, fname)
        wb.save(out)

        return FileResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=fname,
        )
    except Exception as e:
        logger.error("export-tonbag-excel error: %s", e)
        raise HTTPException(500, str(e))


# ── Sales Order 업로드 (MISS-02) ─────────────────────────────────
@router.post("/sales-order-upload", summary="📊 Sales Order 업로드 (MISS-02)")
async def sales_order_upload(file: UploadFile = File(...)):
    """
    Sales Order Excel 파일 업로드 → SalesOrderEngine 처리 → 결과 반환
    v864.2 대응: toolbar_mixin._on_sales_order_upload()
    """
    import sys
    import tempfile

    # 확장자 검사
    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, f"Excel 파일(.xlsx/.xls)만 허용됩니다: {fname}")

    # 임시 파일 저장
    suffix = ".xlsx" if fname.lower().endswith(".xlsx") else ".xls"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        content = await file.read()
        tmp.write(content)
        tmp.flush()
        tmp.close()

        # SQMDatabase 경유 — SalesOrderEngine 요구사항 (fetchall/execute/transaction)
        root = _root()
        db_path = os.path.join(root, "data", "db", "sqm_inventory.db")

        # sys.path에 프로젝트 루트 추가 (엔진 모듈 import용)
        if root not in sys.path:
            sys.path.insert(0, root)

        from engine_modules.database import SQMDatabase
        from features.parsers.sales_order_engine import SalesOrderEngine

        db = SQMDatabase(db_path=db_path)
        engine = SalesOrderEngine(db=db)
        result = engine.process(tmp.name, sales_order_file=fname)

        # SQMDatabase close
        try:
            db.close_all()
        except Exception:
            pass

        return ok_response(data=result)

    except HTTPException:
        raise
    except Exception as e:
        logger.error("sales-order-upload error: %s", e, exc_info=True)
        return err_response(f"Sales Order 처리 중 오류: {e}")
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass


# ── Swap 리포트 조회 (MISS-03) ───────────────────────────────────
def _swap_report_rows(
    start_date: str,
    end_date: str,
    customer: str = "",
    lot_no: str = "",
    operator: str = "",
) -> list:
    """uid_swap_history 기반 Swap 리포트 조회 (v864.2: _query_swap_report_rows 이식)."""
    sql = """
        SELECT
            s.created_at,
            s.lot_no,
            COALESCE(st.customer, t.picked_to, '') AS customer,
            COALESCE(st.created_by, 'barcode_scan_swap') AS operator,
            COALESCE(s.expected_uid, '') AS expected_uid,
            COALESCE(s.scanned_uid, '') AS scanned_uid,
            COALESCE(s.reason, '') AS reason
        FROM uid_swap_history s
        LEFT JOIN inventory_tonbag t ON t.id = s.scanned_tonbag_id
        LEFT JOIN (
            SELECT s1.*
            FROM sold_table s1
            INNER JOIN (
                SELECT tonbag_id, MAX(id) AS max_id
                FROM sold_table
                GROUP BY tonbag_id
            ) m ON m.max_id = s1.id
        ) st ON st.tonbag_id = s.scanned_tonbag_id
        WHERE date(s.created_at) BETWEEN date(?) AND date(?)
          AND (? = '' OR s.lot_no LIKE ?)
          AND (? = '' OR COALESCE(st.customer, t.picked_to, '') LIKE ?)
          AND (? = '' OR COALESCE(st.created_by, 'barcode_scan_swap') LIKE ?)
        ORDER BY s.created_at DESC, s.id DESC
        LIMIT 5000
    """
    lot_like = f"%{lot_no.strip()}%" if lot_no.strip() else ""
    cust_like = f"%{customer.strip()}%" if customer.strip() else ""
    op_like = f"%{operator.strip()}%" if operator.strip() else ""
    try:
        con = _db()
        cur = con.execute(sql, (
            start_date.strip(), end_date.strip(),
            lot_no.strip(), lot_like,
            customer.strip(), cust_like,
            operator.strip(), op_like,
        ))
        rows = [dict(r) for r in cur.fetchall()]
        con.close()
        return rows
    except Exception as e:
        logger.error("swap-report query error: %s", e)
        return []


@router.get("/swap-report", summary="🔁 Swap 리포트 조회 (MISS-03)")
def swap_report(
    start_date: str = QP(..., description="시작일 YYYY-MM-DD"),
    end_date:   str = QP(..., description="종료일 YYYY-MM-DD"),
    customer:   str = QP("",  description="고객사 필터"),
    lot_no:     str = QP("",  description="LOT NO 필터"),
    operator:   str = QP("",  description="작업자 필터"),
):
    """
    Swap 리포트 JSON 반환
    v864.2 대응: outbound_handlers._query_swap_report_rows()
    """
    try:
        rows = _swap_report_rows(start_date, end_date, customer, lot_no, operator)
        return ok_response(data={"rows": rows, "total": len(rows)})
    except Exception as e:
        logger.error("swap-report error: %s", e)
        return err_response(str(e))



@router.get("/swap-report/export", summary="Swap Report Export (MISS-03)")
def swap_report_export(
    start_date: str = QP(...),
    end_date:   str = QP(...),
    customer:   str = QP(""),
    lot_no:     str = QP(""),
    operator:   str = QP(""),
    fmt:        str = QP("xlsx", description="xlsx or csv"),
):
    """Swap report CSV/Excel download. v864.2: _export_swap_report_csv/_xlsx"""
    try:
        rows = _swap_report_rows(start_date, end_date, customer, lot_no, operator)
        headers = ["created_at", "lot_no", "customer", "operator",
                   "expected_uid", "scanned_uid", "reason"]
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        if fmt == "csv":
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(headers)
            for r in rows:
                w.writerow([str(r.get(h, "") or "") for h in headers])
            buf.seek(0)
            fname_dl = f"swap_report_{start_date}_{end_date}_{ts}.csv"
            return StreamingResponse(
                iter([buf.getvalue().encode("utf-8-sig")]),
                media_type="text/csv; charset=utf-8-sig",
                headers={"Content-Disposition": f"attachment; filename={fname_dl}"},
            )
        else:
            import tempfile
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "swap_report"
            ws.append(headers)
            hdr_fill = PatternFill("solid", fgColor="1F4E79")
            hdr_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
            for r in rows:
                ws.append([str(r.get(h, "") or "") for h in headers])
            tmp_dir = tempfile.gettempdir()
            fname_dl = f"swap_report_{start_date}_{end_date}_{ts}.xlsx"
            out = os.path.join(tmp_dir, fname_dl)
            wb.save(out)
            return FileResponse(
                out,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=fname_dl,
            )
    except Exception as e:
        logger.error("swap-report-export error: %s", e)
        raise HTTPException(500, str(e))
