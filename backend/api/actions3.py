"""
SQM v8.6.6 — Phase 4-C 액션 엔드포인트
actions3.py: optimize-db, cleanup-logs, do-update,
             return-create, export-invoice-excel, settings-get

Rule 4: try/except 의무, ok_response/err_response 표준 포맷
"""
import os
import sqlite3
import logging
from datetime import datetime, date
from typing import Optional
from fastapi import APIRouter, Body, HTTPException, Query as QP
from fastapi.responses import FileResponse
from backend.common.errors import ok_response, err_response
from backend.common.excel_alignment import safe_apply_sqm_workbook

router = APIRouter(prefix="/api/action3", tags=["actions3"])
logger = logging.getLogger(__name__)


# ── 공통 헬퍼 ────────────────────────────────────────────────────
def _db_path() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(os.path.dirname(here))
    return os.path.join(root, "data", "db", "sqm_inventory.db")


def _db() -> sqlite3.Connection:
    con = sqlite3.connect(_db_path(), timeout=5, check_same_thread=False)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA busy_timeout=3000")
    return con


def _root() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.dirname(os.path.dirname(here))


# ── DB 최적화 (VACUUM) ───────────────────────────────────────────
@router.post("/optimize-db", summary="🔧 DB 최적화 VACUUM (F071)")
def optimize_db():
    """
    SQLite VACUUM 실행 — DB 파일 조각 정리 + 크기 축소
    실행 시간: DB 크기에 따라 수 초 소요
    """
    try:
        db = _db_path()

        # 최적화 전 크기
        size_before = os.path.getsize(db) if os.path.exists(db) else 0

        # WAL checkpoint 먼저
        con = sqlite3.connect(db, timeout=30)
        con.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        con.close()

        # VACUUM (별도 연결 — autocommit 필요)
        con2 = sqlite3.connect(db, timeout=60, isolation_level=None)
        con2.execute("VACUUM")
        con2.close()

        size_after = os.path.getsize(db) if os.path.exists(db) else 0
        saved_mb = round((size_before - size_after) / 1024 / 1024, 3)

        return ok_response(data={
            "size_before_mb": round(size_before / 1024 / 1024, 3),
            "size_after_mb":  round(size_after  / 1024 / 1024, 3),
            "saved_mb": saved_mb,
            "message": f"DB 최적화 완료 — {saved_mb:+.3f} MB",
        })
    except Exception as e:
        logger.error("optimize-db error: %s", e)
        return err_response(str(e))


# ── 오래된 로그 정리 ──────────────────────────────────────────────
@router.post("/cleanup-logs", summary="🧹 로그 정리 (F072)")
def cleanup_logs(payload: dict = None):
    """
    payload: { days: int (기본 90) }
    audit_log + stock_movement에서 N일 이상 된 레코드 삭제
    """
    if payload is None:
        payload = {}
    days = int(payload.get("days", 90))
    if days < 30:
        return err_response("최소 30일 이상만 정리 가능합니다")

    try:
        con = _db()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # audit_log 정리
        audit_del = con.execute("""
            DELETE FROM audit_log
            WHERE julianday('now') - julianday(created_at) > ?
        """, (days,)).rowcount

        # stock_movement 정리 (참고용 이동 이력만 — INBOUND/OUTBOUND는 보존)
        move_del = con.execute("""
            DELETE FROM stock_movement
            WHERE movement_type NOT IN ('INBOUND','OUTBOUND')
              AND julianday('now') - julianday(created_at) > ?
        """, (days,)).rowcount

        con.commit()
        con.close()
        return ok_response(data={
            "days_threshold": days,
            "audit_log_deleted": audit_del,
            "stock_movement_deleted": move_del,
            "total_deleted": audit_del + move_del,
            "message": f"{days}일 이전 로그 {audit_del + move_del}건 정리 완료",
        })
    except Exception as e:
        logger.error("cleanup-logs error: %s", e)
        return err_response(str(e))


# ── D/O 업데이트 ─────────────────────────────────────────────────
@router.post("/do-update", summary="📄 D/O 정보 업데이트 (F057)")
def do_update(payload: dict):
    """
    payload: { lot_no: str, field: str, value: str }
    document_do 단일 필드 업데이트 (free_time, con_return, warehouse_name 등)
    """
    lot_no = (payload.get("lot_no") or "").strip()
    field  = (payload.get("field")  or "").strip()
    value  = payload.get("value", "")

    # 허용 필드 화이트리스트 (SQL Injection 방지)
    ALLOWED_FIELDS = {
        "free_time", "con_return", "warehouse_name", "warehouse_code",
        "arrival_date", "stock_date", "place_of_delivery", "final_destination"
    }
    if not lot_no or not field:
        raise HTTPException(400, "lot_no, field 필수")
    if field not in ALLOWED_FIELDS:
        return err_response(f"'{field}' 필드는 수정 불가. 허용: {sorted(ALLOWED_FIELDS)}")

    try:
        con = _db()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row = con.execute("SELECT id FROM document_do WHERE lot_no=?", (lot_no,)).fetchone()
        if not row:
            con.close()
            return err_response(f"D/O LOT '{lot_no}' 없음")

        con.execute(
            f"UPDATE document_do SET {field}=?, parsed_at=? WHERE lot_no=?",
            (value, ts, lot_no)
        )
        con.commit()
        con.close()
        return ok_response(data={
            "lot_no": lot_no,
            "field": field,
            "value": value,
            "message": f"{lot_no} D/O {field} 업데이트 완료",
        })
    except Exception as e:
        logger.error("do-update error: %s", e)
        return err_response(str(e))


# ── 반품 생성 ────────────────────────────────────────────────────
@router.post("/return-create", summary="↩️ 반품 등록 (F007-alt)")
def return_create(payload: dict):
    """
    payload: { lot_no, tonbag_uid(optional), reason, weight_kg(optional) }
    return_history 에 반품 레코드 삽입 + inventory 상태 업데이트
    """
    lot_no     = (payload.get("lot_no")     or "").strip()
    tonbag_uid = (payload.get("tonbag_uid") or "").strip() or None
    reason     = (payload.get("reason")     or "고객요청").strip()
    weight_kg  = float(payload.get("weight_kg") or 0)

    if not lot_no:
        raise HTTPException(400, "lot_no 필수")

    try:
        con = _db()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        today = ts[:10]

        # inventory 확인
        row = con.execute("SELECT id, status, current_weight FROM inventory WHERE lot_no=?", (lot_no,)).fetchone()
        if not row:
            con.close()
            return err_response(f"LOT '{lot_no}' 없음")

        # return_history INSERT
        con.execute("""
            INSERT INTO return_history
                (lot_no, tonbag_uid, reason, weight_kg, return_date,
                 status, created_at)
            VALUES (?, ?, ?, ?, ?, 'RETURNED', ?)
        """, (lot_no, tonbag_uid, reason, weight_kg, today, ts))

        # stock_movement 기록
        con.execute("""
            INSERT INTO stock_movement
                (lot_no, movement_type, qty_kg, source_type,
                 actor, remarks, created_at)
            VALUES (?, 'RETURN', ?, 'MANUAL', 'user', ?, ?)
        """, (lot_no, weight_kg, f"반품: {reason}", ts))

        # audit_log
        con.execute("""
            INSERT INTO audit_log
                (event_type, event_data, user_note, created_by, created_at)
            VALUES ('RETURN_CREATE', ?, ?, 'system', ?)
        """, (
            f'{{"lot_no":"{lot_no}","reason":"{reason}","weight_kg":{weight_kg}}}',
            reason, ts
        ))

        con.commit()
        con.close()
        return ok_response(data={
            "lot_no": lot_no,
            "reason": reason,
            "weight_kg": weight_kg,
            "return_date": today,
            "message": f"{lot_no} 반품 등록 완료",
        })
    except Exception as e:
        logger.error("return-create error: %s", e)
        return err_response(str(e))


# ── 거래명세서 Excel 내보내기 ─────────────────────────────────────
@router.get("/export-invoice-excel", summary="🧾 거래명세서 Excel (F045)")
def export_invoice_excel(lot_no: Optional[str] = QP(None)):
    """document_invoice 목록 → Excel FileResponse (openpyxl)"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        con = _db()
        if lot_no:
            rows = con.execute("""
                SELECT invoice_no, salar_invoice_no, invoice_date,
                       lot_no, sap_no, bl_no,
                       customer_code, customer_name, customer_ref,
                       product_name, quantity_mt, unit_price,
                       total_amount, currency,
                       net_weight_kg, vessel, origin, destination,
                       incoterm, payment_term, created_at
                FROM document_invoice
                WHERE lot_no=?
                ORDER BY invoice_date DESC
            """, (lot_no,)).fetchall()
        else:
            rows = con.execute("""
                SELECT invoice_no, salar_invoice_no, invoice_date,
                       lot_no, sap_no, bl_no,
                       customer_code, customer_name, customer_ref,
                       product_name, quantity_mt, unit_price,
                       total_amount, currency,
                       net_weight_kg, vessel, origin, destination,
                       incoterm, payment_term, created_at
                FROM document_invoice
                ORDER BY invoice_date DESC, lot_no
            """).fetchall()
        con.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "거래명세서"

        headers = [
            "Invoice No", "Salar Invoice", "Invoice Date",
            "LOT No", "SAP No", "BL No",
            "고객코드", "고객명", "고객Ref",
            "제품명", "수량(MT)", "단가",
            "총금액", "통화",
            "순중량(kg)", "선박명", "원산지", "목적지",
            "인코텀즈", "결제조건", "등록일"
        ]
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF")
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        row_fill = PatternFill("solid", fgColor="EBF3FB")
        for i, r in enumerate(rows):
            ws.append(list(r))
            if i % 2 == 1:
                for cell in ws[ws.max_row]:
                    cell.fill = row_fill

        # 열 너비
        widths = [18,18,14,16,14,18,10,20,16,18,12,12,14,8,14,16,12,12,10,12,16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        import tempfile
        tmp_dir = tempfile.gettempdir()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"invoice_list_{ts}.xlsx"
        out = os.path.join(tmp_dir, fname)
        safe_apply_sqm_workbook(wb)
        wb.save(out)

        return FileResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=fname,
        )
    except Exception as e:
        logger.error("export-invoice-excel error: %s", e)
        raise HTTPException(500, str(e))


# ── 테스트 DB 초기화 (개발자 전용) ──────────────────────────────
@router.post("/db-reset", summary="🗑️ 테스트 DB 초기화")
def db_reset(body: dict = Body(default={})):
    """
    모든 데이터를 삭제하고 빈 DB로 초기화합니다.
    ⚠️ 위험한 작업 — 되돌릴 수 없습니다.
    """
    import shutil
    try:
        if not body.get("confirm"):
            return err_response("confirm: true 가 필요합니다")

        db_file = _db_path()

        # 초기화 전 자동 백업
        backup_dir = os.path.join(os.path.dirname(os.path.dirname(
            os.path.dirname(os.path.abspath(__file__)))), "backup")
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"sqm_pre_reset_{ts}.db"
        if os.path.exists(db_file):
            shutil.copy2(db_file, os.path.join(backup_dir, backup_name))

        # 모든 사용자 테이블 데이터 삭제
        con = _db()
        tables = [r[0] for r in con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()]

        deleted = {}
        for tbl in tables:
            count = con.execute(f"SELECT COUNT(*) FROM [{tbl}]").fetchone()[0]
            if count > 0:
                con.execute(f"DELETE FROM [{tbl}]")
                deleted[tbl] = count

        con.execute("VACUUM")
        con.commit()
        con.close()

        return ok_response(
            data={"tables_cleared": deleted, "backup": backup_name},
            message=f"✅ DB 초기화 완료 — {len(deleted)}개 테이블 데이터 삭제, 백업: {backup_name}"
        )
    except Exception as e:
        logger.error("db-reset error: %s", e)
        return err_response(str(e))
