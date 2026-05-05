"""
SQM v8.6.6 — Phase 4-C SQL 조회 엔드포인트
queries3.py: sales-order-dn, dn-cross-check, do-status,
             invoice-list, settings-info

모든 응답: ok_response(data=...) 표준 포맷
"""
import os
import re
import sqlite3
import logging
import tempfile
from datetime import date
from pathlib import Path
from fastapi import APIRouter, HTTPException, Query as QP
from fastapi.responses import FileResponse
from backend.common.errors import ok_response, err_response

router = APIRouter(prefix="/api/q3", tags=["queries3"])
logger = logging.getLogger(__name__)


# ── DB 경로 헬퍼 ─────────────────────────────────────────────────
def _db() -> sqlite3.Connection:
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(os.path.dirname(here))
    db_path = os.path.join(root, "data", "db", "sqm_inventory.db")
    con = sqlite3.connect(db_path, timeout=5, check_same_thread=False)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA busy_timeout=3000")
    return con


def _rows(rows) -> list:
    return [dict(r) for r in rows]


def _project_root() -> Path:
    here = Path(__file__).resolve()
    return here.parents[2]


def _sales_order_dn_template_path() -> Path:
    root = _project_root()
    candidates = [
        root / "data" / "templates" / "sales_order_dn_template.xlsx",
        root.parent / "sqm_2_merge_upload" / "data" / "templates" / "sales_order_dn_template.xlsx",
        Path(r"D:\program\SQM_inventory\sqm_2_merge_upload\data\templates\sales_order_dn_template.xlsx"),
    ]
    for p in candidates:
        if p.exists():
            return p
    raise HTTPException(
        status_code=404,
        detail="sales_order_dn_template.xlsx 템플릿을 찾을 수 없습니다.",
    )


def _safe_filename_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "").strip())[:80] or "SalesOrder"


def _fmt_date(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text[:10]


def _fmt_mt(value) -> float:
    try:
        return round(float(value or 0) / 1000.0, 4)
    except Exception:
        return 0.0


def _sales_order_dn_rows(con: sqlite3.Connection, sales_order_no: str) -> list:
    return _rows(con.execute(
        """
        WITH sold_group AS (
            SELECT
                lot_no,
                COALESCE(NULLIF(picking_no, ''), '') AS picking_no,
                COALESCE(NULLIF(sap_no, ''), '') AS sap_no,
                COALESCE(NULLIF(bl_no, ''), '') AS bl_no,
                COALESCE(NULLIF(customer, ''), '') AS customer,
                COALESCE(NULLIF(sku, ''), '') AS sku,
                COALESCE(NULLIF(delivery_date, ''), '') AS delivery_date,
                COALESCE(is_sample, 0) AS is_sample,
                SUM(COALESCE(sold_qty_kg, 0)) AS net_kg,
                SUM(COALESCE(gross_weight_kg, 0)) AS sold_gross_kg,
                COUNT(*) AS ct_plt
            FROM sold_table
            WHERE sales_order_no = ?
              AND COALESCE(status, '') IN ('SOLD', 'OUTBOUND', 'CONFIRMED')
            GROUP BY lot_no, picking_no, sap_no, bl_no, customer, sku, delivery_date, COALESCE(is_sample, 0)
        )
        SELECT
            sg.*,
            i.product,
            i.product_code,
            i.initial_weight AS inv_net_kg,
            i.gross_weight AS inv_gross_kg,
            d.final_destination,
            d.port_of_discharge,
            d.warehouse_name,
            d.gross_weight_kg AS do_gross_kg
        FROM sold_group sg
        LEFT JOIN inventory i ON i.lot_no = sg.lot_no
        LEFT JOIN document_do d ON d.lot_no = sg.lot_no
        ORDER BY sg.delivery_date, sg.lot_no, sg.is_sample, sg.picking_no
        """,
        (sales_order_no,),
    ).fetchall())


def _estimate_gross_kg(row: dict) -> float:
    net = float(row.get("net_kg") or 0)
    sold_gross = float(row.get("sold_gross_kg") or 0)
    if sold_gross > 0:
        return sold_gross
    inv_net = float(row.get("inv_net_kg") or 0)
    inv_gross = float(row.get("inv_gross_kg") or 0)
    if net > 0 and inv_net > 0 and inv_gross > 0:
        return net * (inv_gross / inv_net)
    return net


# ── Sales Order / DN 보고서 ───────────────────────────────────────
@router.get("/sales-order-dn", summary="📋 Sales Order DN 보고서 (F054)")
def get_sales_order_dn():
    """
    allocation_plan + document_do JOIN → Sales Order DN 현황
    allocation_plan이 비어있으면 document_do 단독 조회
    """
    try:
        con = _db()

        # allocation_plan 존재 여부 확인
        ap_count = con.execute("SELECT COUNT(*) FROM allocation_plan").fetchone()[0]

        if ap_count > 0:
            rows = con.execute("""
                SELECT ap.lot_no, ap.sub_lt, ap.customer, ap.sale_ref,
                       ap.qty_mt, ap.outbound_date, ap.status AS ap_status,
                       ap.picking_no, ap.bl_no,
                       d.do_no, d.vessel, d.voyage, d.port_of_discharge,
                       d.arrival_date, d.stock_date, d.free_time
                FROM allocation_plan ap
                LEFT JOIN document_do d ON d.lot_no = ap.lot_no
                ORDER BY ap.outbound_date DESC, ap.lot_no
                LIMIT 200
            """).fetchall()
        else:
            # allocation_plan 비어있음 → document_do 단독
            rows = con.execute("""
                SELECT lot_no, do_no, bl_no, sap_no,
                       vessel, voyage, carrier_id,
                       port_of_loading, port_of_discharge,
                       arrival_date, stock_date, free_time,
                       total_containers, total_packages,
                       gross_weight_kg, warehouse_name,
                       created_at
                FROM document_do
                ORDER BY arrival_date DESC, lot_no
                LIMIT 200
            """).fetchall()

        con.close()
        return ok_response(data={
            "items": _rows(rows),
            "total": len(rows),
            "source": "allocation_plan+document_do" if ap_count > 0 else "document_do",
            "note": "Sales Order DN 현황 보고서",
        })
    except Exception as e:
        logger.error("sales-order-dn error: %s", e)
        return err_response(str(e))


@router.get("/sales-order-nos", summary="📋 Sales Order No 목록")
def get_sales_order_nos(limit: int = QP(100, ge=1, le=500)):
    """sold_table 기준 Sales Order No 선택 목록."""
    try:
        con = _db()
        rows = con.execute(
            """
            SELECT
                sales_order_no,
                COUNT(*) AS row_count,
                ROUND(SUM(COALESCE(sold_qty_kg, 0)) / 1000.0, 4) AS total_mt,
                MAX(COALESCE(sold_date, created_at, '')) AS last_date
            FROM sold_table
            WHERE sales_order_no IS NOT NULL
              AND TRIM(sales_order_no) != ''
              AND COALESCE(status, '') IN ('SOLD', 'OUTBOUND', 'CONFIRMED')
            GROUP BY sales_order_no
            ORDER BY last_date DESC, sales_order_no DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        con.close()
        return ok_response(data={"items": _rows(rows), "total": len(rows)})
    except Exception as e:
        logger.error("sales-order-nos error: %s", e)
        return err_response(str(e))


@router.get("/sales-order-dn-template", summary="📋 Sales Order DN 템플릿 Excel 생성")
def export_sales_order_dn_template(sales_order_no: str = QP(..., min_length=1)):
    """
    Sales Order No 선택 → sold_table 조회 → sales_order_dn_template.xlsx 복사/채움 → 다운로드.
    템플릿은 5행 헤더, 6행부터 데이터 영역으로 사용한다.
    """
    so_no = str(sales_order_no or "").strip()
    if not so_no:
        raise HTTPException(status_code=400, detail="Sales Order No가 필요합니다.")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    con = _db()
    try:
        rows = _sales_order_dn_rows(con, so_no)
    finally:
        con.close()

    if not rows:
        raise HTTPException(status_code=404, detail=f"Sales Order No '{so_no}' 출고 데이터가 없습니다.")

    template = _sales_order_dn_template_path()
    wb = load_workbook(template)
    ws = wb["DN"] if "DN" in wb.sheetnames else wb.active

    ws["B3"] = "Sales order No :"
    ws["C3"] = so_no

    start_row = 6
    for r in range(start_row, max(ws.max_row, start_row) + 1):
        for c in range(2, 14):
            ws.cell(r, c).value = None

    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        net_kg = float(row.get("net_kg") or 0)
        gross_kg = _estimate_gross_kg(row)
        destination = (
            row.get("customer")
            or row.get("final_destination")
            or row.get("port_of_discharge")
            or row.get("warehouse_name")
            or ""
        )
        sku = row.get("sku") or row.get("product_code") or ""
        description = row.get("product") or sku

        ws.cell(excel_row, 2).value = destination
        ws.cell(excel_row, 3).value = _fmt_date(row.get("delivery_date"))
        ws.cell(excel_row, 4).value = row.get("lot_no") or ""
        ws.cell(excel_row, 5).value = row.get("sap_no") or ""
        ws.cell(excel_row, 6).value = row.get("bl_no") or ""
        ws.cell(excel_row, 7).value = so_no
        ws.cell(excel_row, 8).value = row.get("picking_no") or ""
        ws.cell(excel_row, 9).value = sku
        ws.cell(excel_row, 10).value = description
        ws.cell(excel_row, 11).value = _fmt_mt(net_kg)
        ws.cell(excel_row, 12).value = _fmt_mt(gross_kg)
        ws.cell(excel_row, 13).value = int(row.get("ct_plt") or 0)

    out_dir = Path(tempfile.gettempdir()) / "sqm_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"Sales_order_DN_{_safe_filename_part(so_no)}.xlsx"
    out_path = out_dir / out_name
    wb.save(out_path)

    return FileResponse(
        path=str(out_path),
        filename=out_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── DN Cross Check ────────────────────────────────────────────────
@router.get("/dn-cross-check", summary="🔁 DN 교차검증 (F055-alt)")
def get_dn_cross_check():
    """
    inventory vs document_do 교차 비교
    — DO 있는데 재고 없는 LOT, 재고 있는데 DO 없는 LOT 검출
    """
    try:
        con = _db()

        # DO 있는데 inventory 없는 케이스
        do_no_inv = con.execute("""
            SELECT d.lot_no, d.do_no, d.bl_no, d.vessel,
                   d.arrival_date, d.gross_weight_kg,
                   'DO있음_재고없음' AS cross_status
            FROM document_do d
            LEFT JOIN inventory i ON i.lot_no = d.lot_no
            WHERE i.id IS NULL
            ORDER BY d.arrival_date DESC
        """).fetchall()

        # inventory 있는데 DO 없는 케이스
        inv_no_do = con.execute("""
            SELECT i.lot_no, i.sap_no, i.bl_no, i.product,
                   i.status, i.current_weight, i.inbound_date,
                   'DO없음_재고있음' AS cross_status
            FROM inventory i
            LEFT JOIN document_do d ON d.lot_no = i.lot_no
            WHERE d.id IS NULL
            ORDER BY i.inbound_date DESC
            LIMIT 100
        """).fetchall()

        # 매칭된 케이스 (정상)
        matched = con.execute("""
            SELECT COUNT(*) AS cnt
            FROM inventory i
            INNER JOIN document_do d ON d.lot_no = i.lot_no
        """).fetchone()

        con.close()
        return ok_response(data={
            "do_without_inventory": _rows(do_no_inv),
            "inventory_without_do": _rows(inv_no_do),
            "matched_count": matched["cnt"] if matched else 0,
            "issues_count": len(do_no_inv) + len(inv_no_do),
            "note": "교차검증 이슈가 없으면 issues_count=0",
        })
    except Exception as e:
        logger.error("dn-cross-check error: %s", e)
        return err_response(str(e))


# ── D/O 현황 ─────────────────────────────────────────────────────
@router.get("/do-status", summary="📄 D/O 현황 (F056-alt)")
def get_do_status(limit: int = QP(100, ge=1, le=500)):
    """document_do 전체 목록 + inventory JOIN 상태"""
    try:
        con = _db()
        rows = con.execute("""
            SELECT d.id, d.lot_no, d.do_no, d.bl_no, d.sap_no,
                   d.vessel, d.voyage, d.carrier_id,
                   d.port_of_loading, d.port_of_discharge,
                   d.arrival_date, d.stock_date, d.free_time,
                   d.con_return, d.total_containers,
                   d.gross_weight_kg, d.warehouse_name,
                   COALESCE(i.status, '미입고') AS inv_status,
                   i.current_weight, i.location,
                   d.created_at
            FROM document_do d
            LEFT JOIN inventory i ON i.lot_no = d.lot_no
            ORDER BY d.arrival_date DESC, d.lot_no
            LIMIT ?
        """, (limit,)).fetchall()

        # 상태별 집계
        summary = con.execute("""
            SELECT COALESCE(i.status, '미입고') AS status,
                   COUNT(*) AS cnt
            FROM document_do d
            LEFT JOIN inventory i ON i.lot_no = d.lot_no
            GROUP BY status
            ORDER BY cnt DESC
        """).fetchall()
        con.close()

        return ok_response(data={
            "items": _rows(rows),
            "total": len(rows),
            "summary_by_status": _rows(summary),
            "columns": ["lot_no", "do_no", "bl_no", "vessel",
                        "arrival_date", "stock_date", "inv_status",
                        "gross_weight_kg", "warehouse_name"],
        })
    except Exception as e:
        logger.error("do-status error: %s", e)
        return err_response(str(e))


# ── 거래명세서 목록 ─────────────────────────────────────────────
@router.get("/invoice-list", summary="🧾 거래명세서 목록 (F045-alt)")
def get_invoice_list(
    customer: str = QP(None, description="고객명 필터"),
    limit: int = QP(100, ge=1, le=500),
):
    """document_invoice 목록 — 거래명세서 조회"""
    try:
        con = _db()

        if customer:
            rows = con.execute("""
                SELECT id, lot_no, sap_no, invoice_no, salar_invoice_no,
                       invoice_date, customer_code, customer_name,
                       customer_ref, product_name, quantity_mt,
                       unit_price, total_amount, currency,
                       net_weight_kg, vessel, origin, destination,
                       incoterm, payment_term, bl_no, created_at
                FROM document_invoice
                WHERE customer_name LIKE ?
                ORDER BY invoice_date DESC, lot_no
                LIMIT ?
            """, (f"%{customer}%", limit)).fetchall()
        else:
            rows = con.execute("""
                SELECT id, lot_no, sap_no, invoice_no, salar_invoice_no,
                       invoice_date, customer_code, customer_name,
                       customer_ref, product_name, quantity_mt,
                       unit_price, total_amount, currency,
                       net_weight_kg, vessel, origin, destination,
                       incoterm, payment_term, bl_no, created_at
                FROM document_invoice
                ORDER BY invoice_date DESC, lot_no
                LIMIT ?
            """, (limit,)).fetchall()

        # 고객별 집계
        by_customer = con.execute("""
            SELECT customer_name,
                   COUNT(*) AS invoice_cnt,
                   ROUND(SUM(quantity_mt), 3) AS total_mt,
                   ROUND(SUM(total_amount), 2) AS total_amount,
                   MAX(currency) AS currency
            FROM document_invoice
            GROUP BY customer_name
            ORDER BY total_mt DESC
        """).fetchall()
        con.close()

        return ok_response(data={
            "items": _rows(rows),
            "total": len(rows),
            "by_customer": _rows(by_customer),
            "columns": ["invoice_no", "invoice_date", "customer_name",
                        "lot_no", "product_name", "quantity_mt",
                        "total_amount", "currency", "vessel"],
        })
    except Exception as e:
        logger.error("invoice-list error: %s", e)
        return err_response(str(e))


# ── 시스템 설정 정보 ──────────────────────────────────────────────
@router.get("/settings-info", summary="⚙️ 시스템 설정 정보 (F058-alt)")
def get_settings_info():
    """DB 메타데이터 + 테이블별 행수 + 시스템 정보"""
    try:
        con = _db()

        # 테이블별 행수
        tables = [
            "inventory", "inventory_tonbag", "stock_movement",
            "audit_log", "return_history", "allocation_plan",
            "document_do", "document_invoice", "document_bl",
        ]
        table_stats = []
        for t in tables:
            try:
                cnt = con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                table_stats.append({"table": t, "rows": cnt})
            except Exception:
                table_stats.append({"table": t, "rows": -1, "error": "테이블 없음"})

        # DB 파일 크기
        db_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            "data", "db", "sqm_inventory.db"
        )
        try:
            db_size_mb = round(os.path.getsize(db_path) / 1024 / 1024, 3)
        except Exception:
            db_size_mb = -1

        # journal mode 확인
        jm = con.execute("PRAGMA journal_mode").fetchone()[0]
        wal_size = con.execute("PRAGMA page_count").fetchone()[0]

        con.close()
        return ok_response(data={
            "version": "SQM v864.3",
            "db_path": db_path,
            "db_size_mb": db_size_mb,
            "journal_mode": jm,
            "page_count": wal_size,
            "table_stats": table_stats,
            "note": "DB 메타데이터 및 테이블별 행수",
        })
    except Exception as e:
        logger.error("settings-info error: %s", e)
        return err_response(str(e))
