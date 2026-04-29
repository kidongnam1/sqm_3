# -*- coding: utf-8 -*-
"""
SQM v7.9.2 — 출고 템플릿 / Allocation 생성 Mixin
=================================================
v7.9.2: 외부 subprocess → UI 내부 직접 통합
        DB 직접 연결, RESERVED LOT 자동 제외, 샘플행 자동 포함
"""
from gui_app_modular.utils.ui_constants import create_themed_toplevel  # v8.0.9
from gui_app_modular.utils.ui_constants import tc
import logging
import os
import tkinter as tk
from datetime import date
from pathlib import Path
from tkinter import ttk

from ..utils.custom_messagebox import CustomMessageBox

logger = logging.getLogger(__name__)


class OutboundTemplateMixin:
    """출고 템플릿 및 Allocation Table Mixin (v7.9.2: UI 내부 통합)"""

    def _generate_allocation_samples(self) -> None:
        """Allocation Excel N개 생성 (v7.9.2: UI 직접 통합)."""
        root   = getattr(self, 'root', None)
        engine = getattr(self, 'engine', None)
        if not engine or not getattr(engine, 'db', None):
            CustomMessageBox.showwarning(root, "오류", "DB 연결이 없습니다.")
            return

        # ── 입력 다이얼로그 ───────────────────────────────────────────────────
        dlg = create_themed_toplevel(root)
        dlg.title("Allocation Excel 생성")
        dlg.resizable(True, True)
        dlg.grab_set()
        try:
            dlg.geometry("400x300")
            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - 400) // 2
            y = root.winfo_y() + (root.winfo_height() - 300) // 2
            dlg.geometry(f"+{x}+{y}")
        except Exception:
            logger.debug("[SUPPRESSED] exception in outbound_template_mixin.py")  # noqa

        frm = ttk.Frame(dlg, padding=16)
        frm.pack(fill=tk.BOTH, expand=True)

        fields = [
            ("고객사 (Customer):", "PT LBM"),
            ("SALE REF:",          ""),
            ("분할 파일 수:",       "3"),
        ]
        vars_ = []
        for r, (label, default) in enumerate(fields):
            ttk.Label(frm, text=label).grid(row=r, column=0, sticky='w', pady=5)
            v = tk.StringVar(value=default)
            ttk.Entry(frm, textvariable=v, width=28).grid(row=r, column=1, pady=5, padx=(8,0))
            vars_.append(v)

        ttk.Label(frm, text="  ※ SALE REF 입력 시 기존 예약 LOT 자동 제외",
                  foreground=tc('text_muted')).grid(row=3, column=1, sticky='w', padx=(8,0))

        project_root = Path(__file__).resolve().parent.parent.parent
        default_out  = str(project_root / "generated_allocation")
        ttk.Label(frm, text="출력 폴더:").grid(row=4, column=0, sticky='w', pady=5)
        v_out = tk.StringVar(value=default_out)
        ttk.Entry(frm, textvariable=v_out, width=28).grid(row=4, column=1, pady=5, padx=(8,0))

        ok_flag = {'v': False}
        def _ok():
            ok_flag['v'] = True
            dlg.destroy()
        def _cancel():
            dlg.destroy()

        bf = ttk.Frame(frm)
        bf.grid(row=5, column=0, columnspan=2, pady=(12, 0))
        ttk.Button(bf, text="생성", command=_ok,     width=10).pack(side=tk.LEFT, padx=6)
        ttk.Button(bf, text="취소", command=_cancel, width=10).pack(side=tk.LEFT, padx=6)
        dlg.wait_window()

        if not ok_flag['v']:
            return

        customer = vars_[0].get().strip() or "PT LBM"
        sale_ref = vars_[1].get().strip()
        try:
            n_files = max(1, int(vars_[2].get() or 3))
        except ValueError:
            n_files = 3
        out_dir  = Path(v_out.get().strip() or default_out)

        # ── 생성 실행 ─────────────────────────────────────────────────────────
        generated = self._run_alloc_gen(engine, n_files, customer, sale_ref, out_dir)

        if not generated:
            CustomMessageBox.showwarning(root, "결과 없음",
                "생성된 파일이 없습니다.\\nAVAILABLE LOT이 없거나 모두 이미 예약됐습니다.")
            return

        msg = (
            f"Allocation Excel {len(generated)}개 생성 완료\\n\\n"
            + "\\n".join(f"  {Path(p).name}" for p in generated)
            + f"\\n\\n출력 폴더:\\n{out_dir}"
        )
        CustomMessageBox.showinfo(root, "생성 완료", msg)
        try:
            if os.name == 'nt':
                os.startfile(str(out_dir))
        except Exception:
            logger.debug("[SUPPRESSED] exception in outbound_template_mixin.py")  # noqa
        if hasattr(self, '_log'):
            self._log(f"✅ Allocation Excel {len(generated)}개 생성 → {out_dir}")

    def _run_alloc_gen(self, engine, n_files, customer, sale_ref, out_dir):
        """scripts 모듈 우선 → 없으면 내장 로직 폴백."""
        try:
            from scripts.generate_allocation_from_tonbag import generate
            db_path = None
            if hasattr(engine.db, 'db_path'):
                db_path = Path(engine.db.db_path)
            return generate(n_files=n_files, customer=customer,
                            sale_ref=sale_ref, db_path=db_path, out_dir=out_dir)
        except Exception as e:
            logger.warning(f"[ALLOC_GEN] scripts 모듈 폴백: {e}")
            return self._alloc_gen_internal(engine, n_files, customer, sale_ref, out_dir)

    def _alloc_gen_internal(self, engine, n_files, customer, sale_ref, out_dir):
        """내장 Allocation 엑셀 생성."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl 미설치")
            return []

        db = engine.db

        # 기존 예약 LOT 제외
        reserved = set()
        if sale_ref:
            try:
                rows = db.fetchall(
                    "SELECT DISTINCT lot_no FROM allocation_plan "
                    "WHERE sale_ref=? AND status IN ('RESERVED','STAGED','PENDING_APPROVAL')",
                    (str(sale_ref).strip(),))
                reserved = {str(r.get('lot_no') if isinstance(r, dict) else r[0]).strip()
                            for r in (rows or [])}
            except Exception as e:
                logger.debug(f"예약조회 스킵: {e}")

        # AVAILABLE LOT
        try:
            raw = db.fetchall("""
                SELECT i.lot_no, i.sap_no, i.product, i.gross_weight,
                       i.warehouse, i.arrival_date, i.inbound_date
                FROM inventory i
                WHERE i.status='AVAILABLE'
                  AND EXISTS(SELECT 1 FROM inventory_tonbag t
                             WHERE t.lot_no=i.lot_no AND t.status='AVAILABLE'
                               AND COALESCE(t.is_sample,0)=0)
                ORDER BY COALESCE(i.arrival_date, i.inbound_date, i.created_at) ASC, i.lot_no
            """)
        except Exception as e:
            logger.error(f"LOT 조회 실패: {e}")
            return []

        lots = []
        for r in (raw or []):
            d = dict(r) if isinstance(r, dict) else {
                'lot_no': r[0], 'sap_no': r[1], 'product': r[2],
                'gross_weight': r[3], 'warehouse': r[4],
                'arrival_date': r[5], 'inbound_date': r[6]}
            if str(d.get('lot_no','')).split('.')[0].strip() not in reserved:
                lots.append(d)

        if not lots:
            return []

        chunks = [[] for _ in range(n_files)]
        for i, lot in enumerate(lots):
            chunks[i % n_files].append(lot)

        out_dir.mkdir(parents=True, exist_ok=True)
        generated = []
        today = date.today().strftime('%Y-%m-%d')

        TF = PatternFill('solid', start_color='1F3864')
        HF = PatternFill('solid', start_color='2E75B6')
        EF = PatternFill('solid', start_color='DCE6F1')
        OF = PatternFill('solid', start_color='FFFFFF')
        SF = PatternFill('solid', start_color='FFF2CC')
        t  = Side(style='thin', color='B8CCE4')
        BD = Border(left=t, right=t, top=t, bottom=t)
        C  = Alignment(horizontal='center', vertical='center')
        HDR = ['Product','SAP NO','ETA BUSAN','Date in stock',
               'QTY (MT)','Lot No','WH','Customs','GW','SALE REF']
        CW  = [22,14,14,14,10,14,8,10,8,10]

        for fi, chunk in enumerate(chunks, 1):
            if not chunk:
                continue
            mt    = len(chunk) * 5.0
            title = f"Allocation - GY - {customer} ({n_files}분할 {fi}/{n_files}) — {mt:.2f} MT"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Allocation'

            ws.merge_cells('A1:J1')
            ws['A1'].value = title
            ws['A1'].font  = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            ws['A1'].fill  = TF
            ws['A1'].alignment = C
            ws.row_dimensions[1].height = 22

            n_r = len(chunk) * 2
            ws.merge_cells('A2:J2')
            ws['A2'].value = f'=SUM(E4:E{3+n_r})'
            ws['A2'].font  = Font(name='Arial', bold=True, size=10)
            ws['A2'].alignment = C
            ws.row_dimensions[2].height = 18

            for ci, h in enumerate(HDR, 1):
                c = ws.cell(3, ci)
                c.value = h
                c.font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
                c.fill  = HF; c.alignment = C; c.border = BD
            ws.row_dimensions[3].height = 20

            er = 4
            for idx, lot in enumerate(chunk):
                ln  = str(lot.get('lot_no','')).split('.')[0]
                sap = str(lot.get('sap_no','') or '').split('.')[0]
                prd = str(lot.get('product') or 'LITHIUM CARBONATE')
                gw  = float(lot.get('gross_weight') or 0)
                gw_mt = round(gw/1000, 3) if gw > 0 else 5.13
                wh  = str(lot.get('warehouse') or '광양')
                eta = str(lot.get('arrival_date') or lot.get('inbound_date') or '')
                fl  = EF if idx % 2 == 0 else OF

                for ci, v in enumerate([prd,sap,eta,today,5.0,ln,wh,'',gw_mt,sale_ref or ''],1):
                    c = ws.cell(er, ci)
                    c.value = v
                    c.font  = Font(name='Arial', size=10)
                    c.fill  = fl; c.alignment = C; c.border = BD
                ws.row_dimensions[er].height = 18; er += 1

                for ci, v in enumerate([prd,sap,eta,today,0.001,ln,wh,'',0.001,sale_ref or ''],1):
                    c = ws.cell(er, ci)
                    c.value = v
                    c.font  = Font(name='Arial', size=10, italic=True, color='7F6000')
                    c.fill  = SF; c.alignment = C; c.border = BD
                ws.row_dimensions[er].height = 16; er += 1

            for ci, w in enumerate(CW, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            fname = f"Allocation_GY_{customer.replace(' ','_')}_{fi}of{n_files}.xlsx"
            p = out_dir / fname
            wb.save(str(p))
            generated.append(p)
            logger.info(f"[ALLOC_GEN] {fname} ({len(chunk)}개 LOT, {len(chunk)*5}MT)")

        from collections import Counter
        all_ln = [str(lot.get('lot_no','')).split('.')[0]
                  for c in chunks for lot in c]
        dups = {l:n for l,n in Counter(all_ln).items() if n > 1}
        if dups:
            logger.error(f"[ALLOC_GEN] 중복 LOT 감지: {dups}")
        else:
            logger.info(f"[ALLOC_GEN] 파일간 LOT 중복 없음 ({len(all_ln)}개)")
        return generated
