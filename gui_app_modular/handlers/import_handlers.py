# -*- coding: utf-8 -*-
"""
SQM 재고관리 - Excel 입고 처리 핸들러
=====================================

v2.9.91 - gui_app.py에서 분리
v5.8.9: 3가지 매핑(수동 입고/빠른 출고/위치) 시 템플릿 열기 vs 파일 업로드 선택
v6.0: Excel/데이터 입력 원칙 통일 — 프로그램 내장 형식 → 데이터 붙여넣기 또는 파일 업로드 (AGENTS.md)

Excel 파일 입고 처리, 컬럼 자동 인식, 데이터 변환
"""

from gui_app_modular.utils.ui_constants import create_themed_toplevel  # v8.0.9
from engine_modules.constants import STATUS_AVAILABLE
import logging
import os
from ..utils.ui_constants import CustomMessageBox
from typing import Dict, Optional

logger = logging.getLogger(__name__)


def _open_file_with_default_app(file_path: str, parent=None, ask_yes=None) -> bool:
    """OS 기본 앱으로 파일 열기 (엑셀 템플릿 등). 동일 이름 열림 시 순번 붙여서 오픈 제안."""
    from ..utils.excel_file_helper import open_excel_with_fallback
    from ..utils.custom_messagebox import CustomMessageBox
    if parent and ask_yes is None:
        ask_yes = lambda t, m: CustomMessageBox.askyesno(parent, t, m)
    if parent and ask_yes:
        return open_excel_with_fallback(parent, file_path, ask_yes=ask_yes)
    try:
        if os.name == 'nt':
            os.startfile(os.path.normpath(file_path))
            return True
        import subprocess
        if hasattr(os, 'uname') and getattr(os.uname(), 'sysname', '') == 'Darwin':
            subprocess.run(['open', file_path], check=False)
        else:
            subprocess.run(['xdg-open', file_path], check=False)
        return True
    except Exception as e:
        logger.debug(f"파일 열기 실패: {e}")
        return False


class ImportHandlersMixin:
    """
    Excel 입고 처리 Mixin
    
    SQMInventoryApp 클래스에 mix-in 됩니다.
    v5.8.9: 3가지 매핑(수동 입고/빠른 출고/위치) 시 [템플릿 열기] vs [파일 업로드] 선택.
    """
    
    def _show_template_or_upload_choice(self, title: str, kind: str) -> Optional[str]:
        """Excel/데이터 입력 원칙: 데이터 붙여넣기 vs 파일 업로드 선택. 반환: 'template' | 'upload' | None(취소)."""
        import tkinter as tk
        from tkinter import ttk
        from ..utils.ui_constants import (
            apply_modal_window_options,
            setup_dialog_geometry_persistence,
            UPLOAD_CHOICE_HEADER, UPLOAD_CHOICE_PASTE, UPLOAD_CHOICE_UPLOAD,
            UPLOAD_CHOICE_BTN_PASTE, UPLOAD_CHOICE_BTN_UPLOAD,
        )
        result = [None]
        win = create_themed_toplevel(self.root)
        win.title(title)
        apply_modal_window_options(win)
        win.transient(self.root)
        win.grab_set()
        win.minsize(400, 280)
        setup_dialog_geometry_persistence(win, "import_upload_choice_dialog", self.root, "small")
        f = ttk.Frame(win, padding=(20, 20, 20, 32))
        f.pack(fill=tk.BOTH, expand=True)
        ttk.Label(f, text=UPLOAD_CHOICE_HEADER, font=('맑은 고딕', 12, 'bold')).pack(anchor='w', pady=(0, 12))
        ttk.Label(f, text=UPLOAD_CHOICE_PASTE, font=('맑은 고딕', 10), wraplength=400, justify=tk.LEFT).pack(anchor='w', pady=(0, 10))
        ttk.Label(f, text=UPLOAD_CHOICE_UPLOAD, font=('맑은 고딕', 10), wraplength=400, justify=tk.LEFT).pack(anchor='w', pady=(0, 24))
        btn_wrap = ttk.Frame(f)
        btn_wrap.pack(fill=tk.X, pady=(0, 8))
        btn_f = ttk.Frame(btn_wrap)
        btn_f.pack(anchor='center')
        def on_template():
            result[0] = 'template'
            win.destroy()
        def on_upload():
            result[0] = 'upload'
            win.destroy()
        ttk.Button(btn_f, text=UPLOAD_CHOICE_BTN_PASTE, command=on_template, width=22).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_f, text=UPLOAD_CHOICE_BTN_UPLOAD, command=on_upload, width=22).pack(side=tk.LEFT)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.wait_window(win)
        return result[0]
    
    def _create_and_open_template(self, kind: str) -> None:
        """kind: inbound | outbound_tonbag | location. 템플릿 생성 → 저장 → 기본 앱으로 열기."""
        from ..utils.constants import filedialog
        try:
            pass
        except ImportError:
            CustomMessageBox.showerror(self.root, "오류", "openpyxl이 필요합니다.\npip install openpyxl")
            return
        if kind == 'inbound':
            file_path = filedialog.asksaveasfilename(
                parent=self.root,
                title="입고 템플릿 저장 후 열기",
                defaultextension=".xlsx",
                initialfile="입고_템플릿.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if file_path and self._write_inbound_template_to(file_path):
                if _open_file_with_default_app(file_path, parent=self.root):
                    CustomMessageBox.showinfo(self.root, "템플릿 열림",
                        "입고 템플릿을 열었습니다.\n데이터를 채운 뒤 저장하고,\n[Excel 입고] → [엑셀 파일 업로드]에서 해당 파일을 선택하세요.")
        elif kind == 'outbound_tonbag':
            file_path = filedialog.asksaveasfilename(
                parent=self.root,
                title="출고 템플릿 저장 후 열기",
                defaultextension=".xlsx",
                initialfile="출고_톤백수_템플릿.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if file_path and self._write_outbound_tonbag_template_to(file_path):
                if _open_file_with_default_app(file_path, parent=self.root):
                    CustomMessageBox.showinfo(self.root, "템플릿 열림",
                        "출고 템플릿을 열었습니다.\n데이터를 채운 뒤 저장하고,\n[Excel 입고] → [엑셀 파일 업로드]에서 해당 파일을 선택하세요.")
        elif kind == 'location':
            # 톤백 위치 업로드 기능 제거됨
            return

    def _show_inbound_spreadsheet_dialog(self) -> None:
        """입고 형식 스프레드시트를 화면에 띄우고, 붙여넣기 후 DB 반영."""
        from ..utils.paste_table_dialog import show_paste_table_dialog
        columns = [(c[0], c[1], c[2]) for c in self.INVENTORY_TEMPLATE_COLUMNS]
        show_paste_table_dialog(
            self.root,
            title="📝 입고 데이터 (붙여넣기)",
            columns=columns,
            instruction="아래 표에 Excel 등에서 복사한 입고 데이터를 붙여넣기(Ctrl+V) 한 뒤 [DB 반영]을 누르세요. 형식은 입고 템플릿과 동일합니다.",
            confirm_text="DB 반영",
            cancel_text="취소",
            on_confirm=self._on_inbound_paste_confirm,
            min_size=(800, 440),
        )

    def _on_inbound_paste_confirm(self, rows: list) -> None:
        """스프레드시트 붙여넣기 확인 시 DataFrame으로 변환 후 입고 처리."""
        from ..utils.constants import pd, HAS_PANDAS
        if not HAS_PANDAS or not rows:
            CustomMessageBox.showwarning(self.root, "경고", "입력된 데이터가 없습니다.")
            return
        df = pd.DataFrame(rows)
        self._import_inbound_from_dataframe(df)

    def _show_return_inbound_spreadsheet_dialog(self) -> None:
        """v6.9.5 [중복제거]: inbound_handlers 로 위임.
        반품 입고 다이얼로그는 inbound_handlers._show_return_inbound_spreadsheet_dialog()
        가 단일 엔트리 포인트. import_handlers 에서 직접 호출 금지.
        """
        # v6.9.5: 동일 기능 두 곳 중복 구현 → inbound_handlers 단일화
        if hasattr(self, '_inbound_handler_ref') and self._inbound_handler_ref:
            self._inbound_handler_ref._show_return_inbound_spreadsheet_dialog()
        else:
            # fallback: inbound_handlers 가 없는 환경 (단독 실행 등)
            import logging as _lg
            _lg.getLogger(__name__).warning(
                "[DEPRECATED] import_handlers._show_return_inbound_spreadsheet_dialog: "
                "inbound_handlers 를 통해 호출하세요 (v6.9.5 단일화)"
            )
            from ..utils.paste_table_dialog import show_paste_table_dialog
            columns = [(c[0], c[1], c[2]) for c in self.RETURN_TEMPLATE_COLUMNS]
            show_paste_table_dialog(
                self.root,
                title="📝 반품 입고 데이터 (붙여넣기)",
                columns=columns,
                instruction="아래 표에서 반품 데이터를 붙여넣기(Ctrl+V) 한 뒤 [DB 반영]을 누르세요.",
                confirm_text="DB 반영",
                cancel_text="취소",
                on_confirm=self._on_return_inbound_paste_confirm,
                min_size=(800, 440),
            )

    def _bulk_import_inventory_simple(self, file_path: str = None) -> None:
        """
        간단한 Excel 입고 처리.
        v5.8.9: file_path가 None이면 [템플릿 열기] vs [파일 업로드] 선택 후 진행.
        """
        from ..utils.constants import filedialog
        
        if file_path is None:
            choice = self._show_template_or_upload_choice("수동 입고", "inbound")
            if choice is None:
                return
            if choice == 'template':
                self._show_inbound_spreadsheet_dialog()
                return
            file_path = filedialog.askopenfilename(
                parent=self.root,
                title="입고 Excel 파일 선택",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ]
            )
        
        if not file_path:
            return  # 취소됨
        
        # 파일 타입 자동 감지
        file_type = self._detect_excel_type(file_path)
        
        if file_type == 'inbound':
            self._import_inbound_excel_auto(file_path)
        elif file_type == 'outbound':
            self._import_outbound_excel_auto(file_path)
        elif file_type == 'location':
            CustomMessageBox.showinfo(self.root, "안내", "톤백 위치 업로드 기능은 제거되었습니다.")
            return
        else:
            # 타입 불명 - 사용자에게 선택 요청
            if CustomMessageBox.askyesno(self.root, "파일 타입 확인",
                f"파일 타입을 자동 감지할 수 없습니다.\n\n"
                f"파일: {file_path}\n\n"
                "입고 Excel로 처리할까요?\n"
                "(아니오를 선택하면 출고로 처리합니다)"):
                self._import_inbound_excel_auto(file_path)
            else:
                self._import_outbound_excel_auto(file_path)
    
    def _detect_excel_type(self, file_path: str) -> str:
        """
        Excel 파일 타입 자동 감지
        
        Returns:
            'inbound' | 'outbound' | 'location' | 'unknown'
        """
        from ..utils.constants import pd, HAS_PANDAS
        
        if not HAS_PANDAS:
            return 'unknown'
        
        try:
            df = pd.read_excel(file_path, nrows=10)
            columns_lower = [str(c).lower() for c in df.columns]
            
            # 입고 Excel 특징: lot_no, sap_no, product, weight
            inbound_keywords = ['lot', 'sap', 'product', 'weight', 'qty']
            inbound_score = sum(1 for kw in inbound_keywords 
                               if any(kw in c for c in columns_lower))
            
            # 출고 Excel 특징: outbound, customer, destination, tonbag_count(간편 출고 템플릿)
            outbound_keywords = ['outbound', 'customer', 'destination', 'ship', 'deliver', 'tonbag']
            outbound_score = sum(1 for kw in outbound_keywords 
                                if any(kw in c for c in columns_lower))
            
            # 위치 Excel 특징: location, zone, rack
            location_keywords = ['location', 'zone', 'rack', 'position', 'area']
            location_score = sum(1 for kw in location_keywords 
                                if any(kw in c for c in columns_lower))
            
            scores = {
                'inbound': inbound_score,
                'outbound': outbound_score,
                'location': location_score
            }
            
            max_type = max(scores, key=scores.get)
            if scores[max_type] >= 2:
                return max_type
            
            return 'unknown'
            
        except (FileNotFoundError, OSError, PermissionError) as e:
            logger.error(f"Excel 타입 감지 오류: {e}")
            return 'unknown'
    
    def _is_manual_inbound_template(self, df) -> bool:
        """수동 입고 템플릿 여부: 2행이 DB필드명(lot_no, net_weight, mxbg_pallet)인 경우"""
        cols = [str(c).strip().lower() for c in df.columns]
        return ('lot_no' in cols or 'lot no' in cols) and ('net_weight' in cols or 'net weight' in cols) and ('mxbg_pallet' in cols or 'mxbg' in cols)

    def _read_inbound_excel_with_header(self, file_path: str):
        """
        입고 Excel 읽기 — 헤더 행 자동 감지.
        - 프로그램 내보내기: Row1=타이틀, Row2=빈줄, Row3=헤더 → header=2
        - 일반/템플릿: Row1=헤더 → header=0
        """
        from ..utils.constants import pd
        cols_lower = lambda cols: [str(c).strip().lower() for c in cols]

        def _has_lot_header(cols):
            c = cols_lower(cols)
            return any('lot' in x and ('no' in x or x == 'lot_no') for x in c)

        # v8.4.2: SAP NO/LOT NO 앞자리 0 보존 — 해당 컬럼만 str로 강제 읽기
        _STR_COLS = {'SAP NO': str, 'LOT NO': str, 'BL NO': str, 'CONTAINER': str}

        for header_row in [2, 0]:  # 2=내보내기형식(재고리스트/톤백리스트), 0=일반
            try:
                # 1차: str 강제 적용 시도
                df = pd.read_excel(file_path, header=header_row, dtype=_STR_COLS)
                if df.empty or len(df.columns) < 2:
                    continue
                if _has_lot_header(df.columns):
                    return df
                # 2차: str 적용 실패 시 기본 읽기
                df = pd.read_excel(file_path, header=header_row)
                if not df.empty and len(df.columns) >= 2 and _has_lot_header(df.columns):
                    return df
            except (ValueError, KeyError):
                continue
        return pd.read_excel(file_path, dtype=_STR_COLS)

    def _import_inbound_excel_auto(self, file_path: str) -> None:
        """
        입고 Excel 자동 처리
        
        - 수동 입고 템플릿(2행=DB필드명)이면 process_inbound 호출 → 톤백+샘플 자동 생성
        - 그 외 레거시는 add_inventory_from_dict
        - 동일 이름 파일 열림 시 순번 붙여서 읽기 제안
        """
        from ..utils.constants import pd, HAS_PANDAS, HAS_COLUMN_ALIASES, ColumnMapper
        from ..utils.safe_utils import safe_str, safe_float, safe_date
        from ..utils.excel_file_helper import (
            _get_path_with_sequence, _try_copy_open_file,
            _MSG_SAME_OPEN, _MSG_CLOSE_RETRY,
        )
        
        if not HAS_PANDAS:
            CustomMessageBox.showerror(self.root, "오류", "pandas가 설치되지 않았습니다.")
            return
        
        path_to_read = file_path
        while True:
            try:
                # 수동 입고 템플릿: 2행이 헤더(DB필드명)
                df_header1 = pd.read_excel(path_to_read, header=1)
                if not df_header1.empty and self._is_manual_inbound_template(df_header1):
                    self._import_inbound_manual_template(path_to_read, df_header1)
                    return
                
                # 재고리스트/톤백리스트 내보내기 형식(header=2) 또는 일반(header=0) 자동 감지
                df = self._read_inbound_excel_with_header(path_to_read)
                break
            except (OSError, IOError, PermissionError) as e:
                logger.debug(f"Excel 읽기 실패: {e}")
                if not CustomMessageBox.askyesno(self.root, "파일 읽기", _MSG_SAME_OPEN):
                    CustomMessageBox.showwarning(self.root, "파일 읽기", _MSG_CLOSE_RETRY)
                    return
                copy_path = _get_path_with_sequence(path_to_read)
                if not _try_copy_open_file(path_to_read, copy_path):
                    CustomMessageBox.showwarning(self.root, "파일 읽기", _MSG_CLOSE_RETRY)
                    return
                path_to_read = copy_path
        
        # break 후 — 루프 밖에서 처리
        if df.empty:
            CustomMessageBox.showwarning(self.root, "경고", "빈 Excel 파일입니다.")
            return
        
        self._log(f"📥 입고 Excel 로드: {len(df)}행")
        
        # 컬럼 매핑 (Column Alias 시스템 사용)
        if HAS_COLUMN_ALIASES and ColumnMapper:
            mapper = ColumnMapper()
            col_map = {}
            for col in df.columns:
                std_key = mapper.get_standard_key(str(col))
                if std_key:
                    col_map[std_key] = col
        else:
            # 기본 매핑
            col_map = self._get_basic_column_mapping(df.columns)
        
        self._log(f"   컬럼 매핑: {col_map}")
        
        # 필수 컬럼 확인 → 누락 시 매핑 다이얼로그
        required = ['lot_no']
        missing = [r for r in required if r not in col_map]
        if missing:
            try:
                from ..dialogs.column_mapper_dialog import ColumnMapperDialog
                sample_data = [list(df.iloc[i].values) for i in range(min(3, len(df)))]
                mapper_dlg = ColumnMapperDialog(
                    self.root, list(df.columns), sample_data)
                manual_map = mapper_dlg.get_result()
                if manual_map:
                    col_map = manual_map
                    self._log(f"   수동 매핑: {col_map}")
                else:
                    return  # 사용자 취소
            except (ImportError, ModuleNotFoundError) as e:
                logger.debug(f"매핑 다이얼로그 오류: {e}")
                CustomMessageBox.showerror(self.root, "오류", f"필수 컬럼 누락: {missing}")
                return
        
        # 데이터 변환 및 저장
        added_lots = 0
        added_tonbags = 0
        skipped = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                lot_no = safe_str(row.get(col_map.get('lot_no', 'lot_no')))
                if not lot_no:
                    skipped += 1
                    continue
                
                # 데이터 준비 (v3.8.8: 18열 전체 매핑)
                # 중량: net_weight 우선, 없으면 qty_mt * 1000
                _net_raw = row.get(col_map.get('net_weight', ''), None)
                if _net_raw is not None and safe_float(_net_raw) > 0:
                    _net_kg = safe_float(_net_raw)
                else:
                    _net_kg = safe_float(row.get(col_map.get('qty_mt', 'weight'), 0)) * 1000
                
                _gross_raw = row.get(col_map.get('gross_weight', ''), None)
                _gross_kg = safe_float(_gross_raw) if _gross_raw is not None and safe_float(_gross_raw) > 0 else _net_kg
                
                data = {
                    'lot_no': lot_no,
                    'sap_no': safe_str(row.get(col_map.get('sap_no', 'sap_no'), '')),
                    'bl_no': safe_str(row.get(col_map.get('bl_no', 'bl_no'), '')),
                    'container_no': safe_str(row.get(col_map.get('container_no', 'container'), '')),
                    'product': safe_str(row.get(col_map.get('product', 'product'), '')),
                    'product_code': safe_str(row.get(col_map.get('product_code', ''), '')),
                    'lot_sqm': safe_str(row.get(col_map.get('lot_sqm', ''), '')),
                    'mxbg_pallet': int(safe_float(row.get(col_map.get('mxbg_pallet', 'mxbg_pallet'), 10))),
                    'net_weight': _net_kg,
                    'gross_weight': _gross_kg,
                    'initial_weight': _net_kg,
                    'current_weight': _net_kg,
                    'salar_invoice_no': safe_str(row.get(col_map.get('salar_invoice_no', ''), '')),
                    'ship_date': safe_date(row.get(col_map.get('ship_date', ''), '')),
                    'arrival_date': safe_date(row.get(col_map.get('arrival_date', 'arrival_date'), '')),
                    'free_time': int(safe_float(row.get(col_map.get('free_time', ''), 0))),
                    'warehouse': safe_str(row.get(col_map.get('warehouse', 'warehouse'), '광양')),
                    'stock_date': safe_date(row.get(col_map.get('stock_date', 'stock_date'), '')),
                    'location': safe_str(row.get(col_map.get('location', ''), '')),
                    'remark': safe_str(row.get(col_map.get('remark', ''), '')),
                    'status': STATUS_AVAILABLE,
                }
                
                # DB 저장
                result = self.engine.add_inventory_from_dict(data)
                if result.get('success'):
                    added_lots += 1
                    added_tonbags += result.get('tonbags', 0)
                else:
                    errors.append(f"행 {idx+2}: {result.get('message', '알 수 없는 오류')}")
                
            except (ValueError, TypeError, AttributeError) as e:
                errors.append(f"행 {idx+2}: {str(e)}")
        
        # 결과 보고
        self._log(f"✅ 입고 완료: {added_lots}개 LOT, {added_tonbags}개 톤백")
        
        # v3.8.4: 처리 완료 파일 아카이브
        if added_lots > 0 and hasattr(self, '_archive_processed_file'):
            self._archive_processed_file(file_path, 'inbound')
        
        if errors:
            error_msg = '\n'.join(errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... 외 {len(errors)-10}개 오류"
            CustomMessageBox.showwarning(self.root, "일부 오류", 
                f"입고 완료: {added_lots}개 LOT\n\n오류:\n{error_msg}")
        else:
            # 완료 알림
            try:
                import winsound
                winsound.MessageBeep(winsound.MB_ICONASTERISK)
            except (ImportError, RuntimeError) as e:
                logger.debug(f"{type(e).__name__}: {e}")
            
            CustomMessageBox.showinfo(self.root, "✅ 입고 완료",
                f"입고 처리가 완료되었습니다!\n\n"
                f"📦 추가된 LOT: {added_lots:,}개\n"
                f"📦 생성된 톤백: {added_tonbags:,}개\n"
                f"⏭️ 스킵: {skipped:,}개")
        
        # UI 새로고침 (상위 메뉴 → 하단 4개 탭 반영)
        self._safe_refresh()
    def _import_inbound_manual_template(self, file_path: str, df) -> None:
        """수동 입고 템플릿 → process_inbound 호출 (1 LOT = N톤백+1샘플)."""
        from ..utils.safe_utils import safe_str, safe_float, safe_date
        
        df = df.copy()
        df.columns = [str(c).strip().lower().replace(' ', '_').rstrip('*').strip('_') for c in df.columns]
        
        def _get(row, *keys):
            for k in keys:
                for c in df.columns:
                    if c == k or (c and c.replace('_', '') == k.replace('_', '')):
                        return row.get(c)
            return None
        
        def _row_to_packing(row) -> dict:
            net_kg = safe_float(_get(row, 'net_weight', 'netweight'), 0)
            gross_kg = safe_float(_get(row, 'gross_weight', 'grossweight')) or net_kg
            mxbg = int(safe_float(_get(row, 'mxbg_pallet', 'mxbg'), 10) or 10)
            return {
                'lot_no': safe_str(_get(row, 'lot_no', 'lotno') or '').strip(),
                'sap_no': safe_str(_get(row, 'sap_no', 'sapno') or ''),
                'bl_no': safe_str(_get(row, 'bl_no', 'blno') or ''),
                'container_no': safe_str(_get(row, 'container_no', 'containerno') or ''),
                'product': safe_str(_get(row, 'product') or '') or 'LITHIUM CARBONATE',
                'product_code': safe_str(_get(row, 'product_code') or ''),
                'lot_sqm': safe_str(_get(row, 'lot_sqm') or ''),
                'mxbg_pallet': mxbg,
                'net_weight': net_kg,
                'gross_weight': gross_kg,
                'salar_invoice_no': safe_str(_get(row, 'salar_invoice_no') or ''),
                'ship_date': safe_date(_get(row, 'ship_date') or ''),
                'arrival_date': safe_date(_get(row, 'arrival_date') or ''),
                'free_time': int(safe_float(_get(row, 'free_time') or 0) or 0),
                'warehouse': safe_str(_get(row, 'warehouse') or '') or '광양',
            }
        
        rows_valid = []
        missing_rows = []  # 필수 누락 행 번호 (데이터 입력했는데 LOT/순중량 비어 있는 행)
        for idx, row in df.iterrows():
            packing = _row_to_packing(row)
            excel_row = idx + 4
            has_any = bool(packing['lot_no'] or _get(row, 'product') or packing.get('mxbg_pallet') or (packing['net_weight'] or 0) > 0)
            if has_any and (not packing['lot_no'] or (packing['net_weight'] or 0) <= 0):
                missing_rows.append(excel_row)
                continue
            if not packing['lot_no'] or (packing['net_weight'] or 0) <= 0:
                continue
            rows_valid.append((excel_row, packing))
        
        if missing_rows:
            CustomMessageBox.showerror(
                self.root, "필수 항목 누락",
                f"다음 행에 필수 항목(LOT 번호, 순중량)이 비어 있거나 0입니다:\n\n"
                f"행 {', '.join(map(str, missing_rows[:15]))}"
                + (f" 외 {len(missing_rows)-15}건" if len(missing_rows) > 15 else "")
                + "\n\n엑셀에서 확인한 뒤 다시 업로드하세요."
            )
            return
        if not rows_valid:
            CustomMessageBox.showwarning(self.root, "경고", "유효한 입고 데이터가 없습니다.\nLOT 번호와 순중량(NET Kg)을 확인하세요.")
            return
        
        preview_lines = []
        for i, (rnum, p) in enumerate(rows_valid[:5]):
            preview_lines.append(f"  {i+1}. 행{rnum}: LOT {p['lot_no']} | SAP {p['sap_no']} | {p['net_weight']:.0f}kg | 톤백 {p['mxbg_pallet']}개")
        if len(rows_valid) > 5:
            preview_lines.append(f"  ... 외 {len(rows_valid)-5}건")
        msg = f"수동 입고 템플릿 인식: 총 {len(rows_valid)}건\n\n미리보기:\n" + "\n".join(preview_lines) + "\n\nDB에 반영(process_inbound) 하시겠습니까?"
        if not CustomMessageBox.askyesno(self.root, "입고 확인", msg):
            return
        
        self._log(f"📥 수동 입고 템플릿: {len(rows_valid)}건 → process_inbound")
        added_lots = 0
        added_tonbags = 0
        errors = []
        for rnum, packing in rows_valid:
            result = self.engine.process_inbound(packing)
            if result.get('success'):
                added_lots += 1
                added_tonbags += result.get('created_tonbags', 0)
            else:
                errors.append(f"행{rnum} {packing.get('lot_no')}: {result.get('message', '')}; {result.get('errors', [])}")
        
        self._safe_refresh()
        result_msg = f"✅ 입고 완료: {added_lots}개 LOT, {added_tonbags}개 톤백 생성"
        if errors:
            result_msg += f"\n\n❌ 실패 {len(errors)}건:\n" + "\n".join(errors[:5])
        CustomMessageBox.showinfo(self.root, "입고 결과", result_msg)

    def _import_inbound_from_dataframe(self, df) -> None:
        """붙여넣기 테이블에서 받은 DataFrame으로 입고 처리 (파일 없이). 수동 입고 템플릿 형식이면 process_inbound, 아니면 add_inventory_from_dict."""
        from ..utils.constants import HAS_PANDAS
        from ..utils.safe_utils import safe_str, safe_float, safe_date

        if not HAS_PANDAS or df.empty:
            CustomMessageBox.showwarning(self.root, "경고", "유효한 데이터가 없습니다.")
            return
        df = df.copy()
        df.columns = [str(c).strip().lower().replace(' ', '_').rstrip('*').strip('_') for c in df.columns]
        if self._is_manual_inbound_template(df):
            self._import_inbound_manual_template('', df)
            return
        col_map = {c: c for c in df.columns}
        required = ['lot_no']
        missing = [r for r in required if r not in col_map]
        if missing:
            CustomMessageBox.showerror(self.root, "오류", f"필수 컬럼 누락: {missing}\nLOT NO 등 입고 형식에 맞게 붙여넣기 하세요.")
            return
        added_lots = 0
        added_tonbags = 0
        skipped = 0
        errors = []
        for idx, row in df.iterrows():
            try:
                lot_no = safe_str(row.get(col_map.get('lot_no', 'lot_no')))
                if not lot_no:
                    skipped += 1
                    continue
                _net_raw = row.get(col_map.get('net_weight', ''), None)
                _net_kg = safe_float(_net_raw) if _net_raw is not None and safe_float(_net_raw) > 0 else safe_float(row.get(col_map.get('qty_mt', 'weight'), 0)) * 1000
                _gross_raw = row.get(col_map.get('gross_weight', ''), None)
                _gross_kg = safe_float(_gross_raw) if _gross_raw is not None and safe_float(_gross_raw) > 0 else _net_kg
                data = {
                    'lot_no': lot_no,
                    'sap_no': safe_str(row.get(col_map.get('sap_no', 'sap_no'), '')),
                    'bl_no': safe_str(row.get(col_map.get('bl_no', 'bl_no'), '')),
                    'container_no': safe_str(row.get(col_map.get('container_no', 'container'), '')),
                    'product': safe_str(row.get(col_map.get('product', 'product'), '')),
                    'product_code': safe_str(row.get(col_map.get('product_code', ''), '')),
                    'lot_sqm': safe_str(row.get(col_map.get('lot_sqm', ''), '')),
                    'mxbg_pallet': int(safe_float(row.get(col_map.get('mxbg_pallet', 'mxbg_pallet'), 10))),
                    'net_weight': _net_kg,
                    'gross_weight': _gross_kg,
                    'initial_weight': _net_kg,
                    'current_weight': _net_kg,
                    'salar_invoice_no': safe_str(row.get(col_map.get('salar_invoice_no', ''), '')),
                    'ship_date': safe_date(row.get(col_map.get('ship_date', ''), '')),
                    'arrival_date': safe_date(row.get(col_map.get('arrival_date', 'arrival_date'), '')),
                    'free_time': int(safe_float(row.get(col_map.get('free_time', ''), 0))),
                    'warehouse': safe_str(row.get(col_map.get('warehouse', 'warehouse'), '광양')),
                    'stock_date': safe_date(row.get(col_map.get('stock_date', 'stock_date'), '')),
                    'location': safe_str(row.get(col_map.get('location', ''), '')),
                    'remark': safe_str(row.get(col_map.get('remark', ''), '')),
                    'status': STATUS_AVAILABLE,
                }
                result = self.engine.add_inventory_from_dict(data)
                if result.get('success'):
                    added_lots += 1
                    added_tonbags += result.get('tonbags', 0)
                else:
                    errors.append(f"행 {idx+2}: {result.get('message', '알 수 없는 오류')}")
            except (ValueError, TypeError, AttributeError) as e:
                errors.append(f"행 {idx+2}: {str(e)}")
        self._log(f"✅ 입고(붙여넣기) 완료: {added_lots}개 LOT, {added_tonbags}개 톤백")
        if errors:
            error_msg = '\n'.join(errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... 외 {len(errors)-10}개 오류"
            CustomMessageBox.showwarning(self.root, "일부 오류", f"입고 완료: {added_lots}개 LOT\n\n오류:\n{error_msg}")
        else:
            CustomMessageBox.showinfo(self.root, "✅ 입고 완료",
                f"입고 처리가 완료되었습니다!\n\n📦 추가된 LOT: {added_lots:,}개\n📦 생성된 톤백: {added_tonbags:,}개\n⏭️ 스킵: {skipped:,}개")
        self._safe_refresh()
    def _get_basic_column_mapping(self, columns) -> Dict[str, str]:
        """기본 컬럼 매핑 (Column Alias 없을 때)"""
        col_map = {}
        columns_lower = {str(c).lower().strip(): c for c in columns}
        
        mappings = {
            'lot_no': ['lot_no', 'lot no', 'lot', 'lotno'],
            'sap_no': ['sap_no', 'sap no', 'sap', 'sapno'],
            'bl_no': ['bl_no', 'bl no', 'bl', 'b/l', 'blno'],
            'product': ['product', 'material', '제품', '품목'],
            'container_no': ['container', 'container_no', 'cntr', '컨테이너'],
            'qty_mt': ['qty', 'weight', 'qty_mt', 'qty(mt)', '중량', '수량'],
            'arrival_date': ['arrival_date', 'arrival', '입항일'],
            'stock_date': ['stock_date', 'inbound_date', '입고일'],
            'warehouse': ['warehouse', 'wh', '창고'],
        }
        
        for std_key, candidates in mappings.items():
            for candidate in candidates:
                if candidate in columns_lower:
                    col_map[std_key] = columns_lower[candidate]
                    break
        
        return col_map
    
    def _import_outbound_excel_auto(self, file_path: str) -> None:
        """출고 Excel 자동 처리. 톤백 수 컬럼 있으면 FIFO N개 합산→weight_kg, 없으면 LOT 전량."""
        from ..utils.constants import pd, HAS_PANDAS
        from ..utils.safe_utils import safe_str, safe_float
        
        if not HAS_PANDAS:
            CustomMessageBox.showerror(self.root, "오류", "pandas가 설치되지 않았습니다.")
            return
        
        try:
            df = pd.read_excel(file_path)
            if df.empty:
                CustomMessageBox.showwarning(self.root, "경고", "빈 Excel 파일입니다.")
                return
            df.columns = [str(c).strip().lower().replace(' ', '_').rstrip('*').strip('_') for c in df.columns]
            cols = list(df.columns)
            col_lot = next((c for c in cols if c in ('lot_no', 'lotno', 'lot')), None) or (cols[0] if cols else None)
            col_customer = next((c for c in cols if c in ('customer', '고객', 'sold_to', 'destination') or (c and 'customer' in c)), None)
            col_sale_ref = next((c for c in cols if c in ('sale_ref', 'saleref', 'sales_ref', 'reference', 'ref') or (c and 'sale' in c and 'ref' in c)), None)
            col_tonbag = next((c for c in cols if c and ('tonbag' in c or '톤백' in c) and ('count' in c or '수' in c)), None)
            
            # 필수 항목 사전 검증: 데이터가 있는데 LOT 또는 톤백 수가 비어 있으면 에러
            missing_outbound_rows = []
            for idx, row in df.iterrows():
                excel_row = idx + 2
                lot = safe_str(row.get(col_lot, '') if col_lot else '').strip()
                n = int(safe_float(row.get(col_tonbag, 0) or 0)) if col_tonbag and col_tonbag in row.index else 1
                has_any = bool(lot or (col_customer and row.get(col_customer)) or (col_tonbag and row.get(col_tonbag) is not None))
                if has_any and (not lot or (col_tonbag and n <= 0)):
                    missing_outbound_rows.append(excel_row)
            if missing_outbound_rows:
                CustomMessageBox.showerror(
                    self.root, "필수 항목 누락",
                    f"다음 행에 LOT 번호 또는 출고 톤백 수가 비어 있습니다:\n\n"
                    f"행 {', '.join(map(str, missing_outbound_rows[:15]))}"
                    + (f" 외 {len(missing_outbound_rows)-15}건" if len(missing_outbound_rows) > 15 else "")
                    + "\n\n확인한 뒤 다시 업로드하세요."
                )
                return
            
            self._log(f"📤 출고 Excel 로드: {len(df)}행 (톤백수모드={bool(col_tonbag)})")
            processed = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    lot_no = safe_str(row.get(col_lot, '')).strip()
                    if not lot_no:
                        continue
                    customer = safe_str(row.get(col_customer, '')).strip() if col_customer else ''
                    sale_ref = safe_str(row.get(col_sale_ref, '')).strip() if col_sale_ref else ''
                    
                    if col_tonbag and col_tonbag in row.index:
                        n = int(safe_float(row.get(col_tonbag, 0)) or 0)
                        if n <= 0:
                            errors.append(f"행 {idx+2}: 톤백 수 0 — {lot_no}")
                            continue
                        tonbags = self.engine.db.fetchall(
                            """SELECT id, weight FROM inventory_tonbag
                               WHERE lot_no = ? AND status = 'AVAILABLE' AND COALESCE(is_sample, 0) = 0
                               ORDER BY sub_lt DESC LIMIT ?""",
                            (lot_no, n)
                        )
                        if not tonbags or len(tonbags) < n:
                            errors.append(f"행 {idx+2}: 판매가능 톤백 부족 — {lot_no} (요청 {n}개, 판매가능 {len(tonbags) if tonbags else 0}개)")
                            continue
                        weight_kg = sum(float(t.get('weight') or 0) for t in tonbags[:n])
                    else:
                        lot_row = self.engine.db.fetchone(
                            "SELECT current_weight FROM inventory WHERE lot_no = ?", (lot_no,)
                        )
                        if not lot_row:
                            errors.append(f"행 {idx+2}: LOT 없음 — {lot_no}")
                            continue
                        weight_kg = float(lot_row.get('current_weight') or 0)
                        if weight_kg <= 0:
                            errors.append(f"행 {idx+2}: 판매가능 재고 0 — {lot_no}")
                            continue
                    
                    result = self.engine.process_outbound([{
                        'lot_no': lot_no,
                        'weight_kg': weight_kg,
                        'qty_mt': weight_kg / 1000.0,
                        'customer': customer,
                        'sold_to': customer,
                        'sale_ref': sale_ref,
                    }], source='EXCEL', stop_at_picked=False)
                    if result.get('success'):
                        processed += 1
                    else:
                        errors.append(f"행 {idx+2}: {result.get('message')}")
                except (ValueError, TypeError, AttributeError) as e:
                    errors.append(f"행 {idx+2}: {str(e)}")
            
            self._log(f"✅ 출고 완료: {processed}건")
            if errors:
                error_msg = '\n'.join(errors[:10])
                CustomMessageBox.showwarning(self.root, "일부 오류",
                    f"출고 완료: {processed}건\n\n오류:\n{error_msg}")
            else:
                CustomMessageBox.showinfo(self.root, "✅ 출고 완료", f"출고 처리 완료: {processed}건")
            self._safe_refresh()
        except (RuntimeError, ValueError) as e:
            logger.error(f"출고 Excel 처리 오류: {e}")
            CustomMessageBox.showerror(self.root, "오류", f"출고 처리 오류: {e}")

    # ═══════════════════════════════════════════════════════
    # v5.8.9: 출고 템플릿(톤백 수) 다운로드
    # ═══════════════════════════════════════════════════════
    
    def _write_outbound_tonbag_template_to(self, file_path: str) -> bool:
        """출고 템플릿 = 톤백(언로케이션) 리스트와 동일 형식. 필수(lot_no, tonbag_count, customer)만 색상 표시."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "출고 데이터"
            headers = self.TONBAG_TEMPLATE_COLUMNS
            required_outbound = {'lot_no', 'tonbag_count', 'customer'}
            header_font = Font(bold=True, color="FFFFFF", size=11)
            required_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            optional_fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
            sample_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ncols = len(headers)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            ws['A1'] = "출고 템플릿 v5.8.9 — 톤백 리스트와 동일 형식. 진한색 컬럼은 필수(LOT NO, 출고 톤백 수, 고객명). 2행=DB필드명."
            ws['A1'].font = Font(bold=True, size=11, color="2C3E50")
            ws.row_dimensions[1].height = 28
            for col in range(1, ncols + 1):
                ws.cell(row=1, column=col).border = thin_border
            for col, (db_field, display, width, req_out, _) in enumerate(headers, 1):
                ws.cell(row=2, column=col, value=db_field)
                ws.cell(row=2, column=col).font = Font(size=8, color="999999")
                ws.cell(row=2, column=col).border = thin_border
            ws.row_dimensions[2].height = 14
            for col, (db_field, display, width, req_out, _) in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=display + (' *' if db_field in required_outbound else ''))
                cell.font = header_font
                cell.fill = required_fill if db_field in required_outbound else optional_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(width, 24)
            for row in range(4, 8):
                for c in range(1, ncols + 1):
                    cell = ws.cell(row=row, column=c, value='')
                    cell.fill = sample_fill
                    cell.border = thin_border
            wb.save(file_path)
            self._log(f"✅ 출고 템플릿 저장: {file_path}")
            return True
        except ImportError:
            CustomMessageBox.showerror(self.root, "오류", "openpyxl이 필요합니다.\npip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"출고 템플릿 생성 오류: {e}")
            CustomMessageBox.showerror(self.root, "오류", f"템플릿 저장 실패:\n{e}")
            return False
    

    def _write_inbound_template_to(self, file_path: str) -> bool:
        """입고 템플릿 = 재고 리스트와 동일 형식. 필수(lot_no, product, mxbg_pallet, net_weight)만 색상 표시."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "입고 데이터"
            headers = self.INVENTORY_TEMPLATE_COLUMNS
            required_inbound = {'lot_no', 'product', 'mxbg_pallet', 'net_weight'}
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            required_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            optional_fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
            sample_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            ncols = len(headers)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            ws['A1'] = "입고 템플릿 v5.8.9 — 재고 리스트와 동일 형식. 진한색 컬럼은 필수 입력. 2행=DB필드명. 채운 뒤 Excel 입고로 업로드하세요."
            ws['A1'].font = Font(bold=True, size=11, color="2C3E50")
            ws.row_dimensions[1].height = 28
            
            for col, (db_field, display, width, _) in enumerate(headers, 1):
                ws.cell(row=2, column=col, value=db_field)
                ws.cell(row=2, column=col).font = Font(size=8, color="999999")
                ws.cell(row=2, column=col).border = thin_border
            ws.row_dimensions[2].height = 14
            for col in range(1, ncols + 1):
                ws.cell(row=1, column=col).border = thin_border

            for col, (db_field, display, width, _) in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=display + (' *' if db_field in required_inbound else ''))
                cell.font = header_font
                cell.fill = required_fill if db_field in required_inbound else optional_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                col_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[col_letter].width = min(width, 20)

            for row in range(4, 8):
                for c in range(1, ncols + 1):
                    cell = ws.cell(row=row, column=c, value='')
                    cell.fill = sample_fill
                    cell.border = thin_border

            ws2 = wb.create_sheet("안내")
            ws2.cell(row=1, column=1, value="필수 항목(진한색) 없으면 업로드 시 오류가 납니다.")
            ws2.cell(row=1, column=1).font = Font(bold=True)
            for i, (field, display, _, req) in enumerate(headers, 2):
                ws2.cell(row=i, column=1, value=field)
                ws2.cell(row=i, column=2, value=display + (" (필수)" if req else ""))
                ws2.cell(row=i, column=1).border = thin_border
                ws2.cell(row=i, column=2).border = thin_border
            ws2.cell(row=1, column=2).border = thin_border
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 24
            
            try:
                from gui_app_modular.utils.report_footer import add_gy_logistics_footer
                add_gy_logistics_footer(ws)
            except (ImportError, ModuleNotFoundError) as _e:
                logger.debug(f'Suppressed: {_e}')
            wb.save(file_path)
            self._log(f"✅ 입고 템플릿 저장: {file_path}")
            return True
        except ImportError:
            CustomMessageBox.showerror(self.root, "오류", "openpyxl 패키지가 필요합니다.\npip install openpyxl")
            return False
        except (ImportError, ModuleNotFoundError, OSError) as e:
            logger.error(f"입고 템플릿 생성 오류: {e}")
            CustomMessageBox.showerror(self.root, "오류", f"템플릿 생성 실패:\n{e}")
            return False
    

    def _archive_processed_file(self, file_path: str, category: str = 'inbound') -> str:
        """
        처리 완료된 Excel 파일을 archive 폴더로 복사
        
        Args:
            file_path: 원본 파일 경로
            category: 'inbound' | 'outbound'
        
        Returns:
            아카이브 경로
        """
        import os
        import shutil
        from datetime import datetime
        
        try:
            db_dir = os.path.dirname(getattr(self, 'db_path', '') or '')
            if not db_dir:
                db_dir = os.path.dirname(file_path)
            
            archive_dir = os.path.join(db_dir, 'archive', category)
            os.makedirs(archive_dir, exist_ok=True)
            
            base_name = os.path.basename(file_path)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            name, ext = os.path.splitext(base_name)
            archive_name = f"{name}_{timestamp}{ext}"
            archive_path = os.path.join(archive_dir, archive_name)
            
            shutil.copy2(file_path, archive_path)
            self._log(f"📁 아카이브: {archive_name}")
            return archive_path
            
        except (OSError, IOError, PermissionError) as e:
            logger.debug(f"아카이브 실패 (무시): {e}")
            return ''
