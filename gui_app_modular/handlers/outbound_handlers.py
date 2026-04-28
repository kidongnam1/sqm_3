# -*- coding: utf-8 -*-
"""
SQM Inventory - Outbound Handlers
=================================

v2.9.91 - Extracted from gui_app.py

Outbound processing: simple outbound, Excel outbound
"""

from gui_app_modular.utils.ui_constants import create_themed_toplevel  # v8.0.9
from gui_app_modular.utils.ui_constants import tc
import logging
import sqlite3
import csv
import os
# v8.0.6 [PICKING-REVIEW] 피킹리스트 첫 행 검수 패치
try:
    from features.parsers.picking_candidate_patch import enrich_picking_doc_with_review
    _HAS_PICKING_REVIEW = True
except ImportError:
    _HAS_PICKING_REVIEW = False
import hashlib
import json
import shutil
from datetime import datetime, date, timedelta

from ..utils.ui_constants import CustomMessageBox, ThemeColors, apply_tooltip, setup_dialog_geometry_persistence
from utils.path_utils import get_app_base_dir
logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


class OutboundHandlersMixin:
    """
    Outbound handlers mixin
    
    Mixed into SQMInventoryApp class
    """

    def _refresh_after_outbound_action(self, reason: str) -> None:
        """출고/취소 계열 액션 이후 표준 새로고침 진입점.
        v8.1.8: Return / Move 탭 갱신 추가.
        """
        if hasattr(self, 'refresh_bus_deferred'):
            self.refresh_bus_deferred(reason=reason, delay_ms=50)
            return
        self._safe_refresh()
        for fn_name in (
            '_refresh_inventory',
            '_refresh_tonbag',
            '_refresh_allocation',
            '_refresh_picked',
            '_refresh_sold',
            '_refresh_dashboard',
            '_refresh_return_tab',   # v8.1.8: Return 탭
            '_refresh_move_tab',     # v8.1.8: Move 탭
        ):
            fn = getattr(self, fn_name, None)
            if callable(fn):
                try:
                    fn()
                except Exception as _e:
                    logger.debug(f"[refresh_after_outbound] {fn_name}: {_e}")

    def _s1_get_proof_base_dir(self) -> str:
        """S1 근거문서 저장 폴더(일자별) 경로를 반환한다."""
        day_dir = os.path.join(
            get_app_base_dir(),
            "data",
            "proof_docs",
            date.today().isoformat(),
        )
        os.makedirs(day_dir, exist_ok=True)
        return day_dir

    def _s1_cleanup_old_proof_docs(self, retention_days: int = 90) -> dict:
        """보관기간 초과 근거문서 폴더를 정리한다."""
        base_dir = os.path.join(get_app_base_dir(), "data", "proof_docs")
        result = {"removed_dirs": 0, "removed_files": 0}
        if not os.path.isdir(base_dir):
            return result

        cutoff = date.today() - timedelta(days=retention_days)
        for name in os.listdir(base_dir):
            target_dir = os.path.join(base_dir, name)
            if not os.path.isdir(target_dir):
                continue
            try:
                folder_day = date.fromisoformat(name)
            except ValueError:
                continue
            if folder_day >= cutoff:
                continue
            file_count = 0
            for _, _, files in os.walk(target_dir):
                file_count += len(files)
            shutil.rmtree(target_dir, ignore_errors=True)
            result["removed_dirs"] += 1
            result["removed_files"] += file_count
        return result

    def _s1_get_audit_columns(self) -> set:
        """audit_log 테이블 컬럼 목록을 안전하게 조회한다."""
        try:
            rows = self.engine.db.fetchall("PRAGMA table_info(audit_log)")
            cols = set()
            for row in rows or []:
                if isinstance(row, dict):
                    name = row.get("name")
                else:
                    name = row[1] if len(row) > 1 else None
                if name:
                    cols.add(str(name))
            return cols
        except Exception:
            return set()

    def _s1_write_audit(self, event_type: str, payload=None, **meta) -> None:
        """환경별 audit_log 스키마 차이를 흡수해 이벤트를 기록한다."""
        cols = self._s1_get_audit_columns()
        if not cols:
            return
        if "event_type" not in cols or "created_at" not in cols:
            return
        payload_json = json.dumps(payload or {}, ensure_ascii=False)
        record = {
            "event_type": event_type,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        if "payload" in cols:
            record["payload"] = payload_json
        if "event_data" in cols:
            record["event_data"] = payload_json
        for key, value in meta.items():
            if key in cols:
                record[key] = value
        col_names = list(record.keys())
        placeholders = ", ".join(["?"] * len(col_names))
        sql = f"INSERT INTO audit_log ({', '.join(col_names)}) VALUES ({placeholders})"
        try:
            self.engine.db.execute(sql, tuple(record[c] for c in col_names))
            if hasattr(self.engine.db, "conn") and self.engine.db.conn:
                self.engine.db.conn.commit()
        except Exception as e:
            logger.debug(f"[S1] audit 기록 스킵: {e}")

    def _s1_open_audit_viewer(self) -> None:
        """S1 감사 로그 뷰어를 연다."""
        from ..utils.constants import tk, ttk, BOTH, LEFT, RIGHT, X, END, filedialog

        cols = self._s1_get_audit_columns()
        if not cols:
            CustomMessageBox.showwarning(self.root, "안내", "audit_log 테이블을 찾지 못했습니다.")
            return

        note_col = "user_note" if "user_note" in cols else ("event_data" if "event_data" in cols else "payload")
        dialog = create_themed_toplevel(self.root)
        dialog.title("감사 로그")
        dialog.transient(self.root)
        setup_dialog_geometry_persistence(dialog, "audit_log_dialog", self.root, "large")

        top = ttk.Frame(dialog)
        top.pack(fill=X, padx=8, pady=6)
        ttk.Label(top, text="이벤트").pack(side=LEFT, padx=4)
        event_var = tk.StringVar(value="전체")
        ttk.Entry(top, textvariable=event_var, width=24).pack(side=LEFT, padx=4)
        ttk.Label(top, text="시작일").pack(side=LEFT, padx=4)
        from_var = tk.StringVar(value=date.today().isoformat())
        ttk.Entry(top, textvariable=from_var, width=12).pack(side=LEFT, padx=2)
        ttk.Label(top, text="종료일").pack(side=LEFT, padx=4)
        to_var = tk.StringVar(value=date.today().isoformat())
        ttk.Entry(top, textvariable=to_var, width=12).pack(side=LEFT, padx=2)

        tree = ttk.Treeview(dialog, columns=("id", "event_type", "note", "created_at"), show="headings", height=18)
        tree.heading("id", text="ID", anchor='center')
        tree.heading("event_type", text="EVENT", anchor='center')
        tree.heading("note", text="NOTE", anchor='center')
        tree.heading("created_at", text="TIME", anchor='center')
        tree.column("id", width=60, anchor="center")
        tree.column("event_type", width=220, anchor="w")
        tree.column("note", width=430, anchor="w")
        tree.column("created_at", width=180, anchor="center")
        tree.pack(fill=BOTH, expand=True, padx=8, pady=4)
        status_var = tk.StringVar(value="")
        ttk.Label(dialog, textvariable=status_var).pack(fill=X, padx=8, pady=(0, 6))

        cache_rows = {"rows": []}

        def _search():
            tree.delete(*tree.get_children())
            sql = (
                f"SELECT id, event_type, {note_col} AS note, created_at "
                f"FROM audit_log WHERE 1=1"
            )
            params = []
            event_text = event_var.get().strip()
            if event_text and event_text != "전체":
                sql += " AND event_type = ?"
                params.append(event_text)
            if "created_at" in cols and from_var.get().strip():
                sql += " AND created_at >= ?"
                params.append(f"{from_var.get().strip()} 00:00:00")
            if "created_at" in cols and to_var.get().strip():
                sql += " AND created_at <= ?"
                params.append(f"{to_var.get().strip()} 23:59:59")
            sql += " ORDER BY id DESC LIMIT 500"
            rows = self.engine.db.fetchall(sql, tuple(params))
            cache_rows["rows"] = rows
            for row in rows:
                note = (row.get("note") or "") if isinstance(row, dict) else ""
                tree.insert(
                    "",
                    END,
                    values=(
                        row.get("id", ""),
                        row.get("event_type", ""),
                        str(note)[:90],
                        row.get("created_at", ""),
                    ),
                )
            status_var.set(f"조회 {len(rows)}건")

        def _export_csv():
            rows = cache_rows["rows"]
            if not rows:
                CustomMessageBox.showwarning(dialog, "안내", "내보낼 데이터가 없습니다.")
                return
            out_path = filedialog.asksaveasfilename(
                parent=dialog,
                title="감사 로그 CSV 저장",
                defaultextension=".csv",
                initialfile=f"audit_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                filetypes=[("CSV 파일", "*.csv"), ("모든 파일", "*.*")],
            )
            if not out_path:
                return
            with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "EVENT", "NOTE", "TIME"])
                for row in rows:
                    writer.writerow([
                        row.get("id", ""),
                        row.get("event_type", ""),
                        row.get("note", ""),
                        row.get("created_at", ""),
                    ])
            status_var.set(f"CSV 저장 완료: {os.path.basename(out_path)}")
            self._s1_write_audit(
                "AUDIT_EXPORT",
                {"count": len(rows), "file": os.path.basename(out_path)},
                user_note=f"감사 로그 CSV 내보내기 {len(rows)}건",
            )

        btn = ttk.Frame(dialog)
        btn.pack(fill=X, padx=8, pady=6)
        ttk.Button(top, text="조회", command=_search).pack(side=LEFT, padx=6)
        ttk.Button(btn, text="CSV 저장", command=_export_csv).pack(side=LEFT, padx=4)
        ttk.Button(btn, text="닫기", command=dialog.destroy).pack(side=RIGHT, padx=4)
        _search()
    
    def _on_simple_outbound(self) -> None:
        """Simple outbound dialog - enter LOT and quantity (v4.0.3: UI 분리, v8.6.4: SRP 분해)"""
        from ..utils.constants import tk, ttk, LEFT, RIGHT, X, END, filedialog
        from ..utils.constants import HAS_TTKBOOTSTRAP

        # v4.0.3: UI 위젯 생성을 별도 메서드로 분리
        w = self._build_simple_outbound_ui()
        dialog, lot_text, preview_tree = w['dialog'], w['lot_text'], w['preview_tree']
        summary_var, customer_var = w['summary_var'], w['customer_var']
        sale_ref_var, btn_frame = w['sale_ref_var'], w['btn_frame']

        # 공유 상태 번들
        st = {
            "dialog": dialog, "lot_text": lot_text, "preview_tree": preview_tree,
            "summary_var": summary_var, "customer_var": customer_var,
            "sale_ref_var": sale_ref_var, "btn_frame": btn_frame,
            "proof_docs": [], "proof_hashes": set(),
            "tonbag_meta_by_item": {},
            "out_scan_state": {"loaded": False, "records": [], "matched": [], "unmatched": []},
            "proof_status_var": tk.StringVar(value="첨부 없음"),
            "out_scan_status_var": tk.StringVar(value="파일 미선택"),
            "unmatched_var": tk.StringVar(value=""),
            "proof_base_dir": self._s1_get_proof_base_dir(),
            "proof_listbox": None,  # set after widget creation
        }

        # v6.3.0: 근거문서 자동정리
        cleanup_result = self._s1_cleanup_old_proof_docs(retention_days=90)
        if cleanup_result.get("removed_dirs", 0) > 0:
            self._s1_write_audit(
                "PROOF_CLEANUP",
                cleanup_result,
                user_note=f"근거문서 자동정리 {cleanup_result['removed_dirs']}개 폴더",
            )

        # 클로저 → sub-method 호출 래퍼
        def _attach_proof_doc():
            self._sob_attach_proof_doc(st)

        def _preview_proof_doc(event=None):
            self._sob_preview_proof_doc(st)

        def _load_out_scan_file():
            self._sob_load_out_scan_file(st)

        def on_preview():
            self._sob_on_preview(st)

        def on_execute():
            self._sob_on_execute(st)

        def on_get_selected_lot():
            self._sob_on_get_selected_lot(st)

        # v6.3.0: 근거문서 + OUT 스캔 업로드 바
        self._sob_build_aux_widgets(st, _attach_proof_doc, _preview_proof_doc, _load_out_scan_file)

        # 버튼 및 키바인딩
        self._sob_build_buttons(
            st, on_preview, on_execute, on_get_selected_lot, HAS_TTKBOOTSTRAP,
        )

    # ------------------------------------------------------------------
    # _on_simple_outbound sub-methods (v8.6.4 [SRP])
    # ------------------------------------------------------------------

    @staticmethod
    def _sob_normalize_tonbag_key(raw):  # v8.6.4 [SRP]
        """톤백 키 정규화."""
        if raw is None:
            return ""
        return str(raw).strip().upper().replace(" ", "").replace("-", "").replace("_", "")

    @staticmethod
    def _sob_hash_file(path):  # v8.6.4 [SRP]
        """SHA-256 해시 계산."""
        h = hashlib.sha256()
        with open(path, 'rb') as f:
            while True:
                chunk = f.read(65536)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()

    def _sob_attach_proof_doc(self, st):  # v8.6.4 [SRP]
        """근거문서 첨부 처리."""
        from ..utils.constants import END, filedialog
        dialog = st["dialog"]
        proof_docs = st["proof_docs"]
        proof_hashes = st["proof_hashes"]
        proof_base_dir = st["proof_base_dir"]
        proof_listbox = st["proof_listbox"]
        proof_status_var = st["proof_status_var"]

        paths = filedialog.askopenfilenames(
            parent=dialog,
            title="근거문서 선택",
            filetypes=[
                ("지원 파일", "*.pdf *.png *.jpg *.jpeg *.xlsx *.csv *.txt *.docx"),
                ("모든 파일", "*.*"),
            ],
        )
        for fpath in paths:
            try:
                fhash = self._sob_hash_file(fpath)
                if fhash in proof_hashes:
                    CustomMessageBox.showwarning(dialog, "중복 파일",
                                                 f"{os.path.basename(fpath)}\n이미 첨부된 문서입니다.")
                    continue
                fname = os.path.basename(fpath)
                stored_name = f"{fhash[:8]}_{fname}"
                stored_path = os.path.join(proof_base_dir, stored_name)
                if not os.path.exists(stored_path):
                    shutil.copy2(fpath, stored_path)
                proof_hashes.add(fhash)
                proof_docs.append({
                    "id": fhash[:16],
                    "name": fname,
                    "path": stored_path,
                    "original_path": fpath,
                    "size": os.path.getsize(fpath),
                    "hash": fhash,
                    "added_at": datetime.now().strftime("%H:%M:%S"),
                })
                self._s1_write_audit(
                    "PROOF_ATTACH",
                    {
                        "name": fname,
                        "hash": fhash,
                        "size": os.path.getsize(fpath),
                        "stored_path": stored_path,
                    },
                    user_note=f"근거문서 첨부: {fname}",
                )
            except Exception as e:
                logger.error(f"[SimpleOutbound] 근거문서 첨부 실패: {fpath} -> {e}")
                CustomMessageBox.showerror(dialog, "첨부 오류", str(e))
        proof_listbox.delete(0, END)
        for d in proof_docs:
            proof_listbox.insert(END, f"📄 {d['name']} ({d['size']/1024:.1f}KB) [{d['added_at']}]")
        proof_status_var.set(f"📎 {len(proof_docs)}건 첨부" if proof_docs else "첨부 없음")

    def _sob_preview_proof_doc(self, st):  # v8.6.4 [SRP]
        """근거문서 미리보기."""
        from ..utils.constants import tk
        dialog = st["dialog"]
        proof_docs = st["proof_docs"]
        proof_listbox = st["proof_listbox"]

        sel = proof_listbox.curselection()
        if not sel:
            return
        doc = proof_docs[sel[0]]
        path = doc["path"]
        if not os.path.exists(path):
            CustomMessageBox.showwarning(dialog, "파일 없음", path)
            return
        ext = os.path.splitext(path)[1].lower()
        if ext == ".pdf":
            try:
                os.startfile(path)
            except Exception:
                CustomMessageBox.showinfo(dialog, "문서 정보", f"PDF 파일:\n{path}")
            return
        if ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"):
            try:
                from PIL import Image, ImageTk
            except Exception:
                CustomMessageBox.showwarning(dialog, "미리보기 불가", "Pillow 미설치로 이미지 미리보기를 열 수 없습니다.")
                return
            win = create_themed_toplevel(dialog)
            win.title(f"미리보기 - {doc['name']}")
            img = Image.open(path)
            img.thumbnail((800, 600), Image.Resampling.LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS)
            ph = ImageTk.PhotoImage(img)
            lb = tk.Label(win, image=ph)
            lb.image = ph
            lb.pack()
            return
        CustomMessageBox.showinfo(
            dialog, "문서 정보",
            f"파일: {doc['name']}\n크기: {doc['size']/1024:.1f}KB\nhash: {doc['hash'][:24]}..."
        )

    def _sob_parse_out_scan_file(self, path):  # v8.6.4 [SRP]
        """OUT 스캔 파일 파싱 (xlsx/csv/tsv)."""
        ext = os.path.splitext(path)[1].lower()
        records = []
        if ext in (".xlsx", ".xls") and HAS_OPENPYXL:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return []
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
            key_idx = next((i for i, h in enumerate(header) if any(k in h for k in ["tonbag_id", "tonbag", "tb_id", "uid", "id", "톤백"])), None)
            wt_idx = next((i for i, h in enumerate(header) if any(k in h for k in ["weight", "kg", "무게", "중량"])), None)
            start = 1 if key_idx is not None else 0
            if key_idx is None:
                key_idx, wt_idx = 0, 1
            for row in rows[start:]:
                if not row or len(row) <= key_idx:
                    continue
                key = str(row[key_idx]).strip() if row[key_idx] is not None else ""
                if not key:
                    continue
                w = 0.0
                if wt_idx is not None and len(row) > wt_idx and row[wt_idx] is not None:
                    try:
                        w = float(str(row[wt_idx]).replace(",", "").strip())
                    except Exception:
                        w = 0.0
                records.append({"raw_key": key, "key": self._sob_normalize_tonbag_key(key), "weight": w})
            return records

        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
            except Exception:
                dialect = csv.excel
            rows = list(csv.reader(f, dialect))
        if not rows:
            return []
        header = [str(c).strip().lower() for c in rows[0]]
        key_idx = next((i for i, h in enumerate(header) if any(k in h for k in ["tonbag_id", "tonbag", "tb_id", "uid", "id", "톤백"])), None)
        wt_idx = next((i for i, h in enumerate(header) if any(k in h for k in ["weight", "kg", "무게", "중량"])), None)
        start = 1 if key_idx is not None else 0
        if key_idx is None:
            key_idx, wt_idx = 0, 1
        for row in rows[start:]:
            if not row or len(row) <= key_idx:
                continue
            key = str(row[key_idx]).strip()
            if not key:
                continue
            w = 0.0
            if wt_idx is not None and len(row) > wt_idx:
                try:
                    w = float(str(row[wt_idx]).replace(",", "").strip() or 0)
                except Exception:
                    w = 0.0
            records.append({"raw_key": key, "key": self._sob_normalize_tonbag_key(key), "weight": w})
        return records

    def _sob_load_out_scan_file(self, st):  # v8.6.4 [SRP]
        """OUT 스캔 파일 불러오기 및 매칭."""
        from ..utils.constants import filedialog
        dialog = st["dialog"]
        preview_tree = st["preview_tree"]
        tonbag_meta_by_item = st["tonbag_meta_by_item"]
        out_scan_state = st["out_scan_state"]
        out_scan_status_var = st["out_scan_status_var"]
        unmatched_var = st["unmatched_var"]

        path = filedialog.askopenfilename(
            parent=dialog,
            title="OUT 스캔 파일 선택",
            filetypes=[("스캔 파일", "*.csv *.tsv *.txt *.xlsx *.xls"), ("모든 파일", "*.*")],
        )
        if not path:
            return
        try:
            recs = self._sob_parse_out_scan_file(path)
        except Exception as e:
            CustomMessageBox.showerror(dialog, "파싱 오류", f"OUT 파일 파싱 실패:\n{e}")
            return
        if not recs:
            CustomMessageBox.showwarning(dialog, "경고", "유효한 스캔 데이터가 없습니다.")
            return

        selected_items = preview_tree.selection()
        selected_keys = set()
        for iid in selected_items:
            meta = tonbag_meta_by_item.get(iid, {})
            if meta.get("key"):
                selected_keys.add(meta["key"])
        if not selected_keys:
            CustomMessageBox.showwarning(dialog, "안내", "Preview 후 출고 톤백을 먼저 선택하세요.")
            return

        matched = [r for r in recs if r["key"] in selected_keys]
        unmatched = [r for r in recs if r["key"] not in selected_keys]
        out_scan_state["loaded"] = True
        out_scan_state["records"] = recs
        out_scan_state["matched"] = matched
        out_scan_state["unmatched"] = unmatched

        out_scan_status_var.set(
            f"📊 {os.path.basename(path)} | 전체 {len(recs)}건 | 매칭 {len(matched)} | 미매칭 {len(unmatched)}"
        )
        if unmatched:
            preview_ids = ", ".join(r["raw_key"] for r in unmatched[:5])
            more = f" 외 {len(unmatched)-5}건" if len(unmatched) > 5 else ""
            unmatched_var.set(f"⛔ 미매칭 {len(unmatched)}건 (무단 출고 의심): {preview_ids}{more}")
            self._s1_write_audit(
                "UNMATCHED_SCAN",
                {
                    "file": os.path.basename(path),
                    "count": len(unmatched),
                    "sample": [r.get("raw_key", "") for r in unmatched[:20]],
                },
                user_note=f"OUT 스캔 미매칭 {len(unmatched)}건",
            )
        else:
            unmatched_var.set("")

    def _sob_on_preview(self, st):  # v8.6.4 [SRP]
        """Preview: LOT별 톤백 상세 표시 (v3.8.4)."""
        from ..utils.constants import END
        preview_tree = st["preview_tree"]
        lot_text = st["lot_text"]
        summary_var = st["summary_var"]
        tonbag_meta_by_item = st["tonbag_meta_by_item"]

        preview_tree.delete(*preview_tree.get_children())
        tonbag_meta_by_item.clear()

        lines_input = lot_text.get("1.0", END).strip().split('\n')
        total_kg = 0
        tonbag_count = 0
        warnings = []
        lot_requests = {}

        for line in lines_input:
            line = line.strip()
            if not line:
                continue

            parts = [p.strip() for p in line.replace('\t', ',').split(',')]
            if len(parts) < 2:
                warnings.append(f"형식 오류: {line}")
                continue

            lot_no = parts[0]
            try:
                qty_mt = float(parts[1])
            except ValueError:
                warnings.append(f"수량 오류: {line}")
                continue
            lot_requests[lot_no] = lot_requests.get(lot_no, 0.0) + (qty_mt * 1000)

        # v8.2.0 N+1 최적화: inventory + tonbag 일괄 pre-fetch
        _inv_map_oh, _tb_map_oh = self._sob_prefetch_lot_data(lot_requests)

        for lot_no, qty_kg in lot_requests.items():
            t_kg, t_cnt = self._sob_preview_single_lot(
                st, lot_no, qty_kg, _inv_map_oh, _tb_map_oh, warnings,
            )
            total_kg += t_kg
            tonbag_count += t_cnt

        # 스타일
        preview_tree.tag_configure('error', foreground=tc('danger'))
        preview_tree.tag_configure('lot_header', background=tc('shipped'), font=('', 13, 'bold'))
        preview_tree.tag_configure('tonbag', foreground=ThemeColors.get('text_primary', False))

        summary_var.set(f"톤백 {tonbag_count}개 / {total_kg/1000:.3f} MT 출고 예정")

        if warnings:
            CustomMessageBox.showwarning(self.root, "확인 필요", "\n".join(warnings[:10]))

    def _sob_prefetch_lot_data(self, lot_requests):  # v8.6.4 [SRP]
        """inventory + tonbag 일괄 pre-fetch (N+1 최적화)."""
        from collections import defaultdict as _ddict
        _lot_keys = list(lot_requests.keys())
        _ph_oh = ','.join('?' * len(_lot_keys))
        _inv_rows_oh = self.engine.db.fetchall(
            f"SELECT lot_no, current_weight, product FROM inventory WHERE lot_no IN ({_ph_oh})",
            tuple(_lot_keys)
        ) or [] if _lot_keys else []
        _inv_map_oh = {
            (r.get('lot_no') if isinstance(r, dict) else r[0]): r
            for r in _inv_rows_oh
        }
        _tb_rows_oh = self.engine.db.fetchall(
            f"SELECT id, lot_no, sub_lt, weight, status, location "
            f"FROM inventory_tonbag "
            f"WHERE lot_no IN ({_ph_oh}) AND status = 'AVAILABLE' "
            f"ORDER BY lot_no, sub_lt DESC",
            tuple(_lot_keys)
        ) or [] if _lot_keys else []
        _tb_map_oh = _ddict(list)
        for _r in _tb_rows_oh:
            _k = _r.get('lot_no') if isinstance(_r, dict) else _r[1]
            _tb_map_oh[_k].append(_r)
        return _inv_map_oh, _tb_map_oh

    def _sob_preview_single_lot(self, st, lot_no, qty_kg, inv_map, tb_map, warnings):  # v8.6.4 [SRP]
        """단일 LOT에 대한 프리뷰 행 생성. (total_kg, tonbag_count) 반환."""
        from ..utils.constants import END
        preview_tree = st["preview_tree"]
        tonbag_meta_by_item = st["tonbag_meta_by_item"]
        total_kg = 0
        tonbag_count = 0

        # LOT 존재 확인 (cache)
        lot_info = inv_map.get(lot_no)

        if not lot_info:
            preview_tree.insert('', END, values=(
                lot_no, '-', '-', '-', '❌ 미발견', '-'
            ), tags=('error',))
            warnings.append(f"LOT 미발견: {lot_no}")
            return total_kg, tonbag_count

        avail_kg = (lot_info.get('current_weight') if isinstance(lot_info, dict)
                    else lot_info[1]) or 0
        product = (lot_info.get('product') if isinstance(lot_info, dict)
                   else lot_info[2]) or '-'

        if avail_kg < qty_kg - 0.01:
            warnings.append(f"재고 부족: {lot_no} (판매가능: {avail_kg:.0f}kg, 요청: {qty_kg:.0f}kg)")

        # 판매가능 톤백 (cache)
        tonbags = tb_map.get(lot_no, [])

        # LOT 헤더 행
        preview_tree.insert('', END, iid=f"LOT_{lot_no}",
            values=(f"📦 {lot_no}", '', product, f"{avail_kg:,.0f}",
                    f"요청: {qty_kg:,.0f}kg", ''),
            tags=('lot_header',))

        remaining = qty_kg
        for tb in tonbags:
            if remaining <= 0.01:
                break
            tb_weight = tb['weight'] or 0
            sub_lt = tb['sub_lt']
            loc = tb['location'] or ''
            tb_id = str(tb.get('id') or f"{lot_no}-{sub_lt}")
            normalized_key = self._sob_normalize_tonbag_key(tb_id)

            status = '✅ 출고' if remaining >= tb_weight else '⚠️ 초과'

            item_id = preview_tree.insert('', END, values=(
                f"  └ {lot_no}", str(sub_lt), product,
                f"{tb_weight:,.0f}", status, loc
            ), tags=('tonbag',))
            tonbag_meta_by_item[item_id] = {
                "lot_no": lot_no,
                "tonbag_id": tb_id,
                "key": normalized_key,
                "weight": float(tb_weight or 0),
            }

            # 자동 선택 (요청 수량 만큼)
            preview_tree.selection_add(item_id)

            remaining -= tb_weight
            total_kg += tb_weight
            tonbag_count += 1

        return total_kg, tonbag_count

    def _sob_on_execute(self, st):  # v8.6.4 [SRP]
        """Execute outbound (v3.8.4: 선택된 톤백 기반)."""
        from ..utils.constants import END
        dialog = st["dialog"]
        preview_tree = st["preview_tree"]
        lot_text = st["lot_text"]
        customer_var = st["customer_var"]
        sale_ref_var = st["sale_ref_var"]
        proof_docs = st["proof_docs"]
        tonbag_meta_by_item = st["tonbag_meta_by_item"]
        out_scan_state = st["out_scan_state"]

        customer = customer_var.get().strip()
        sale_ref = sale_ref_var.get().strip()

        if not customer:
            CustomMessageBox.showwarning(self.root, "입력 필요", "고객명을 입력하세요.")
            return

        # 선택된 톤백 항목 수집
        allocation_items = self._sob_collect_allocation_items(
            st, customer, sale_ref,
        )
        if allocation_items is None:
            return  # 중단됨 (에러 메시지 이미 표시)

        if not allocation_items:
            CustomMessageBox.showwarning(self.root, "입력 필요", "Preview 후 톤백을 선택하거나 LOT를 입력하세요.")
            return

        # v5.0.9: 톤백/샘플 구분 카운트
        from ..dialogs.allocation_preview import _is_sample_item
        tonbag_items = [i for i in allocation_items if not _is_sample_item(i)]
        sample_items = [i for i in allocation_items if _is_sample_item(i)]
        tonbag_qty = sum(i.get('qty_mt', 0) for i in tonbag_items)
        sample_qty = sum(i.get('qty_mt', 0) for i in sample_items)
        total_qty = tonbag_qty + sample_qty

        # v6.1.0: 빠른 출고 8개 톤백 제한 (초과 시 일반 출고 전환 안내)
        if len(tonbag_items) > 8:
            go_normal = CustomMessageBox.askyesno(
                self.root, "수량 초과",
                f"빠른 출고는 최대 8개 톤백까지 가능합니다.\n"
                f"(선택: {len(tonbag_items)}개)\n\n"
                f"일반 출고(배정표)로 전환하시겠습니까?"
            )
            if go_normal and hasattr(self, '_on_allocation_dialog'):
                dialog.destroy()
                self._on_allocation_dialog()
            return

        # v6.1.0: source='QUICK' 마킹 (allocation_plan 추적용)
        for _qi in allocation_items:
            _qi['source'] = 'QUICK'

        # Confirm (v6.1.0: 판매화물 결정 용어 + PICKED 멈춤 안내)
        confirm_msg = (
            f"판매화물 결정을 진행할까요?\n\n"
            f"고객: {customer}\n"
            f"📦 톤백: {len(tonbag_items)}개 ({tonbag_qty:.3f} MT)\n"
            f"🧪 샘플: {len(sample_items)}개 ({sample_qty:.3f} MT)\n"
            f"📎 근거문서: {len(proof_docs)}건\n"
            f"합계: {len(allocation_items)}건 ({total_qty:.3f} MT)\n\n"
            f"※ 현장 출고 후 [출고 확정]으로 최종 처리하세요."
        )
        if not CustomMessageBox.askyesno(self.root, "출고 확인", confirm_msg):
            return

        # Execute
        self._sob_run_outbound_engine(st, allocation_items, customer, sale_ref)

    def _sob_collect_allocation_items(self, st, customer, sale_ref):  # v8.6.4 [SRP]
        """선택된 톤백 / fallback 텍스트에서 allocation_items 수집.
        Returns list on success, None if hard-stopped."""
        from ..utils.constants import END
        preview_tree = st["preview_tree"]
        lot_text = st["lot_text"]
        proof_docs = st["proof_docs"]
        tonbag_meta_by_item = st["tonbag_meta_by_item"]
        out_scan_state = st["out_scan_state"]

        selected = preview_tree.selection()
        allocation_items = []

        if selected:
            # 선택된 톤백에서 LOT별 수량 집계
            lot_weights = {}
            selected_keys = set()
            selected_count = 0
            for item_id in selected:
                values = preview_tree.item(item_id)['values']
                tags = preview_tree.item(item_id).get('tags', ())

                if 'lot_header' in tags:
                    continue  # LOT 헤더는 건너뜀
                selected_count += 1
                meta = tonbag_meta_by_item.get(item_id, {})
                lot_no = str(meta.get("lot_no") or str(values[0]).replace('└', '').strip())
                weight = float(meta.get("weight") or 0)
                if weight <= 0:
                    try:
                        weight = float(str(values[3]).replace(',', ''))
                    except (ValueError, IndexError):
                        continue
                if meta.get("key"):
                    selected_keys.add(meta["key"])
            if selected_count > 0 and len(selected_keys) < selected_count:
                CustomMessageBox.showerror(
                    self.root,
                    "출고 중단",
                    "선택 톤백 키가 중복되어 출고를 중단합니다.\n"
                    "톤백 선택을 해제 후 다시 선택해 주세요.",
                )
                return None

                if lot_no not in lot_weights:
                    lot_weights[lot_no] = 0
                lot_weights[lot_no] += weight

            for lot_no, weight_kg in lot_weights.items():
                allocation_items.append({
                    'lot_no': lot_no,
                    'weight_kg': weight_kg,
                    'qty_mt': weight_kg / 1000.0,
                    'sold_to': customer,
                    'customer': customer,
                    'sale_ref': sale_ref,
                    'proof_doc_ids': [d['id'] for d in proof_docs],
                })

            # OUT 스캔 파일이 있으면 미매칭 하드스톱
            if out_scan_state.get("loaded"):
                unmatched = out_scan_state.get("unmatched", [])
                if unmatched:
                    CustomMessageBox.showerror(
                        self.root, "출고 중단",
                        f"OUT 스캔 파일에 미매칭 톤백 {len(unmatched)}건이 있어 출고를 중단합니다.\n"
                        f"(무단 출고 의심)\n미매칭을 정리한 후 다시 진행하세요."
                    )
                    return None

        if not allocation_items:
            # Fallback: 텍스트 입력에서 추출
            lines = lot_text.get("1.0", END).strip().split('\n')
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                parts = [p.strip() for p in line.replace('\t', ',').split(',')]
                if len(parts) < 2:
                    continue
                lot_no = parts[0]
                try:
                    qty_mt = float(parts[1])
                except ValueError:
                    continue
                allocation_items.append({
                    'lot_no': lot_no,
                    'qty_mt': qty_mt,
                    'sold_to': customer,
                    'customer': customer,
                    'sale_ref': sale_ref,
                    'proof_doc_ids': [d['id'] for d in proof_docs],
                })

        return allocation_items

    def _sob_run_outbound_engine(self, st, allocation_items, customer, sale_ref):  # v8.6.4 [SRP]
        """엔진 호출 및 결과 처리 (v3.8.4: All-or-Nothing)."""
        dialog = st["dialog"]
        proof_docs = st["proof_docs"]
        out_scan_state = st["out_scan_state"]

        try:
            if hasattr(self.engine, 'process_outbound_safe'):
                try:
                    result = self.engine.process_outbound_safe(
                        allocation_items, source='QUICK', stop_at_picked=True
                    )
                except TypeError:
                    result = self.engine.process_outbound(
                        allocation_items, source='QUICK', stop_at_picked=True
                    )
                except (ValueError, RuntimeError, sqlite3.OperationalError) as pf_err:
                    err_msg = str(pf_err)
                    display_msg = err_msg[:500] + '...' if len(err_msg) > 500 else err_msg
                    self._log(f"❌ 출고 검증 실패 (All-or-Nothing): {display_msg[:200]}")
                    CustomMessageBox.showerror(self.root, "출고 검증 실패",
                        f"All-or-Nothing 검증에서 오류가 발견되어\n전체 출고가 중단되었습니다.\n\n{display_msg}")
                    return
            else:
                result = self.engine.process_outbound(
                    allocation_items, source='QUICK', stop_at_picked=True
                )

            if result.get('success') or result.get('processed', 0) > 0:
                processed = result.get('lots_processed', result.get('processed', 0))
                picked = result.get('total_picked', 0)
                msg = (f"판매화물 결정 완료!\n\n"
                       f"처리: {processed}건\n"
                       f"총 중량: {picked:.3f} MT\n\n"
                       f"현장 출고 확인 후 [출고 확정]을 실행하세요.")

                if result.get('warnings'):
                    msg += "\n\n경고:\n" + "\n".join(result['warnings'][:5])

                CustomMessageBox.showinfo(self.root, "완료", msg)
                self._log(f"✅ 빠른 출고: {processed}건, {picked:.3f} MT")
                self._s1_write_audit(
                    "OUTBOUND_EXECUTE",
                    {
                        "customer": customer,
                        "sale_ref": sale_ref,
                        "lots": len(allocation_items),
                        "picked_mt": picked,
                        "proof_docs": len(proof_docs),
                        "out_scan_loaded": bool(out_scan_state.get("loaded")),
                    },
                    user_note=f"빠른 출고 실행: {customer}, {picked:.3f} MT",
                )

                dialog.destroy()
                self._refresh_after_outbound_action("SIMPLE_OUTBOUND_EXECUTE")
            else:
                errs = '\n'.join(result.get('errors', ['알 수 없는 오류']))
                CustomMessageBox.showerror(self.root, "출고 실패", f"출고 처리 실패:\n{errs}")

        except (ValueError, RuntimeError, KeyError, sqlite3.OperationalError, sqlite3.IntegrityError) as e:
            logger.error(f"출고 오류: {e}")
            err_msg = str(e)[:500]
            CustomMessageBox.showerror(self.root, "출고 오류", f"출고 처리 중 오류:\n\n{err_msg}")

    def _sob_build_aux_widgets(self, st, attach_cmd, preview_cmd, load_scan_cmd):  # v8.6.4 [SRP]
        """근거문서 + OUT 스캔 업로드 위젯 생성."""
        from ..utils.constants import tk, ttk, LEFT, X
        dialog = st["dialog"]
        btn_frame = st["btn_frame"]
        proof_status_var = st["proof_status_var"]
        out_scan_status_var = st["out_scan_status_var"]
        unmatched_var = st["unmatched_var"]

        aux_frame = ttk.Frame(btn_frame.master)
        aux_frame.pack(fill=X, pady=(0, 6))

        proof_frame = ttk.LabelFrame(aux_frame, text="📎 근거문서 (선택)")
        proof_frame.pack(fill=X, pady=(0, 4))
        proof_btn_row = ttk.Frame(proof_frame)
        proof_btn_row.pack(fill=X, padx=6, pady=4)
        ttk.Button(proof_btn_row, text="+ 파일 첨부", command=attach_cmd).pack(side=LEFT, padx=2)
        ttk.Label(proof_btn_row, textvariable=proof_status_var).pack(side=LEFT, padx=8)
        proof_listbox = tk.Listbox(proof_frame, height=3, font=("Consolas", 9))
        proof_listbox.pack(fill=X, padx=6, pady=(0, 6))
        proof_listbox.bind("<Double-1>", lambda e: preview_cmd(e))
        st["proof_listbox"] = proof_listbox

        out_frame = ttk.LabelFrame(aux_frame, text="📊 OUT 스캔 파일")
        out_frame.pack(fill=X)
        out_btn_row = ttk.Frame(out_frame)
        out_btn_row.pack(fill=X, padx=6, pady=4)
        ttk.Button(out_btn_row, text="📂 파일 불러오기 (csv/xlsx)", command=load_scan_cmd).pack(side=LEFT, padx=2)
        ttk.Button(out_btn_row, text="📋 감사 로그", command=self._s1_open_audit_viewer).pack(side=LEFT, padx=2)
        ttk.Label(out_btn_row, textvariable=out_scan_status_var).pack(side=LEFT, padx=8)
        ttk.Label(out_frame, textvariable=unmatched_var, foreground=tc('danger')).pack(fill=X, padx=6, pady=(0, 6))

    def _sob_build_buttons(self, st, on_preview, on_execute, on_get_selected_lot, HAS_TTKBOOTSTRAP):  # v8.6.4 [SRP]
        """메인 버튼 및 키바인딩 생성."""
        from ..utils.constants import ttk, LEFT, RIGHT
        dialog = st["dialog"]
        btn_frame = st["btn_frame"]
        lot_text = st["lot_text"]

        btn_style = {"bootstyle": "info"} if HAS_TTKBOOTSTRAP else {}
        btn_style_success = {"bootstyle": "success"} if HAS_TTKBOOTSTRAP else {}
        btn_style_secondary = {"bootstyle": "secondary"} if HAS_TTKBOOTSTRAP else {}
        btn_style_outline = {"bootstyle": "outline"} if HAS_TTKBOOTSTRAP else {}

        _bp = ttk.Button(btn_frame, text="Preview", command=on_preview, **btn_style)
        _bp.pack(side=LEFT, padx=5)
        apply_tooltip(_bp, "입력한 LOT·수량·출고처로 출고 미리보기를 표시합니다. DB에는 반영되지 않습니다.")
        _be = ttk.Button(btn_frame, text="Execute", command=on_execute, **btn_style_success)
        _be.pack(side=LEFT, padx=5)
        apply_tooltip(_be, "미리보기한 내용으로 실제 출고를 실행합니다. 재고가 차감되고 출고 이력에 기록됩니다.")
        _bc = ttk.Button(btn_frame, text="Cancel", command=dialog.destroy, **btn_style_secondary)
        _bc.pack(side=RIGHT, padx=5)
        apply_tooltip(_bc, "출고 대화상자를 닫습니다. 실행하지 않은 출고는 반영되지 않습니다.")

        _badd = ttk.Button(btn_frame, text="Add Selected", command=on_get_selected_lot,
                           **btn_style_outline)
        _badd.pack(side=LEFT, padx=20)
        apply_tooltip(_badd, "LOT 리스트에서 선택한 LOT를 출고 목록에 추가합니다. 여러 LOT를 쉼표로 구분해 넣을 수 있습니다.")

        # Center dialog
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 700) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 600) // 2
        dialog.geometry(f"+{x}+{y}")

        # Key bindings
        dialog.bind('<Escape>', lambda e: dialog.destroy())

    def _sob_on_get_selected_lot(self, st):  # v8.6.4 [SRP]
        """Get selected LOT from inventory list."""
        from ..utils.constants import END
        lot_text = st["lot_text"]
        selected = self.tree_inventory.selection()
        if selected:
            values = self.tree_inventory.item(selected[0])['values']
            lot_no = values[0]
            current_text = lot_text.get("1.0", END).strip()
            if current_text:
                lot_text.insert(END, f"\n{lot_no}, ")
            else:
                lot_text.insert(END, f"{lot_no}, ")
            lot_text.focus_set()

    def _on_go_allocation_tab(self) -> None:
        """판매 배정 탭으로 이동 (메뉴 공통 진입점) — v8.1.6: 위젯 참조 단일화."""
        notebook = getattr(self, 'notebook', None)
        if not notebook:
            return
        target_tab = getattr(self, 'tab_allocation', None)
        if target_tab is None:
            logger.debug("[출고UI] tab_allocation 미등록 — 탭 이동 스킵")
            return
        try:
            notebook.select(target_tab)
        except Exception as e:
            logger.debug(f"[출고UI] 탭 선택 실패: {e}")
    
    def _on_allocation_stress_test(self) -> None:
        """🧪 Allocation 7-Gate Stress Test 다이얼로그 (v7.1.2)."""
        try:
            from ..dialogs.Claude_allocation_stress_test_dialog_v712 import AllocationStressTestDialog
            AllocationStressTestDialog(self, self.engine)
        except Exception as e:
            logger.error(f"Stress Test 다이얼로그 오류: {e}")

    def _on_go_scan_tab(self) -> None:
        """📷 스캔 탭으로 이동 (v7.4.0) — v8.1.6: 위젯 참조 단일화."""
        notebook = getattr(self, 'notebook', None)
        if not notebook:
            return
        target_tab = getattr(self, 'tab_scan', None)
        if target_tab is None:
            logger.debug("[스캔탭] tab_scan 미등록 — 탭 이동 스킵")
            return
        try:
            notebook.select(target_tab)
        except Exception as e:
            logger.debug(f"[스캔탭] 탭 선택 실패: {e}")

    def _on_outbound_click(self) -> None:
        """v4.0.5 Phase2: 파일 선택 → 미리보기 팝업 → 사용자 확인 → DB 반영"""
        from ..utils.constants import filedialog
        
        files = filedialog.askopenfilenames(
            parent=self.root,
            title="출고 Allocation Excel 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not files:
            return
        
        for file_path in files:
            self._preview_outbound(file_path)
    
    def _preview_outbound(self, excel_path: str) -> None:
        """v4.0.5: 출고 Excel → 파싱 → 미리보기 팝업"""
        import os
        
        self._log(f"📤 출고 파일 읽기: {os.path.basename(excel_path)}")
        
        try:
            from parsers.allocation_parser import AllocationParser
            
            parser = AllocationParser()
            alloc_data = parser.parse(excel_path)
            
            if not alloc_data or not alloc_data.rows:
                self._log("⚠️ 출고 데이터 없음")
                CustomMessageBox.showwarning(self.root, "경고", "출고 데이터가 없습니다.")
                return
            
            # AllocationRow → dict 변환 (미리보기용, v5.1.0: 용어 통일)
            preview_items = []
            for row in alloc_data.rows:
                preview_items.append({
                    'lot_no': row.lot_no,
                    'sap_no': row.sap_no,
                    'product': row.product,
                    'qty_mt': row.qty_mt,
                    'sold_to': row.sold_to,           # DB 호환
                    'customer': row.sold_to,           # v5.1.0 표준
                    'sale_ref': row.sale_ref,
                    'sub_lt': row.sub_lt,              # DB 호환
                    'tonbag_no': row.sub_lt,           # v5.1.0 표준
                    'warehouse': row.warehouse,
                    'customs': row.customs,
                    'gross_weight': row.gross_weight,
                })
            
            self._log(f"📋 출고 미리보기: {len(preview_items)}건, {alloc_data.total_qty:.3f} MT")
            
            # 미리보기 팝업 표시 → Execute 클릭 시 _execute_outbound 호출
            self._show_outbound_preview(
                preview_items,
                callback=lambda items: self._execute_outbound(items, alloc_data)
            )
            
        except ImportError:
            self._log("⚠️ AllocationParser 모듈 없음")
            CustomMessageBox.showwarning(self.root, "모듈 없음", "출고 파서 모듈이 필요합니다.")
        except (RuntimeError, ValueError, TypeError) as e:
            logger.error(f"출고 파일 읽기 실패: {e}")
            self._log(f"❌ 출고 파일 오류: {e}")
            CustomMessageBox.show_detailed_error(
                self.root, "출고 파일 오류", 
                f"Excel 파일을 읽는 중 오류가 발생했습니다.\n\n{e}",
                exception=e
            )
    
    def _show_outbound_preview(self, preview_items, callback):
        """
        v5.0.4: Allocation 출고 미리보기 다이얼로그 표시
        
        Args:
            preview_items: Allocation 데이터 리스트
            callback: 확인 시 콜백 함수
        """
        try:
            from ..dialogs.allocation_preview import AllocationPreviewDialog
            
            _ = AllocationPreviewDialog(
                self.root,
                preview_items,
                on_confirm=callback,
                on_cancel=lambda: self._log("❌ 출고 취소됨")
            )
            
        except ImportError as e:
            self._log(f"⚠️ AllocationPreviewDialog 로딩 실패: {e}")
            # Fallback: 기존 방식
            if callback:
                callback(preview_items)
    
    def _execute_outbound(self, preview_items, alloc_data) -> None:
        """v4.0.5: 사용자 확인 후 실제 DB 반영. v5.9.92: AllocationRow → dict 변환 후 process_outbound(EXCEL)."""
        try:
            # AllocationRow → dict 리스트 변환 (process_outbound는 dict 기대)
            if hasattr(alloc_data, 'rows'):
                items = []
                for row in alloc_data.rows:
                    items.append({
                        'lot_no': getattr(row, 'lot_no', ''),
                        'weight_kg': (getattr(row, 'qty_mt', 0) or 0) * 1000.0,
                        'qty_mt': getattr(row, 'qty_mt', 0),
                        'customer': getattr(row, 'sold_to', '') or getattr(row, 'customer', ''),
                        'sold_to': getattr(row, 'sold_to', ''),
                        'sale_ref': getattr(row, 'sale_ref', ''),
                    })
            else:
                items = list(preview_items) if preview_items else []
            
            if not items:
                self._log("⚠️ 출고할 항목 없음")
                CustomMessageBox.showwarning(self.root, "출고", "출고할 항목이 없습니다.")
                return
            
            if hasattr(self, 'do_action_tx'):
                result = self.do_action_tx(
                    "EXECUTE_OUTBOUND_EXCEL",
                    lambda: self.engine.process_outbound(items, source='EXCEL', stop_at_picked=False),
                    parent=self.root,
                    refresh_mode="deferred",
                )
            else:
                result = self.engine.process_outbound(items, source='EXCEL', stop_at_picked=False)
            processed = result.get('lots_processed', result.get('processed', 0))
            
            if not result.get('success') and result.get('errors'):
                self._log(f"⚠️ 출고 오류: {result['errors'][:3]}")
                CustomMessageBox.showwarning(
                    self.root, "출고 완료",
                    f"처리: {processed}건\n오류: {result['errors'][0]}")
            
            # 화면 새로고침 (do_action_tx가 있는 경우 이미 처리됨)
            if not hasattr(self, 'do_action_tx'):
                self._refresh_after_outbound_action("EXECUTE_OUTBOUND_EXCEL")
            
            self._log(f"✅ 출고 완료: {processed}건")
            CustomMessageBox.showinfo(self.root, "출고 완료", 
                f"출고 처리가 완료되었습니다.\n\n처리: {processed}건")
            
        except (ValueError, RuntimeError, KeyError, sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as pf_err:
            err_msg = str(pf_err)
            display_msg = err_msg[:500] + '...' if len(err_msg) > 500 else err_msg
            self._log(f"❌ 출고 실패: {display_msg[:200]}")
            CustomMessageBox.show_detailed_error(
                self.root, "출고 처리 실패",
                f"출고 처리 중 오류가 발생했습니다.\n\n{display_msg}",
                exception=pf_err)
    

    def _build_simple_outbound_ui(self):
        """v4.0.3: Simple Outbound UI 위젯 생성 (~80줄 추출)"""
        from ..utils.constants import tk, ttk, VERTICAL, BOTH, LEFT, RIGHT, X, Y, W

        dialog = create_themed_toplevel(self.root)
        dialog.title("Simple Outbound")
        dialog.transient(self.root)
        dialog.grab_set()
        setup_dialog_geometry_persistence(dialog, "simple_outbound_dialog", self.root, "medium")

        main_frame = ttk.Frame(dialog, padding=10)
        main_frame.pack(fill=BOTH, expand=True)

        # Input area
        input_frame = ttk.LabelFrame(main_frame, text="Outbound Information")
        input_frame.pack(fill=X, pady=(0, 10))

        ttk.Label(input_frame, text="Customer:").grid(row=0, column=0, sticky=W, pady=2)
        customer_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=customer_var, width=30).grid(row=0, column=1, sticky=W, pady=2)

        ttk.Label(input_frame, text="Sales Ref:").grid(row=0, column=2, sticky=W, padx=(20, 0), pady=2)
        sale_ref_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=sale_ref_var, width=20).grid(row=0, column=3, sticky=W, pady=2)

        ttk.Label(input_frame, text="Enter LOT number and quantity (one per line):",
                  font=('', 16)).grid(row=1, column=0, columnspan=4, sticky=W, pady=(10, 2))
        ttk.Label(input_frame, text="Format: LOT_NO, Qty(MT)  e.g.) 1234567890, 5.0",
                  foreground=tc('text_muted')).grid(row=2, column=0, columnspan=4, sticky=W)

        # LOT text
        lot_frame = ttk.Frame(main_frame)
        lot_frame.pack(fill=BOTH, expand=True, pady=(0, 10))
        lot_text = tk.Text(lot_frame, height=10, width=60)
        lot_sb = tk.Scrollbar(lot_frame, orient=VERTICAL, command=lot_text.yview)
        lot_text.configure(yscrollcommand=lot_sb.set)
        lot_text.pack(side=LEFT, fill=BOTH, expand=True)
        lot_sb.pack(side=RIGHT, fill=Y)
        try:
            from gui_app_modular.utils.tree_enhancements import TreeviewTotalFooter as _TTF
            _TTF(frm, tree, [], {}, {}).pack(fill='x')
        except Exception as e:
            logger.warning(f'[UI] outbound_handlers: {e}')
        # Preview tree
        preview_frame = ttk.LabelFrame(main_frame, text="Outbound Preview")
        preview_frame.pack(fill=BOTH, expand=True, pady=(0, 10))
        columns = ('lot_no', 'sub_lt', 'product', 'weight_kg', 'status', 'location')
        preview_tree = ttk.Treeview(preview_frame, columns=columns, show='headings',
                                     height=10, selectmode='extended')
        for col, text, w in [('lot_no', 'LOT No', 110), ('sub_lt', '톤백#', 55),
                              ('product', 'Product', 80), ('weight_kg', '중량(kg)', 90),
                              ('status', '상태', 80), ('location', '위치', 70)]:
            preview_tree.heading(col, text=text, anchor='center')
            anchor = 'e' if col == 'weight_kg' else 'center' if col in ('sub_lt', 'status') else 'w'
            preview_tree.column(col, width=w, anchor=anchor)
        pv_sb = tk.Scrollbar(preview_frame, orient=VERTICAL, command=preview_tree.yview)
        preview_tree.configure(yscrollcommand=pv_sb.set)
        preview_tree.pack(side=LEFT, fill=BOTH, expand=True)
        pv_sb.pack(side=RIGHT, fill=Y)

        ttk.Label(main_frame, text="💡 Preview 후 출고할 톤백을 선택하세요 (미선택 시 자동 배정)",
                  foreground=tc('text_muted'), font=('', 16)).pack(pady=(0, 5))
        summary_var = tk.StringVar(value="Click Preview to check outbound details")
        ttk.Label(main_frame, textvariable=summary_var, font=('', 13, 'bold')).pack(pady=5)
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=X, pady=10)

        return {
            'dialog': dialog, 'lot_text': lot_text, 'preview_tree': preview_tree,
            'summary_var': summary_var, 'customer_var': customer_var,
            'sale_ref_var': sale_ref_var, 'btn_frame': btn_frame,
        }

    def _on_allocation_input_unified(self, initial_file: str = None) -> None:
        """Allocation 입력 통합: 파일 불러오기 vs 템플릿 붙여넣기. initial_file 있으면 선택 없이 해당 파일로 열기(드래그 등)."""
        from ..utils.constants import filedialog
        from ..utils.ui_constants import DialogSize, center_dialog

        import tkinter as tk
        from tkinter import ttk

        if initial_file:
            try:
                from ..dialogs.allocation_dialog import AllocationDialog
                dlg = AllocationDialog(self, self.engine)
                dlg.show(initial_file=initial_file)
            except (ImportError, AttributeError) as e:
                logger.error(f"Allocation 다이얼로그 오류: {e}", exc_info=True)
                CustomMessageBox.showerror(self.root, "오류", f"Allocation 열기 실패:\n{e}")
            return

        result = [None]
        win = create_themed_toplevel(self.root)
        win.title("Allocation 입력")
        pass  # modal
        win.transient(self.root)
        win.grab_set()
        win.geometry(DialogSize.get_geometry(self.root, 'small'))
        win.minsize(420, 260)
        center_dialog(win, self.root)
        f = ttk.Frame(win, padding=(20, 20, 20, 32))
        f.pack(fill=tk.BOTH, expand=True)
        from ..utils.ui_constants import (
            UPLOAD_CHOICE_HEADER, UPLOAD_CHOICE_PASTE, UPLOAD_CHOICE_UPLOAD,
            UPLOAD_CHOICE_BTN_PASTE, UPLOAD_CHOICE_BTN_UPLOAD,
        )
        ttk.Label(f, text=UPLOAD_CHOICE_HEADER, font=('맑은 고딕', 12, 'bold')).pack(anchor='w', pady=(0, 12))
        ttk.Label(f, text=UPLOAD_CHOICE_PASTE, font=('맑은 고딕', 10), wraplength=400, justify=tk.LEFT).pack(anchor='w', pady=(0, 10))
        ttk.Label(f, text=UPLOAD_CHOICE_UPLOAD, font=('맑은 고딕', 10), wraplength=400, justify=tk.LEFT).pack(anchor='w', pady=(0, 24))
        btn_f = ttk.Frame(f)
        btn_f.pack(anchor='center')
        def on_file():
            result[0] = 'file'
            win.destroy()
        def on_paste():
            result[0] = 'paste'
            win.destroy()
        ttk.Button(btn_f, text=UPLOAD_CHOICE_BTN_UPLOAD, command=on_file, width=22).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_f, text=UPLOAD_CHOICE_BTN_PASTE, command=on_paste, width=22).pack(side=tk.LEFT)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.wait_window(win)

        choice = result[0]
        if not choice:
            return
        try:
            from ..dialogs.allocation_dialog import AllocationDialog
            dlg = AllocationDialog(self, self.engine)
            if choice == 'file':
                path = filedialog.askopenfilename(
                    parent=self.root, title="Allocation Excel 선택",
                    filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
                )
                if path:
                    dlg.show(initial_file=path)
                return
            if choice == 'paste':
                from ..utils.paste_table_dialog import show_paste_table_dialog
                ALLOC_PASTE_COLUMNS = [
                    ('lot_no', 'LOT NO', 110),
                    ('sap_no', 'SAP NO', 100),
                    ('product', 'Product', 140),
                    ('qty_mt', 'QTY (MT)', 80),
                    ('sold_to', 'CUSTOMER', 130),
                    ('sale_ref', 'SALE REF', 120),
                    ('outbound_date', 'OUTBOUND DATE', 100),
                    ('warehouse', 'WH', 60),
                ]

                def on_paste_confirm(rows: list):
                    if not rows:
                        CustomMessageBox.showwarning(self.root, "경고", "붙여넣기 데이터가 없습니다.")
                        return
                    _hdr_kw = {'lot_no', 'lot no', 'lotno', 'sap_no', 'product', 'qty'}
                    if rows and str(rows[0].get('lot_no', '')).strip().lower().replace(' ', '_') in _hdr_kw:
                        rows = rows[1:]

                    parsed = []
                    parse_errors = []
                    for idx, r in enumerate(rows, 1):
                        lot_no = str(r.get('lot_no', '')).strip()
                        if not lot_no:
                            continue
                        raw_qty = str(r.get('qty_mt', '0')).replace(',', '').replace(' ', '').strip()
                        try:
                            qty = float(raw_qty) if raw_qty else 0.0
                        except (ValueError, TypeError):
                            parse_errors.append(f"행 {idx}: QTY 파싱 실패 '{r.get('qty_mt', '')}'")
                            qty = 0.0
                        if qty <= 0:
                            parse_errors.append(f"행 {idx}: {lot_no} - QTY <= 0 (건너뜀)")
                            continue
                        row = dict(r)
                        row['lot_no'] = lot_no
                        row['qty_mt'] = qty
                        parsed.append(row)

                    if parse_errors:
                        logger.warning(f"빠른 출고 파싱 경고: {parse_errors}")
                    if not parsed:
                        msg = "유효한 LOT NO · QTY 행이 없습니다."
                        if parse_errors:
                            msg += "\n\n" + "\n".join(parse_errors[:8])
                        CustomMessageBox.showwarning(self.root, "경고", msg)
                        return

                    from collections import OrderedDict
                    lot_merged = OrderedDict()
                    for row in parsed:
                        ln = row['lot_no']
                        if ln in lot_merged:
                            lot_merged[ln]['qty_mt'] += row['qty_mt']
                        else:
                            lot_merged[ln] = dict(row)
                    merged = list(lot_merged.values())

                    lot_nos = [r['lot_no'] for r in merged]
                    validation_errors = []
                    try:
                        placeholders = ",".join("?" * len(lot_nos))
                        avail_rows = self.engine.db.fetchall(
                            f"SELECT lot_no, COUNT(*) as avail_cnt "
                            f"FROM inventory_tonbag "
                            f"WHERE lot_no IN ({placeholders}) AND status='AVAILABLE' "
                            f"AND COALESCE(is_sample,0)=0 "
                            f"GROUP BY lot_no",
                            tuple(lot_nos),
                        )
                        avail_map = {r['lot_no']: r['avail_cnt'] for r in avail_rows}
                        for row in merged:
                            if row['lot_no'] not in avail_map:
                                validation_errors.append(f"❌ {row['lot_no']}: LOT 미등록 또는 가용 톤백 없음")
                    except Exception as e:
                        logger.warning(f"빠른 출고 사전 검증 실패 (계속 진행): {e}")

                    if validation_errors:
                        warn_msg = "다음 LOT에 문제가 있습니다:\n\n" + "\n".join(validation_errors[:10])
                        warn_msg += "\n\n문제 LOT을 제외하고 계속 진행할까요?"
                        if not CustomMessageBox.askyesno(self.root, "LOT 검증 경고", warn_msg):
                            return
                        problem_lots = {e.split(':')[0].replace('❌ ', '').strip() for e in validation_errors}
                        merged = [r for r in merged if r['lot_no'] not in problem_lots]
                        if not merged:
                            CustomMessageBox.showwarning(self.root, "경고", "유효한 LOT이 없습니다.")
                            return

                    for row in merged:
                        try:
                            from engine_modules.constants import get_tonbag_unit_weight
                            unit_w = get_tonbag_unit_weight(self.engine.db, row['lot_no'])
                        except Exception:
                            from engine_modules.constants import DEFAULT_TONBAG_WEIGHT
                            unit_w = DEFAULT_TONBAG_WEIGHT  # v8.6.1
                        unit_mt = unit_w / 1000.0
                        row['sublot_count'] = max(1, int(row['qty_mt'] / unit_mt + 0.001))

                    dlg = AllocationDialog(self, self.engine)
                    dlg.show_with_data(merged)

                show_paste_table_dialog(
                    self.root,
                    title="📋 Allocation 데이터 (붙여넣기)",
                    columns=ALLOC_PASTE_COLUMNS,
                    instruction="아래 표에 Excel 등에서 복사한 Allocation 데이터를 붙여넣기(Ctrl+V) 한 뒤 [확인]을 누르세요. LOT NO, QTY (MT), CUSTOMER 등.",
                    confirm_text="확인",
                    cancel_text="취소",
                    on_confirm=on_paste_confirm,
                    min_size=(800, 440),
                )
        except (ImportError, AttributeError) as e:
            logger.error(f"Allocation 입력 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(self.root, "오류", f"Allocation 입력 실패:\n{e}")

    def _on_quick_outbound_paste(self) -> None:
        """빠른 출고: 가운데 선택 창 없이 바로 붙여넣기 테이블만 열기. 컬럼 유지, 확인 시 Allocation 미리보기 → 예약."""
        try:
            from ..dialogs.allocation_dialog import AllocationDialog
            from ..utils.paste_table_dialog import show_paste_table_dialog

            ALLOC_PASTE_COLUMNS = [
                ('lot_no', 'LOT NO', 110),
                ('sap_no', 'SAP NO', 100),
                ('product', 'Product', 140),
                ('qty_mt', 'QTY (MT)', 80),
                ('sold_to', 'CUSTOMER', 130),
                ('sale_ref', 'SALE REF', 120),
                ('outbound_date', 'OUTBOUND DATE', 100),
                ('warehouse', 'WH', 60),
            ]

            def on_paste_confirm(rows: list):
                if not rows:
                    CustomMessageBox.showwarning(self.root, "경고", "붙여넣기 데이터가 없습니다.")
                    return
                _hdr_kw = {'lot_no', 'lot no', 'lotno', 'sap_no', 'product', 'qty'}
                if rows and str(rows[0].get('lot_no', '')).strip().lower().replace(' ', '_') in _hdr_kw:
                    rows = rows[1:]

                parsed = []
                parse_errors = []
                for idx, r in enumerate(rows, 1):
                    lot_no = str(r.get('lot_no', '')).strip()
                    if not lot_no:
                        continue
                    raw_qty = str(r.get('qty_mt', '0')).replace(',', '').replace(' ', '').strip()
                    try:
                        qty = float(raw_qty) if raw_qty else 0.0
                    except (ValueError, TypeError):
                        parse_errors.append(f"행 {idx}: QTY 파싱 실패 '{r.get('qty_mt', '')}'")
                        qty = 0.0
                    if qty <= 0:
                        parse_errors.append(f"행 {idx}: {lot_no} - QTY <= 0 (건너뜀)")
                        continue
                    row = dict(r)
                    row['lot_no'] = lot_no
                    row['qty_mt'] = qty
                    parsed.append(row)

                if parse_errors:
                    logger.warning(f"빠른 출고 파싱 경고: {parse_errors}")
                if not parsed:
                    msg = "유효한 LOT NO · QTY 행이 없습니다."
                    if parse_errors:
                        msg += "\n\n" + "\n".join(parse_errors[:8])
                    CustomMessageBox.showwarning(self.root, "경고", msg)
                    return

                from collections import OrderedDict
                lot_merged = OrderedDict()
                for row in parsed:
                    ln = row['lot_no']
                    if ln in lot_merged:
                        lot_merged[ln]['qty_mt'] += row['qty_mt']
                    else:
                        lot_merged[ln] = dict(row)
                merged = list(lot_merged.values())

                lot_nos = [r['lot_no'] for r in merged]
                validation_errors = []
                try:
                    placeholders = ",".join("?" * len(lot_nos))
                    avail_rows = self.engine.db.fetchall(
                        f"SELECT lot_no, COUNT(*) as avail_cnt "
                        f"FROM inventory_tonbag "
                        f"WHERE lot_no IN ({placeholders}) AND status='AVAILABLE' "
                        f"AND COALESCE(is_sample,0)=0 "
                        f"GROUP BY lot_no",
                        tuple(lot_nos),
                    )
                    avail_map = {r['lot_no']: r['avail_cnt'] for r in avail_rows}
                    for row in merged:
                        if row['lot_no'] not in avail_map:
                            validation_errors.append(f"❌ {row['lot_no']}: LOT 미등록 또는 가용 톤백 없음")
                except Exception as e:
                    logger.warning(f"빠른 출고 사전 검증 실패 (계속 진행): {e}")

                if validation_errors:
                    warn_msg = "다음 LOT에 문제가 있습니다:\n\n" + "\n".join(validation_errors[:10])
                    warn_msg += "\n\n문제 LOT을 제외하고 계속 진행할까요?"
                    if not CustomMessageBox.askyesno(self.root, "LOT 검증 경고", warn_msg):
                        return
                    problem_lots = {e.split(':')[0].replace('❌ ', '').strip() for e in validation_errors}
                    merged = [r for r in merged if r['lot_no'] not in problem_lots]
                    if not merged:
                        CustomMessageBox.showwarning(self.root, "경고", "유효한 LOT이 없습니다.")
                        return

                for row in merged:
                    try:
                        from engine_modules.constants import get_tonbag_unit_weight
                        unit_w = get_tonbag_unit_weight(self.engine.db, row['lot_no'])
                    except Exception:
                        unit_w = 500.0
                    unit_mt = unit_w / 1000.0
                    row['sublot_count'] = max(1, int(row['qty_mt'] / unit_mt + 0.001))

                dlg = AllocationDialog(self, self.engine)
                dlg.show_with_data(merged)

            show_paste_table_dialog(
                self.root,
                title="📤 빠른 출고 (붙여넣기)",
                columns=ALLOC_PASTE_COLUMNS,
                instruction="아래 표에 Excel 등에서 복사한 출고 데이터를 붙여넣기(Ctrl+V) 한 뒤 [확인]을 누르세요. LOT NO, QTY (MT), CUSTOMER 등.",
                confirm_text="확인",
                cancel_text="취소",
                on_confirm=on_paste_confirm,
                min_size=(800, 440),
            )
        except (ImportError, AttributeError) as e:
            logger.error(f"빠른 출고 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(self.root, "오류", f"빠른 출고 열기 실패:\n{e}")

    def _on_allocation_dialog(self) -> None:
        """Allocation 출고 예약 다이얼로그 열기 (v5.9.5). 통합 메뉴에서는 _on_allocation_input_unified 사용."""
        try:
            from ..dialogs.allocation_dialog import AllocationDialog
            dlg = AllocationDialog(self, self.engine)
            dlg.show()
        except (ImportError, AttributeError) as e:
            logger.error(f"Allocation 다이얼로그 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(
                self.root, "오류",
                f"Allocation 다이얼로그를 열 수 없습니다:\n{e}"
            )

    def _show_allocation_approval_queue(self) -> None:
        """출고 > Allocation 승인 대기 화면."""
        try:
            from ..dialogs.allocation_approval_dialog import AllocationApprovalDialog
            AllocationApprovalDialog(self).show_queue()
        except Exception as e:
            logger.error(f"Allocation 승인 대기 화면 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(self.root, "오류", f"승인 대기 화면을 열 수 없습니다:\n{e}")

    def _show_allocation_approval_history(self) -> None:
        """출고 > 승인 이력 조회 화면."""
        try:
            from ..dialogs.allocation_approval_dialog import AllocationApprovalDialog
            AllocationApprovalDialog(self).show_history()
        except Exception as e:
            logger.error(f"Allocation 승인 이력 화면 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(self.root, "오류", f"승인 이력 화면을 열 수 없습니다:\n{e}")

    def _apply_approved_allocation(self) -> None:
        """승인 완료(STAGED/APPROVED) 건을 RESERVED로 반영."""
        if not hasattr(self, "engine") or not self.engine:
            CustomMessageBox.showwarning(self.root, "확인", "엔진이 초기화되지 않았습니다.")
            return
        if not hasattr(self.engine, "apply_approved_allocation_reservations"):
            CustomMessageBox.showwarning(self.root, "기능 없음", "승인분 예약 반영 엔진이 없습니다.")
            return
        if not CustomMessageBox.askyesno(
            self.root,
            "예약 반영",
            "승인 완료된 Allocation(STAGED)을 RESERVED로 반영하시겠습니까?",
        ):
            return
        try:
            if hasattr(self, "do_action_tx"):
                result = self.do_action_tx(
                    "APPLY_APPROVED_ALLOCATION",
                    lambda: self.engine.apply_approved_allocation_reservations(),
                    parent=self.root,
                    refresh_mode="deferred",
                )
            else:
                result = self.engine.apply_approved_allocation_reservations()
            if result.get("success"):
                msg = f"예약 반영 완료: {result.get('applied', 0)}건"
                errs = result.get("errors", [])
                if errs:
                    msg += "\n\n미반영 사유:\n" + "\n".join(errs[:5])
                CustomMessageBox.showinfo(self.root, "완료", msg)
                if not hasattr(self, "do_action_tx"):
                    self._refresh_after_outbound_action("APPLY_APPROVED_ALLOCATION")
            else:
                errs = "\n".join(result.get("errors", [])[:8]) or "반영된 건이 없습니다."
                CustomMessageBox.showwarning(self.root, "예약 반영 결과", errs)
        except Exception as e:
            logger.error(f"승인분 예약 반영 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(self.root, "오류", f"예약 반영 중 오류가 발생했습니다.\n{e}")

    def _on_picking_list_upload(self) -> None:
        """
        v6.9.1: 피킹 리스트 업로드.
        ① PDF 파일 선택 (취소 시 즉시 종료 — UX 개선)
        ② 피킹 템플릿 선택 (고객사 프로파일 팝업)
        ③ Gate-1 교차검증 → RESERVED→PICKED
        """
        # ── v6.9.1 [FIX-4]: PDF 먼저 선택 → 취소 시 템플릿 팝업 안 뜸 ──
        from ..utils.constants import filedialog

        path = filedialog.askopenfilename(
            parent=self.root,
            title="Picking List PDF 선택",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if not path or not path.strip():
            return

        # ── 템플릿 선택 팝업 (PDF 선택 후) ──────────────────────────────
        _picking_tpl = [None]

        def _on_tpl_chosen(t: dict):
            _picking_tpl[0] = t

        try:
            from ..dialogs.picking_template_dialog import PickingTemplateDialog
            current_theme = getattr(self, '_current_theme', 'darkly')
            PickingTemplateDialog(
                self.root, self.engine,
                current_theme=current_theme,
                on_select_callback=_on_tpl_chosen,
            )
        except Exception as _tpl_err:
            logger.warning(f"[outbound] 피킹 템플릿 팝업 오류(무시): {_tpl_err}")

        if _picking_tpl[0] is None:
            return

        _tpl = _picking_tpl[0]
        logger.info(
            f"[outbound] 피킹 템플릿 선택: {_tpl.get('template_id','?')}: "
            f"/ {_tpl.get('customer','')} / {_tpl.get('bag_weight_kg',500)}kg"
        )

        # ── 이하 기존 흐름 (Gate-1) ────────────────────────────
        # v6.1.0: Gate-1 경로 (피킹 파서 → 교차검증 → RESERVED→PICKED)
        try:
            from parsers.document_parser_modular.picking_mixin import PickingListParserMixin
            parser = PickingListParserMixin()
            picking_result = parser.parse_picking_list(path)

            # ── v7.4.0: 파싱 후 meta에 템플릿 고정값 주입 ──
            meta = picking_result.meta
            _inject = [
                ('contact_person',  'contact_person'),
                ('contact_email',   'contact_email'),
                ('port_loading',    'port_loading'),
                ('port_discharge',  'port_discharge'),
                ('delivery_terms',  'delivery_terms'),
            ]
            for tpl_key, meta_attr in _inject:
                tpl_val = _tpl.get(tpl_key, '')
                if tpl_val and not getattr(meta, meta_attr, ''):
                    setattr(meta, meta_attr, tpl_val)
            # customer / bag_weight는 outbound_mixin에서 직접 사용
            # v8.6.1 [BAG-WEIGHT-PRIORITY]: PDF 감지값 > 템플릿값 > fallback(500)
            _meta_bag = getattr(getattr(picking_result, 'meta', None), 'bag_weight_kg', 0) or 0
            _tpl_bag  = int(_tpl.get('bag_weight_kg', 500) or 500)
            _resolved_bag = _meta_bag if _meta_bag >= 100 else _tpl_bag
            if not hasattr(picking_result, '_tpl_customer'):
                picking_result._tpl_customer    = _tpl.get('customer', '')
                picking_result._tpl_cust_code   = _tpl.get('customer_code', '')
                picking_result._tpl_bag_weight  = _resolved_bag
                picking_result._tpl_storage     = _tpl.get('storage_location',
                                                             '1001 GY logistics')
            if not picking_result.success:
                errs = '\n'.join(picking_result.errors[:5])
                CustomMessageBox.showerror(
                    self.root, '피킹리스트 파싱 실패',
                    f'PDF 파싱 중 오류:\n\n{errs}'
                )
                return
            warnings = list(getattr(picking_result, 'warnings', []))
            if warnings:
                top_warnings = '\n'.join(f'- {w}' for w in warnings[:5])
                if not CustomMessageBox.askyesno(
                    self.root, '피킹리스트 경고',
                    f'파싱 경고 {len(warnings)}건이 확인되었습니다.\n\n'
                    f'{top_warnings}\n\n'
                    f'경고를 확인하고 계속 진행하시겠습니까?'
                ):
                    return
            summary = picking_result.summary
            meta = picking_result.meta
            if not CustomMessageBox.askyesno(
                self.root, '피킹리스트 확인',
                f'[피킹리스트 파싱 완료]\n\n'
                f'피킹 No    : {getattr(meta, "picking_no", "")}\n'
                f'Sales Order: {getattr(meta, "sales_order", "")}\n'
                f'총 LOT     : {summary.get("total_lots", 0)}개\n'
                f'총 중량    : {summary.get("total_mt", 0):.1f} MT\n\n'
                f'Gate-1 교차검증을 진행하시겠습니까?'
            ):
                return
            if not hasattr(self.engine, 'gate1_verify_picking'):
                CustomMessageBox.showerror(
                    self.root, '기능 없음',
                    'gate1_verify_picking() 미구현'
                )
                return
            gate1 = self.engine.gate1_verify_picking(
                picking_result, getattr(meta, 'picking_no', '')
            )

            # v6.12.1: Gate-1 결과 JSON 저장 (감사 추적용)
            self._save_gate1_result_json(gate1, getattr(meta, 'picking_no', ''))

            if not gate1['passed']:
                # Gate-1 실패 → 팝업 다이얼로그(읽기 전용, '진행' 버튼 없음)
                try:
                    from ..dialogs.gate1_result_dialog import Gate1ResultDialog
                    current_theme = getattr(self, '_current_theme', 'darkly')
                    Gate1ResultDialog(
                        self.root, gate1,
                        picking_no=getattr(meta, 'picking_no', ''),
                        on_proceed=None,
                        current_theme=current_theme,
                    )
                except ImportError:
                    CustomMessageBox.showerror(
                        self.root, 'Gate-1 교차검증 실패',
                        gate1['error_report'][:800]
                    )
                self._save_gate1_report(gate1, getattr(meta, 'picking_no', ''))
                return

            # Gate-1 통과 → 팝업 다이얼로그('진행' 버튼 포함)
            _proceed_flag = [False]

            def _do_execute():
                _proceed_flag[0] = True

            try:
                from ..dialogs.gate1_result_dialog import Gate1ResultDialog
                current_theme = getattr(self, '_current_theme', 'darkly')
                Gate1ResultDialog(
                    self.root, gate1,
                    picking_no=getattr(meta, 'picking_no', ''),
                    on_proceed=_do_execute,
                    current_theme=current_theme,
                )
            except ImportError:
                # fallback: 기존 텍스트 확인
                matched = len(gate1['matched_lots'])
                if CustomMessageBox.askyesno(
                    self.root, '판매화물 결정 실행',
                    f'Gate-1 통과\n\n매칭된 LOT: {matched}개\n\n'
                    f'{matched}개 LOT을 [판매화물 결정] 상태로 전환합니다.\n계속하시겠습니까?'
                ):
                    _proceed_flag[0] = True

            if not _proceed_flag[0]:
                return

            allow_qty_mismatch = False
            approval_reason = ''
            if gate1.get('requires_approval'):
                configured_code = str(os.environ.get('SQM_ADMIN_CODE', '')).strip()
                if not configured_code:
                    CustomMessageBox.showerror(
                        self.root,
                        '관리자 코드 미설정',
                        '수량 불일치 승인 진행을 위해 SQM_ADMIN_CODE 환경변수를 설정하세요.'
                    )
                    return
                entered = CustomMessageBox.askstring(
                    self.root,
                    '관리자 승인',
                    'Gate-1 수량 불일치 승인 코드 입력:',
                    show='*',
                )
                if entered is None:
                    return
                if str(entered).strip() != configured_code:
                    CustomMessageBox.showerror(self.root, '승인 실패', '관리자 코드가 올바르지 않습니다.')
                    return
                reason = CustomMessageBox.askstring(
                    self.root,
                    '승인 사유',
                    '승인 사유를 입력하세요 (필수):',
                )
                if reason is None or not str(reason).strip():
                    CustomMessageBox.showwarning(self.root, '승인 중단', '승인 사유는 필수입니다.')
                    return
                approval_reason = str(reason).strip()
                allow_qty_mismatch = True

            if hasattr(self, 'do_action_tx'):
                exec_result = self.do_action_tx(
                    "EXECUTE_FROM_PICKING",
                    lambda: self.engine.gate1_apply_picking_result(
                        picking_result,
                        picking_no=getattr(meta, 'picking_no', ''),
                        sales_order=getattr(meta, 'sales_order', ''),
                        allow_qty_mismatch=allow_qty_mismatch,
                        approval_reason=approval_reason,
                    ),
                    parent=self.root,
                    refresh_mode="deferred",
                )
            else:
                exec_result = self.engine.gate1_apply_picking_result(
                    picking_result,
                    picking_no=getattr(meta, 'picking_no', ''),
                    sales_order=getattr(meta, 'sales_order', ''),
                    allow_qty_mismatch=allow_qty_mismatch,
                    approval_reason=approval_reason,
                )
            if exec_result.get('success'):
                CustomMessageBox.showinfo(
                    self.root, '판매화물 결정 완료',
                    f'처리: {exec_result.get("executed", 0)}개 LOT\n현장 출고 완료 후 [출고 확정]을 실행하세요.'
                )
                if not hasattr(self, 'do_action_tx'):
                    self._refresh_after_outbound_action("EXECUTE_FROM_PICKING")
            else:
                errs = '\n'.join(exec_result.get('errors', [])[:3])
                CustomMessageBox.showerror(self.root, '실행 실패', errs)
            return
        except ImportError:
            logger.debug("[SUPPRESSED] exception in outbound_handlers.py")  # noqa

        parse_picking_list_pdf = None
        try:
            from features.parsers.picking_list_parser import parse_picking_list_pdf as _parse
            parse_picking_list_pdf = _parse
        except ImportError:
            try:
                from parsers import parse_picking_list_pdf as _parse
                parse_picking_list_pdf = _parse
            except ImportError:
                try:
                    from parsers.picking_list_parser import parse_picking_list_pdf as _parse
                    parse_picking_list_pdf = _parse
                except ImportError:
                    logger.debug("[SUPPRESSED] exception in outbound_handlers.py")  # noqa

        if not parse_picking_list_pdf:
            CustomMessageBox.showerror(
                self.root,
                "Picking List 파서 없음",
                "features.parsers 또는 parsers.picking_list_parser를 불러올 수 없습니다.",
            )
            return

        try:
            doc = parse_picking_list_pdf(path)
            # v8.0.6 [PICKING-REVIEW]: 첫 행 검수 + 열 매핑 품질 점수 추가
            if _HAS_PICKING_REVIEW:
                _log_fn = getattr(self, '_append_log', None) or getattr(self, '_log_safe', None)
                doc = enrich_picking_doc_with_review(doc, path, log_fn=_log_fn)
        except Exception as e:
            logger.exception("Picking List PDF 파싱 오류")
            CustomMessageBox.show_detailed_error(
                self.root,
                "파싱 오류",
                "PDF 파싱 중 오류가 발생했습니다.",
                exception=e,
            )
            return

        on_apply = None
        try:
            from features.parsers.picking_engine import apply_picking_list_to_db
            def _apply(d, p):
                apply_picking_list_to_db(self.engine, d, p)
                CustomMessageBox.showinfo(self.root, "완료", "DB 반영이 완료되었습니다.")
            on_apply = _apply
        except ImportError:
            logger.debug("features.parsers.picking_engine 없음 — DB 반영 버튼 비표시")

        from ..dialogs.picking_list_preview_dialog import PickingListPreviewDialog
        PickingListPreviewDialog(self.root, doc, path, on_apply_clicked=on_apply)

    def _save_gate1_result_json(self, gate1: dict, picking_no: str) -> None:
        """v6.12.1: Gate-1 결과를 JSON 파일로 저장 (감사 추적용)."""
        import os
        import json
        from datetime import datetime
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'Gate1_{picking_no}_{ts}.json'
        fpath = os.path.join(os.path.expanduser('~'), 'Desktop', fname)
        try:
            # set은 JSON 직렬화 불가 → list 변환
            serializable = {}
            for k, v in gate1.items():
                if isinstance(v, set):
                    serializable[k] = sorted(v)
                elif k == 'lot_details':
                    serializable[k] = v  # list[dict] — 이미 직렬화 가능
                else:
                    serializable[k] = v
            with open(fpath, 'w', encoding='utf-8') as f:
                json.dump(serializable, f, ensure_ascii=False, indent=2)
            if hasattr(self, '_log'):
                self._log(f'Gate-1 JSON 저장: {fpath}')
            logger.info(f'[Gate-1] JSON 저장 완료: {fpath}')
        except (OSError, TypeError) as e:
            logger.debug(f'Gate-1 JSON 저장 실패: {e}')

    def _save_gate1_report(self, gate1: dict, picking_no: str) -> None:
        """v6.1.0: Gate-1 실패 에러 리포트를 바탕화면에 텍스트 파일로 저장."""
        import os
        from datetime import datetime
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'Gate1_실패_{picking_no}_{ts}.txt'
        fpath = os.path.join(os.path.expanduser('~'), 'Desktop', fname)
        try:
            with open(fpath, 'w', encoding='utf-8') as f:
                f.write(gate1.get('error_report', ''))
            if hasattr(self, '_log'):
                self._log(f'Gate-1 에러 리포트 저장: {fpath}')
        except OSError:
            logger.debug(f'Gate-1 리포트 저장 실패: {fpath}')

    def _on_revert_picked_to_reserved(self) -> None:
        """판매화물 결정 취소: PICKED → RESERVED(판매 배정).
        v8.1.8: 탭에서 LOT 선택 시 선택분만, 미선택 시 다이얼로그.
        """
        engine = getattr(self, 'engine', None)
        if not engine or not hasattr(engine, 'revert_picked_to_reserved'):
            CustomMessageBox.showwarning(
                self.root, '기능 없음',
                'revert_picked_to_reserved()를 사용할 수 없습니다.'
            )
            return

        # ── 탭에서 선택된 LOT 먼저 확인 ───────────────────────────────────
        pre_selected = []
        for attr in ('tree_picked', 'tree_picked_detail'):
            tree = getattr(self, attr, None)
            if tree:
                for iid in tree.selection():
                    vals = tree.item(iid).get('values', [])
                    lot_no = str(vals[0]).strip() if vals else ''
                    if lot_no and lot_no not in pre_selected:
                        pre_selected.append(lot_no)
                if pre_selected:
                    break

        # ── 전체 대상 목록 조회 ────────────────────────────────────────────
        try:
            rows = engine.db.fetchall(
                "SELECT DISTINCT lot_no FROM allocation_plan "
                "WHERE status = 'EXECUTED' ORDER BY lot_no"
            )
        except Exception:
            rows = []
        lot_list = [str(r.get('lot_no', '')).strip() for r in (rows or []) if r.get('lot_no')]

        if not lot_list:
            CustomMessageBox.showinfo(
                self.root, '대상 없음',
                '되돌릴 판매화물 결정(PICKED) 건이 없습니다.'
            )
            return

        # 탭 선택 항목이 있으면 바로 실행 ─────────────────────────────────
        if pre_selected:
            msg = (f"선택한 {len(pre_selected)}개 LOT를\n"
                   "판매 배정(RESERVED)으로 되돌립니다.\n계속하시겠습니까?")
            if not CustomMessageBox.askyesno(self.root, '판매화물 결정 취소', msg):
                return
            total, result_msg = self._run_revert_picked_to_reserved(engine, pre_selected)
            CustomMessageBox.showinfo(self.root, '취소 완료', result_msg)
            self._refresh_after_outbound_action("REVERT_PICKED_SELECTED")
            return

        self._show_revert_lot_dialog(
            title='판매화물 결정 취소 (→ 판매 배정)',
            lot_list=lot_list,
            confirm_message='선택한 LOT을 판매 배정(RESERVED)으로 되돌립니다.',
            revert_all_message='전체를 판매 배정으로 되돌립니다.',
            revert_fn=lambda lot_nos: self._run_revert_picked_to_reserved(engine, lot_nos),
        )

    def _run_revert_picked_to_reserved(self, engine, lot_nos):
        total = 0
        for lot_no in lot_nos:
            r = engine.revert_picked_to_reserved(lot_no=lot_no)
            total += r.get('reverted', 0)
        return total, f"{total}건 → 판매 배정(RESERVED)"

    def _on_revert_outbound_to_available(self) -> None:
        """출고 취소: OUTBOUND/SOLD → AVAILABLE 직접 복귀.
        v8.1.8 BUG-A 수정: 'SOLD' 하드코딩 → STATUS_OUTBOUND + 'SOLD' 병행 조회.
        탭에서 LOT 선택 시 선택분만, 미선택 시 다이얼로그 표시.
        """
        engine = getattr(self, 'engine', None)
        if not engine or not hasattr(engine, 'revert_sold_to_picked'):
            CustomMessageBox.showwarning(
                self.root, '기능 없음',
                'revert_sold_to_picked()를 사용할 수 없습니다.'
            )
            return

        # ── 탭에서 선택된 LOT 먼저 확인 (Q2: 선택 우선 방식) ──────────────
        pre_selected = []
        for attr in ('tree_sold', 'tree_sold_detail'):
            tree = getattr(self, attr, None)
            if tree:
                for iid in tree.selection():
                    vals = tree.item(iid).get('values', [])
                    lot_no = str(vals[0]).strip() if vals else ''
                    if lot_no and lot_no not in pre_selected:
                        pre_selected.append(lot_no)
                if pre_selected:
                    break

        # ── 전체 대상 목록 조회 (OUTBOUND + SOLD 모두) ─────────────────────
        try:
            rows = engine.db.fetchall(
                "SELECT DISTINCT lot_no FROM inventory_tonbag "
                "WHERE status IN ('OUTBOUND', 'SOLD') ORDER BY lot_no"
            )
        except Exception:
            rows = []
        lot_list = [str(r.get('lot_no', '')).strip() for r in (rows or []) if r.get('lot_no')]

        if not lot_list:
            CustomMessageBox.showinfo(
                self.root, '대상 없음',
                '되돌릴 출고(OUTBOUND/SOLD) 건이 없습니다.'
            )
            return

        # 탭 선택 항목이 있으면 바로 확인 후 실행 ──────────────────────────
        if pre_selected:
            msg = (f"선택한 {len(pre_selected)}개 LOT의 출고를 취소하여\n"
                   "AVAILABLE(판매가능)으로 되돌립니다.\n계속하시겠습니까?")
            if not CustomMessageBox.askyesno(self.root, '출고 취소', msg):
                return
            total, result_msg = self._run_revert_sold_to_picked(engine, pre_selected)
            CustomMessageBox.showinfo(self.root, '취소 완료', result_msg)
            self._refresh_after_outbound_action("REVERT_OUTBOUND_SELECTED")
            return

        # 선택 없으면 다이얼로그 표시 ─────────────────────────────────────
        self._show_revert_lot_dialog(
            title='출고 취소 (→ AVAILABLE)',
            lot_list=lot_list,
            confirm_message='선택한 LOT을 AVAILABLE(판매가능)으로 되돌립니다.',
            revert_all_message='전체 출고를 취소하여 AVAILABLE로 되돌립니다.',
            revert_fn=lambda lot_nos: self._run_revert_sold_to_picked(engine, lot_nos),
        )

    # 하위 호환 alias — sold_tab의 _safe_call('_on_revert_sold_to_picked') 유지
    _on_revert_sold_to_picked = _on_revert_outbound_to_available

    def _run_revert_sold_to_picked(self, engine, lot_nos):
        """출고 취소 실행 — OUTBOUND/SOLD → AVAILABLE."""
        total = 0
        for lot_no in lot_nos:
            r = engine.revert_sold_to_picked(lot_no=lot_no)
            total += r.get('reverted', 0)
        return total, f"{total}건 → AVAILABLE (출고 취소 완료)"

    def _show_revert_lot_dialog(
        self,
        title,
        lot_list,
        confirm_message,
        revert_all_message,
        revert_fn,
    ) -> None:
        """LOT 목록 다중 선택 다이얼로그 — 일부/전체 취소 공통."""
        import tkinter as tk
        from tkinter import ttk

        d = create_themed_toplevel(self.root)
        d.title(title)
        d.transient(self.root)
        d.grab_set()
        f = ttk.Frame(d, padding=10)
        f.pack(fill=tk.BOTH, expand=True)
        ttk.Label(f, text="취소할 LOT를 선택하세요 (일부 또는 [전체 선택] 후 선택 취소).").pack(anchor=tk.W)
        lb_frame = ttk.Frame(f)
        lb_frame.pack(fill=tk.BOTH, expand=True, pady=(4, 8))
        scroll = tk.Scrollbar(lb_frame)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        lb = tk.Listbox(lb_frame, selectmode=tk.EXTENDED, height=12, yscrollcommand=scroll.set, font=('Consolas', 10))
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.config(command=lb.yview)
        for lot in lot_list:
            lb.insert(tk.END, lot)

        def select_all():
            lb.selection_set(0, tk.END)

        def do_revert():
            sel = lb.curselection()
            lot_nos = [lot_list[i] for i in sel] if sel else []
            if not lot_nos:
                CustomMessageBox.showwarning(
                    d, '선택 필요',
                    'LOT을 선택하거나 [전체 선택] 버튼으로 전부 선택한 뒤 [선택 취소]를 누르세요.'
                )
                return
            if len(lot_nos) == len(lot_list):
                msg = revert_all_message
            else:
                msg = f"선택한 {len(lot_nos)}개 LOT에 대해 취소합니다.\n{confirm_message}"
            if not CustomMessageBox.askyesno(d, '확인', msg + '\n계속하시겠습니까?'):
                return
            total, result_msg = revert_fn(lot_nos)
            d.destroy()
            CustomMessageBox.showinfo(self.root, '취소 완료', result_msg)
            self._refresh_after_outbound_action("REVERT_LOT_DIALOG_ACTION")

        btn_f = ttk.Frame(f)
        btn_f.pack(fill=tk.X, pady=(0, 4))
        ttk.Button(btn_f, text="전체 선택", command=select_all).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_f, text="선택 취소", command=do_revert).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_f, text="닫기", command=d.destroy).pack(side=tk.LEFT, padx=2)
        try:
            from ..utils.ui_constants import setup_dialog_geometry_persistence
            setup_dialog_geometry_persistence(d, "revert_lot_dialog", self.root, "large")
        except Exception:
            d.geometry("500x400")
        d.update_idletasks()
        try:
            from ..utils.ui_constants import center_dialog
            center_dialog(d, self.root)
        except Exception as e:
            logger.debug(f"[출고UI] 다이얼로그 센터링 실패: {e}")

    def _on_barcode_scan_upload(self) -> None:
        """v6.12 Stage3: 바코드 스캔 파일 업로드 → UID 대조 + PICKED→SOLD"""
        from tkinter import filedialog
        import tkinter.messagebox as mb
        import os

        file_path = filedialog.askopenfilename(
            parent=self.root,
            title="바코드 스캔 파일 선택 (CSV/Excel/TXT)",
            filetypes=[("스캔 파일", "*.csv;*.xlsx;*.xls;*.txt"), ("모든 파일", "*.*")]
        )
        if not file_path:
            return

        try:
            from core.barcode_scan_engine import BarcodeScanEngine
            scanner = BarcodeScanEngine(self.engine.db, engine=self.engine)  # v8.0.2 P2

            # Phase3: 랜덤 출고 모드(리오님 운영) — 스캔 즉시 확정(OUT=SOLD)
            #   - 환경변수로 강제 가능: SQM_OUTBOUND_MODE=random_scan_confirm
            outbound_mode = (os.environ.get('SQM_OUTBOUND_MODE', 'random_scan_confirm') or '').strip().lower()

            selected_sale_ref = None
            sale_refs = scanner.get_picked_sale_refs()
            if len(sale_refs) == 1:
                selected_sale_ref = sale_refs[0]
            elif len(sale_refs) > 1:
                pick_hint = ", ".join(sale_refs[:10])
                if len(sale_refs) > 10:
                    pick_hint += f" ... 외 {len(sale_refs) - 10}건"
                selected = CustomMessageBox.askstring(
                    self.root,
                    "SALE REF 선택",
                    "이번 바코드 스캔 대상 SALE REF를 정확히 1개 입력하세요.\n"
                    f"(선택 가능: {pick_hint})",
                )
                if not selected:
                    return
                selected = selected.strip()
                if selected not in sale_refs:
                    mb.showerror(
                        "SALE REF 오류",
                        "유효하지 않은 SALE REF 입니다. 목록의 값 중 1개를 정확히 입력하세요.",
                        parent=self.root,
                    )
                    return
                selected_sale_ref = selected

            scanned_codes = scanner.read_scan_file(file_path)
            if not scanned_codes:
                mb.showwarning("스캔 파일 비어있음", "스캔 파일에 유효한 UID가 없습니다.", parent=self.root)
                return

            # ------------------------------
            # Phase 3 (RUBI) : Random Scan Confirm
            # ------------------------------
            if outbound_mode in ('random_scan_confirm', 'random', 'scan_confirm'):
                try:
                    # All-or-Nothing 트랜잭션. Target 초과/UID 오류 1건이라도 있으면 전체 롤백.
                    result = scanner.process_barcode_scan_confirm_out(scanned_codes, sale_ref=selected_sale_ref)
                    if not isinstance(result, dict) or not result.get('success'):
                        err = (result.get('errors') or ['출고 확정 실패'])[0] if isinstance(result, dict) else '출고 확정 실패'
                        mb.showerror('출고 확정 실패', str(err), parent=self.root)
                        return

                    confirmed = int(result.get('confirmed', 0) or 0)
                    # LOT별 진행 현황 요약
                    lot_cnt = len({r.get('lot_no') for r in (result.get('rows') or []) if isinstance(r, dict) and r.get('lot_no')})
                    mb.showinfo(
                        '바코드 스캔 출고 확정 완료',
                        f"✅ 스캔 즉시 확정 완료\n\n- 확정 톤백: {confirmed}건\n- LOT 수: {lot_cnt}개\n\n(Phase3 규칙: 스캔=확정)",
                        parent=self.root,
                    )

                                        # Phase4: 스캔 확정 리포트 자동 저장(CSV)
                    try:
                        from core.config import OUTPUT_DIR
                        report_path = scanner.export_scan_confirm_report_csv(
                            result.get('rows') or [],
                            output_dir=str(OUTPUT_DIR),
                            prefix=f"OUTBOUND_SCAN_{(selected_sale_ref or 'NO_SALEREF')}",
                        )
                        if report_path:
                            logger.info(f"[Phase4] outbound scan report saved: {report_path}")
                    except Exception as _re:
                        logger.debug(f"[Phase4] report export skipped: {_re}")

                    if hasattr(self, '_refresh_after_outbound_action'):
                        self._refresh_after_outbound_action('PHASE3_SCAN_CONFIRM_OUT')
                    if hasattr(self, '_refresh_tonbag_list'):
                        self._refresh_tonbag_list()
                    return
                except Exception as e:
                    # BarcodeScanEngine 내부에서 RuntimeError(json)로 실패 상세를 올 수 있음.
                    mb.showerror('출고 확정 실패', str(e), parent=self.root)
                    return

            expected_uids = scanner.get_picked_uids(sale_ref=selected_sale_ref)
            if not expected_uids:
                lot_reserved = 0
                try:
                    lot_reserved = scanner.get_lot_mode_reserved_count()
                except Exception:
                    lot_reserved = 0
                if lot_reserved <= 0:
                    mb.showwarning(
                        "PICKED 톤백 없음",
                        "PICKED 상태 톤백이 없습니다.\n출고 실행을 먼저 진행하세요.",
                        parent=self.root
                    )
                    return
                if not mb.askyesno(
                    "LOT 단위 예약 스캔",
                    f"PICKED 톤백은 없지만 LOT 단위 예약 {lot_reserved}건이 있습니다.\n"
                    "이번 스캔을 LOT 단위 예약 확정(SOLD)으로 처리할까요?",
                    parent=self.root,
                ):
                    return
                if hasattr(self, 'do_action_tx'):
                    sold_result = self.do_action_tx(
                        "BARCODE_LOT_MODE_TO_SOLD",
                        lambda: scanner.process_barcode_scan_for_lot_mode(file_path),
                        parent=self.root,
                        refresh_mode="deferred",
                    )
                else:
                    sold_result = scanner.process_barcode_scan_for_lot_mode(file_path)

                if not isinstance(sold_result, dict):
                    mb.showerror("오류", "바코드 스캔 결과 형식이 올바르지 않습니다.", parent=self.root)
                    return
                if sold_result.get('success') is False and sold_result.get('sold', 0) <= 0:
                    err = (sold_result.get('errors') or ['처리 실패'])[0]
                    mb.showerror("오류", f"LOT 단위 스캔 처리 실패:\n{err}", parent=self.root)
                    return
                msg = (
                    f"LOT 단위 스캔 출고 완료: {sold_result.get('sold', 0)}건 SOLD 전환\n"
                    f"잔여 LOT 예약: {sold_result.get('remaining_lot_reserved', 0)}건"
                )
                if sold_result.get('not_found'):
                    msg += f"\n⚠️ 미매칭 UID: {len(sold_result.get('not_found', []))}건"
                if sold_result.get('no_plan'):
                    msg += f"\n⚠️ 예약 계획 없는 UID: {len(sold_result.get('no_plan', []))}건"
                mb.showinfo("LOT 단위 스캔 출고 완료", msg, parent=self.root)
                if not hasattr(self, 'do_action_tx'):
                    self._refresh_after_outbound_action("BARCODE_LOT_MODE_TO_SOLD")
                if hasattr(self, '_refresh_tonbag_list'):
                    self._refresh_tonbag_list()
                return

            verify = scanner.verify_outbound_scan(
                expected_uids=expected_uids,
                scanned_uids_raw=scanned_codes,
                outbound_ref=f"SCAN-{os.path.basename(file_path)}",
                scan_file_name=os.path.basename(file_path),
                sale_ref=selected_sale_ref or '',
            )

            if verify['result'] == 'FAIL':
                msg = verify['message'] + "\n\n"
                if verify['missing']:
                    msg += f"누락 UID ({len(verify['missing'])}개):\n"
                    for u in verify['missing'][:10]: msg += f"  - {u}\n"
                if verify['extra']:
                    msg += f"\n초과 UID ({len(verify['extra'])}개):\n"
                    for u in verify['extra'][:10]: msg += f"  - {u}\n"
                mb.showerror("UID 대조 실패 — 출고 중단", msg, parent=self.root)
                return

            if verify.get('result') == 'PASS_SWAP':
                lot_preview = ", ".join((verify.get('swap_lots') or [])[:8])
                if len(verify.get('swap_lots') or []) > 8:
                    lot_preview += f" ... 외 {len(verify.get('swap_lots') or []) - 8}개"
                verify_msg = (
                    f"{verify['message']}\n\n"
                    f"대상 LOT: {lot_preview or '-'}\n"
                    f"같은 LOT 내부에서만 스왑 출고가 허용됩니다."
                )
            else:
                verify_msg = verify['message']

            if not mb.askyesno("UID 대조 통과", f"{verify_msg}\n\nPICKED → SOLD 전환하시겠습니까?", parent=self.root):
                return

            # TOCTOU 방지: 실행 직전에 동일 스캔 데이터로 재검증
            reverify = scanner.verify_outbound_scan(
                expected_uids=expected_uids,
                scanned_uids_raw=scanned_codes,
                outbound_ref=f"SCAN-RECHECK-{os.path.basename(file_path)}",
                scan_file_name=os.path.basename(file_path),
                sale_ref=selected_sale_ref or '',
            )
            if reverify.get('result') == 'FAIL':
                mb.showerror(
                    "재검증 실패",
                    "실행 직전 상태가 변경되어 출고를 중단했습니다.\n"
                    "다시 조회 후 재시도하세요.",
                    parent=self.root,
                )
                return

            if hasattr(self, 'do_action_tx'):
                sold_result = self.do_action_tx(
                    "BARCODE_SCAN_TO_SOLD",
                    lambda: scanner.process_barcode_scan_to_sold(scanned_codes, sale_ref=selected_sale_ref),
                    parent=self.root,
                    refresh_mode="deferred",
                )
            else:
                sold_result = scanner.process_barcode_scan_to_sold(scanned_codes, sale_ref=selected_sale_ref)

            if not isinstance(sold_result, dict):
                mb.showerror("오류", "바코드 스캔 결과 형식이 올바르지 않습니다.", parent=self.root)
                return
            if sold_result.get('success') is False and sold_result.get('sold', 0) <= 0:
                err = (sold_result.get('errors') or ['처리 실패'])[0]
                mb.showerror("오류", f"바코드 스캔 처리 실패:\n{err}", parent=self.root)
                return

            msg = f"출고 완료: {sold_result.get('sold', 0)}건 SOLD 전환\n"
            if sold_result.get('swap_count', 0) > 0:
                msg += f"\n🔁 LOT 내부 Swap 적용: {sold_result.get('swap_count', 0)}건\n"
            if sold_result.get('not_found'):
                msg += f"\n⚠️ 미매칭: {len(sold_result.get('not_found', []))}건\n"
            if sold_result.get('remaining_picked', 0) > 0:
                msg += f"\n⚠️ 잔여 PICKED: {sold_result.get('remaining_picked', 0)}건\n"
            mb.showinfo("바코드 스캔 출고 완료", msg, parent=self.root)

            if not hasattr(self, 'do_action_tx'):
                self._refresh_after_outbound_action("BARCODE_SCAN_TO_SOLD")
            if hasattr(self, '_refresh_tonbag_list'):
                self._refresh_tonbag_list()

        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"바코드 스캔 오류: {e}", exc_info=True)
            mb.showerror("오류", f"바코드 스캔 처리 중 오류:\n{e}", parent=self.root)

    def _on_barcode_live_scan(self) -> None:
        """(Phase4) 실시간 바코드 스캔(USB 스캐너 Enter) → 즉시 확정(OUT=SOLD)
        - 파일 업로드 없이, 스캐너가 입력+Enter를 치면 즉시 확정
        - Undo(최근 1건), 리포트 저장 지원
        """

        import tkinter as tk
        from tkinter import ttk
        import tkinter.messagebox as mb

        try:
            from core.barcode_scan_engine import BarcodeScanEngine
            scanner = BarcodeScanEngine(self.engine.db, engine=self.engine)  # v8.0.2 P2
        except Exception as e:
            mb.showerror("오류", f"바코드 엔진 로드 실패: {e}", parent=self.root)
            return

        # SALE REF 선택(선택 강제: Target 체크 정확도)
        selected_sale_ref = None
        sale_refs = scanner.get_picked_sale_refs()
        if len(sale_refs) == 1:
            selected_sale_ref = sale_refs[0]
        elif len(sale_refs) > 1:
            pick_hint = ", ".join(sale_refs[:10])
            if len(sale_refs) > 10:
                pick_hint += f" ... 외 {len(sale_refs) - 10}건"
            selected = CustomMessageBox.askstring(
                self.root,
                "SALE REF 선택",
                "실시간 스캔 대상 SALE REF를 정확히 1개 입력하세요.\n"
                f"(선택 가능: {pick_hint})",
            )
            if not selected:
                return
            selected = selected.strip()
            if selected not in sale_refs:
                mb.showerror("SALE REF 오류", "유효하지 않은 SALE REF 입니다.", parent=self.root)
                return
            selected_sale_ref = selected

        d = create_themed_toplevel(self.root)
        d.title("📟 실시간 바코드 스캔 (Phase4: 스캔=즉시 확정)")
        d.transient(self.root)
        d.grab_set()
        setup_dialog_geometry_persistence(d, "barcode_live_scan_dialog", self.root, "medium")

        rows_confirmed = []
        var_status = tk.StringVar(value="대기: 바코드를 스캔하면 Enter로 입력됩니다.")
        var_cnt = tk.StringVar(value="확정 0 / 실패 0")

        top = ttk.Frame(d, padding=10)
        top.pack(fill=tk.BOTH, expand=True)

        ttk.Label(top, text="바코드 스캐너 입력창(자동 포커스):").pack(anchor="w")
        ent = ttk.Entry(top, width=40)
        ent.pack(fill=tk.X, pady=(4, 8))
        ent.focus_set()

        ttk.Label(top, textvariable=var_status).pack(anchor="w", pady=(0, 6))

        lb = tk.Listbox(top, height=12)
        lb.pack(fill=tk.BOTH, expand=True)

        btns = ttk.Frame(top)
        btns.pack(fill=tk.X, pady=(8, 0))

        def _update_counter():
            ok = sum(1 for r in rows_confirmed if r.get("success"))
            bad = sum(1 for r in rows_confirmed if not r.get("success"))
            var_cnt.set(f"확정 {ok} / 실패 {bad}")

        def _append_line(msg: str):
            lb.insert(tk.END, msg)
            lb.yview_moveto(1.0)

        def _confirm_from_entry(event=None):
            uid = (ent.get() or "").strip()
            ent.delete(0, tk.END)
            if not uid:
                return
            res = scanner.confirm_one_uid_live(uid, sale_ref=selected_sale_ref, source="live_scan")
            rows_confirmed.append(res)
            if res.get("success"):
                _append_line(f"✅ {res.get('uid')}  LOT={res.get('lot_no')}  {res.get('weight_kg')}kg")
                var_status.set(f"확정 OK: {res.get('uid')}")
            else:
                _append_line(f"❌ {res.get('uid')}  REASON={res.get('reason')}")
                var_status.set(f"확정 FAIL: {res.get('uid')} ({res.get('reason')})")
            _update_counter()
            try:
                if hasattr(self, '_refresh_tonbag_list'):
                    self._refresh_tonbag_list()
            except Exception:
                logger.debug("[SUPPRESSED] exception in outbound_handlers.py")  # noqa
            ent.focus_set()

        def _undo_last():
            res = scanner.undo_last_scan_confirm(sale_ref=selected_sale_ref)
            if res.get("success"):
                _append_line(f"↩️ UNDO OK: {res.get('uid','')}")
                var_status.set("최근 1건 Undo 완료")
                try:
                    if hasattr(self, '_refresh_tonbag_list'):
                        self._refresh_tonbag_list()
                except Exception:
                    logger.debug("[SUPPRESSED] exception in outbound_handlers.py")  # noqa
            else:
                _append_line(f"⚠️ UNDO FAIL: {res.get('message','')}")
                var_status.set(res.get("message","Undo 실패"))
            ent.focus_set()

        def _save_report():
            ok_rows = []
            for r in rows_confirmed:
                if isinstance(r, dict) and r.get("success"):
                    ok_rows.append(r)
            if not ok_rows:
                mb.showwarning("리포트 없음", "확정된 스캔이 없습니다.", parent=d)
                return
            try:
                from core.config import OUTPUT_DIR
                path = scanner.export_scan_confirm_report_csv(
                    rows=[{
                        "sale_ref": rr.get("sale_ref",""),
                        "customer": rr.get("customer",""),
                        "lot_no": rr.get("lot_no",""),
                        "tonbag_id": rr.get("tonbag_id",""),
                        "uid": rr.get("uid",""),
                        "weight_kg": rr.get("weight_kg",""),
                    } for rr in ok_rows],
                    output_dir=str(OUTPUT_DIR),
                    prefix=f"OUTBOUND_LIVE_{(selected_sale_ref or 'NO_SALEREF')}",
                )
                if path:
                    mb.showinfo("저장 완료", f"리포트 저장 완료:\n{path}", parent=d)
                else:
                    mb.showwarning("저장 실패", "리포트 저장에 실패했습니다.", parent=d)
            except Exception as e:
                mb.showerror("오류", f"리포트 저장 실패: {e}", parent=d)
            ent.focus_set()

        ttk.Button(btns, text="Undo(최근 1건)", command=_undo_last).pack(side=tk.LEFT, padx=2)
        ttk.Button(btns, text="리포트 저장(CSV)", command=_save_report).pack(side=tk.LEFT, padx=2)
        ttk.Button(btns, text="닫기", command=d.destroy).pack(side=tk.RIGHT, padx=2)

        ttk.Label(top, textvariable=var_cnt).pack(anchor="e", pady=(6, 0))

        ent.bind("<Return>", _confirm_from_entry)
        ent.bind("<KP_Enter>", _confirm_from_entry)

        d.update_idletasks()

    def _query_swap_report_rows(
        self,
        start_date: str,
        end_date: str,
        customer: str = "",
        lot_no: str = "",
        operator: str = "",
    ) -> list[dict]:
        """uid_swap_history 기반 Swap 리포트 조회 (기간/고객사/LOT/작업자 필터)."""
        sql = """
            SELECT
                s.created_at,
                s.lot_no,
                COALESCE(st.customer, t.picked_to, '') AS customer,
                COALESCE(st.created_by, 'barcode_scan_swap') AS operator,
                COALESCE(s.expected_uid, '') AS expected_uid,
                COALESCE(s.scanned_uid, '') AS scanned_uid,
                COALESCE(s.reason, '') AS reason
            FROM uid_swap_history s
            LEFT JOIN inventory_tonbag t ON t.id = s.scanned_tonbag_id
            LEFT JOIN (
                SELECT s1.*
                FROM sold_table s1
                INNER JOIN (
                    SELECT tonbag_id, MAX(id) AS max_id
                    FROM sold_table
                    GROUP BY tonbag_id
                ) m ON m.max_id = s1.id
            ) st ON st.tonbag_id = s.scanned_tonbag_id
            WHERE date(s.created_at) BETWEEN date(?) AND date(?)
              AND (? = '' OR s.lot_no LIKE ?)
              AND (? = '' OR COALESCE(st.customer, t.picked_to, '') LIKE ?)
              AND (? = '' OR COALESCE(st.created_by, 'barcode_scan_swap') LIKE ?)
            ORDER BY s.created_at DESC, s.id DESC
            LIMIT 5000
        """
        lot_like = f"%{lot_no.strip()}%" if lot_no.strip() else ""
        customer_like = f"%{customer.strip()}%" if customer.strip() else ""
        operator_like = f"%{operator.strip()}%" if operator.strip() else ""
        try:
            rows = self.engine.db.fetchall(
                sql,
                (
                    start_date.strip(),
                    end_date.strip(),
                    lot_no.strip(),
                    lot_like,
                    customer.strip(),
                    customer_like,
                    operator.strip(),
                    operator_like,
                ),
            )
            return rows or []
        except Exception as e:
            logger.error(f"Swap 리포트 조회 실패: {e}", exc_info=True)
            return []

    def _get_swap_report_save_dir(self, base_date_text: str) -> str:
        """Swap 리포트 저장 경로 고정: reports/swap/YYYY-MM."""
        month_token = datetime.now().strftime("%Y-%m")
        try:
            dt = datetime.strptime(base_date_text.strip(), "%Y-%m-%d")
            month_token = dt.strftime("%Y-%m")
        except (ValueError, TypeError) as e:
            logger.debug(f"[출고UI] 날짜 파싱 실패: {e}")
        root = get_app_base_dir()
        out_dir = os.path.join(root, "reports", "swap", month_token)
        os.makedirs(out_dir, exist_ok=True)
        return out_dir

    def _export_swap_report_csv(self, rows: list[dict], out_path: str) -> None:
        """Swap 리포트 CSV 저장."""
        headers = ["created_at", "lot_no", "customer", "operator", "expected_uid", "scanned_uid", "reason"]
        with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow([str(r.get(h, "") or "") for h in headers])

    def _export_swap_report_xlsx(self, rows: list[dict], out_path: str) -> None:
        """Swap 리포트 XLSX 저장."""
        try:
            from openpyxl import Workbook
        except Exception as e:
            raise RuntimeError(f"openpyxl 미설치로 XLSX 저장 불가: {e}")
        wb = Workbook()
        ws = wb.active
        ws.title = "swap_report"
        headers = ["created_at", "lot_no", "customer", "operator", "expected_uid", "scanned_uid", "reason"]
        ws.append(headers)
        for r in rows:
            ws.append([str(r.get(h, "") or "") for h in headers])
        wb.save(out_path)

    def _show_swap_report_dialog(self) -> None:
        """Swap 리포트 팝업 (기간/고객사/LOT/작업자 필터 + CSV/XLSX 저장)."""
        import tkinter as tk
        from tkinter import ttk

        if not hasattr(self, "engine") or not getattr(self.engine, "db", None):
            CustomMessageBox.showwarning(self.root, "경고", "DB 연결이 없어 Swap 리포트를 열 수 없습니다.")
            return
        try:
            from core.barcode_scan_engine import BarcodeScanEngine
            BarcodeScanEngine(self.engine.db, engine=self.engine)  # v8.0.2 P2  # uid_swap_history 테이블 보장
        except Exception as e:
            logger.debug(f"Swap 리포트 테이블 준비 스킵: {e}")

        dlg = create_themed_toplevel(self.root)
        dlg.title("🔁 Swap 리포트")
        dlg.transient(self.root)
        dlg.grab_set()
        setup_dialog_geometry_persistence(dlg, "swap_report_dialog", self.root, "large")
        frm = ttk.Frame(dlg, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        today = datetime.now().strftime("%Y-%m-%d")
        start_var = tk.StringVar(value='')
        end_var = tk.StringVar(value='')
        customer_var = tk.StringVar()
        lot_var = tk.StringVar()
        operator_var = tk.StringVar()
        summary_var = tk.StringVar(value="조회 전")

        filter_frm = ttk.LabelFrame(frm, text="조회 조건")
        filter_frm.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(filter_frm, text="시작일(YYYY-MM-DD)").grid(row=0, column=0, padx=4, pady=4, sticky="w")
        _e_start = ttk.Entry(filter_frm, textvariable=start_var, width=14)
        _e_start.grid(row=0, column=1, padx=4, pady=4, sticky="w")
        ttk.Label(filter_frm, text="종료일").grid(row=0, column=2, padx=4, pady=4, sticky="w")
        _e_end = ttk.Entry(filter_frm, textvariable=end_var, width=14)
        _e_end.grid(row=0, column=3, padx=4, pady=4, sticky="w")
        # v8.1.8: 플레이스홀더 연결
        try:
            from gui_app_modular.utils.tree_enhancements import attach_date_placeholder
            attach_date_placeholder(_e_start, start_var)
            attach_date_placeholder(_e_end,   end_var)
        except Exception as e:
            logger.warning(f'[UI] outbound_handlers: {e}')
        ttk.Label(filter_frm, text="고객사").grid(row=0, column=4, padx=4, pady=4, sticky="w")
        ttk.Entry(filter_frm, textvariable=customer_var, width=18).grid(row=0, column=5, padx=4, pady=4, sticky="w")
        ttk.Label(filter_frm, text="LOT").grid(row=0, column=6, padx=4, pady=4, sticky="w")
        ttk.Entry(filter_frm, textvariable=lot_var, width=16).grid(row=0, column=7, padx=4, pady=4, sticky="w")
        ttk.Label(filter_frm, text="작업자").grid(row=0, column=8, padx=4, pady=4, sticky="w")
        ttk.Entry(filter_frm, textvariable=operator_var, width=14).grid(row=0, column=9, padx=4, pady=4, sticky="w")

        cols = ("created_at", "lot_no", "customer", "operator", "expected_uid", "scanned_uid", "reason")
        tree = ttk.Treeview(frm, columns=cols, show="headings", height=16)
        headings = {
            "created_at": "SWAP_AT",
            "lot_no": "LOT NO",
            "customer": "CUSTOMER",
            "operator": "OPERATOR",
            "expected_uid": "EXPECTED UID",
            "scanned_uid": "SCANNED UID",
            "reason": "REASON",
        }
        widths = {
            "created_at": 140,
            "lot_no": 110,
            "customer": 130,
            "operator": 110,
            "expected_uid": 180,
            "scanned_uid": 180,
            "reason": 220,
        }
        for c in cols:
            tree.heading(c, text=headings[c], anchor='center')
            tree.column(c, width=widths[c], anchor="center")
        ysb = tk.Scrollbar(frm, orient="vertical", command=tree.yview)
        xsb = tk.Scrollbar(frm, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        ysb.pack(fill=tk.Y, side=tk.RIGHT)
        xsb.pack(fill=tk.X)
        ttk.Label(frm, textvariable=summary_var).pack(fill=tk.X, pady=(4, 0))

        state = {"rows": []}

        def _load_rows():
            from gui_app_modular.utils.tree_enhancements import parse_date_range
            d_from, d_to = parse_date_range(
                start_var.get().strip(), end_var.get().strip()
            )
            # v8.1.8: 둘 다 비면 전체 기간. 하나라도 있으면 유효성 검사
            if d_from is None and d_to is None:
                # 전체 기간 — BETWEEN 조건 없이 조회
                rows = self._query_swap_report_rows(
                    start_date='2000-01-01',
                    end_date='2099-12-31',
                    customer=customer_var.get().strip(),
                    lot_no=lot_var.get().strip(),
                    operator=operator_var.get().strip(),
                )
            else:
                # 기간 입력 있으면 유효성 검사 후 조회
                _s = d_from or '2000-01-01'
                _e = d_to   or '2099-12-31'
                rows = self._query_swap_report_rows(
                    start_date=_s,
                    end_date=_e,
                    customer=customer_var.get().strip(),
                    lot_no=lot_var.get().strip(),
                    operator=operator_var.get().strip(),
                )
            state["rows"] = rows
            tree.delete(*tree.get_children(""))
            for r in rows:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        r.get("created_at", ""),
                        r.get("lot_no", ""),
                        r.get("customer", ""),
                        r.get("operator", ""),
                        r.get("expected_uid", ""),
                        r.get("scanned_uid", ""),
                        r.get("reason", ""),
                    ),
                )
            summary_var.set(f"총 {len(rows)}건")

        def _save_csv():
            rows = state.get("rows", [])
            if not rows:
                CustomMessageBox.showwarning(dlg, "저장 불가", "저장할 데이터가 없습니다. 먼저 조회하세요.")
                return
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_dir = self._get_swap_report_save_dir(start_var.get().strip() or today)
            out_name = f"swap_report_{start_var.get().strip()}_{end_var.get().strip()}_{ts}.csv"
            out_path = os.path.join(out_dir, out_name)
            self._export_swap_report_csv(rows, out_path)
            CustomMessageBox.showinfo(dlg, "저장 완료", f"CSV 저장 완료\n{out_path}")

        def _save_xlsx():
            rows = state.get("rows", [])
            if not rows:
                CustomMessageBox.showwarning(dlg, "저장 불가", "저장할 데이터가 없습니다. 먼저 조회하세요.")
                return
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_dir = self._get_swap_report_save_dir(start_var.get().strip() or today)
            out_name = f"swap_report_{start_var.get().strip()}_{end_var.get().strip()}_{ts}.xlsx"
            out_path = os.path.join(out_dir, out_name)
            try:
                self._export_swap_report_xlsx(rows, out_path)
                CustomMessageBox.showinfo(dlg, "저장 완료", f"XLSX 저장 완료\n{out_path}")
            except Exception as e:
                CustomMessageBox.showerror(dlg, "저장 실패", f"XLSX 저장 실패:\n{e}")

        btn_frm = ttk.Frame(frm)
        btn_frm.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(btn_frm, text="조회", command=_load_rows).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frm, text="CSV 저장", command=_save_csv).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frm, text="XLSX 저장", command=_save_xlsx).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frm, text="닫기", command=dlg.destroy).pack(side=tk.RIGHT, padx=2)

        _load_rows()

    def _on_s1_onestop_outbound(self) -> None:
        """S1 원스톱 출고: 4단계 워크플로우 (v6.3.1)
        입력(붙여넣기) → 톤백선택(LOT일괄/랜덤/수동) → 스캔검증(하드스톱) → 확정
        """
        try:
            from ..dialogs.onestop_outbound import S1OneStopOutboundDialog
            dlg = S1OneStopOutboundDialog(self, self.engine)
            dlg.show()
        except (ImportError, AttributeError) as e:
            logger.error(f"S1 원스톱 출고 오류: {e}", exc_info=True)
            from ..utils.ui_constants import CustomMessageBox
            CustomMessageBox.showerror(self.root, "오류", f"S1 원스톱 출고 열기 실패:\n{e}")