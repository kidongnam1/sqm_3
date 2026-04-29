"""
SQM 재고관리 시스템 - 톤백 위치 업로드 유틸리티
===============================================

v4.2.3: 바코드 스캔 Excel → 톤백 위치 자동 업데이트
v5.9.8: 로케이션 4파트 지원 — 약식 A-01-01-10 (구역-열-층-칸)

작성자: Ruby
"""

import logging
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)


class TonbagLocationUploader:
    """톤백 위치 Excel 업로드 처리"""

    def __init__(self, db_engine):
        """
        Args:
            db_engine: SQMDatabase 인스턴스 또는 SQMInventoryEngineV3(엔진 전달 시 .db 사용)
        """
        # 엔진이 전달되면 engine.db, DB가 전달되면 그대로 사용
        self.db = getattr(db_engine, 'db', db_engine)
        if self.db is None:
            raise ValueError("db_engine에 DB가 없습니다 (engine.db가 None)")

    def parse_excel(self, file_path: str) -> Tuple[bool, str, List[Dict]]:
        """
        Excel 파일 파싱
        
        Args:
            file_path: Excel 파일 경로
            
        Returns:
            (성공여부, 메시지, 데이터 리스트)
            
        Expected Excel Format:
        ┌──────────────────┬────────┐
        │ UID              │ 위치   │
        ├──────────────────┼────────┤
        │ 1125072340-01    │ A-1-3  │
        │ 1125072340-02    │ A-1-4  │
        └──────────────────┴────────┘
        
        v5.6.1: lot_no + tonbag_no + location 양식도 지원
        v5.6.9: 로케이션 엑셀 양식 확정
        v5.8.x: 양식2에 uid 컬럼 선택 포함 가능 (있으면 해당 값으로 매칭)
        현장(입고/출고)은 UID를 모름 — LOT NO + 톤백 NO + LOCATION만 전달, 프로그램이 UID 자동 매칭
        ┌────────┬────────┬────────┬──────────┬───────────┬──────────────┬──────────┐
        │ 순번   │ 입고일 │ BL No  │ lot_no   │ tonbag_no│ uid (선택)   │ location │
        ├────────┼────────┼────────┼──────────┼───────────┼──────────────┼──────────┤
        │ (자동) │ 선택   │ 선택   │ 필수     │ 필수     │ 선택         │ 필수     │
        │ 1      │ 2025-01│ B/L123 │ 112508.. │ 3         │ 1125081447-03│ A-01-01  │
        └────────┴────────┴────────┴──────────┴───────────┴──────────────┴──────────┘
        로케이션 약식: 3파트 A-01-01 (구역-열-층) 또는 4파트 A-01-01-10 (구역-열-층-칸)
        """
        try:
            # 매핑 파일 형식: 1행=타이틀, 2행=요약(선택), 3행=헤더, 4행~=데이터
            peek = pd.read_excel(file_path, header=None, nrows=4)
            first_cell = str(peek.iloc[0, 0]) if peek.size else ""
            header_row = 2 if ("SQM" in first_cell and "로케이션" in first_cell) else 0
            df = pd.read_excel(file_path, header=header_row)
            # 빈 행 제거(제목/요약만 있는 파일 시)
            df = df.dropna(how='all').reset_index(drop=True)

            from core.column_registry import normalize_header
            df.columns = [normalize_header(c) for c in df.columns]

            # 3행=헤더로 읽었는데 필수 컬럼이 없으면 2행을 헤더로 재시도 (1행 타이틀만 있는 경우)
            # 톤백번호: column_registry가 TONBAG NO → sub_lt 로 정규화하므로 sub_lt도 후보에 포함
            if header_row == 2:
                has_required = (
                    any(c in df.columns for c in ('lot_no', 'lot', 'lotno')) and
                    any(c in df.columns for c in ('tonbag_no', 'tonbag', 'tb_no', 'sub_lt')) and
                    any(c in df.columns for c in ('location', '위치'))
                )
                if not has_required and len(peek) > 1:
                    df2 = pd.read_excel(file_path, header=1)
                    df2 = df2.dropna(how='all').reset_index(drop=True)
                    df2.columns = [normalize_header(c) for c in df2.columns]
                    if any(c in df2.columns for c in ('lot_no', 'lot', 'lotno')) and any(c in df2.columns for c in ('location', '위치')):
                        df = df2
                        header_row = 1

            # v5.6.1: 양식 자동 감지
            # 양식 1: UID + 위치 (기존)
            # 양식 2: lot_no + tonbag_no + location (신규)
            # TONBAG NO 헤더는 column_registry에서 sub_lt로 정규화되므로 sub_lt도 톤백번호 후보
            has_uid = 'uid' in df.columns or 'tonbag_uid' in df.columns
            has_lot = any(c in df.columns for c in ('lot_no', 'lot', 'lotno'))
            has_tb = any(c in df.columns for c in ('tonbag_no', 'tonbag', 'tb_no', 'sub_lt'))
            has_loc = any(c in df.columns for c in ('location', '위치'))

            if has_lot and has_tb and has_loc:
                # 양식 2: lot_no + tonbag_no + location (입고일, BL No 선택 컬럼 무시)
                ok, msg, data = self._parse_lot_tonbag_format(df)
                if ok and data and header_row > 0:
                    for item in data:
                        item['row_num'] += header_row
                return ok, msg, data
            elif has_uid and has_loc:
                # 양식 1: UID + 위치 (기존)
                ok, msg, data = self._parse_uid_format(df)
                if ok and data and header_row > 0:
                    for item in data:
                        item['row_num'] += header_row
                return ok, msg, data
            else:
                return False, (
                    "❌ 지원하는 양식이 아닙니다.\n\n"
                    "양식1: uid + 위치\n"
                    "양식2: lot_no + tonbag_no + location\n\n"
                    f"발견된 컬럼: {list(df.columns)}"
                ), []

        except (ValueError, TypeError, KeyError, OSError) as e:
            logger.error(f"Excel 파싱 실패: {e}")
            return False, f"❌ Excel 파싱 실패: {e}", []

    def parse_pasted_text(self, text: str) -> Tuple[bool, str, List[Dict]]:
        """
        붙여넣은 텍스트(TSV/CSV) 파싱. 헤더 포함 시 첫 줄을 컬럼명으로 사용.
        양식: lot_no, tonbag_no, uid(선택), location 또는 uid, 위치
        """
        try:
            from core.column_registry import normalize_header
            lines = [ln.strip() for ln in text.strip().splitlines() if ln.strip()]
            if not lines:
                return False, "❌ 붙여넣은 데이터가 비어 있습니다.", []
            # 구분자: 탭 우선, 없으면 쉼표
            first = lines[0]
            delim = "\t" if "\t" in first else ","
            parts_list = [ln.split(delim) for ln in lines]
            max_cols = max(len(p) for p in parts_list)
            # 컬럼 수 맞추기
            for p in parts_list:
                while len(p) < max_cols:
                    p.append("")
            # 첫 줄이 헤더인지 판단 (lot_no, uid, location 등 키워드 포함)
            first_row = [str(x).strip() for x in parts_list[0]]
            normalized_first = [normalize_header(c) for c in first_row]
            has_lot = any(c in normalized_first for c in ("lot_no", "lot", "lotno"))
            has_tb = any(c in normalized_first for c in ("tonbag_no", "tonbag", "tb_no", "sub_lt"))
            has_loc = any(c in normalized_first for c in ("location", "위치"))
            has_uid = any(c in normalized_first for c in ("uid", "tonbag_uid"))
            if has_lot and has_tb and has_loc:
                header_row = normalized_first
                data_start = 1
            elif has_uid and has_loc:
                header_row = normalized_first
                data_start = 1
            else:
                # 헤더 없음: 2열 = uid, location / 3열 = lot_no, tonbag_no, location / 4열 = + uid
                if max_cols >= 4:
                    header_row = ["lot_no", "tonbag_no", "uid", "location"][:max_cols]
                elif max_cols >= 3:
                    header_row = ["lot_no", "tonbag_no", "location"][:max_cols]
                else:
                    header_row = ["uid", "location"]
                data_start = 0
            # DataFrame 생성
            import pandas as pd
            col_count = min(len(header_row), max_cols)
            header_row = header_row[:col_count]
            data_rows = []
            for i in range(data_start, len(parts_list)):
                row = parts_list[i][:col_count]
                data_rows.append(row)
            if not data_rows:
                return False, "❌ 데이터 행이 없습니다.", []
            df = pd.DataFrame(data_rows, columns=header_row)
            df.columns = [normalize_header(c) for c in df.columns]
            has_lot = any(c in df.columns for c in ("lot_no", "lot", "lotno"))
            has_tb = any(c in df.columns for c in ("tonbag_no", "tonbag", "tb_no", "sub_lt"))
            has_loc = any(c in df.columns for c in ("location", "위치"))
            has_uid = any(c in df.columns for c in ("uid", "tonbag_uid"))
            if has_lot and has_tb and has_loc:
                return self._parse_lot_tonbag_format(df)
            if has_uid and has_loc:
                return self._parse_uid_format(df)
            return False, (
                "❌ 지원하는 양식이 아닙니다. 헤더: lot_no, tonbag_no, location 또는 uid, 위치"
            ), []
        except Exception as e:
            logger.error(f"붙여넣기 파싱 실패: {e}")
            return False, f"❌ 파싱 실패: {e}", []

    def _parse_uid_format(self, df) -> Tuple[bool, str, List[Dict]]:
        """양식 1: UID + 위치"""
        if 'uid' not in df.columns:
            if 'tonbag_uid' in df.columns:
                df.rename(columns={'tonbag_uid': 'uid'}, inplace=True)
            else:
                return False, "❌ 'UID' 컬럼을 찾을 수 없습니다", []

        if '위치' not in df.columns:
            if 'location' in df.columns:
                df.rename(columns={'location': '위치'}, inplace=True)
            else:
                return False, "❌ '위치' 컬럼을 찾을 수 없습니다", []

        df = df.dropna(subset=['uid', '위치'])
        if len(df) == 0:
            return False, "❌ 유효한 데이터가 없습니다", []

        data = []
        for idx, row in df.iterrows():
            uid = str(row['uid']).strip()
            location = str(row['위치']).strip()
            if not uid or uid.lower() == 'nan':
                continue
            if not location or location.lower() == 'nan':
                continue
            data.append({
                'uid': uid,
                'location': location,
                'row_num': idx + 2
            })

        if len(data) == 0:
            return False, "❌ 유효한 데이터가 없습니다", []
        return True, f"✅ {len(data)}개 데이터 파싱 완료 (UID 양식)", data

    def _parse_lot_tonbag_format(self, df) -> Tuple[bool, str, List[Dict]]:
        """양식 2: (LOT 번호, 톤백 번호) 한 쌍 + location. 현장 업로드 = 이 쌍으로 재고와 일치시킨 뒤 로케이션 반영."""
        # 컬럼명 정규화 (입고일, BL No, uid 등 선택 컬럼 포함)
        # TONBAG NO → sub_lt 로 정규화되므로 sub_lt도 톤백번호 컬럼으로 인정
        col_map = {}
        for c in df.columns:
            c_lower = c.lower().strip()
            if c_lower in ('lot_no', 'lot', 'lotno'):
                col_map['lot_no'] = c
            elif c_lower in ('tonbag_no', 'tonbag', 'tb_no', 'sub_lt'):
                col_map['tonbag_no'] = c
            elif c_lower in ('location', '위치'):
                col_map['location'] = c
            elif c_lower in ('uid', 'tonbag_uid'):
                col_map['uid'] = c

        lot_col = col_map.get('lot_no')
        tb_col = col_map.get('tonbag_no')
        loc_col = col_map.get('location')
        uid_col = col_map.get('uid')

        if not lot_col or not tb_col or not loc_col:
            return False, "❌ 필수 컬럼 필요: lot_no, tonbag_no, location", []

        df = df.dropna(subset=[lot_col, tb_col, loc_col])
        if len(df) == 0:
            return False, "❌ 유효한 데이터가 없습니다", []

        def _norm_lot(s):
            s = str(s).strip()
            if s.lower() == 'nan' or not s:
                return s
            try:
                if '.' in s and s.endswith('.0'):
                    return str(int(float(s)))
            except (ValueError, TypeError) as e:
                logger.warning(f"[_norm_lot] Suppressed: {e}")
            return s

        data = []
        for idx, row in df.iterrows():
            lot_no = _norm_lot(row[lot_col])
            tonbag_no = str(row[tb_col]).strip()
            location = str(row[loc_col]).strip()
            excel_uid = ''
            if uid_col and uid_col in row.index:
                v = row[uid_col]
                if pd.notna(v) and str(v).strip() and str(v).lower() != 'nan':
                    excel_uid = str(v).strip()
            if excel_uid and excel_uid.lower() == 'nan':
                excel_uid = ''

            if not lot_no or lot_no.lower() == 'nan':
                continue
            if not location or location.lower() == 'nan':
                continue
            # v5.6.9/v5.9.8: 로케이션 형식 검증 (3파트 A-01-01 또는 4파트 A-01-01-10)
            valid, msg = validate_location_format(location)
            if not valid:
                logger.warning(f"행 {idx + 2}: location '{location}' — {msg}")

            # UID: 엑셀에 uid(또는 tonbag_uid) 값이 있으면 사용, 없으면 lot_no+tonbag_no로 조회/생성
            if excel_uid:
                uid = excel_uid
            else:
                try:
                    tb_num = int(float(tonbag_no))
                    tonbag = self.db.fetchone(
                        "SELECT tonbag_uid FROM inventory_tonbag WHERE lot_no = ? AND sub_lt = ?",
                        (lot_no, tb_num))
                    if tonbag:
                        uid = tonbag['tonbag_uid'] if isinstance(tonbag, dict) else tonbag[0]
                    else:
                        uid = f"{lot_no}-{tb_num:02d}"
                        logger.warning(f"톤백 미발견: {lot_no}/sub_lt={tb_num}, UID 추정: {uid}")
                except (ValueError, TypeError):
                    uid = f"{lot_no}-{tonbag_no}"

            data.append({
                'uid': uid,
                'location': location,
                'lot_no': lot_no,
                'tonbag_no': tonbag_no,
                'row_num': idx + 2
            })

        if len(data) == 0:
            return False, "❌ 유효한 데이터가 없습니다", []
        # 3행=헤더인데 첫 행이 헤더 문자열로 들어온 경우 한 행만 제거 (헤더가 데이터로 포함된 경우)
        if len(data) > 0:
            r0 = data[0]
            lot0 = str(r0.get('lot_no', '')).strip().lower()
            tb0 = str(r0.get('tonbag_no', '')).strip().lower()
            if (lot0 in ('lot_no', 'lot', 'lotno', 'lot no') and
                    tb0 in ('tonbag_no', 'tonbag', 'tb_no', 'tonbag no', 'sub_lt')):
                data = data[1:]
        if len(data) == 0:
            return False, "❌ 유효한 데이터가 없습니다 (헤더만 있음)", []
        return True, f"✅ {len(data)}개 데이터 파싱 완료 (LOT+톤백 양식)", data

    def validate_and_match(self, data: List[Dict]) -> Dict:
        """
        톤백 리스트(inventory_tonbag) 전용. 재고 리스트(LOT 리스트) DB는 사용하지 않음.
        매칭 키: (LOT 번호, 톤백 번호) 한 쌍 — 현장 (LOT, 톤백번호)와 톤백 리스트 (lot_no, sub_lt) 일치 후
        해당 톤백 행의 location 컬럼에 업로드한 로케이션 반영.
        
        Returns:
            matched, not_found, total, success_count, fail_count
        """
        result = {
            'matched': [],
            'not_found': [],
            'total': len(data),
            'success_count': 0,
            'fail_count': 0
        }

        def _norm_lot_no(val):
            if val is None:
                return ''
            s = str(val).strip()
            if not s or s.lower() == 'nan':
                return ''
            try:
                f = float(s)
                if f == int(f):
                    return str(int(f))
            except (ValueError, TypeError) as e:
                logger.warning(f"[_norm_lot_no] Suppressed: {e}")
            return s

        for item in data:
            raw_lot = item.get('lot_no')
            lot_no = _norm_lot_no(raw_lot)
            tonbag_no = item.get('tonbag_no', '')
            location = item['location']
            row_num = item['row_num']
            uid = item.get('uid', '')
            tb_str = str(tonbag_no).strip()
            tonbag = None

            # UID가 있으면 LOT/톤백번호보다 먼저 매칭 시도 (현장 업로드 안정성 향상)
            if uid and str(uid).strip().lower() != 'nan':
                tonbag = self.db.fetchone("""
                    SELECT t.id, t.tonbag_uid, t.lot_no, t.sub_lt, COALESCE(t.is_sample,0) AS is_sample,
                           t.location AS current_location, i.product
                    FROM inventory_tonbag t
                    LEFT JOIN inventory i ON t.inventory_id = i.id
                    WHERE t.tonbag_uid = ?
                """, (uid,))

            # UID 매칭 실패 시 기존 (LOT, 톤백번호) 매칭 시도
            sub_lt_val = None
            if tonbag is None and tb_str not in ('', 'nan'):
                tb_upper = tb_str.upper()
                if tb_upper in ('S00', 'S0'):
                    sub_lt_val = 0
                elif tb_upper.startswith('S') and len(tb_upper) > 1 and tb_upper[1:].strip().isdigit():
                    try:
                        sub_lt_val = int(tb_upper[1:].strip())
                    except (ValueError, TypeError):
                        sub_lt_val = 0
                else:
                    try:
                        sub_lt_val = int(float(tb_str))
                    except (ValueError, TypeError):
                        sub_lt_val = None

                if sub_lt_val is not None and lot_no:
                    tonbag = self.db.fetchone("""
                        SELECT t.id, t.tonbag_uid, t.lot_no, t.sub_lt,
                               t.location AS current_location, i.product
                        FROM inventory_tonbag t
                        LEFT JOIN inventory i ON t.inventory_id = i.id
                        WHERE t.lot_no = ? AND t.sub_lt = ?
                    """, (lot_no, sub_lt_val))
                    # DB에 공백/숫자형 차이 있을 수 있어 fallback: trim(lot_no), sub_lt 문자 비교
                    if tonbag is None:
                        tonbag = self.db.fetchone("""
                            SELECT t.id, t.tonbag_uid, t.lot_no, t.sub_lt,
                                   t.location AS current_location, i.product
                            FROM inventory_tonbag t
                            LEFT JOIN inventory i ON t.inventory_id = i.id
                            WHERE trim(cast(t.lot_no as text)) = ? AND (t.sub_lt = ? OR cast(t.sub_lt as text) = ?)
                        """, (lot_no, sub_lt_val, str(sub_lt_val)))

            if tonbag is None:
                if uid and str(uid).strip().lower() != 'nan':
                    reason = f'UID 톤백 리스트에 없음: {uid}'
                elif tb_str in ('', 'nan'):
                    reason = f'톤백 번호 없음 (LOT: {lot_no})'
                elif sub_lt_val is None:
                    reason = f'톤백 번호 형식 오류: {tonbag_no}'
                else:
                    reason = f'(LOT·톤백번호 쌍) 톤백 리스트에 없음: LOT {lot_no} / 톤백 {sub_lt_val}'
                result['not_found'].append({
                    'uid': uid, 'location': location, 'row_num': row_num,
                    'reason': reason
                })
                result['fail_count'] += 1
                continue

            # v6.6.0 ① 하드스톱 검증 ─────────────────────────────────
            tb_status = (tonbag.get('status') or '').strip().upper()
            _BLOCKED = {'PICKED', 'SOLD', 'DEPLETED', 'SHIPPED'}
            db_loc   = (tonbag.get('current_location') or '').strip()
            to_loc_upper = (location or '').strip().upper()

            hard_stop = None
            # §5-②: 이동 금지 상태
            if tb_status in _BLOCKED:
                hard_stop = f"이동 금지 상태: {tb_status}"

            # v9.0 [SAMPLE-MOVE-BLOCK]: 샘플 톤백 엑셀 업로드 이동 차단
            # 단건 update_tonbag_location은 is_sample 체크 있으나
            # 엑셀 업로드 경로(직접 UPDATE)는 체크 없었음 → 여기서 명시 차단
            if not hard_stop:
                _is_smp = int(tonbag.get('is_sample') or 0)
                _sub    = int(tonbag.get('sub_lt') or 0)
                if _is_smp == 1 or _sub == 0:
                    hard_stop = (
                        f"[SAMPLE-MOVE-BLOCK] 샘플 톤백 이동 불가: "
                        f"{tonbag.get('lot_no','?')}-sub_lt={_sub} "
                        f"(is_sample=1 또는 sub_lt=0 — 고정 보관 정책)"
                    )
            # §5-③: 동일 위치
            elif db_loc and db_loc.upper() == to_loc_upper:
                hard_stop = f"현재 위치와 동일: {to_loc_upper}"
            # §5-④: 위치 형식
            elif to_loc_upper:
                ok, msg = validate_location_format(to_loc_upper)
                if not ok:
                    hard_stop = f"위치 형식 오류: {msg}"
            # §5-⑤: Capacity=1
            if hard_stop is None and to_loc_upper:
                try:
                    occ = self.db.fetchone(
                        """SELECT lot_no, sub_lt FROM inventory_tonbag
                           WHERE UPPER(COALESCE(location,'')) = ?
                             AND id != ?
                             AND status NOT IN ('SOLD','DEPLETED','SHIPPED')
                           LIMIT 1""",
                        (to_loc_upper, tonbag['id'])
                    )
                    if occ:
                        hard_stop = (
                            f"Capacity=1 위반: {to_loc_upper}에 "
                            f"LOT {occ.get('lot_no','?')}/톤백{occ.get('sub_lt','?')} 점유"
                        )
                except Exception:
                    logger.debug("[SUPPRESSED] exception in tonbag_location_uploader.py")  # noqa

            if hard_stop:
                result['not_found'].append({
                    'uid': uid, 'location': location, 'row_num': row_num,
                    'reason': f'[HARD-STOP] {hard_stop}'
                })
                result['fail_count'] += 1
                continue
            # ─────────────────────────────────────────────────────────────

            result['matched'].append({
                'uid': tonbag.get('tonbag_uid') or uid,
                'location': location,
                'target_location': location,
                'row_num': row_num,
                'tonbag_id': tonbag['id'],
                'lot_no': tonbag['lot_no'],
                'sub_lt': tonbag['sub_lt'],
                'product': tonbag['product'] or '',
                'db_current_location': db_loc,
                'current_location': db_loc or location,
                'move_1': '',
                'move_2': '',
                'move_3': '',
                'location_changed': db_loc.upper() != to_loc_upper if db_loc else True,
                'tonbag_status': tb_status,
            })
            result['success_count'] += 1

        # 최근 이동 이력(최대 3회) 채우기
        for item in result['matched']:
            moves = self._get_recent_move_locations(item.get('lot_no', ''), item.get('sub_lt', 0))
            item['move_1'] = moves[0] if len(moves) > 0 else ''
            item['move_2'] = moves[1] if len(moves) > 1 else ''
            item['move_3'] = moves[2] if len(moves) > 2 else ''

        return result

    def update_locations(self, matched_data: List[Dict]) -> Tuple[bool, str]:
        """
        매칭된 톤백의 위치 업데이트 (v7.0.1: stock_movement 이력 기록)
        
        Args:
            matched_data: 매칭 성공한 데이터 리스트
            
        Returns:
            (성공여부, 메시지)
        """
        if not matched_data:
            return False, "❌ 업데이트할 데이터가 없습니다"

        try:
            from datetime import datetime
            updated = 0
            relocated = 0
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            with self.db.transaction():
                for item in matched_data:
                    tonbag_id = item['tonbag_id']
                    from_loc = (item.get('db_current_location') or '').strip()
                    display_current = (item.get('current_location') or '').strip()
                    lot_no = item.get('lot_no', '')
                    sub_lt = item.get('sub_lt', 0)
                    move_1 = (item.get('move_1') or '').strip()
                    move_2 = (item.get('move_2') or '').strip()
                    move_3 = (item.get('move_3') or '').strip()

                    # 이동 경로 구성:
                    # - move_1~3 입력이 있으면 순차 이동
                    # - 없으면 target/location/current 기준 단일 반영
                    move_targets = [v for v in (move_1, move_2, move_3) if v and v != '-']
                    if move_targets:
                        route_targets = move_targets
                    else:
                        fallback_target = (
                            (item.get('target_location') or '').strip() or
                            (item.get('location') or '').strip() or
                            display_current
                        )
                        route_targets = [fallback_target] if fallback_target else []

                    # 실제 시작 위치: DB 현재 위치 우선, 없으면 화면 현재 위치 사용
                    route_from = from_loc or display_current

                    # 톤백 리스트(inventory_tonbag)만 갱신 — 재고 리스트(inventory)는 건드리지 않음
                    # location_updated_at, updated_at 함께 갱신 (v4.2.3 마이그레이션 컬럼)
                    final_location = route_from
                    for target in route_targets:
                        if not target:
                            continue
                        if final_location == target:
                            continue
                        # 최초 매핑(DB 기존 위치 없음)에서 첫 배치는 이동 이력으로 보지 않음
                        is_initial_set = (not from_loc) and (final_location == display_current)
                        if not is_initial_set:
                            # 톤백 weight 조회
                            tb_row = self.db.fetchone(
                                "SELECT weight FROM inventory_tonbag WHERE id = ?", (tonbag_id,)
                            )
                            tb_weight = (tb_row.get('weight') or 0) if tb_row else 0
                            self.db.execute("""
                                INSERT INTO stock_movement 
                                (lot_no, movement_type, qty_kg, from_location, to_location, remarks, created_at)
                                VALUES (?, 'RELOCATE', ?, ?, ?, ?, ?)
                            """, (lot_no, tb_weight, final_location, target,
                                  f"sub_lt={sub_lt}, source=EXCEL_UPLOAD", now))
                            # v8.1.7: tonbag_move_log 동시 기록
                            try:
                                _tonbag_no = item.get('tonbag_no') or ''
                                self.db.execute("""
                                    INSERT INTO tonbag_move_log
                                        (lot_no, sub_lt, tonbag_no, from_location, to_location,
                                         move_date, status, operator, source_type, source_file, created_at)
                                    VALUES (?, ?, ?, ?, ?, date(?), 'COMPLETED', 'system',
                                            'EXCEL_UPLOAD', ?, ?)
                                """, (lot_no, sub_lt, _tonbag_no, final_location, target,
                                      now, item.get('source_file', ''), now))
                            except Exception as _mle:
                                logger.debug(f"[TonbagUploader] move_log 기록 스킵: {_mle}")
                            relocated += 1
                        final_location = target

                    self.db.execute("""
                        UPDATE inventory_tonbag
                        SET location = ?,
                            location_updated_at = ?,
                            updated_at = ?
                        WHERE id = ?
                    """, (final_location, now, now, tonbag_id))

                    updated += 1

            msg = f"✅ {updated}개 톤백 위치 업데이트 완료"
            if relocated > 0:
                msg += f" (이동 이력 {relocated}건 기록)"

            # [P4] v6.8.1: 위치 배정 audit_log 기록
            # 최초 배치(LOCATION_INITIAL_ASSIGN) / 위치 변경(LOCATION_CHANGE) 구분
            try:
                import json as _json
                _initial_cnt = sum(
                    1 for item in matched_data
                    if not (item.get('db_current_location') or '').strip()
                )
                _change_cnt = updated - _initial_cnt
                _event = 'LOCATION_INITIAL_ASSIGN' if _initial_cnt > 0 and _change_cnt == 0 \
                         else 'LOCATION_CHANGE' if _change_cnt > 0 \
                         else 'LOCATION_MIXED'
                _payload = _json.dumps({
                    'total': updated,
                    'initial_assign': _initial_cnt,
                    'changed': _change_cnt,
                    'relocated_steps': relocated,
                    'operator': getattr(self.db, '_current_user', 'GUI'),
                }, ensure_ascii=False)
                self.db.execute(
                    "INSERT INTO audit_log(event_type, event_data, created_at) VALUES (?, ?, ?)",
                    (_event, _payload, now)
                )
                logger.info(f"[P4 audit_log] {_event}: {updated}건 ({_initial_cnt}배치+{_change_cnt}변경)")
            except Exception as _ae:
                logger.debug(f"[P4 audit_log] 기록 스킵: {_ae}")

            return True, msg

        except (ValueError, TypeError, KeyError, OSError) as e:
            # with 블록에서 자동 롤백 완료
            logger.error(f"위치 업데이트 실패: {e}")
            return False, f"❌ 업데이트 실패: {e}"

    def _location_header_match(self, cell_value) -> bool:
        """엑셀 셀 값이 로케이션 열 헤더인지 판별."""
        if cell_value is None:
            return False
        s = str(cell_value).strip()
        if not s:
            return False
        if s in ('위치', 'location', 'LOCATION', 'Location', 'LOC'):
            return True
        try:
            from core.column_registry import normalize_header
            return normalize_header(s) == 'location'
        except Exception as e:
            logger.debug(f"location 헤더 판별: {e}")
            return False

    def _fetch_reflected_locations(self, matched_data: List[Dict]) -> Dict[int, Dict]:
        """매칭 행의 tonbag_id 기준 DB 최종 location·식별자."""
        id_to_loc: Dict[int, Dict] = {}
        for item in matched_data:
            tid = item.get('tonbag_id')
            if tid is None:
                continue
            try:
                row = self.db.fetchone(
                    "SELECT location, tonbag_uid, lot_no, sub_lt FROM inventory_tonbag WHERE id = ?",
                    (tid,),
                )
            except Exception as e:
                logger.debug(f"tonbag 조회 skip id={tid}: {e}")
                continue
            if row:
                id_to_loc[int(tid)] = {
                    'location': (row.get('location') or '').strip(),
                    'tonbag_uid': row.get('tonbag_uid') or '',
                    'lot_no': row.get('lot_no') or '',
                    'sub_lt': row.get('sub_lt'),
                }
        return id_to_loc

    def _export_location_new_workbook(
        self,
        matched_data: List[Dict],
        id_to_loc: Dict[int, Dict],
        ts: str,
    ) -> Tuple[bool, str, Optional[str]]:
        """원본 없거나 .xls 등일 때 lot_no / tonbag_no / uid / location 신규 xlsx."""
        try:
            from openpyxl import Workbook
        except ImportError as e:
            return False, f"openpyxl 필요: {e}", None

        wb = Workbook()
        ws = wb.active
        ws.title = "tonbag_location"
        ws.append(['lot_no', 'tonbag_no', 'uid', 'location'])
        for item in matched_data:
            tid = item.get('tonbag_id')
            if tid is None or int(tid) not in id_to_loc:
                continue
            info = id_to_loc[int(tid)]
            sub = info.get('sub_lt')
            ws.append([info['lot_no'], sub, info['tonbag_uid'], info['location']])
        out = Path(tempfile.gettempdir()) / f"SQM_톤백로케이션반영_{ts}.xlsx"
        wb.save(str(out))
        return True, str(out), str(out)

    def export_location_reflected_workbook(
        self,
        source_path: Optional[str],
        matched_data: List[Dict],
    ) -> Tuple[bool, str, Optional[str]]:
        """
        DB 반영 직후 최종 location으로 엑셀 저장.
        - source_path가 .xlsx이면 동일 시트에서 location 열·행(row_num) 기준으로 덮어쓴 복사본 저장.
        - 그 외(붙여넣기·xls)는 임시 폴더에 신규 xlsx.
        """
        if not matched_data:
            return False, "저장할 매칭 행이 없습니다.", None

        from datetime import datetime

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        id_to_loc = self._fetch_reflected_locations(matched_data)
        if not id_to_loc:
            return False, "DB에서 반영 위치를 읽지 못했습니다.", None

        src = (source_path or '').strip()
        if src.lower().endswith('.xlsx') and Path(src).is_file():
            try:
                from openpyxl import load_workbook
            except ImportError:
                return self._export_location_new_workbook(matched_data, id_to_loc, ts)

            wb = None
            try:
                wb = load_workbook(src)
                ws = wb.active
                max_r = min(ws.max_row or 0, 8000)
                max_c = min(ws.max_column or 0, 60)
                if max_r < 1 or max_c < 1:
                    return self._export_location_new_workbook(matched_data, id_to_loc, ts)

                loc_col = None
                for r in range(1, min(12, max_r + 1)):
                    for c in range(1, max_c + 1):
                        v = ws.cell(row=r, column=c).value
                        if self._location_header_match(v):
                            loc_col = c
                            break
                    if loc_col:
                        break

                if not loc_col:
                    return self._export_location_new_workbook(matched_data, id_to_loc, ts)

                for item in matched_data:
                    tid = item.get('tonbag_id')
                    if tid is None or int(tid) not in id_to_loc:
                        continue
                    try:
                        rn = int(item.get('row_num') or 0)
                    except (TypeError, ValueError):
                        continue
                    if rn < 1:
                        continue
                    ws.cell(row=rn, column=loc_col, value=id_to_loc[int(tid)]['location'])

                out = Path(src).parent / f"{Path(src).stem}_로케이션반영_{ts}.xlsx"
                wb.save(str(out))
                return True, str(out), str(out)
            except Exception as e:
                logger.warning(f"원본 xlsx 로케이션 반영 실패, 신규 파일로 저장: {e}")
                return self._export_location_new_workbook(matched_data, id_to_loc, ts)
            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception as e:
                        logger.debug(f"workbook close: {e}")

        return self._export_location_new_workbook(matched_data, id_to_loc, ts)

    def _get_recent_move_locations(self, lot_no: str, sub_lt: int) -> List[str]:
        """해당 톤백의 최근 RELOCATE 도착 위치를 최대 3건 반환."""
        if not lot_no and sub_lt is None:
            return []
        try:
            rows = self.db.fetchall("""
                SELECT to_location
                FROM stock_movement
                WHERE movement_type = 'RELOCATE'
                  AND lot_no = ?
                  AND remarks LIKE ?
                  AND COALESCE(to_location, '') <> ''
                ORDER BY created_at DESC, id DESC
                LIMIT 3
            """, (str(lot_no), f"%sub_lt={sub_lt}%"))
            return [str(r.get('to_location', '')).strip() for r in rows if str(r.get('to_location', '')).strip()]
        except (ValueError, TypeError, KeyError, OSError):
            return []

    def get_location_summary(self) -> Dict:
        """
        위치별 톤백 통계
        
        Returns:
            {
                'A-1-3': {'count': 5, 'total_weight': 125000},
                'A-1-4': {'count': 3, 'total_weight': 75000},
                ...
            }
        """
        try:
            rows = self.db.fetchall("""
                SELECT 
                    location,
                    COUNT(*) AS count,
                    SUM(weight) AS total_weight
                FROM inventory_tonbag
                WHERE location IS NOT NULL
                  AND location != ''
                  AND status = 'AVAILABLE'
                GROUP BY location
                ORDER BY location
            """)

            summary = {}
            for row in rows:
                summary[row['location']] = {
                    'count': row['count'],
                    'total_weight': row['total_weight'] or 0
                }

            return summary

        except (ValueError, TypeError, KeyError, OSError) as e:
            logger.error(f"위치 통계 조회 실패: {e}")
            return {}


# v8.6.4: validate_location_format → engine_modules.constants 로 이동
from engine_modules.constants import validate_location_format  # noqa


# 테스트
if __name__ == '__main__':
    test_cases = [
        ("A-1-3", True),
        ("A-01-01", True),
        ("A-01-01-10", True),   # 4파트 약식
        ("C-02-01-15", True),
        ("B-2-5", True),
        ("A-10-1", True),
        ("AA-1-3", False),
        ("A-B-3", False),
        ("A-1", False),
        ("A-1-2-3-4", False),   # 5파트 차단
        ("A-01-03-AB", False),  # 4번째 문자 차단
        ("", False),
    ]
    for location, expected in test_cases:
        valid, msg = validate_location_format(location)
        status = "✅" if valid == expected else "❌"
        logger.debug(f"{status} '{location}': {valid} - {msg}")
