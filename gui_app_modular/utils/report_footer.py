"""
SQM v4.0.0 — Excel/PDF 보고서 공통 유틸
=========================================

모든 보고서 하단에 "(주) 지와이로지스    년   월   일" 추가
"""

import logging

logger = logging.getLogger(__name__)
from datetime import datetime


def add_gy_logistics_footer(ws, start_row: int = None, max_col: int = None) -> int:
    """
    Excel 워크시트에 GY Logistics 푸터 추가
    
    Args:
        ws: openpyxl Worksheet
        start_row: 푸터 시작 행 (None이면 마지막 데이터 + 2)
        max_col: 최대 열 (None이면 자동 감지)
    
    Returns:
        실제 삽입된 행 번호
    """
    try:
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return 0

    if start_row is None:
        start_row = ws.max_row + 2

    if max_col is None:
        max_col = ws.max_column or 10

    now = datetime.now()
    footer_text = f"(주) 지와이로지스          {now.year}년  {now.month:2d}월  {now.day:2d}일"

    # 오른쪽 하단에 배치
    footer_col = max(max_col - 2, 1)
    cell = ws.cell(row=start_row, column=footer_col, value=footer_text)
    cell.font = Font(name='맑은 고딕', size=10, color='2F5496')
    cell.alignment = Alignment(horizontal='right')

    return start_row


def add_gy_logistics_footer_pdf(elements, styles=None):
    """
    PDF 보고서에 GY Logistics 푸터 추가 (reportlab)
    
    Args:
        elements: reportlab elements 리스트
        styles: reportlab styles (없으면 자동 생성)
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, Spacer
    except ImportError:
        return

    if styles is None:
        styles = getSampleStyleSheet()

    now = datetime.now()
    footer_text = f"(주) 지와이로지스&nbsp;&nbsp;&nbsp;&nbsp;{now.year}년 {now.month:2d}월 {now.day:2d}일"

    elements.append(Spacer(1, 10 * mm))

    footer_style = ParagraphStyle(
        'GYFooter', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#2F5496'),
        alignment=2,  # RIGHT
    )
    try:
        footer_style.fontName = 'MalgunGothic'
    except (ValueError, TypeError, AttributeError) as _e:
        logger.debug(f'Suppressed: {_e}')

    elements.append(Paragraph(footer_text, footer_style))
