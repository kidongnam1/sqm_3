# -*- coding: utf-8 -*-
"""
SQM v7.7.0 — 정합성 검증 리포트 다이얼로그
==========================================
v7.5.0~v7.6.0 신규 검증 항목 시각화:
  검증  8: 샘플 무게 오류 (ERROR)
  검증  9: 샘플 잔류 (WARNING)
  검증 11: 부분 출고 잔류 (WARNING)
  검증 12: allocation mismatch (ERROR/WARNING)

표시 구성:
  ① 상단 요약 카드 (ERROR / WARNING / OK 건수)
  ② LOT별 결과 Treeview (색상 코딩)
  ③ 하단 선택 LOT 상세 패널
  ④ [새로고침] [Excel 저장] [닫기]
"""

from gui_app_modular.utils.ui_constants import create_themed_toplevel  # v8.0.9
from gui_app_modular.utils.ui_constants import tc
import logging
import tkinter as tk
from tkinter import ttk, END

logger = logging.getLogger(__name__)

# ─── 색상 상수 ────────────────────────────────────────────────────────────────
_CLR_ERROR   = '#E74C3C'   # 빨강
_CLR_WARN    = '#F39C12'   # 주황
_CLR_OK      = '#27AE60'   # 초록
_CLR_INFO    = '#2980B9'   # 파랑
_CLR_BG      = tc('bg_primary')   # v9.1: tc() 자동 테마
_CLR_CARD    = '#253545'   # 카드 배경
_CLR_HDR     = '#17202A'   # 헤더 배경
_CLR_FG      = tc('text_primary')   # v9.1: tc() 자동 테마
_CLR_FG_DIM  = '#95A5A6'   # 흐린 텍스트
_CLR_SEL     = '#2471A3'   # 선택 배경


class IntegrityV760Dialog:
    """v7.7.0 정합성 검증 리포트 다이얼로그"""

    # Treeview 컬럼 정의
    _COLS = (
        ('lot_no',      'LOT NO',         160),
        ('sample_ok',   '샘플 상태',       80),
        ('partial_out', '부분 출고',       80),
        ('alloc_stat',  'Allocation',     120),
        ('errors',      '오류',           240),
        ('warnings',    '경고',           240),
    )

    def __init__(self, parent: tk.Widget, engine):
        self.parent = parent
        self.engine = engine
        self._rows: list[dict] = []          # 검증 결과 캐시
        self._iid_lot: dict[str, str] = {}   # iid → lot_no

        self._build_ui()
        self._run_check()

    # ─── UI 구성 ──────────────────────────────────────────────────────────────
    def _build_ui(self) -> None:
        self.win = create_themed_toplevel(self.parent)
        self.win.title("📋 정합성 검증 리포트 v7.7.0")
        try:
            from gui_app_modular.utils.ui_constants import setup_dialog_geometry_persistence as _sgp
            _sgp(self.win, "integrity_v760_dialog", None, "large")
        except Exception as e:
            logger.warning(f'[UI] integrity_v760_dialog: {e}')
        self.win.geometry("1060x660")
        self.win.resizable(True, True)
        self.win.configure(bg=_CLR_BG)
        self.win.grab_set()

        # ── 헤더 바
        hdr = tk.Frame(self.win, bg=_CLR_HDR, height=46)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        tk.Label(
            hdr, text="📋  SQM 정합성 검증 리포트  (v7.5.0 ~ v7.6.0 강화 검증 포함)",
            bg=_CLR_HDR, fg=_CLR_FG, font=('맑은 고딕', 12, 'bold')
        ).pack(side='left', padx=16, pady=10)

        self._lbl_ts = tk.Label(hdr, text="", bg=_CLR_HDR, fg=_CLR_FG_DIM,
                                 font=('맑은 고딕', 10))
        self._lbl_ts.pack(side='right', padx=14)

        # ── 요약 카드 행
        card_row = tk.Frame(self.win, bg=_CLR_BG)
        card_row.pack(fill='x', padx=12, pady=(10, 4))

        self._card_total   = self._make_card(card_row, "전체 LOT",    "—", _CLR_INFO)
        self._card_error   = self._make_card(card_row, "🔴 오류",     "—", _CLR_ERROR)
        self._card_warn    = self._make_card(card_row, "🟡 경고",     "—", _CLR_WARN)
        self._card_ok      = self._make_card(card_row, "✅ 정상",     "—", _CLR_OK)
        self._card_partial = self._make_card(card_row, "⚠️ 부분 출고", "—", _CLR_WARN)
        self._card_alloc   = self._make_card(card_row, "📊 Alloc 이상","—", _CLR_WARN)

        # ── Treeview 영역
        tree_frame = tk.Frame(self.win, bg=_CLR_BG)
        tree_frame.pack(fill='both', expand=True, padx=12, pady=(0, 4))

        style = ttk.Style()
        style.configure('IntV760.Treeview',
                        background=_CLR_CARD, foreground=_CLR_FG,
                        fieldbackground=_CLR_CARD, rowheight=32,
                        font=('맑은 고딕', 10))
        style.configure('IntV760.Treeview.Heading',
                        background=_CLR_HDR, foreground=_CLR_FG,
                        font=('맑은 고딕', 10, 'bold'),
                        anchor='center')
        style.map('IntV760.Treeview',
                  background=[('selected', _CLR_SEL)])

        self.tree = ttk.Treeview(
            tree_frame,
            columns=[c[0] for c in self._COLS],
            show='headings',
            style='IntV760.Treeview',
            selectmode='browse',
        )
        for cid, label, width in self._COLS:
            self.tree.heading(cid, text=label, anchor='center')
            anchor = 'center' if cid in ('sample_ok', 'partial_out') else 'w'
            self.tree.column(cid, width=width, anchor=anchor, stretch=(cid == 'errors'))

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        # v8.1.9: TreeviewTotalFooter
        try:
            from gui_app_modular.utils.tree_enhancements import TreeviewTotalFooter as _TTF
            _f = _TTF(tree_frame, self.tree, summable_column_ids=[])
            _f.frame.grid(row=2, column=0, columnspan=2, sticky='ew')
        except Exception as e:
            logger.warning(f'[UI] integrity_v760_dialog: {e}')
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.tree.tag_configure('error',   foreground=_CLR_ERROR)
        self.tree.tag_configure('warning', foreground=_CLR_WARN)
        self.tree.tag_configure('ok',      foreground=_CLR_OK)

        self.tree.bind('<<TreeviewSelect>>', self._on_select)

        # ── 상세 패널
        detail_frame = tk.LabelFrame(
            self.win, text="  선택 LOT 상세  ",
            bg=_CLR_CARD, fg=_CLR_FG_DIM,
            font=('맑은 고딕', 10), bd=1, relief='flat'
        )
        detail_frame.pack(fill='x', padx=12, pady=(0, 4))

        self._txt_detail = tk.Text(
            detail_frame, height=5, bg=_CLR_CARD, fg=_CLR_FG,
            font=('Consolas', 10), relief='flat', state='disabled',
            wrap='word'
        )
        self._txt_detail.pack(fill='x', padx=8, pady=6)
        self._txt_detail.tag_configure('err_tag',  foreground=_CLR_ERROR)
        self._txt_detail.tag_configure('warn_tag', foreground=_CLR_WARN)
        self._txt_detail.tag_configure('ok_tag',   foreground=_CLR_OK)
        self._txt_detail.tag_configure('hdr_tag',  foreground=_CLR_INFO,
                                        font=('Consolas', 10, 'bold'))

        # ── 하단 버튼 바
        btn_bar = tk.Frame(self.win, bg=_CLR_HDR, height=44)
        btn_bar.pack(fill='x', side='bottom')
        btn_bar.pack_propagate(False)

        self._lbl_status = tk.Label(btn_bar, text="", bg=_CLR_HDR,
                                     fg=_CLR_FG_DIM, font=('맑은 고딕', 10))
        self._lbl_status.pack(side='left', padx=14, pady=10)

        for txt, cmd, fg in [
            ("닫기",      self.win.destroy, _CLR_FG_DIM),
            ("Excel 저장", self._on_export,  _CLR_INFO),
            ("새로고침",   self._run_check,   _CLR_OK),
        ]:
            tk.Button(
                btn_bar, text=txt, command=cmd,
                bg=_CLR_CARD, fg=fg, relief='flat',
                font=('맑은 고딕', 10, 'bold'),
                padx=14, pady=6, cursor='hand2',
                activebackground=_CLR_SEL, activeforeground=_CLR_FG,
                bd=0
            ).pack(side='right', padx=6, pady=8)

    def _make_card(self, parent, title: str, value: str,
                   accent: str) -> tk.Label:
        """요약 카드 위젯 생성, 값 레이블 반환"""
        f = tk.Frame(parent, bg=_CLR_CARD, bd=0, relief='flat')
        f.pack(side='left', padx=5, pady=4, ipadx=12, ipady=8)
        tk.Label(f, text=title, bg=_CLR_CARD, fg=_CLR_FG_DIM,
                 font=('맑은 고딕', 8)).pack()
        lbl = tk.Label(f, text=value, bg=_CLR_CARD, fg=accent,
                       font=('맑은 고딕', 18, 'bold'))
        lbl.pack()
        return lbl

    # ─── 검증 실행 ────────────────────────────────────────────────────────────
    def _run_check(self) -> None:
        from datetime import datetime
        self._lbl_status.config(text="⏳ 검증 중…")
        self.win.update_idletasks()

        self.tree.delete(*self.tree.get_children())
        self._iid_lot.clear()
        self._rows.clear()

        try:
            result = self.engine.verify_all_integrity()
        except Exception as e:
            logger.error(f"[IntegrityV760] verify_all_integrity 오류: {e}", exc_info=True)
            self._lbl_status.config(text=f"❌ 검증 오류: {e}")
            return

        total = result.get('total_lots', 0)
        error_lots  = {r['lot_no']: r['errors']   for r in result.get('error_lots', [])}
        warning_lots = {r['lot_no']: r['warnings'] for r in result.get('warning_lots', [])}

        # 전체 LOT 목록 수집 (errors / warnings 합집합)
        all_lots: list[str] = []
        try:
            rows = self.engine.db.fetchall("SELECT lot_no FROM inventory ORDER BY lot_no")
            all_lots = [r['lot_no'] if isinstance(r, dict) else r[0] for r in rows]
        except Exception as _e:
            logger.debug(f"LOT 목록 조회 오류: {_e}")
            all_lots = list(set(list(error_lots.keys()) + list(warning_lots.keys())))

        n_error = n_warn = n_ok = n_partial = n_alloc = 0

        for lot_no in all_lots:
            errs  = error_lots.get(lot_no, [])
            warns = warning_lots.get(lot_no, [])

            # 검증 11 — 부분 출고 잔류
            partial = any('부분 출고 잔류' in m for m in warns)
            # 검증 12 — allocation mismatch
            alloc_err  = any('allocation 초과' in m for m in errs)
            alloc_warn = any('allocation 미완결' in m for m in warns)
            alloc_stat = ('❌ 초과' if alloc_err
                          else ('⚠️ 미완결' if alloc_warn else '✅'))
            # 검증 8 — 샘플 무게 오류
            sample_err = any('샘플 무게 오류' in m for m in errs)
            # 검증 9 — 샘플 잔류 경고
            sample_warn = any('샘플 잔류 경고' in m for m in warns)
            sample_ok = ('❌ 무게오류' if sample_err
                         else ('⚠️ 잔류' if sample_warn else '✅'))

            has_error = bool(errs)
            has_warn  = bool(warns)

            if has_error:
                tag = 'error'; n_error += 1
            elif has_warn:
                tag = 'warning'; n_warn += 1
            else:
                tag = 'ok'; n_ok += 1

            if partial: n_partial += 1
            if alloc_err or alloc_warn: n_alloc += 1

            err_txt  = ' | '.join(errs[:3])  + ('…' if len(errs) > 3 else '')
            warn_txt = ' | '.join(warns[:3]) + ('…' if len(warns) > 3 else '')

            row_data = {
                'lot_no': lot_no, 'errs': errs, 'warns': warns,
                'sample_ok': sample_ok, 'partial': partial,
                'alloc_stat': alloc_stat, 'tag': tag,
            }
            self._rows.append(row_data)

            iid = self.tree.insert('', END, tags=(tag,), values=(
                lot_no, sample_ok,
                '⚠️ 잔류' if partial else '✅',
                alloc_stat,
                err_txt, warn_txt,
            ))
            self._iid_lot[iid] = lot_no

        # 카드 업데이트
        self._card_total.config(text=str(total or len(all_lots)))
        self._card_error.config(text=str(n_error))
        self._card_warn.config(text=str(n_warn))
        self._card_ok.config(text=str(n_ok))
        self._card_partial.config(text=str(n_partial))
        self._card_alloc.config(text=str(n_alloc))

        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self._lbl_ts.config(text=f"마지막 검증: {ts}")
        self._lbl_status.config(
            text=f"전체 {total}개 LOT — 🔴 오류 {n_error}건 | 🟡 경고 {n_warn}건 | ✅ 정상 {n_ok}건"
        )

    # ─── 선택 상세 ────────────────────────────────────────────────────────────
    def _on_select(self, _event=None) -> None:
        sel = self.tree.selection()
        if not sel:
            return
        lot_no = self._iid_lot.get(sel[0], '')
        row = next((r for r in self._rows if r['lot_no'] == lot_no), None)
        if not row:
            return

        self._txt_detail.config(state='normal')
        self._txt_detail.delete('1.0', END)
        self._txt_detail.insert(END, f"LOT: {lot_no}\n", 'hdr_tag')

        if row['errs']:
            self._txt_detail.insert(END, "🔴 오류\n", 'err_tag')
            for e in row['errs']:
                self._txt_detail.insert(END, f"   • {e}\n", 'err_tag')
        if row['warns']:
            self._txt_detail.insert(END, "🟡 경고\n", 'warn_tag')
            for w in row['warns']:
                self._txt_detail.insert(END, f"   • {w}\n", 'warn_tag')
        if not row['errs'] and not row['warns']:
            self._txt_detail.insert(END, "✅ 모든 검증 통과\n", 'ok_tag')

        self._txt_detail.config(state='disabled')

    # ─── Excel 내보내기 ───────────────────────────────────────────────────────
    def _on_export(self) -> None:
        if not self._rows:
            return
        from datetime import datetime
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            parent=self.win,
            title="정합성 리포트 저장",
            defaultextension=".xlsx",
            initialfile=f"SQM-Integrity-{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "정합성 검증"

            headers = ['LOT NO', '샘플 상태', '부분 출고', 'Allocation', '오류', '경고']
            ws.append(headers)
            for cell in ws[1]:
                cell.font      = Font(bold=True, color='FFFFFF')
                cell.fill      = PatternFill('solid', fgColor='17202A')
                cell.alignment = Alignment(horizontal='center')

            fill_err  = PatternFill('solid', fgColor='FADBD8')
            fill_warn = PatternFill('solid', fgColor='FDEBD0')
            fill_ok   = PatternFill('solid', fgColor='D5F5E3')

            for r in self._rows:
                partial_str = '⚠️ 잔류' if r['partial'] else '✅'
                row = [r['lot_no'], r['sample_ok'], partial_str,
                       r['alloc_stat'],
                       ' | '.join(r['errs']), ' | '.join(r['warns'])]
                ws.append(row)
                fill = (fill_err if r['tag'] == 'error'
                        else fill_warn if r['tag'] == 'warning'
                        else fill_ok)
                for cell in ws[ws.max_row]:
                    cell.fill = fill

            ws.column_dimensions['A'].width = 24
            ws.column_dimensions['E'].width = 60
            ws.column_dimensions['F'].width = 60

            wb.save(path)
            from ..utils.custom_messagebox import CustomMessageBox
            CustomMessageBox.showinfo(self.win, "저장 완료",
                                      f"정합성 리포트 저장:\n{path}")
        except Exception as e:
            logger.error(f"[IntegrityV760] Excel 저장 오류: {e}", exc_info=True)
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showerror(self.win, "오류", f"저장 실패:\n{e}")
            except Exception:
                logger.debug("[SUPPRESSED] exception in integrity_v760_dialog.py")  # noqa
