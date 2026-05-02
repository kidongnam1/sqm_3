"""
SQM v5.9.4 — 원스톱 입고: DB 업로드 + Excel 내보내기 Mixin
==========================================================

onestop_inbound.py에서 분리 (1869줄 → ~1300 + ~500).
DB 저장 로직(_save_to_db), 업로드 스레드(_upload_thread),
중복 체크(_on_upload), Excel 내보내기(_export_to_excel)를 담당.
"""
import logging
import sqlite3
import threading
from datetime import datetime
from tkinter import filedialog

from core.constants import DEFAULT_WAREHOUSE
from core.types import safe_float

logger = logging.getLogger(__name__)


class InboundUploadMixin:
    """DB 업로드 + Excel 내보내기 로직 (OneStopInboundDialog에 MRO 합성)"""

    def _preflight_validate_preview_data(self):
        """DB 반영 전 미리보기 데이터 검증 (오류 리스트 반환)."""
        rows = getattr(self, 'preview_data', []) or []
        errors = []
        seen_lots = {}
        for idx, row in enumerate(rows, 1):
            lot_no = str(row.get('lot_no', '') or '').strip()
            product = str(row.get('product', '') or '').strip()
            if not lot_no:
                errors.append(f"{idx}행: LOT NO 필수")
            if not product:
                errors.append(f"{idx}행: PRODUCT 필수")
            try:
                nw = safe_float(row.get('net_weight', 0))
                if nw <= 0:
                    errors.append(f"{idx}행: NET(Kg) 0 초과 필요")
            except Exception:
                errors.append(f"{idx}행: NET(Kg) 숫자 형식 오류")
            try:
                mx = int(float(str(row.get('mxbg_pallet', '0')).replace(',', '') or 0))
                if mx <= 0:
                    errors.append(f"{idx}행: MXBG 1 이상 필요")
            except Exception:
                errors.append(f"{idx}행: MXBG 숫자 형식 오류")
            arr = str(row.get('arrival_date', '') or '').strip()
            cr = str(row.get('con_return', '') or '').strip()
            if arr and (len(arr) != 10 or arr.count('-') != 2):
                errors.append(f"{idx}행: ARRIVAL 날짜 형식 오류(YYYY-MM-DD)")
            if cr and (len(cr) != 10 or cr.count('-') != 2):
                errors.append(f"{idx}행: CON RETURN 날짜 형식 오류(YYYY-MM-DD)")
            if arr and cr:
                try:
                    arr_d = datetime.strptime(arr[:10], '%Y-%m-%d').date()
                    cr_d = datetime.strptime(cr[:10], '%Y-%m-%d').date()
                    if cr_d < arr_d:
                        errors.append(f"{idx}행: CON RETURN은 ARRIVAL 이상이어야 함")
                except ValueError as e:
                    logger.warning(f"[_preflight_validate_preview_data] Suppressed: {e}")
            if lot_no:
                if lot_no in seen_lots:
                    errors.append(f"{idx}행: LOT 중복({lot_no}) - {seen_lots[lot_no]}행과 중복")
                else:
                    seen_lots[lot_no] = idx
        return errors

    def _on_upload(self) -> None:
        """DB 업로드 (v3.8.8: 중복 LOT 사전 경고 + 위젯 안전 처리)"""
        if not self.preview_data:
            return
        from .onestop_inbound import DOC_TYPES
        if not self._has_required_docs():
            missing = [name for (dt, name, req) in DOC_TYPES if req and dt not in self.file_paths]
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showwarning(
                    self.dialog, "필수 서류 누락",
                    "DB 업로드를 하려면 다음 3종 서류가 모두 필요합니다:\n\n"
                    "  • ① Packing List (포장명세서)\n"
                    "  • ② Invoice, FA (송장)\n"
                    "  • ③ Bill of Loading (선하증권)\n\n"
                    f"누락: {', '.join(missing)}\n\n"
                    "Delivery Order(인도지시서)는 선택사항이며, 나중에 [📋 D/O 후속 연결] 메뉴로 보충할 수 있습니다."
                )
            except (ImportError, ModuleNotFoundError):
                from ..utils.ui_constants import CustomMessageBox
                CustomMessageBox.showwarning(
                    self.dialog, "필수 서류 누락",
                    "Packing List, Invoice/FA, Bill of Loading 3종 모두 필요합니다."
                )
            return

        preflight_errors = self._preflight_validate_preview_data()

        # v6.2.1: 크로스 체크 CRITICAL 항목 차단(사용자 최종 확인 포함)
        xc = getattr(self, '_cross_check_result', None)
        if xc and xc.has_critical:
            critical_msgs = [str(i) for i in xc.items if i.level.value >= 3]
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                proceed = CustomMessageBox.askyesno(
                    self.dialog, "🚫 크로스 체크 심각 불일치",
                    f"4종 서류 교차 검증에서 심각한 불일치 {xc.critical_count}건이 발견되었습니다.\n\n"
                    + "\n".join(critical_msgs[:8])
                    + ("\n... 외 추가 항목" if len(critical_msgs) > 8 else "")
                    + "\n\n⚠️ 서류를 재확인하시기 바랍니다.\n그래도 업로드를 진행하시겠습니까?"
                )
                if not proceed:
                    return
            except (ImportError, ModuleNotFoundError):
                from tkinter import messagebox as msgbox
                proceed = msgbox.askyesno(
                    "크로스 체크 경고",
                    f"심각한 불일치 {xc.critical_count}건. 진행하시겠습니까?"
                )
                if not proceed:
                    return
        if preflight_errors:
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showerror(
                    self.dialog, "입력 검증 실패",
                    "미리보기 데이터 검증에서 오류가 발견되었습니다.\n\n"
                    + "\n".join(preflight_errors[:12])
                    + (f"\n... 외 {len(preflight_errors) - 12}건" if len(preflight_errors) > 12 else "")
                )
            except (ImportError, ModuleNotFoundError):
                from tkinter import messagebox as msgbox
                msgbox.showerror("입력 검증 실패", "\n".join(preflight_errors[:12]))
            return

        dup_lots = []
        if hasattr(self.engine, '_check_lot_exists') or hasattr(self.engine, 'db'):
            try:
                db = getattr(self.engine, 'db', None)
                if db:
                    for row in self.preview_data:
                        lot_no = row.get('lot_no', '')
                        if lot_no:
                            existing = db.fetchone(
                                "SELECT 1 FROM inventory WHERE lot_no = ?", (lot_no,))
                            if existing:
                                dup_lots.append(lot_no)
            except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
                logger.debug(f"중복 체크 오류: {e}")

        if dup_lots:
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                dup_msg = ', '.join(dup_lots[:5])
                if len(dup_lots) > 5:
                    dup_msg += f" 외 {len(dup_lots) - 5}건"
                ok = CustomMessageBox.askyesno(
                    self.dialog, "⚠️ 중복 LOT 경고",
                    f"다음 {len(dup_lots)}개 LOT가 이미 DB에 존재합니다:\n\n"
                    f"{dup_msg}\n\n"
                    f"중복 LOT는 건너뛰고 나머지만 입고합니다.\n계속하시겠습니까?"
                )
            except (ImportError, ModuleNotFoundError):
                from tkinter import messagebox as msgbox
                ok = msgbox.askyesno("⚠️ 중복 LOT 경고",
                    f"{len(dup_lots)}개 LOT 중복! 건너뛰고 계속?")
            if not ok:
                return

        try:
            from ..utils.custom_messagebox import CustomMessageBox
            edited_cnt = len(getattr(self, '_edited_rows', set()) or set())
            edited_suffix = f"\n수정된 행: {edited_cnt}건" if edited_cnt else ""
            ok = CustomMessageBox.askyesno(
                self.dialog, "DB 업로드 확인",
                f"{len(self.preview_data)}개 LOT를 데이터베이스에 저장합니다.\n\n"
                f"이 작업은 되돌릴 수 없습니다.{edited_suffix}\n계속하시겠습니까?"
            )
        except (ImportError, ModuleNotFoundError):
            from tkinter import messagebox as msgbox
            ok = msgbox.askyesno("DB 업로드 확인",
                f"{len(self.preview_data)}개 LOT 저장?")

        if not ok:
            return

        try:
            if self.btn_upload and self.btn_upload.winfo_exists():
                self.btn_upload.config(state='disabled')
            if self.btn_excel and self.btn_excel.winfo_exists():
                self.btn_excel.config(state='disabled')
        except (RuntimeError, ValueError) as _e:
            logger.debug(f'Suppressed: {_e}')

        self._show_progress_inline()
        thread = threading.Thread(target=self._upload_thread, daemon=True)
        thread.start()

    def _upload_thread(self) -> None:
        """백그라운드 DB 업로드"""
        try:
            self._update_progress(0, "📤 DB 업로드 시작...")

            pl = self.parsed_results.get('packing_list')
            invoice = self.parsed_results.get('invoice')
            bl = self.parsed_results.get('bl')
            do = self.parsed_results.get('do')

            if not pl or not getattr(pl, 'lots', None):
                self._update_progress(0, "❌ Packing List 없음")
                self._enable_buttons()
                return
            if not invoice:
                self._update_progress(0, "❌ FA(송장) 필수 — 3종(PL+FA+BL) 모두 필요")
                self._enable_buttons()
                return
            if not bl:
                self._update_progress(0, "❌ B/L(선하증권) 필수 — 3종(PL+FA+BL) 모두 필요")
                self._enable_buttons()
                return

            success, failed_rows = self._save_to_db(pl, invoice, bl, do)

            if success:
                total = len(self.preview_data)
                self._update_progress(100, f"✅ 업로드 완료: {total} LOT")
                self._log_safe(f"✅ DB 업로드 완료: {total} LOT")
                self.upload_success = True
                self._show_success_and_close(total)
            else:
                self._update_progress(0, "❌ 업로드 실패")
                try:
                    from ..utils.upload_error_dialog import show_upload_error_dialog
                    from ..utils.upload_error_template import UploadErrorTemplate
                    rows_for_msg = failed_rows if failed_rows else [{'row': '?', 'value': '업로드 실패', 'column': ''}]
                    err_type = (rows_for_msg[0].get('type', 'missing_required') if rows_for_msg else 'missing_required')
                    error_msg = UploadErrorTemplate.format_multiple_errors(
                        errors=[{'type': err_type, 'rows': rows_for_msg}],
                        total_rows=len(self.preview_data)
                    )
                    show_upload_error_dialog(self.dialog, "입고 업로드 실패", error_msg)
                except (ImportError, Exception):
                    from ..utils.ui_constants import CustomMessageBox
                    CustomMessageBox.showerror(
                        self.dialog, "업로드 실패",
                        "입고 처리 중 오류가 발생했습니다.\n로그를 확인하세요."
                    )
                self._enable_buttons()

        except (ValueError, TypeError, AttributeError) as e:
            self._update_progress(0, f"❌ 오류: {e}")
            self._log_safe(f"❌ 업로드 오류: {e}")
            logger.error(f"업로드 오류: {e}", exc_info=True)
            try:
                from ..utils.upload_error_dialog import show_upload_error_dialog
                from ..utils.upload_error_template import UploadErrorTemplate
                error_msg = UploadErrorTemplate.format_multiple_errors(
                    errors=[{'type': 'file_format', 'rows': [{'row': '?', 'value': str(e), 'column': ''}]}],
                    total_rows=len(self.preview_data) if hasattr(self, 'preview_data') else 0
                )
                show_upload_error_dialog(self.dialog, "입고 처리 오류", error_msg)
            except (ImportError, Exception):
                from ..utils.ui_constants import CustomMessageBox
                CustomMessageBox.showerror(self.dialog, "오류", f"입고 처리 오류:\n{e}")
            self._enable_buttons()

    def _save_to_db(self, pl, invoice, bl, do):
        """engine.process_inbound를 LOT별로 호출하여 DB 저장

        Returns:
            (success: bool, failed_rows: list)
        """
        try:
            if not hasattr(self.engine, 'process_inbound'):
                self._log_safe("❌ engine.process_inbound 메서드 없음")
                return False, []

            if hasattr(self, '_get_upload_rows_for_db'):
                _rows = list(self._get_upload_rows_for_db() or [])
            else:
                _rows = list(getattr(self, 'preview_data', []) or [])
            total = len(_rows)
            if total == 0:
                return False, []

            created_lots = []
            skipped_lots = []
            errors = []
            failed_rows = []
            _last_idx = -1  # DB 예외 시 행 번호 표시용

            for idx, row in enumerate(_rows):
                _last_idx = idx
                pct = 10 + int(80 * (idx + 1) / max(1, total))
                lot_no = str(row.get('lot_no', '') or '')
                self._update_progress(pct, f"📦 LOT {idx+1}/{total}: {lot_no}")

                if lot_no:
                    try:
                        existing = self.engine.db.fetchone(
                            "SELECT 1 FROM inventory WHERE lot_no = ?", (lot_no,))
                        if existing:
                            self._log_safe(f"  ⏭ LOT {lot_no}: 이미 존재 (건너뜀)")
                            skipped_lots.append(lot_no)
                            continue
                    except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as _e:
                        logger.debug(f'Suppressed: {_e}')

                _tonbag = row.get('mxbg_pallet', row.get('tonbag_count', 10))
                try:
                    _tonbag = int(float(str(_tonbag).replace(',', '') or 0))
                except (TypeError, ValueError):
                    _tonbag = 10

                try:
                    from utils.date_utils import normalize_date_str as _norm_date_str
                except Exception:
                    _norm_date_str = None
                _arrival_raw = str(row.get('arrival_date', '') or '').strip()
                _con_raw = str(row.get('con_return', '') or '').strip()
                _arrival = (_norm_date_str(_arrival_raw) if _norm_date_str else None) or _arrival_raw[:10]
                _con_return = (_norm_date_str(_con_raw) if _norm_date_str else None) or _con_raw[:10]
                _free_time = 0
                _ft_raw = str(row.get('free_time', '') or '').strip()
                if _ft_raw:
                    try:
                        _free_time = int(float(_ft_raw.replace(',', '')))
                    except Exception:
                        _free_time = 0
                if not _con_return and do:
                    ft_infos = getattr(do, 'free_time_info', []) or []
                    for ft in ft_infos:
                        ftd = getattr(ft, 'free_time_date', '') or (ft.get('free_time_date', '') if isinstance(ft, dict) else '')
                        if ftd:
                            _con_return = (_norm_date_str(ftd) if _norm_date_str else None) or str(ftd)[:10]
                            break
                if _con_return and _arrival and not _ft_raw:
                    try:
                        from utils.date_utils import calculate_free_days as _calc_free_days
                        _days = _calc_free_days(_arrival, _con_return)
                        _free_time = int(_days) if _days is not None else 0
                    except Exception:
                        _free_time = 0

                packing_dict = {
                    'lot_no': lot_no,
                    'lot_sqm': str(row.get('lot_sqm', '') or ''),
                    'sap_no': str(row.get('sap_no', '') or getattr(pl, 'sap_no', '') or (getattr(invoice, 'sap_no', '') if invoice else '') or ''),
                    'bl_no': self._format_bl(
                        str(row.get('bl_no', '') or '') or
                        (getattr(bl, 'bl_no', '') if bl else '') or
                        (getattr(do, 'bl_no', '') if do else '') or ''
                    ),
                    'container_no': str(row.get('container_no', '') or ''),
                    'product': str(row.get('product', '') or getattr(pl, 'product', '') or 'LITHIUM CARBONATE'),
                    'product_code': str(row.get('product_code', '') or getattr(pl, 'code', '') or ''),
                    'net_weight': safe_float(row.get('net_weight', 0) or 0),
                    'gross_weight': safe_float(row.get('gross_weight', 0) or 0),
                    'mxbg_pallet': _tonbag,
                    'tonbag_count': _tonbag,
                    'salar_invoice_no': str(row.get('salar_invoice_no', '') or (getattr(invoice, 'salar_invoice_no', '') if invoice else '') or ''),
                    'ship_date': str(row.get('ship_date', '') or self._date_str(getattr(bl, 'ship_date', None) if bl else None) or self._date_str(getattr(invoice, 'invoice_date', None) if invoice else None) or ''),
                    'arrival_date': _arrival,
                    'free_time': _free_time,
                    'free_time_date': _con_return,
                    'con_return': _con_return,
                    # v8.7.0 [FIX]: DOData 실제 필드는 warehouse_name — 기존 'warehouse' getattr는 항상 ''
                    'warehouse': str(
                        row.get('warehouse', '')
                        or (getattr(do, 'warehouse_name', '') if do else '')
                        or (getattr(do, 'warehouse_code', '') if do else '')
                        or DEFAULT_WAREHOUSE
                    ),
                    'vessel': getattr(pl, 'vessel', '') or '',
                    # v8.7.0 [U-1 b]: PL 원본 lots[] 참조 보존 — document_pl.lots_json 저장용
                    #   per-LOT packing_dict에는 개별 LOT만 담기므로, 동일 PL의 전체 lots[]를
                    #   밑줄 prefix 키로 부착(인벤토리 컬럼으로는 해석되지 않음)
                    '_pl_lots_raw': getattr(pl, 'lots', []) or [],
                }

                missing_display = []
                if not (str(packing_dict.get('lot_no', '') or '').strip()):
                    missing_display.append('LOT NO')
                if not (str(packing_dict.get('product', '') or '').strip()):
                    missing_display.append('PRODUCT')
                try:
                    nw = packing_dict.get('net_weight', 0)
                    if nw is None or (isinstance(nw, (int, float)) and float(nw) <= 0):
                        missing_display.append('NET(Kg)')
                except (TypeError, ValueError):
                    missing_display.append('NET(Kg)')
                try:
                    mx = packing_dict.get('mxbg_pallet', 0)
                    if mx is None or (isinstance(mx, (int, float)) and int(float(mx)) <= 0):
                        missing_display.append('MXBG')
                except (TypeError, ValueError):
                    missing_display.append('MXBG')
                if missing_display:
                    display_row = idx + 2  # Excel/미리보기 1-based 행 번호
                    failed_rows.append({
                        'row': display_row, 'row_num': display_row,
                        'value': '비어 있음',
                        'column': ', '.join(missing_display),
                        'missing_columns': missing_display,
                    })
                    errors.append(f"행 {idx + 2}: {', '.join(missing_display)} 누락")
                    continue

                # v8.7.0 Phase 3-B: invoice/bl 공용 숫자 정규화 — invoice 블록 밖으로 승격
                def _num_safe(v):
                    try:
                        return float(v) if v not in (None, '') else 0.0
                    except (TypeError, ValueError):
                        return 0.0

                inv_dict = None
                if invoice:
                    # v8.7.0 Phase 2: Invoice 전 필드 전달 — 기존엔 sap/salar/date 3개만 전달
                    # v8.7.0 Phase 3-B: document_invoice 테이블 보존용 전필드 확장
                    inv_dict = {
                        'sap_no': getattr(invoice, 'sap_no', '') or '',
                        'salar_invoice_no': getattr(invoice, 'salar_invoice_no', '') or '',
                        'invoice_no': getattr(invoice, 'invoice_no', '') or '',
                        'invoice_date': self._date_str(getattr(invoice, 'invoice_date', None)) or '',
                        'total_amount': _num_safe(getattr(invoice, 'total_amount', 0)),
                        'currency': str(getattr(invoice, 'currency', '') or '').upper().strip(),
                        'unit_price': _num_safe(getattr(invoice, 'unit_price', 0)),
                        # v8.7.0 Phase 3-B: 고객/제품/거래조건 추가 필드
                        'customer_name': getattr(invoice, 'customer_name', '') or '',
                        'customer_code': getattr(invoice, 'customer_code', '') or '',
                        'customer_address': getattr(invoice, 'customer_address', '') or '',
                        'product_code': getattr(invoice, 'product_code', '') or '',
                        'product_name': getattr(invoice, 'product_name', '') or '',
                        'quantity_mt': _num_safe(getattr(invoice, 'quantity_mt', 0)),
                        'payment_term': getattr(invoice, 'payment_term', '') or '',
                        'incoterm': getattr(invoice, 'incoterm', '') or '',
                        # v8.7.0 [FIX U-3]: package_type 추가 — document_invoice 컬럼 존재하나 dict 경로 누락
                        'package_type': getattr(invoice, 'package_type', '') or '',
                    }

                bl_dict = None
                if bl:
                    # v8.7.0 Phase 2: BL.voyage 포함
                    # v8.7.0 Phase 3-B: document_bl 테이블 보존용 전필드 확장
                    bl_dict = {
                        'bl_no': self._format_bl(getattr(bl, 'bl_no', '') or ''),
                        'ship_date': self._date_str(getattr(bl, 'ship_date', None)) or self._date_str(getattr(bl, 'shipped_on_board_date', None)) or '',
                        'vessel': getattr(bl, 'vessel', '') or '',
                        'voyage': str(getattr(bl, 'voyage', '') or '').strip(),
                        # v8.7.0 Phase 3-B: 당사자/항구/요약/선사/운임
                        'booking_no': getattr(bl, 'booking_no', '') or '',
                        'scac': getattr(bl, 'scac', '') or '',
                        'carrier_id': getattr(bl, 'carrier_id', '') or '',
                        'carrier_name': getattr(bl, 'carrier_name', '') or '',
                        'svc_contract': getattr(bl, 'svc_contract', '') or '',
                        'sap_no': getattr(bl, 'sap_no', '') or '',
                        'shipper_name': getattr(bl, 'shipper_name', '') or '',
                        'shipper_address': getattr(bl, 'shipper_address', '') or '',
                        'consignee_name': getattr(bl, 'consignee_name', '') or '',
                        'consignee_address': getattr(bl, 'consignee_address', '') or '',
                        'notify_party': getattr(bl, 'notify_party', '') or '',
                        'port_of_loading': getattr(bl, 'port_of_loading', '') or '',
                        'port_of_discharge': getattr(bl, 'port_of_discharge', '') or '',
                        'place_of_receipt': getattr(bl, 'place_of_receipt', '') or '',
                        'place_of_delivery': getattr(bl, 'place_of_delivery', '') or '',
                        'place_of_issue': getattr(bl, 'place_of_issue', '') or '',
                        'issue_date': self._date_str(getattr(bl, 'issue_date', None)) or '',
                        'product_name': getattr(bl, 'product_name', '') or '',
                        'total_containers': _num_safe(getattr(bl, 'total_containers', 0)),
                        'total_packages': _num_safe(getattr(bl, 'total_packages', 0)),
                        'net_weight_kg': _num_safe(getattr(bl, 'net_weight_kg', 0)),
                        'gross_weight_kg': _num_safe(getattr(bl, 'gross_weight_kg', 0)),
                        'total_cbm': _num_safe(getattr(bl, 'total_cbm', 0)),
                        'freight_terms': getattr(bl, 'freight_terms', '') or '',
                        'total_freight_usd': _num_safe(getattr(bl, 'total_freight_usd', 0)),
                        'total_freight_krw': _num_safe(getattr(bl, 'total_freight_krw', 0)),
                    }

                do_dict = None
                if do:
                    _con_return = ''
                    ft_infos = getattr(do, 'free_time_info', []) or []
                    for ft in ft_infos:
                        ftd = getattr(ft, 'free_time_date', '') or (ft.get('free_time_date', '') if isinstance(ft, dict) else '')
                        if ftd:
                            try:
                                from utils.date_utils import normalize_date_str as _norm_date_str
                                _con_return = _norm_date_str(ftd) or str(ftd)[:10]
                            except Exception:
                                _con_return = str(ftd)[:10]
                            break
                    _do_arr = getattr(do, 'arrival_date', None)
                    _do_arrival = (_do_arr.isoformat() if hasattr(_do_arr, 'isoformat') else str(_do_arr or '')) if _do_arr and str(_do_arr) != 'None' else ''
                    # v8.7.0 [FIX]: DOData 실제 필드는 warehouse_name / stock_date도 포함
                    _stk = getattr(do, 'stock_date', None)
                    _stock_date_str = (
                        _stk.isoformat() if hasattr(_stk, 'isoformat') else str(_stk or '')
                    ) if _stk and str(_stk) != 'None' else ''
                    # v8.7.0 [FIX D-1.3 CRITICAL]: DOData에 free_time 필드 없음.
                    #   free_time_info[0].storage_free_days에서 일수 추출.
                    _ft_days = ''
                    try:
                        _fti_list = getattr(do, 'free_time_info', []) or []
                        for _fti in _fti_list:
                            _sfd = getattr(_fti, 'storage_free_days', 0) or 0
                            if _sfd and int(_sfd) > 0:
                                _ft_days = str(int(_sfd))
                                break
                    except (TypeError, ValueError, AttributeError) as _fte:
                        logger.debug(f"free_time 추출 실패(무시): {_fte}")
                    if not _ft_days and _do_arrival and _con_return:
                        try:
                            from utils.date_utils import calculate_free_days as _calc_free_days
                            _days = _calc_free_days(_do_arrival, _con_return)
                            if _days is not None:
                                _ft_days = str(int(_days))
                        except Exception as _fte:
                            logger.debug(f"free_time 계산 실패(무시): {_fte}")
                    do_dict = {
                        'bl_no': str(getattr(do, 'bl_no', '') or ''),
                        'arrival_date': _do_arrival,
                        'stock_date': _stock_date_str,
                        'free_time_date': _con_return,
                        'free_time': _ft_days,
                        'warehouse': str(
                            getattr(do, 'warehouse_name', '')
                            or getattr(do, 'warehouse_code', '')
                            or ''
                        ),
                        # v8.7.0 Phase 2: D/O 번호 전달
                        'do_no': str(getattr(do, 'do_no', '') or '').strip(),
                        # v8.7.0 Phase 3-A: 원본 리스트 그대로 전달 (엔진이 정규화)
                        #   엔진 측에서 container_info / freetime_info 테이블에 1:N 저장
                        'containers': getattr(do, 'containers', []) or [],
                        'free_time_info': getattr(do, 'free_time_info', []) or [],
                        # v8.7.0 Phase 3-B: document_do 테이블 보존용 추가 필드
                        'sap_no': str(getattr(do, 'sap_no', '') or ''),
                        'vessel': str(getattr(do, 'vessel', '') or ''),
                        'voyage': str(getattr(do, 'voyage', '') or ''),
                        'flag': str(getattr(do, 'flag', '') or ''),
                        'port_of_loading': str(getattr(do, 'port_of_loading', '') or ''),
                        'port_of_discharge': str(getattr(do, 'port_of_discharge', '') or ''),
                        'final_destination': str(getattr(do, 'final_destination', '') or ''),
                        'place_of_delivery': str(getattr(do, 'place_of_delivery', '') or ''),
                        'issue_date': self._date_str(getattr(do, 'issue_date', None)) or '',
                        'warehouse_code': str(getattr(do, 'warehouse_code', '') or ''),
                        'warehouse_name': str(getattr(do, 'warehouse_name', '') or ''),
                        'con_return': _con_return,
                        'total_containers': len(getattr(do, 'containers', []) or []),
                        'total_packages': _num_safe(getattr(do, 'total_packages', 0)),
                        'gross_weight_kg': _num_safe(getattr(do, 'gross_weight_kg', 0)),
                        'measurement_cbm': _num_safe(getattr(do, 'measurement_cbm', 0)),
                        'mrn': str(getattr(do, 'mrn', '') or ''),
                        'msn': str(getattr(do, 'msn', '') or ''),
                    }

                try:
                    result = self.engine.process_inbound(
                        packing_data=packing_dict, invoice_data=inv_dict,
                        bl_data=bl_dict, do_data=do_dict
                    )
                    if result.get('success'):
                        created_lots.append(lot_no)
                    else:
                        err_msg = result.get('message', '') or ', '.join(result.get('errors', []))
                        errors.append(f"LOT {lot_no}: {err_msg}")
                        failed_rows.append({
                            'row': idx + 2, 'row_num': idx + 2,
                            'value': err_msg, 'column': 'LOT NO',
                            'missing_columns': [],
                        })
                except (ValueError, TypeError, AttributeError) as e:
                    errors.append(f"LOT {lot_no}: {e}")
                    failed_rows.append({
                        'row': idx + 2, 'row_num': idx + 2,
                        'value': str(e), 'column': 'LOT NO',
                        'missing_columns': [],
                    })

            if errors:
                self._log_safe(f"⚠️ 일부 오류: {len(errors)}건")
                for e in errors[:5]:
                    self._log_safe(f"  - {e}")
            if skipped_lots:
                self._log_safe(f"⏭ 중복 건너뜀: {len(skipped_lots)}건")

            if created_lots:
                self._log_safe(f"✅ 저장 완료: {len(created_lots)}건")
                return True, []
            elif skipped_lots and not errors:
                # 모든 LOT가 중복 — 개별 행으로 반환하여 실패 건수 정확히 표시
                self._log_safe(f"⏭ 모든 LOT 중복 ({len(skipped_lots)}건) — 신규 LOT 없음")
                return False, [
                    {
                        'row': '?', 'row_num': '?',
                        'value': f'LOT {lot} 이미 DB에 존재',
                        'column': '', 'missing_columns': [],
                        'type': 'all_duplicate_lot'
                    }
                    for lot in skipped_lots
                ]
            else:
                self._log_safe(f"❌ 저장된 LOT 없음 (오류 {len(errors)}건, 건너뜀 {len(skipped_lots)}건)")
                return False, failed_rows

        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            logger.error(f"DB 저장 실패: {e}", exc_info=True)
            self._log_safe(f"❌ DB 저장 실패: {e}")
            msg = str(e)
            err_type = 'db_schema' if 'no column named' in msg.lower() or 'no such column' in msg.lower() else 'db_error'
            try:
                row_num = _last_idx + 2 if _last_idx >= 0 else '?'
            except NameError:
                row_num = '?'
            return False, [{'row': row_num, 'row_num': row_num, 'value': msg, 'column': '', 'missing_columns': [], 'type': err_type}]

    def _export_to_excel(self) -> None:
        """미리보기 데이터 Excel 내보내기"""
        from .onestop_inbound import PREVIEW_COLUMNS
        if not self.preview_data:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            save_path = filedialog.asksaveasfilename(
                parent=self.dialog, title="Excel 내보내기",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"입고미리보기_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not save_path:
                return
            try:
                from ..utils.excel_file_helper import get_unique_excel_path
                save_path = get_unique_excel_path(save_path)
            except ImportError as e:
                logger.warning(f"[_export_to_excel] Suppressed: {e}")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "입고 미리보기"

            headers = [col[1] for col in PREVIEW_COLUMNS]
            hfill = PatternFill(start_color="2c6fbb", end_color="2c6fbb", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True, size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            for ri, row_data in enumerate(self.preview_data, 2):
                for ci, (col_id, _, _, _) in enumerate(PREVIEW_COLUMNS, 1):
                    cell = ws.cell(row=ri, column=ci, value=row_data.get(col_id, ''))
                    cell.border = border

            for ci, (_, h, w, _) in enumerate(PREVIEW_COLUMNS, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(w / 7, len(h) + 2)

            try:
                from utils.sqm_excel_alignment import apply_sqm_workbook_alignment
                apply_sqm_workbook_alignment(wb)
            except Exception:
                pass
            wb.save(save_path)
            self._log_safe(f"📥 Excel 저장: {save_path}")

        except (ValueError, TypeError, AttributeError) as e:
            self._log_safe(f"❌ Excel 오류: {e}")
