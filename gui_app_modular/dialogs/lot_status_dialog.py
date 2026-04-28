# -*- coding: utf-8 -*-
"""LOT별 통합 현황 다이얼로그 — v8.1.5 Phase 1-G

판매배정 탭 → 📊 LOT별 현황 버튼에서 호출.
LOT 단위로 AVAILABLE / RESERVED / PICKED / OUTBOUND / 샘플 상태를 통합 표시.
"""

import logging
import tkinter as tk
from tkinter import BOTH, LEFT, RIGHT, X, Y, END, W, E, N, S
from datetime import datetime

try:
    import ttkbootstrap as ttk
    from ttkbootstrap.constants import *
except ImportError:
    import tkinter.ttk as ttk

from ..utils.tree_enhancements import TreeviewTotalFooter

logger = logging.getLogger(__name__)

# ── 컬럼 정의 ──────────────────────────────────────────
COLUMNS = [
    ('lot_no',       'LOT NO',      120, 'center'),
    ('sap_no',       'SAP NO',       95, 'center'),
    ('total_tb',     '총톤백',        55, 'center'),
    ('avail_tb',     'AVAILABLE',    75, 'center'),
    ('reserved_tb',  'RESERVED',     85, 'center'),
    ('reserved_mt',  'RSV(MT)',       70, 'center'),
    ('picked_tb',    'PICKED',        65, 'center'),
    ('out_tb',       'OUTBOUND',      70, 'center'),
    ('sample_stat',  '샘플(1kg)',     80, 'center'),
    ('rsv_pct',      '배정%',         55, 'center'),
    ('sale_refs',    'SALE REF',      80, 'center'),
    ('lot_status',   '상태',          80, 'center'),
]

# ── 상태 색상 (Treeview tag) ──────────────────────────
STATUS_TAGS = {
    'AVAILABLE':    {'background': '#d1fae5', 'foreground': '#065f46'},
    'PARTIAL':      {'background': '#dbeafe', 'foreground': '#1e40af'},
    'FULL_RSV':     {'background': '#c7d2fe', 'foreground': '#3730a3'},
    'PICKED':       {'background': '#fef3c7', 'foreground': '#92400e'},
    'OUTBOUND':     {'background': '#f3f4f6', 'foreground': '#374151'},
}


def _calc_lot_status(row: dict) -> str:
    """LOT 상태 자동 계산."""
    tot = int(row.get('total_tb', 0) or 0)
    out = int(row.get('out_tb', 0) or 0)
    pick = int(row.get('picked_tb', 0) or 0)
    rsv = int(row.get('reserved_tb', 0) or 0)
    if tot > 0 and out >= tot:
        return 'OUTBOUND'
    if pick > 0:
        return 'PICKED'
    if rsv > 0 and rsv >= tot and tot > 0:
        return 'FULL_RSV'
    if rsv > 0:
        return 'PARTIAL'
    return 'AVAILABLE'


class LotStatusDialog:
    """LOT별 통합 현황 팝업 다이얼로그."""

    QUERY = """
        SELECT
            i.lot_no, i.sap_no,
            SUM(CASE WHEN COALESCE(t.is_sample,0)=0 THEN 1 ELSE 0 END) AS total_tb,
            SUM(CASE WHEN COALESCE(t.is_sample,0)=0 AND t.status='AVAILABLE'
                THEN 1 ELSE 0 END) AS avail_tb,
            SUM(CASE WHEN COALESCE(t.is_sample,0)=0 AND t.status='PICKED'
                THEN 1 ELSE 0 END) AS picked_tb,
            SUM(CASE WHEN COALESCE(t.is_sample,0)=0
                AND t.status IN ('OUTBOUND','SOLD') THEN 1 ELSE 0 END) AS out_tb,
            COALESCE((SELECT CAST(SUM(ap.qty_mt/0.5) AS INT)
                FROM allocation_plan ap
                WHERE ap.lot_no=i.lot_no AND ap.status='RESERVED'), 0) AS reserved_tb,
            COALESCE((SELECT SUM(ap.qty_mt)
                FROM allocation_plan ap
                WHERE ap.lot_no=i.lot_no AND ap.status='RESERVED'), 0) AS reserved_mt,
            COALESCE((SELECT GROUP_CONCAT(DISTINCT ap.sale_ref)
                FROM allocation_plan ap
                WHERE ap.lot_no=i.lot_no AND ap.status='RESERVED'), '') AS sale_refs,
            MAX(CASE WHEN t.is_sample=1 THEN t.status ELSE NULL END) AS sample_status,
            SUM(CASE WHEN t.is_sample=1 THEN 1 ELSE 0 END) AS sample_cnt
        FROM inventory i
        LEFT JOIN inventory_tonbag t ON t.lot_no = i.lot_no
        GROUP BY i.lot_no
        ORDER BY i.lot_no
    """

    def __init__(self, parent, engine):
        self.parent = parent
        self.engine = engine
        self._rows = []
        self._show_sample = tk.BooleanVar(value=True)
        self.dialog = None

    def show(self):
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title("📊 LOT별 통합 현황")
        try:
            from gui_app_modular.utils.ui_constants import setup_dialog_geometry_persistence as _sgp
            _sgp(self.dialog, "lot_status_dialog", None, "large")
        except Exception as e:
            logger.warning(f'[UI] lot_status_dialog: {e}')
        self.dialog.geometry("1100x650")
        self.dialog.resizable(True, True)
        self.dialog.minsize(800, 400)
        try:
            self.dialog.transient(self.parent)
        except Exception as e:
            logger.warning(f'[UI] lot_status_dialog: {e}')
        self.dialog.grab_set()
        self.dialog.lift()
        self.dialog.focus_force()

        self._build_ui()
        self._load_data()

    def _build_ui(self):
        dlg = self.dialog

        # ── 상단 툴바 ──
        toolbar = ttk.Frame(dlg)
        toolbar.pack(fill=X, padx=8, pady=(8, 4))

        ttk.Button(toolbar, text="🔄 새로고침", command=self._load_data,
                   bootstyle="info-outline", width=12).pack(side=LEFT, padx=2)

        # 검색
        ttk.Label(toolbar, text="LOT 검색:").pack(side=LEFT, padx=(12, 4))
        self._search_var = tk.StringVar()
        self._search_entry = ttk.Entry(toolbar, textvariable=self._search_var, width=16)
        self._search_entry.pack(side=LEFT, padx=2)
        self._search_entry.bind('<Return>', lambda e: self._apply_filter())
        ttk.Button(toolbar, text="🔍", command=self._apply_filter,
                   bootstyle="secondary-outline", width=4).pack(side=LEFT, padx=2)

        # 상태 필터
        ttk.Label(toolbar, text="상태:").pack(side=LEFT, padx=(12, 4))
        self._status_filter = ttk.Combobox(toolbar, width=12, state='readonly',
                                           values=['전체', 'AVAILABLE', 'PARTIAL', 'FULL_RSV', 'PICKED', 'OUTBOUND'])
        self._status_filter.set('전체')
        self._status_filter.pack(side=LEFT, padx=2)
        self._status_filter.bind('<<ComboboxSelected>>', lambda e: self._apply_filter())

        # 샘플 토글
        ttk.Checkbutton(toolbar, text="샘플 포함", variable=self._show_sample,
                        command=self._toggle_sample_column,
                        bootstyle="round-toggle").pack(side=LEFT, padx=(12, 2))

        # Excel 저장
        ttk.Button(toolbar, text="📥 Excel 저장", command=self._export_excel,
                   bootstyle="success-outline", width=12).pack(side=RIGHT, padx=2)

        # ── Treeview ──
        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill=BOTH, expand=True, padx=8, pady=4)

        col_ids = [c[0] for c in COLUMNS]
        self.tree = ttk.Treeview(tree_frame, columns=col_ids, show='headings',
                                 selectmode='browse')

        for cid, heading, width, anchor in COLUMNS:
            self.tree.heading(cid, text=heading)
            self.tree.column(cid, width=width, minwidth=40, anchor=anchor)

        # 스크롤바
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=RIGHT, fill=Y)
        # v8.1.9: TreeviewTotalFooter
        try:
            from gui_app_modular.utils.tree_enhancements import TreeviewTotalFooter as _TTF
            _footer = _TTF(
                tree_frame, self.tree,
                summable_column_ids=[],
                column_display_names={},
                column_formats={},
            )
            _footer.pack(fill='x')
            self._footer_tree = _footer
        except Exception as e:
            logger.warning(f'[UI] lot_status_dialog: {e}')
        hsb.pack(side='bottom', fill=X)
        self.tree.pack(fill=BOTH, expand=True)

        # 상태 태그
        for tag, style in STATUS_TAGS.items():
            self.tree.tag_configure(tag, **style)

        # v8.1.5: 합계 Footer
        self._lot_status_footer = TreeviewTotalFooter(
            dlg, self.tree,
            ['total_tb', 'avail_tb', 'reserved_tb', 'picked_tb', 'out_tb'],
            column_display_names={
                'total_tb': '총톤백', 'avail_tb': 'AVAILABLE',
                'reserved_tb': 'RESERVED', 'picked_tb': 'PICKED', 'out_tb': 'OUTBOUND',
            }
        )
        self._lot_status_footer.pack(fill=X, padx=8)

        # ── 하단 요약 ──
        self._summary_var = tk.StringVar(value="")
        summary_lbl = ttk.Label(dlg, textvariable=self._summary_var,
                                font=('맑은 고딕', 10))
        summary_lbl.pack(fill=X, padx=8, pady=(0, 4))

        formula_lbl = ttk.Label(dlg,
                                text="배정% = (RESERVED+PICKED+OUT kg) ÷ 일반톤백 총중량 × 100 (샘플 제외)",
                                font=('맑은 고딕', 8), bootstyle="secondary")
        formula_lbl.pack(fill=X, padx=8, pady=(0, 8))

    def _load_data(self):
        try:
            rows = self.engine.db.fetchall(self.QUERY)
            self._rows = [dict(r) if hasattr(r, 'keys') else dict(zip(
                ['lot_no','sap_no','total_tb','avail_tb','picked_tb','out_tb',
                 'reserved_tb','reserved_mt','sale_refs','sample_status','sample_cnt'], r
            )) for r in (rows or [])]
        except Exception as e:
            logger.warning(f"[LotStatus] 데이터 로드 실패: {e}")
            self._rows = []
        self._apply_filter()

    def _apply_filter(self):
        search = self._search_var.get().strip().upper()
        status_f = self._status_filter.get()

        self.tree.delete(*self.tree.get_children())

        cnt_total = cnt_avail = cnt_partial = cnt_full = cnt_samp = 0

        for row in self._rows:
            lot_status = _calc_lot_status(row)

            # 필터
            if search and search not in str(row.get('lot_no', '')).upper():
                continue
            if status_f != '전체' and lot_status != status_f:
                continue

            tot = int(row.get('total_tb', 0) or 0)
            rsv = int(row.get('reserved_tb', 0) or 0)
            rsv_mt = float(row.get('reserved_mt', 0) or 0)
            pick = int(row.get('picked_tb', 0) or 0)
            out = int(row.get('out_tb', 0) or 0)
            samp_cnt = int(row.get('sample_cnt', 0) or 0)
            samp_stat = row.get('sample_status', '') or '-'

            pct = int((rsv + pick + out) / max(tot, 1) * 100) if tot > 0 else 0

            # v8.6.3: 샘플은 본품 전량 출고 후에만 처리 — 본품 잔여 시 '대기'
            # 본품 전량 출고 완료 = avail_tb==0 AND reserved_tb==0 AND picked_tb==0
            _avail_tb = int(row.get('avail_tb', 0) or 0)
            _all_normal_done = (_avail_tb == 0 and rsv == 0 and pick == 0)

            if samp_cnt == 0:
                sample_display = "—"
            elif not _all_normal_done:
                # 본품 아직 남아 있음 → 샘플 대기 표시
                sample_display = f"{samp_cnt}개 (본품 출고 후)"
            else:
                # 본품 전량 완료 → 샘플 상태 표시
                _stat_label = {
                    'AVAILABLE': 'AVAIL',
                    'RESERVED':  'RSV',
                    'PICKED':    'PICKED',
                    'OUTBOUND':  'DONE',
                    'SOLD':      'DONE',
                }.get(samp_stat.upper(), samp_stat)
                sample_display = f"{samp_cnt}개 {_stat_label}" 

            values = (
                row.get('lot_no', ''),
                row.get('sap_no', ''),
                tot,
                int(row.get('avail_tb', 0) or 0),
                rsv,
                f"{rsv_mt:.3f}",
                pick,
                out,
                sample_display,
                f"{pct}%",
                row.get('sale_refs', '') or '',
                lot_status,
            )
            self.tree.insert('', END, values=values, tags=(lot_status,))

            cnt_total += 1
            if lot_status == 'AVAILABLE':
                cnt_avail += 1
            elif lot_status == 'PARTIAL':
                cnt_partial += 1
            elif lot_status == 'FULL_RSV':
                cnt_full += 1
            if samp_cnt > 0:
                cnt_samp += 1

        self._summary_var.set(
            f"총 {cnt_total} LOT  |  미배정: {cnt_avail}  |  "
            f"부분배정: {cnt_partial}  |  완전배정: {cnt_full}  |  "
            f"샘플보유: {cnt_samp}"
        )
        if hasattr(self, '_lot_status_footer'):
            self._lot_status_footer.update_totals()

    def _toggle_sample_column(self):
        col_ids = [c[0] for c in COLUMNS]
        if not self._show_sample.get():
            col_ids = [c for c in col_ids if c != 'sample_stat']
        self.tree['displaycolumns'] = col_ids

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from tkinter import filedialog
        except ImportError:
            from ..utils.ui_constants import CustomMessageBox
            CustomMessageBox.showerror(self.dialog, "오류", "openpyxl 패키지가 필요합니다.\npip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            parent=self.dialog,
            title="LOT별 현황 Excel 저장",
            defaultextension=".xlsx",
            initialfile=f"LOT_Status_{datetime.now():%Y%m%d_%H%M}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LOT별현황"

        hdr_fill = PatternFill('solid', fgColor='1F3864')
        hdr_font = Font(bold=True, color='FFFFFF', size=10, name='맑은 고딕')
        smp_fill = PatternFill('solid', fgColor='FFF2CC')

        HEADERS = [
            ('LOT NO', 14), ('SAP NO', 13), ('총톤백', 7), ('AVAILABLE', 10),
            ('RESERVED(개)', 12), ('RESERVED(MT)', 12), ('PICKED', 8),
            ('OUTBOUND', 9), ('샘플수', 7), ('샘플상태', 10),
            ('배정%', 7), ('SALE REF', 10), ('LOT 상태', 11),
        ]

        for c, (h, w) in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = smp_fill if c in (9, 10) else hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(c)].width = w

        for ri, row in enumerate(self._rows, 2):
            lot_status = _calc_lot_status(row)
            tot = int(row.get('total_tb', 0) or 0)
            rsv = int(row.get('reserved_tb', 0) or 0)
            pick = int(row.get('picked_tb', 0) or 0)
            out = int(row.get('out_tb', 0) or 0)
            pct = f"{int((rsv + pick + out) / max(tot, 1) * 100)}%" if tot > 0 else "0%"

            data = [
                row.get('lot_no', ''), row.get('sap_no', ''),
                tot, int(row.get('avail_tb', 0) or 0),
                rsv, round(float(row.get('reserved_mt', 0) or 0), 3),
                pick, out,
                int(row.get('sample_cnt', 0) or 0),
                row.get('sample_status', '') or '-',
                pct, row.get('sale_refs', '') or '',
                lot_status,
            ]
            bg = 'FFFFFF' if ri % 2 == 0 else 'F2F2F2'
            for c, v in enumerate(data, 1):
                cell = ws.cell(row=ri, column=c, value=v)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill('solid', fgColor=bg)

        last = len(self._rows) + 2
        ws.cell(row=last, column=1, value='합계').font = Font(bold=True)
        ws.cell(row=last, column=3, value=f'=SUM(C2:C{last-1})').font = Font(bold=True)
        ws.cell(row=last, column=5, value=f'=SUM(E2:E{last-1})').font = Font(bold=True)
        ws.cell(row=last, column=6, value=f'=SUM(F2:F{last-1})').font = Font(bold=True)

        ws.freeze_panes = 'A2'
        wb.save(path)

        try:
            from ..utils.ui_constants import CustomMessageBox
            CustomMessageBox.showinfo(
                self.dialog, "저장 완료",
                f"LOT별 현황 Excel 저장 완료\n{path}"
            )
        except Exception as e:
            logger.warning(f'[UI] lot_status_dialog: {e}')
        try:
            import os
            os.startfile(path)
        except Exception as e:
            logger.warning(f'[UI] lot_status_dialog: {e}')