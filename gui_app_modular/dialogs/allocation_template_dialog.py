"""
SQM 재고관리 시스템 - Allocation 양식 미리보기 다이얼로그

Author: Ruby
Version: 1.0.0

기능:
  - Song 양식 / Woo 양식 탭 전환 미리보기
  - 본 제품 행 + 샘플 행 각각 독립 줄로 표시 (구분 태그만 다름)
  - "이 양식 다운로드" 버튼으로 Excel 저장
  - resources/templates/ 폴더 내 파일을 읽어 표시 (방법 B)
  - 파일 없으면 내장 샘플 데이터로 fallback (방법 C)

사용법:
from gui_app_modular.utils.ui_constants import create_themed_toplevel  # v8.0.9
    from gui_app_modular.dialogs.allocation_template_dialog import AllocationTemplateDialog
    AllocationTemplateDialog(parent)
"""

import json
import logging
import os
import shutil
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import List, Optional, Tuple

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────
# UI 유틸 import (없으면 fallback)
# ─────────────────────────────────────────
try:
    from gui_app_modular.utils.ui_constants import (
        create_themed_toplevel,
        DialogSize,
        apply_modal_window_options,
        center_dialog,
        setup_dialog_geometry_persistence,
        tc,
    )
    _HAS_UI_UTILS = True
except ImportError:
    _HAS_UI_UTILS = False
    def create_themed_toplevel(parent): return tk.Toplevel(parent)
    def tc(key, dark=None): return '#888888'  # fallback


# ─────────────────────────────────────────
# 색상 상수 (다크 테마 기준)
# ─────────────────────────────────────────
C_BG        = tc('bg_secondary')
C_BG2       = tc('bg_card')
C_FG        = tc('text_primary')
C_FG2       = tc('text_secondary')
C_ACCENT    = tc('accent')
C_SUCCESS   = tc('success')
C_WARNING   = tc('warning')
C_DANGER    = tc('danger')
C_STRIPE    = tc('bg_tertiary')
C_SELECT    = tc('accent')
C_SAMPLE    = '#ff9100'
C_BTN_DL    = '#2979ff'
C_BTN_CLOSE = tc('bg_tertiary')


# ─────────────────────────────────────────
# 내장 샘플 데이터 (파일 없을 때 fallback)
# ─────────────────────────────────────────
_SONG_COLUMNS = [
    'Product', 'SAP NO', 'Date in stock', 'QTY (MT)',
    'Lot No', 'WH', 'Customs', 'SOLD TO', 'SALE REF', 'GW',
]

_SONG_ROWS = [
    # (Product, SAP NO, Date in stock, QTY, Lot No, WH, Customs, SOLD TO, SALE REF, GW)
    ('MIC9000',        '2200032713', '2025-09-05', '5',     '1125062056', 'GY', 'uncleared', 'LBM AP - January 550MT Semarang', '2903', '5.13'),
    ('MIC9000 sample', '2200032713', '2025-09-05', '0.001', '1125062056', 'GY', 'uncleared', 'LBM AP - January 550MT Semarang', '2903', '0.00125'),
    ('MIC9000',        '2200032713', '2025-09-05', '5',     '1125062057', 'GY', 'uncleared', 'LBM AP - January 550MT Semarang', '2903', '5.13'),
    ('MIC9000 sample', '2200032713', '2025-09-05', '0.001', '1125062057', 'GY', 'uncleared', 'LBM AP - January 550MT Semarang', '2903', '0.00125'),
    ('MIC9000',        '2200032991', '2025-09-09', '5',     '1125070745', 'GY', 'uncleared', 'LBM AP - January 550MT Semarang', '2903', '5.13'),
    ('MIC9000 sample', '2200032991', '2025-09-09', '0.001', '1125070745', 'GY', 'uncleared', 'LBM AP - January 550MT Semarang', '2903', '0.00125'),
    ('MIC9000',        '2200032712', '2025-09-18', '5',     '1125070924', 'GY', 'uncleared', 'LBM AP - January 550MT Semarang', '2903', '5.13'),
    ('MIC9000 sample', '2200032712', '2025-09-18', '0.001', '1125070924', 'GY', 'uncleared', 'LBM AP - January 550MT Semarang', '2903', '0.00125'),
]

_WOO_COLUMNS = [
    'Product', 'SAP NO', 'Date in stock', 'QTY (MT)',
    'Lot No', 'WH', 'Customs', 'Export', 'SOLD TO',
    'SALE REF', 'Balance', 'GW', 'Remark',
]

_WOO_ROWS = [
    ('MIC9000',        '2200032902', '2025-10-13', '5',     '1125080535', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026  1st 550mt', '3184', '5',     '5.13', ''),
    ('MIC9000 Sample', '2200032902', '2025-10-13', '0.001', '1125080535', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026  1st 550mt', '3184', '0.001', '0.00125', ''),
    ('MIC9000',        '2200032902', '2025-10-13', '5',     '1125080539', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026  1st 550mt', '3184', '5',     '5.13', ''),
    ('MIC9000 Sample', '2200032902', '2025-10-13', '0.001', '1125080539', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026  1st 550mt', '3184', '0.001', '0.00125', ''),
    ('MIC9000',        '2200033015', '2025-10-13', '5',     '1125080713', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026  1st 550mt', '3184', '5',     '5.13', ''),
    ('MIC9000 Sample', '2200033015', '2025-10-13', '0.001', '1125080713', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026  1st 550mt', '3184', '0.001', '0.00125', ''),
    ('MIC9000',        '2200032904', '2025-10-14', '5',     '1125081218', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026  1st 550mt', '3184', '5',     '5.13', ''),
    ('MIC9000 Sample', '2200032904', '2025-10-14', '0.001', '1125081218', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026  1st 550mt', '3184', '0.001', '0.00125', ''),
]


def _is_sample_row(product_name: str, qty_str: str) -> bool:
    """샘플 행 여부 판단"""
    prod_lower = str(product_name).lower()
    if 'sample' in prod_lower:
        return True
    try:
        return float(qty_str) < 0.01
    except (ValueError, TypeError):
        return False


def _find_template_file(name: str) -> Optional[Path]:
    """
    resources/templates/ 폴더에서 양식 파일 탐색.
    없으면 None 반환 → 내장 데이터 fallback.
    """
    if not name:
        return None
    direct = Path(str(name))
    if direct.is_absolute() and direct.exists():
        return direct
    candidates = [
        _allocation_template_dir() / name,
        Path(__file__).parent.parent / 'resources' / 'templates' / name,
        Path(__file__).parent.parent.parent / 'resources' / 'templates' / name,
        Path(os.getcwd()) / 'resources' / 'templates' / name,
        Path(os.getcwd()) / 'resources' / 'templates' / 'allocation' / name,
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _allocation_template_dir() -> Path:
    """외부 Allocation 양식/매핑 저장 폴더."""
    root = Path(__file__).resolve().parents[2]
    path = root / 'resources' / 'templates' / 'allocation'
    path.mkdir(parents=True, exist_ok=True)
    return path


def _safe_id(value: str) -> str:
    import re
    base = re.sub(r'[^A-Za-z0-9_.-]+', '_', str(value or '').strip()).strip('._')
    return (base or 'allocation_template')[:80]


def _find_header_row_in_xlsx(filepath: Path) -> Tuple[str, int, List[str]]:
    """엑셀에서 Product/Lot No 기준 헤더 행을 찾는다. 반환 header_row는 1-base."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    best = (wb.sheetnames[0], 1, [])
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 12) + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            norm = [str(v).strip() for v in vals if v is not None and str(v).strip()]
            upper = ' '.join(norm).upper()
            if 'PRODUCT' in upper and 'LOT' in upper:
                return ws.title, r, norm
            if 'LOT' in upper and ('QTY' in upper or 'SAP' in upper):
                best = (ws.title, r, norm)
    return best


def _mapping_from_excel(filepath: Path, label: str = '') -> dict:
    """가져온 엑셀 파일에서 기본 매핑 JSON을 생성한다."""
    sheet, header_row, columns = _find_header_row_in_xlsx(filepath)
    file_id = _safe_id(label or filepath.stem)
    return {
        'id': file_id,
        'tab_label': f'📄 {label or filepath.stem}',
        'template_file': f'{file_id}.xlsx',
        'sheet': sheet,
        'header_row': header_row,
        'data_start_row': header_row + 1,
        'description': (
            f'외부 Allocation 양식 │ 시트: {sheet} │ 헤더: {header_row}행 │ '
            f'컬럼: {" / ".join(columns)}'
        ),
        'columns': columns,
        'sample_rule': {
            'field': 'Product',
            'contains': 'sample',
            'qty_mt_max': 0.01,
        },
    }


def _load_external_formats() -> List[dict]:
    """resources/templates/allocation/*.json 매핑을 읽어 FORMATS 앞에 붙인다."""
    base = _allocation_template_dir()
    formats = []
    for json_path in sorted(base.glob('*.json')):
        try:
            data = json.loads(json_path.read_text(encoding='utf-8-sig'))
            if not isinstance(data, dict):
                continue
            tpl_name = data.get('template_file')
            if tpl_name and not Path(str(tpl_name)).is_absolute():
                data['template_file'] = str(base / str(tpl_name))
            data.setdefault('tab_label', f"📄 {data.get('id') or json_path.stem}")
            data.setdefault('description', f"외부 Allocation 양식: {json_path.name}")
            data.setdefault('mapping_file', str(json_path))
            formats.append(data)
        except Exception as e:
            logger.warning("[AllocationTemplateDialog] 외부 매핑 로드 실패 %s: %s", json_path, e)
    return formats


def _load_excel_preview(filepath: Path, max_rows: int = 30) -> Tuple[List[str], List[tuple]]:
    """
    Excel 파일에서 미리보기용 컬럼/행 로드.
    파싱 실패 시 ([], []) 반환.
    """
    try:
        import pandas as pd
        # allocation_parser와 동일한 시트 자동 선택 로직
        xf = pd.ExcelFile(str(filepath))
        best_sheet = xf.sheet_names[0]
        best_score = -1
        for sh in xf.sheet_names:
            df_sh = pd.read_excel(str(filepath), sheet_name=sh, header=None, nrows=10)
            score = 0
            has_lot = has_product = has_data = False
            for i in range(min(8, len(df_sh))):
                row_str = ' '.join(str(v).upper() for v in df_sh.iloc[i].values if str(v) != 'nan')
                if 'LOT' in row_str and 'SUB' not in row_str:
                    has_lot = True
                if 'PRODUCT' in row_str:
                    has_product = True
                for v in df_sh.iloc[i].values:
                    s = str(v).strip().split('.')[0]
                    if s.isdigit() and 8 <= len(s) <= 11:
                        has_data = True
            if has_lot and has_product:
                score = 3
            elif has_lot:
                score = 2
            if has_data:
                score += 1
            if score > best_score:
                best_score = score
                best_sheet = sh

        df = pd.read_excel(str(filepath), sheet_name=best_sheet, header=None)

        # 헤더 행 탐지
        header_idx = None
        for i in range(min(10, len(df))):
            row_str = ' '.join(str(v).upper() for v in df.iloc[i].values if str(v) != 'nan')
            has_lot = 'LOT' in row_str and 'SUB' not in row_str
            has_product = 'PRODUCT' in row_str
            if has_lot and (has_product or 'QTY' in row_str or 'SAP' in row_str):
                header_idx = i
                break
        if header_idx is None:
            return [], []

        # 컬럼명: None/피벗 집계 제외
        raw_headers = list(df.iloc[header_idx].values)
        columns = []
        valid_col_indices = []
        for idx, h in enumerate(raw_headers):
            if h is None or str(h).strip() in ('', 'nan'):
                continue
            h_str = str(h).strip()
            # 피벗 집계 컬럼(숫자) 제외
            try:
                float(h_str)
                continue
            except ValueError:
                logger.debug("[SUPPRESSED] exception in allocation_template_dialog.py")  # noqa
            columns.append(h_str)
            valid_col_indices.append(idx)

        # 데이터 행
        rows = []
        for i in range(header_idx + 1, len(df)):
            if len(rows) >= max_rows:
                break
            row_vals = list(df.iloc[i].values)
            # LOT 컬럼 확인
            lot_col_local = None
            for j, h in enumerate(columns):
                if 'LOT' in h.upper() and 'SUB' not in h.upper():
                    lot_col_local = j
                    break
            if lot_col_local is None:
                continue
            lot_val = row_vals[valid_col_indices[lot_col_local]] if lot_col_local < len(valid_col_indices) else None
            if str(lot_val) in ('', 'nan', 'None') or lot_val is None:
                continue
            # 피벗 집계 행 제외 (lot이 숫자지만 8자리 미만)
            s = str(lot_val).strip().split('.')[0]
            if not (s.isdigit() and 8 <= len(s) <= 11):
                continue
            row_data = tuple(
                str(row_vals[valid_col_indices[j]]).strip()
                if j < len(valid_col_indices) and valid_col_indices[j] < len(row_vals)
                else ''
                for j in range(len(columns))
            )
            rows.append(row_data)

        return columns, rows
    except Exception as e:
        logger.warning(f"[AllocationTemplateDialog] Excel 로드 실패: {e}")
        return [], []


class AllocationTemplateDialog:
    """
    Allocation 양식 미리보기 다이얼로그 (v1.0.0)

    레이아웃:
    ┌─────────────────────────────────────────┐
    │  📋 Allocation 양식 미리보기             │
    ├─────────────────────────────────────────┤
    │  [Song 양식 (250MT)] [Woo 양식 (550MT)] │  ← 탭
    ├─────────────────────────────────────────┤
    │  ℹ 양식 설명 (1줄)                       │
    ├─────────────────────────────────────────┤
    │  Treeview (컬럼 + 행)                   │
    │  ⬜ 일반 행  🟧 샘플 행(주황)           │
    ├─────────────────────────────────────────┤
    │  [⬇ 이 양식 다운로드]  [✕ 닫기]         │
    └─────────────────────────────────────────┘
    """

    # 양식 정의 (파일명, 내장 데이터)
    FORMATS = [
        {
            'tab_label': '📄 Song 양식 (250MT)',
            'template_file': 'allocation_template_song.xlsx',
            'description': (
                'Song 양식 │ 컬럼: Product / SAP NO / Date in stock / QTY (MT) / '
                'Lot No / WH / Customs / SOLD TO / SALE REF / GW  │  '
                '샘플 행: Product에 "sample" 포함, QTY = 0.001 MT'
            ),
            'columns': _SONG_COLUMNS,
            'rows': _SONG_ROWS,
        },
        {
            'tab_label': '📄 Woo 양식 (550MT)',
            'template_file': 'allocation_template_woo.xlsx',
            'description': (
                'Woo 양식 │ 컬럼: Product / SAP NO / Date in stock / QTY (MT) / '
                'Lot No / WH / Customs / Export / SOLD TO / SALE REF / Balance / GW / Remark  │  '
                '샘플 행: Product에 "Sample" 포함, QTY = 0.001 MT  │  Export: 반송 / 일반수출'
            ),
            'columns': _WOO_COLUMNS,
            'rows': _WOO_ROWS,
        },
    ]

    def __init__(self, parent):
        self.parent = parent
        self._current_tab = 0   # 현재 선택 탭 인덱스
        self.FORMATS = _load_external_formats() + list(self.FORMATS)
        self._loaded_data = [None] * len(self.FORMATS)   # 탭별 캐시 (columns, rows)

        self.dialog = create_themed_toplevel(parent)
        self.dialog.title('📋 Allocation 양식 미리보기')
        self.dialog.configure(bg=C_BG)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        if _HAS_UI_UTILS and setup_dialog_geometry_persistence:
            setup_dialog_geometry_persistence(self.dialog, "allocation_template_dialog", parent, "large")
        elif _HAS_UI_UTILS:
            self.dialog.geometry(DialogSize.get_geometry(parent, 'large'))
            apply_modal_window_options(self.dialog)
            center_dialog(self.dialog, parent)
        else:
            self.dialog.geometry('1100x680')
            self.dialog.resizable(True, True)

        self._build_ui()
        self._load_tab(0)

    # ──────────────────────────────────────────
    # UI 구성
    # ──────────────────────────────────────────

    def _build_ui(self):
        # ── 탭 버튼 영역 ──
        tab_frame = tk.Frame(self.dialog, bg=C_BG2, pady=6)
        tab_frame.pack(fill='x', padx=0, pady=(0, 0))

        self._tab_btns = []
        for i, fmt in enumerate(self.FORMATS):
            btn = tk.Button(
                tab_frame,
                text=fmt['tab_label'],
                font=('Malgun Gothic', 10, 'bold'),
                bg=C_BG2, fg=C_FG2,
                activebackground=C_SELECT,
                activeforeground=C_FG,
                relief='flat', bd=0,
                padx=18, pady=6,
                cursor='hand2',
                command=lambda idx=i: self._on_tab_click(idx),
            )
            btn.pack(side='left', padx=4)
            self._tab_btns.append(btn)

        import_btn = tk.Button(
            tab_frame,
            text='＋ 양식 가져오기',
            font=('Malgun Gothic', 10, 'bold'),
            bg=C_BTN_DL, fg=C_FG,
            activebackground=tc('info'),
            relief='flat', bd=0,
            padx=14, pady=6,
            cursor='hand2',
            command=self._on_import_template,
        )
        import_btn.pack(side='right', padx=8)

        # ── 설명 레이블 ──
        self._desc_var = tk.StringVar()
        desc_lbl = tk.Label(
            self.dialog,
            textvariable=self._desc_var,
            font=('맑은 고딕', 10),
            bg=C_BG, fg=C_FG2,
            anchor='w', padx=12, pady=4,
            wraplength=1050,
        )
        desc_lbl.pack(fill='x')

        # ── 범례 ──
        legend_frame = tk.Frame(self.dialog, bg=C_BG)
        legend_frame.pack(fill='x', padx=12, pady=(0, 4))
        tk.Label(legend_frame, text='■', fg=C_BG2, bg=C_BG,
                 font=('Courier', 11)).pack(side='left')
        tk.Label(legend_frame, text=' 본 제품 행',
                 fg=C_FG2, bg=C_BG, font=('맑은 고딕', 10)).pack(side='left')
        tk.Label(legend_frame, text='   ■', fg=C_SAMPLE, bg=C_BG,
                 font=('Courier', 11)).pack(side='left')
        tk.Label(legend_frame, text=' 샘플 행 (QTY=0.001 MT = 1 kg)',
                 fg=C_FG2, bg=C_BG, font=('맑은 고딕', 10)).pack(side='left')

        # ── Treeview 영역 ──
        tree_frame = tk.Frame(self.dialog, bg=C_BG)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=(0, 6))

        # 스타일
        style = ttk.Style()
        # v9.0: theme_use('default') 제거
        style.configure('Template.Treeview',
                        background=C_BG2,
                        fieldbackground=C_BG2,
                        foreground=C_FG,
                        rowheight=32,
                        font=('맑은 고딕', 10))
        style.configure('Template.Treeview.Heading',
                        background=C_BG,
                        foreground=C_ACCENT,
                        font=('맑은 고딕', 10, 'bold'),
                        relief='flat',
                        anchor='center')
        style.map('Template.Treeview',
                  background=[('selected', C_SELECT)],
                  foreground=[('selected', C_FG)])

        self._tree = ttk.Treeview(
            tree_frame,
            style='Template.Treeview',
            selectmode='browse',
            show='headings',
        )
        self._tree.tag_configure('normal', background=C_BG2, foreground=C_FG)
        self._tree.tag_configure('stripe', background=C_STRIPE, foreground=C_FG)
        self._tree.tag_configure('sample', background=tc('bg_secondary'), foreground=C_SAMPLE)
        self._tree.tag_configure('sample_stripe', background=tc('bg_secondary'), foreground=C_SAMPLE)

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')
        self._tree.pack(fill='both', expand=True)

        # 행 수 표시
        self._row_count_var = tk.StringVar(value='')
        tk.Label(self.dialog, textvariable=self._row_count_var,
                 bg=C_BG, fg=C_FG2, font=('Malgun Gothic', 8),
                 anchor='e', padx=12).pack(fill='x')

        # ── 버튼 영역 ──
        btn_frame = tk.Frame(self.dialog, bg=C_BG2, pady=8)
        btn_frame.pack(fill='x')

        self._dl_btn = tk.Button(
            btn_frame,
            text='⬇  이 양식 다운로드 (Excel)',
            font=('Malgun Gothic', 10, 'bold'),
            bg=C_BTN_DL, fg=C_FG,
            activebackground=tc('info'),
            relief='flat', bd=0,
            padx=20, pady=8,
            cursor='hand2',
            command=self._on_download,
        )
        self._dl_btn.pack(side='left', padx=16)

        tk.Button(
            btn_frame,
            text='✕  닫기',
            font=('Malgun Gothic', 10),
            bg=C_BTN_CLOSE, fg=C_FG,
            activebackground=tc('info'),
            relief='flat', bd=0,
            padx=20, pady=8,
            cursor='hand2',
            command=self.dialog.destroy,
        ).pack(side='right', padx=16)

    # ──────────────────────────────────────────
    # 탭 전환
    # ──────────────────────────────────────────

    def _on_tab_click(self, idx: int):
        self._current_tab = idx
        self._load_tab(idx)

    def _update_tab_styles(self, active_idx: int):
        for i, btn in enumerate(self._tab_btns):
            if i == active_idx:
                btn.configure(bg=C_SELECT, fg=C_FG, relief='solid', bd=1)
            else:
                btn.configure(bg=C_BG2, fg=C_FG2, relief='flat', bd=0)

    def _load_tab(self, idx: int):
        """탭 idx의 데이터를 Treeview에 로드"""
        if idx < 0 or idx >= len(self.FORMATS):
            return
        self._update_tab_styles(idx)
        fmt = self.FORMATS[idx]
        self._desc_var.set(fmt['description'])

        # 캐시 없으면 로드
        if self._loaded_data[idx] is None:
            self._loaded_data[idx] = self._get_data(idx)

        columns, rows = self._loaded_data[idx]
        self._populate_tree(columns, rows)

    def _get_data(self, idx: int) -> Tuple[List[str], List[tuple]]:
        """
        우선순위:
        1. resources/templates/ 파일 (방법 B)
        2. 내장 샘플 데이터 (방법 C fallback)
        """
        fmt = self.FORMATS[idx]
        filepath = _find_template_file(fmt['template_file'])
        if filepath:
            cols, rows = _load_excel_preview(filepath, max_rows=50)
            if cols and rows:
                logger.info(f"[AllocationTemplateDialog] 파일 로드: {filepath}")
                return cols, rows
            logger.warning("[AllocationTemplateDialog] 파일 로드 실패, 내장 데이터 사용")

        return list(fmt['columns']), list(fmt['rows'])

    def _rebuild_tabs(self):
        """외부 양식 import 후 탭 버튼을 다시 만든다."""
        self._loaded_data = [None] * len(self.FORMATS)
        self._tab_btns.clear()
        # 첫 번째 child가 tab_frame이다.
        tab_frame = self.dialog.winfo_children()[0]
        for child in tab_frame.winfo_children():
            child.destroy()

        for i, fmt in enumerate(self.FORMATS):
            btn = tk.Button(
                tab_frame,
                text=fmt['tab_label'],
                font=('Malgun Gothic', 10, 'bold'),
                bg=C_BG2, fg=C_FG2,
                activebackground=C_SELECT,
                activeforeground=C_FG,
                relief='flat', bd=0,
                padx=18, pady=6,
                cursor='hand2',
                command=lambda idx=i: self._on_tab_click(idx),
            )
            btn.pack(side='left', padx=4)
            self._tab_btns.append(btn)

        tk.Button(
            tab_frame,
            text='＋ 양식 가져오기',
            font=('Malgun Gothic', 10, 'bold'),
            bg=C_BTN_DL, fg=C_FG,
            activebackground=tc('info'),
            relief='flat', bd=0,
            padx=14, pady=6,
            cursor='hand2',
            command=self._on_import_template,
        ).pack(side='right', padx=8)

    def _on_import_template(self):
        """고객 Allocation xlsx를 외부 템플릿으로 가져오고 JSON 매핑을 저장한다."""
        src = filedialog.askopenfilename(
            parent=self.dialog,
            title='Allocation 양식 파일 선택',
            filetypes=[('Excel 파일', '*.xlsx *.xls'), ('모든 파일', '*.*')],
        )
        if not src:
            return
        src_path = Path(src)
        if src_path.name.startswith('~$'):
            messagebox.showwarning(
                '임시 파일 제외',
                '~$ 로 시작하는 Excel 임시 잠금 파일은 양식으로 가져올 수 없습니다.\n'
                '원본 .xlsx 파일을 선택하세요.',
                parent=self.dialog,
            )
            return

        default_label = src_path.stem.replace('_', ' ')
        label = simpledialog.askstring(
            '양식 이름',
            '탭에 표시할 양식 이름을 입력하세요.',
            initialvalue=default_label,
            parent=self.dialog,
        )
        if not label:
            return

        try:
            mapping = _mapping_from_excel(src_path, label)
            base = _allocation_template_dir()
            template_name = mapping['template_file']
            json_name = f"{mapping['id']}.json"
            dst_xlsx = base / template_name
            dst_json = base / json_name
            if dst_xlsx.exists() or dst_json.exists():
                if not messagebox.askyesno(
                    '덮어쓰기 확인',
                    f'이미 같은 이름의 양식이 있습니다.\n\n{template_name}\n{json_name}\n\n덮어쓸까요?',
                    parent=self.dialog,
                ):
                    return

            shutil.copy2(str(src_path), str(dst_xlsx))
            dst_json.write_text(
                json.dumps(mapping, ensure_ascii=False, indent=2),
                encoding='utf-8',
            )

            external = _load_external_formats()
            builtin = [f for f in self.__class__.FORMATS]
            self.FORMATS = external + builtin
            self._current_tab = next(
                (i for i, f in enumerate(self.FORMATS)
                 if str(f.get('mapping_file', '')).endswith(json_name)),
                0,
            )
            self._rebuild_tabs()
            self._load_tab(self._current_tab)
            messagebox.showinfo(
                '가져오기 완료',
                f'Allocation 양식을 저장했습니다.\n\n{dst_xlsx}\n{dst_json}\n\n다음 실행 시에도 목록에 표시됩니다.',
                parent=self.dialog,
            )
        except Exception as e:
            logger.error("[AllocationTemplateDialog] 양식 가져오기 실패: %s", e, exc_info=True)
            messagebox.showerror('가져오기 실패', f'양식 가져오기 실패:\n{e}', parent=self.dialog)

    # ──────────────────────────────────────────
    # Treeview 채우기
    # ──────────────────────────────────────────

    def _populate_tree(self, columns: List[str], rows: List[tuple]):
        """Treeview 컬럼/행 전체 재구성"""
        # 초기화
        self._tree.delete(*self._tree.get_children())
        self._tree['columns'] = columns
        self._tree['show'] = 'headings'

        # 컬럼 헤더 설정
        col_widths = {
            'Product': 130, 'SAP NO': 100, 'Date in stock': 100,
            'QTY (MT)': 70, 'Lot No': 110, 'WH': 40,
            'Customs': 80, 'Export': 60, 'SOLD TO': 200,
            'SALE REF': 70, 'Balance': 70, 'GW': 65, 'Remark': 80,
        }
        for col in columns:
            w = col_widths.get(col, 90)
            self._tree.heading(col, text=col, anchor='center')
            self._tree.column(col, width=w, minwidth=40, anchor='center', stretch=True)

        # QTY 컬럼 인덱스 찾기 (샘플 판정용)
        qty_idx = next((i for i, c in enumerate(columns)
                        if 'QTY' in c.upper() and 'BALANCE' not in c.upper()), None)
        product_idx = next((i for i, c in enumerate(columns)
                            if 'PRODUCT' in c.upper()), None)

        # 행 삽입
        for row_num, row in enumerate(rows):
            # 샘플 판정
            qty_str = str(row[qty_idx]) if qty_idx is not None and qty_idx < len(row) else '0'
            prod_str = str(row[product_idx]) if product_idx is not None and product_idx < len(row) else ''
            is_sample = _is_sample_row(prod_str, qty_str)

            if is_sample:
                tag = 'sample' if row_num % 2 == 0 else 'sample_stripe'
            else:
                tag = 'normal' if row_num % 2 == 0 else 'stripe'

            self._tree.insert('', 'end', values=row, tags=(tag,))

        # 행 수 표시
        total = len(rows)
        sample_cnt = sum(
            1 for r in rows
            if _is_sample_row(
                str(r[product_idx]) if product_idx is not None and product_idx < len(r) else '',
                str(r[qty_idx]) if qty_idx is not None and qty_idx < len(r) else '0',
            )
        )
        normal_cnt = total - sample_cnt
        self._row_count_var.set(
            f'총 {total}행 표시 중  │  본 제품 {normal_cnt}행  │  샘플 {sample_cnt}행  '
            f'│  (실제 파일에는 더 많은 행이 있을 수 있습니다)'
        )

    # ──────────────────────────────────────────
    # 다운로드
    # ──────────────────────────────────────────

    def _on_download(self):
        """현재 탭 양식을 Excel로 저장"""
        fmt = self.FORMATS[self._current_tab]

        # 기존 파일이 있으면 복사, 없으면 내장 데이터로 생성
        filepath = _find_template_file(fmt['template_file'])
        if filepath:
            self._download_copy(filepath)
        else:
            self._download_builtin(fmt)

    def _download_copy(self, src_path: Path):
        """기존 Excel 파일을 복사해서 저장"""
        from tkinter import filedialog
        import shutil
        save_path = filedialog.asksaveasfilename(
            parent=self.dialog,
            title='양식 저장',
            defaultextension='.xlsx',
            filetypes=[('Excel 파일', '*.xlsx')],
            initialfile=src_path.name,
        )
        if not save_path:
            return
        try:
            shutil.copy2(str(src_path), save_path)
            messagebox.showinfo('저장 완료',
                                f'양식 파일을 저장했습니다:\n{save_path}',
                                parent=self.dialog)
        except Exception as e:
            messagebox.showerror('저장 오류', f'저장 실패: {e}', parent=self.dialog)

    def _download_builtin(self, fmt: dict):
        """내장 데이터로 Excel 생성 후 저장"""
        from tkinter import filedialog
        save_path = filedialog.asksaveasfilename(
            parent=self.dialog,
            title='양식 저장',
            defaultextension='.xlsx',
            filetypes=[('Excel 파일', '*.xlsx')],
            initialfile=fmt['template_file'],
        )
        if not save_path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Allocation'

            # 타이틀 행
            ws.cell(row=1, column=1,
                    value=f"Allocation Template - {fmt['tab_label']}")
            ws.cell(row=1, column=1).font = Font(bold=True, size=11)

            # 헤더
            header_fill = PatternFill('solid', fgColor='1a237e')
            header_font = Font(bold=True, color='ffffff')
            for col_idx, col_name in enumerate(fmt['columns'], start=1):
                cell = ws.cell(row=2, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # 데이터 행
            sample_fill = PatternFill('solid', fgColor='3e2a00')
            normal_fill = PatternFill('solid', fgColor='283593')
            stripe_fill = PatternFill('solid', fgColor='303f9f')

            qty_idx = next((i for i, c in enumerate(fmt['columns'])
                            if 'QTY' in c.upper() and 'BALANCE' not in c.upper()), None)
            product_idx = next((i for i, c in enumerate(fmt['columns'])
                                if 'PRODUCT' in c.upper()), None)

            for row_num, row in enumerate(fmt['rows'], start=3):
                prod_str = str(row[product_idx]) if product_idx is not None else ''
                qty_str = str(row[qty_idx]) if qty_idx is not None else '0'
                is_sample = _is_sample_row(prod_str, qty_str)

                for col_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    if is_sample:
                        cell.fill = sample_fill
                        cell.font = Font(color='ff9100')
                    else:
                        fill = normal_fill if (row_num % 2 == 0) else stripe_fill
                        cell.fill = fill
                        cell.font = Font(color='ffffff')
                    cell.alignment = Alignment(horizontal='center')

            # 컬럼 너비
            for col_idx, col_name in enumerate(fmt['columns'], start=1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = max(12, len(col_name) + 2)

            try:
                from utils.sqm_excel_alignment import apply_sqm_workbook_alignment
                apply_sqm_workbook_alignment(wb)
            except Exception:
                pass
            wb.save(save_path)
            messagebox.showinfo('저장 완료',
                                f'양식 파일을 생성했습니다:\n{save_path}',
                                parent=self.dialog)
        except ImportError:
            messagebox.showerror('오류',
                                 'openpyxl 라이브러리가 없습니다.\npip install openpyxl',
                                 parent=self.dialog)
        except Exception as e:
            messagebox.showerror('저장 오류', f'생성 실패: {e}', parent=self.dialog)
