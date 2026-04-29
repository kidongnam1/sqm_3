"""
SQM v6.12.1 — 반품 사유 통계 리포트 다이얼로그 v2
====================================================
기간 필터 (DateEntry/텍스트) + 4탭 통계 + 추이 차트 + Excel 내보내기
"""

from gui_app_modular.utils.ui_constants import create_themed_toplevel  # v8.0.9
from gui_app_modular.utils.ui_constants import is_dark  # v8.0.9
from gui_app_modular.utils.ui_constants import tc  # v8.1.3: top-level import (gui_bootstrap 종속 제거)
import logging
import tkinter as tk
from datetime import date, timedelta
from tkinter import ttk
from tkinter.constants import BOTH, END, LEFT, RIGHT, VERTICAL, X, Y

logger = logging.getLogger(__name__)

try:
    from gui_app_modular.utils.gui_bootstrap import (
        HAS_DATEENTRY,
        DateEntry,
    )
    from gui_app_modular.utils.ui_constants import (
        DialogSize, apply_modal_window_options, center_dialog,
        setup_dialog_geometry_persistence,
    )
except ImportError:
    DialogSize = center_dialog = apply_modal_window_options = None
    DateEntry = None
    HAS_DATEENTRY = False
    setup_dialog_geometry_persistence = None

try:
    from gui_app_modular.utils.custom_messagebox import CustomMessageBox
except ImportError:
    CustomMessageBox = None

try:
    from gui_app_modular.utils.theme_colors import ThemeColors
except ImportError:
    ThemeColors = None


# 반품 사유 표준 코드
RETURN_REASON_CODES = [
    '품질 불량', '수량 오류', '고객 취소', '배송 문제',
    '파손/변질', '규격 불일치', '기타',
]


class ReturnStatisticsDialog:
    """반품 사유 통계 리포트 — 기간 필터 + 동적 새로고침."""

    def __init__(self, parent, engine, current_theme: str = 'darkly'):
        self.parent = parent
        self.engine = engine
        self.stats = {}

        dark_mode = is_dark()
        self.bg = tc('bg_card')
        self.fg = tc('text_primary')

        popup = create_themed_toplevel(parent)
        popup.title("📊 반품 사유 통계")
        popup.transient(parent)
        popup.resizable(True, True)  # v9.0: 크기 조절 허용
        popup.minsize(400, 300)  # v9.0: 최소 크기
        popup.grab_set()
        popup.resizable(True, True)  # v9.0: 크기 조절 허용
        popup.minsize(400, 300)  # v9.0: 최소 크기
        if setup_dialog_geometry_persistence:
            setup_dialog_geometry_persistence(popup, "return_statistics_dialog", parent, "large")
        elif DialogSize:
            popup.geometry(DialogSize.get_geometry(parent, 'large'))
            if apply_modal_window_options:
                apply_modal_window_options(popup)
            if center_dialog:
                center_dialog(popup, parent)
        else:
            popup.geometry("900x600")
        popup.configure(bg=self.bg)
        self.popup = popup

        # ═══ 기간 필터 바 ═══
        filter_frame = tk.Frame(popup, bg=tc('bg_secondary'), padx=10, pady=8)
        filter_frame.pack(fill=X)

        tk.Label(filter_frame, text="📅 기간:", font=('맑은 고딕', 10, 'bold'),
                 bg=tc('bg_secondary'), fg=tc('text_primary')).pack(side=LEFT, padx=(0, 5))

        today = date.today()
        default_start = (today - timedelta(days=90)).strftime('%Y-%m-%d')
        default_end = today.strftime('%Y-%m-%d')

        # v8.6.3: 공통 날짜 범위 바 통일 적용
        try:
            from gui_app_modular.utils.tree_enhancements import make_date_range_bar
            _db, _sv_s, _sv_e = make_date_range_bar(
                filter_frame, self._on_refresh,
                default_from=default_start, default_to=default_end)
            _db.pack(side=LEFT, padx=2)
            # date_start/date_end 호환 래퍼
            class _VarEntry:
                def __init__(self, sv): self._sv = sv
                def get(self): return self._sv.get()
            self.date_start = _VarEntry(_sv_s)
            self.date_end   = _VarEntry(_sv_e)
        except Exception as _e:
            logging.getLogger(__name__).warning(f"[반품통계] make_date_range_bar: {_e}")
            self.date_start = ttk.Entry(filter_frame, width=12)
            self.date_start.pack(side=LEFT, padx=2)
            self.date_start.insert(0, default_start)
            tk.Label(filter_frame, text="~").pack(side=LEFT, padx=3)
            self.date_end = ttk.Entry(filter_frame, width=12)
            self.date_end.pack(side=LEFT, padx=2)
            self.date_end.insert(0, default_end)

        ttk.Button(filter_frame, text="🔍 조회",
                   command=self._on_search).pack(side=LEFT, padx=8)

        for label, days in [("1개월", 30), ("3개월", 90), ("6개월", 180), ("전체", 0)]:
            ttk.Button(filter_frame, text=label,
                       command=lambda d=days: self._quick_range(d)).pack(side=LEFT, padx=2)

        # ═══ 헤더 (동적) ═══
        self.header_frame = tk.Frame(popup, bg=tc('bg_secondary'), padx=15, pady=8)
        self.header_frame.pack(fill=X)
        self.header_title = tk.Label(self.header_frame, text="📊 반품 사유 통계 리포트",
                                     font=('맑은 고딕', 13, 'bold'), bg=tc('bg_secondary'), fg=tc('text_primary'))
        self.header_title.pack(anchor='w')
        self.header_summary = tk.Label(self.header_frame, text="",
                                        font=('맑은 고딕', 10), bg=tc('bg_secondary'), fg=tc('text_primary'))
        self.header_summary.pack(anchor='w', pady=(2, 0))

        # ═══ Notebook ═══
        self.nb = ttk.Notebook(popup)
        self.nb.pack(fill=BOTH, expand=True, padx=10, pady=5)

        self._build_tab1_reason()
        self._build_tab2_lot()
        self._build_tab3_monthly()
        self._build_tab4_customer()

        # ═══ 하단 ═══
        btn_bar = tk.Frame(popup, bg=self.bg, pady=8)
        btn_bar.pack(fill=X, padx=10)
        ttk.Button(btn_bar, text="닫기", command=popup.destroy).pack(side=RIGHT)
        ttk.Button(btn_bar, text="📥 Excel 내보내기",
                   command=self._on_export).pack(side=RIGHT, padx=5)
        ttk.Button(btn_bar, text="📄 PDF 리포트",
                   command=self._on_export_pdf).pack(side=RIGHT, padx=5)

        self._on_search()
        popup.wait_window()

    def _build_tab1_reason(self):
        tab = tk.Frame(self.nb, bg=self.bg)
        self.nb.add(tab, text="  📋 사유별  ")
        cols = ('no', 'reason', 'count', 'weight_kg', 'pct')
        self.tree1 = ttk.Treeview(tab, columns=cols, show='headings', height=12)
        for c, h, w, a in [('no', '#', 35, 'center'), ('reason', '반품 사유', 200, 'w'),
                            ('count', '건수', 70, 'e'), ('weight_kg', '중량(kg)', 90, 'e'),
                            ('pct', '비율', 60, 'e')]:
            self.tree1.heading(c, text=h, anchor='center')
            self.tree1.column(c, width=w, anchor=a)
        frame = ttk.Frame(tab)
        frame.pack(fill=BOTH, expand=True)
        sb1_y = tk.Scrollbar(frame, orient=VERTICAL, command=self.tree1.yview)
        sb1_x = tk.Scrollbar(frame, orient='horizontal', command=self.tree1.xview)
        self.tree1.configure(yscrollcommand=sb1_y.set, xscrollcommand=sb1_x.set)
        self.tree1.pack(side=LEFT, fill=BOTH, expand=True)
        sb1_y.pack(side=RIGHT, fill=Y)
        # v8.1.9: TreeviewTotalFooter
        try:
            from gui_app_modular.utils.tree_enhancements import TreeviewTotalFooter as _TTF
            _footer = _TTF(
                tab, self.tree1,
                summable_column_ids=[],
                column_display_names={},
                column_formats={},
            )
            _footer.pack(fill='x')
            self._footer_tree1 = _footer
        except Exception as e:
            logger.warning(f'[UI] return_statistics_dialog: {e}')
        sb1_x.pack(side=tk.BOTTOM, fill=X)

    def _build_tab2_lot(self):
        tab = tk.Frame(self.nb, bg=self.bg)
        self.nb.add(tab, text="  📦 LOT별  ")
        cols = ('no', 'lot_no', 'count', 'weight_kg', 'reasons')
        self.tree2 = ttk.Treeview(tab, columns=cols, show='headings', height=12)
        for c, h, w, a in [('no', '#', 35, 'center'), ('lot_no', 'LOT NO', 100, 'center'),
                            ('count', '반품 건수', 70, 'e'), ('weight_kg', '중량(kg)', 90, 'e'),
                            ('reasons', '사유', 250, 'w')]:
            self.tree2.heading(c, text=h, anchor='center')
            self.tree2.column(c, width=w, anchor=a)
        frame = ttk.Frame(tab)
        frame.pack(fill=BOTH, expand=True)
        sb2_y = tk.Scrollbar(frame, orient=VERTICAL, command=self.tree2.yview)
        sb2_x = tk.Scrollbar(frame, orient='horizontal', command=self.tree2.xview)
        self.tree2.configure(yscrollcommand=sb2_y.set, xscrollcommand=sb2_x.set)
        self.tree2.pack(side=LEFT, fill=BOTH, expand=True)
        sb2_y.pack(side=RIGHT, fill=Y)
        sb2_x.pack(side=tk.BOTTOM, fill=X)

    def _build_tab3_monthly(self):
        tab = tk.Frame(self.nb, bg=self.bg)
        self.nb.add(tab, text="  📅 월별 추이  ")
        cols = ('month', 'count', 'weight_kg')
        self.tree3 = ttk.Treeview(tab, columns=cols, show='headings', height=8)
        for c, h, w, a in [('month', '월', 100, 'center'), ('count', '건수', 80, 'e'),
                            ('weight_kg', '중량(kg)', 100, 'e')]:
            self.tree3.heading(c, text=h, anchor='center')
            self.tree3.column(c, width=w, anchor=a)
        frame = ttk.Frame(tab)
        frame.pack(fill=X, padx=5, pady=(5, 0))
        sb3_y = tk.Scrollbar(frame, orient=VERTICAL, command=self.tree3.yview)
        sb3_x = tk.Scrollbar(frame, orient='horizontal', command=self.tree3.xview)
        self.tree3.configure(yscrollcommand=sb3_y.set, xscrollcommand=sb3_x.set)
        self.tree3.pack(side=LEFT, fill=X, expand=True)
        sb3_y.pack(side=RIGHT, fill=Y)
        sb3_x.pack(side=tk.BOTTOM, fill=X)
        self.chart_canvas = tk.Canvas(tab, bg=tc('bg_card'), height=140, highlightthickness=0)
        self.chart_canvas.pack(fill=BOTH, expand=True, padx=5, pady=5)

    def _build_tab4_customer(self):
        tab = tk.Frame(self.nb, bg=self.bg)
        self.nb.add(tab, text="  👤 고객별  ")
        cols = ('no', 'customer', 'count')
        self.tree4 = ttk.Treeview(tab, columns=cols, show='headings', height=12)
        for c, h, w, a in [('no', '#', 35, 'center'), ('customer', '고객', 200, 'w'),
                            ('count', '반품 건수', 80, 'e')]:
            self.tree4.heading(c, text=h, anchor='center')
            self.tree4.column(c, width=w, anchor=a)
        frame = ttk.Frame(tab)
        frame.pack(fill=BOTH, expand=True)
        sb4_y = tk.Scrollbar(frame, orient=VERTICAL, command=self.tree4.yview)
        sb4_x = tk.Scrollbar(frame, orient='horizontal', command=self.tree4.xview)
        self.tree4.configure(yscrollcommand=sb4_y.set, xscrollcommand=sb4_x.set)
        self.tree4.pack(side=LEFT, fill=BOTH, expand=True)
        sb4_y.pack(side=RIGHT, fill=Y)
        sb4_x.pack(side=tk.BOTTOM, fill=X)

    # ─── 기간 헬퍼 ───

    def _get_date_str(self, widget) -> str:
        try:
            if HAS_DATEENTRY and hasattr(widget, 'entry'):
                return widget.entry.get().strip()
            return widget.get().strip()
        except Exception:
            return ''

    def _set_date_str(self, widget, val: str):
        try:
            if HAS_DATEENTRY and hasattr(widget, 'entry'):
                widget.entry.delete(0, END)
                widget.entry.insert(0, val)
            else:
                widget.delete(0, END)
                widget.insert(0, val)
        except Exception as _we:
            logging.getLogger(__name__).debug(f"[반품통계] 위젯 값 설정 실패: {_we}")

    def _quick_range(self, days: int):
        today = date.today()
        if days == 0:
            self._set_date_str(self.date_start, '2020-01-01')
        else:
            self._set_date_str(self.date_start, (today - timedelta(days=days)).strftime('%Y-%m-%d'))
        self._set_date_str(self.date_end, today.strftime('%Y-%m-%d'))
        self._on_search()

    # ─── 조회 ───

    def _on_search(self):
        start = self._get_date_str(self.date_start)
        end = self._get_date_str(self.date_end)
        # v8.1.8: parse_date_range — 빈 문자열이면 전체 기간
        try:
            from gui_app_modular.utils.tree_enhancements import parse_date_range as _pdr
            _s, _e = _pdr(start or '', end or '')
            start = _s or ''
            end   = _e or ''
        except Exception as e:
            logger.warning(f'[UI] return_statistics_dialog: {e}')
        if hasattr(self.engine, 'get_return_statistics'):
            self.stats = self.engine.get_return_statistics(start_date=start, end_date=end)
        else:
            self.stats = {}
        self._refresh_all()

    def _refresh_all(self):
        stats = self.stats
        self.header_summary.config(
            text=f"전체 반품: {stats.get('total_count', 0):,}건 | "
                 f"총 중량: {stats.get('total_weight_kg', 0):,.0f} kg | "
                 f"기간: {self._get_date_str(self.date_start)} ~ {self._get_date_str(self.date_end)}"
        )
        total_cnt = stats.get('total_count', 1) or 1

        self.tree1.delete(*self.tree1.get_children())
        for idx, r in enumerate(stats.get('by_reason', []), 1):
            pct = r['count'] / total_cnt * 100
            self.tree1.insert('', END, values=(idx, r['reason'], f"{r['count']:,}",
                                                f"{r['weight_kg']:,.0f}", f"{pct:.1f}%"))

        self.tree2.delete(*self.tree2.get_children())
        for idx, r in enumerate(stats.get('by_lot', []), 1):
            self.tree2.insert('', END, values=(idx, r['lot_no'], f"{r['count']:,}",
                                                f"{r['weight_kg']:,.0f}", r['reasons']))

        self.tree3.delete(*self.tree3.get_children())
        for r in stats.get('by_month', []):
            self.tree3.insert('', END, values=(r['month'], f"{r['count']:,}", f"{r['weight_kg']:,.0f}"))
        self._draw_trend_chart()

        self.tree4.delete(*self.tree4.get_children())
        for idx, r in enumerate(stats.get('top_customers', []), 1):
            self.tree4.insert('', END, values=(idx, r['customer'], f"{r['count']:,}"))

    # ─── 추이 차트 ───

    def _draw_trend_chart(self):
        """월별 반품 건수 바+꺾은선 차트."""
        c = self.chart_canvas
        c.delete('all')
        c.update_idletasks()
        w = c.winfo_width() or 500
        h = c.winfo_height() or 140

        months = self.stats.get('by_month', [])
        if not months:
            c.create_text(w // 2, h // 2, text="데이터 없음", fill='#999', font=('맑은 고딕', 11))
            return

        ml, mr, mt, mb = 50, 20, 15, 30
        cw, ch = w - ml - mr, h - mt - mb
        max_val = max((m['count'] for m in months), default=1) or 1

        # Y축 그리드
        for i in range(5):
            y = mt + (ch * i // 4)
            c.create_line(ml, y, w - mr, y, fill='#eee', dash=(2, 2))
            val = max_val - (max_val * i // 4)
            c.create_text(ml - 5, y, text=str(int(val)), anchor='e', fill='#999', font=('', 8))

        n = len(months)
        bar_w = max(8, min(30, cw // (n * 2)))
        pts = []

        for i, m in enumerate(months):
            x = ml + int((i + 0.5) / n * cw)
            ratio = m['count'] / max_val
            bar_h = int(ratio * ch)
            y_top = mt + ch - bar_h

            c.create_rectangle(x - bar_w // 2, y_top, x + bar_w // 2, mt + ch,
                               fill='#3498DB', outline='#2980B9')
            if bar_h > 15:
                c.create_text(x, y_top - 8, text=str(m['count']), fill='#2C3E50', font=('', 8))

            label = m['month'][-5:] if len(m['month']) > 5 else m['month']
            c.create_text(x, mt + ch + 12, text=label, fill='#666', font=('', 8))
            pts.append((x, y_top))

        if len(pts) >= 2:
            for i in range(len(pts) - 1):
                c.create_line(pts[i][0], pts[i][1], pts[i + 1][0], pts[i + 1][1],
                              fill='#E74C3C', width=2)
            for px, py in pts:
                c.create_oval(px - 3, py - 3, px + 3, py + 3, fill='#E74C3C', outline='white')

        # 범례
        lx = w - mr - 100
        c.create_rectangle(lx, mt, lx + 10, mt + 10, fill='#3498DB', outline='')
        c.create_text(lx + 15, mt + 5, text='건수', anchor='w', fill='#333', font=('', 8))
        c.create_line(lx, mt + 18, lx + 10, mt + 18, fill='#E74C3C', width=2)
        c.create_text(lx + 15, mt + 18, text='추이', anchor='w', fill='#333', font=('', 8))

    # ─── Excel 내보내기 ───

    def _on_export(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            if CustomMessageBox:
                CustomMessageBox.showerror(self.popup, "오류", "openpyxl 라이브러리가 필요합니다.")
            return

        import os
        from datetime import datetime
        from tkinter import filedialog

        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = filedialog.asksaveasfilename(
            parent=self.popup, title="반품 통계 Excel 저장",
            initialdir=desktop, initialfile=f"반품통계_{ts}.xlsx",
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        hf = Font(bold=True, size=11, color='FFFFFF')
        hfill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        ha = Alignment(horizontal='center', vertical='center')
        tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

        def _ws(ws, title, headers, rows):
            ws.title = title
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, tb
            for ri, row_data in enumerate(rows, 2):
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border = tb
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0' if isinstance(val, int) else '#,##0.0'
            for ci in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18

        stats = self.stats
        total_cnt = stats.get('total_count', 1) or 1

        ws1 = wb.active
        _ws(ws1, '사유별', ['#', '반품 사유', '건수', '중량(kg)', '비율'],
            [(i, r['reason'], r['count'], r['weight_kg'], f"{r['count']/total_cnt*100:.1f}%")
             for i, r in enumerate(stats.get('by_reason', []), 1)])

        _ws(wb.create_sheet(), 'LOT별', ['#', 'LOT NO', '반품 건수', '중량(kg)', '사유'],
            [(i, r['lot_no'], r['count'], r['weight_kg'], r['reasons'])
             for i, r in enumerate(stats.get('by_lot', []), 1)])

        _ws(wb.create_sheet(), '월별 추이', ['월', '건수', '중량(kg)'],
            [(r['month'], r['count'], r['weight_kg']) for r in stats.get('by_month', [])])

        _ws(wb.create_sheet(), '고객별', ['#', '고객', '반품 건수'],
            [(i, r['customer'], r['count']) for i, r in enumerate(stats.get('top_customers', []), 1)])

        wb.save(path)
        if CustomMessageBox:
            CustomMessageBox.showinfo(self.popup, "완료", f"반품 통계가 저장되었습니다.\n\n{path}")

    def _on_export_pdf(self):
        """반품 원인 분석 PDF 생성."""
        import os
        from datetime import datetime as _dt
        from tkinter import filedialog

        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        ts = _dt.now().strftime('%Y%m%d_%H%M%S')
        path = filedialog.asksaveasfilename(
            parent=self.popup, title="반품 분석 PDF 저장",
            initialdir=desktop, initialfile=f"반품분석_{ts}.pdf",
            defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")]
        )
        if not path:
            return

        try:
            from features.reports.return_report_pdf import generate_return_report_pdf
            start = self._get_date_str(self.date_start)
            end = self._get_date_str(self.date_end)
            generate_return_report_pdf(self.engine, path,
                                        start_date=start, end_date=end)
            if CustomMessageBox:
                CustomMessageBox.showinfo(self.popup, "완료",
                    f"반품 분석 PDF가 저장되었습니다.\n\n{path}")
        except ImportError:
            if CustomMessageBox:
                CustomMessageBox.showerror(self.popup, "오류",
                    "reportlab 라이브러리가 필요합니다.\npip install reportlab")
        except Exception as e:
            if CustomMessageBox:
                CustomMessageBox.showerror(self.popup, "오류", f"PDF 생성 오류:\n{e}")
