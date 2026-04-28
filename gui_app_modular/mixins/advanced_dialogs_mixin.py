"""
SQM 재고관리 - 고급 다이얼로그 Mixin
======================================
v3.8.4 - advanced_features_mixin에서 분리

기능:
- 반품 처리 다이얼로그
- 수동 입고 다이얼로그
- 문서 변환 다이얼로그
- 출고 이력 조회
- 스냅샷 차트
- 출고 인보이스 생성
"""

# ══════════════════════════════════════════════════════════════
# 🎨 색상 사용 원칙 (v3.8.0 — 절대 준수)
# ══════════════════════════════════════════════════════════════
# ✅ 올바른 방법: tc() 함수 사용 (라이트/다크 자동 전환)
#     from gui_app_modular.utils.ui_constants import tc
#     label.config(fg=tc('text_primary'), bg=tc('bg_primary'))
#     entry.config(fg=tc('text_primary'), bg=tc('bg_entry'))
#     frame.config(bg=tc('bg_secondary'))
#
# ❌ 금지: 하드코딩 색상 사용
#     label.config(fg=tc('text_primary'))         # 다크 배경에서 안 보임
#     label.config(fg=tc('text_primary'))         # 라이트 배경에서 안 보임
#     label.config(bg=tc('bg_primary'))       # 라이트 모드에서 검은 배경
#
# 📋 주요 tc() 키:
#     text_primary   — 일반 텍스트 (라이트: #2c3e50  다크: #FF8C00)
#     text_secondary — 보조 텍스트
#     text_muted     — 비활성 텍스트 (gray 대체)
#     bg_primary     — 기본 배경
#     bg_secondary   — 보조 배경
#     bg_card        — 카드/패널 배경
#     bg_entry       — 입력창 배경
#     success        — 성공 (녹색 계열)
#     warning        — 경고 (주황 계열)
#     danger         — 오류 (빨간 계열)
#     info           — 정보 (파란 계열)
# ══════════════════════════════════════════════════════════════

from gui_app_modular.utils.ui_constants import create_themed_toplevel  # v8.0.9
from gui_app_modular.utils.ui_constants import tc
import logging
import os
import sqlite3

from ..utils.ui_constants import CustomMessageBox, apply_modal_window_options

# v6.8.6: STATUS 상수 전역 import (L599 undefined name 수정)
try:
    from engine_modules.constants import (
        STATUS_AVAILABLE, STATUS_RESERVED, STATUS_PICKED,
        STATUS_SOLD, STATUS_DEPLETED, STATUS_PARTIAL,
    )
except ImportError:
    STATUS_AVAILABLE  = 'AVAILABLE'
    STATUS_RESERVED   = 'RESERVED'
    STATUS_PICKED     = 'PICKED'
    STATUS_SOLD       = 'SOLD'
    STATUS_DEPLETED   = 'DEPLETED'

logger = logging.getLogger(__name__)


class AdvancedDialogsMixin:
    """고급 다이얼로그 Mixin"""

    def _on_return_inbound_upload(self) -> None:
        """v6.0: 반품 입고 — [템플릿 열기(붙여넣기)] vs [파일 업로드] 선택 후 파싱·DB 반영 (입고 형식과 동일)."""
        choice = self._show_template_or_upload_choice("반품 입고", "return_inbound")
        if choice is None:
            return
        if choice == "template":
            self._show_return_inbound_spreadsheet_dialog()
            return
        from ..utils.constants import filedialog
        path = filedialog.askopenfilename(
            parent=self.root,
            title="반품 입고 Excel 선택",
            filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")],
        )
        if not path or not path.strip():
            return
        self._apply_return_inbound_after_parse(path, source_file=path)

    def _on_return_inbound_paste_confirm(self, rows: list) -> None:
        """반품 입고 붙여넣기 확인 시 파싱 후 DB 반영."""
        from features.parsers.return_inbound_parser import (
            parse_return_inbound_from_rows,
        )
        parse_result = parse_return_inbound_from_rows(rows)
        if not parse_result.get("parse_ok") or parse_result.get("errors"):
            err_msg = "\n".join(parse_result.get("errors", ["파싱 실패"]))
            CustomMessageBox.showerror(
                self.root,
                "반품 입고 — 검증 실패",
                "필수 항목(LOT NO, NET(Kg), PICKING NO, 반품사유) 누락 또는 오류.\n\n" + err_msg,
            )
            return
        n = len(parse_result.get("items", []))
        if n == 0:
            CustomMessageBox.showwarning(self.root, "반품 입고", "처리할 행이 없습니다.")
            return
        if not CustomMessageBox.askyesno(
            self.root,
            "반품 입고 확인",
            f"총 {n}건을 반품 입고 처리합니다.\nPICKING NO 매칭 실패 시 전체가 중단됩니다.\n계속하시겠습니까?",
        ):
            return
        self._apply_return_inbound_after_parse(parse_result, source_file="(붙여넣기)", skip_confirm=True)

    def _apply_return_inbound_after_parse(
        self, parse_result_or_path, source_file: str = "", skip_confirm: bool = False
    ) -> None:
        """반품 파싱 결과 또는 Excel 경로를 받아 검증 후 DB 반영. skip_confirm=True면 확인 대화상자 생략(붙여넣기에서 이미 표시)."""
        if isinstance(parse_result_or_path, str):
            path = parse_result_or_path
            try:
                from features.parsers.return_inbound_parser import (
                    parse_return_inbound_excel,
                )
                parse_result = parse_return_inbound_excel(path)
            except ImportError:
                CustomMessageBox.showerror(
                    self.root,
                    "반품 입고",
                    "features.parsers.return_inbound_parser를 불러올 수 없습니다.",
                )
                return
            source_file = path
        else:
            parse_result = parse_result_or_path
        if not parse_result.get("parse_ok") or parse_result.get("errors"):
            err_msg = "\n".join(parse_result.get("errors", ["파싱 실패"]))
            CustomMessageBox.showerror(
                self.root,
                "반품 입고 — 검증 실패",
                "필수 항목(LOT NO, WEIGHT(MT) 또는 NET(Kg), PICKING NO, REASON) 누락 또는 오류.\n\n" + err_msg,
            )
            return
        n = len(parse_result.get("items", []))
        if n == 0:
            CustomMessageBox.showwarning(self.root, "반품 입고", "처리할 행이 없습니다.")
            return
        if not skip_confirm:
            # v6.12.1: 미리보기 편집 다이얼로그
            _confirmed_items = [None]

            def _on_preview_confirm(edited_items):
                _confirmed_items[0] = edited_items

            try:
                from ..dialogs.return_inbound_preview import ReturnInboundPreviewDialog
                current_theme = getattr(self, '_current_theme', 'darkly')
                ReturnInboundPreviewDialog(
                    self.root,
                    parse_result.get('items', []),
                    on_confirm=_on_preview_confirm,
                    current_theme=current_theme,
                )
            except ImportError:
                # fallback: 기존 텍스트 확인
                if not CustomMessageBox.askyesno(
                    self.root,
                    "반품 입고 확인",
                    f"총 {n}건을 반품 입고 처리합니다.\nPICKING NO 매칭 실패 시 전체가 중단됩니다.\n계속하시겠습니까?",
                ):
                    return
                _confirmed_items[0] = parse_result.get('items', [])

            if _confirmed_items[0] is None:
                return  # 취소
            parse_result['items'] = _confirmed_items[0]  # 편집된 데이터로 교체

            # v6.12.2: 대량 반품 관리자 확인
            try:
                from engine_modules.constants import RETURN_AUTO_APPROVE_MAX_TONBAGS  # STATUS_* 상단 import 사용
                _max_auto = RETURN_AUTO_APPROVE_MAX_TONBAGS
            except ImportError:
                _max_auto = 5
            _ = len(parse_result.get('items', []))  # total_items: 미사용
            total_tb = sum(it.get('tonbag_count', 1) for it in parse_result.get('items', []))
            if total_tb > _max_auto:
                if not CustomMessageBox.askyesno(
                    self.root, "⚠️ 대량 반품 관리자 확인",
                    f"반품 톤백 {total_tb}개 (자동승인 기준 {_max_auto}개 초과)\n\n"
                    f"대량 반품은 되돌릴 수 없습니다.\n정말 진행하시겠습니까?"
                ):
                    return
        try:
            from features.parsers.return_inbound_engine import (
                apply_return_inbound_to_db,
            )
            r = apply_return_inbound_to_db(self.engine, parse_result, source_file)
            self._log(f"✅ 반품 입고 완료: {r.get('returned', 0)}건")
            CustomMessageBox.showinfo(
                self.root,
                "반품 입고 완료",
                f"반품 입고 처리 완료\n\n반영: {r.get('returned', 0)}건",
            )
            self._safe_refresh()
            if hasattr(self, "_refresh_dashboard"):
                self._refresh_dashboard()
        except RuntimeError as e:
            CustomMessageBox.showerror(self.root, "반품 입고 오류", str(e))
        except Exception as e:
            logger.exception("반품 입고 처리 중 오류")
            CustomMessageBox.showerror(self.root, "반품 입고 오류", str(e))

    def _show_return_dialog(self, initial_tab: int = 0) -> None:
        """v4.1.4: 반품 처리 다이얼로그 — initial_tab: 0=소량(단건), 1=다량(Excel)."""
        from ..utils.constants import BOTH, tk, ttk

        dialog = create_themed_toplevel(self.root)
        dialog.title("🔄 반품 처리")
        try:
            from gui_app_modular.utils.ui_constants import setup_dialog_geometry_persistence as _sgp
            _sgp(dialog, "adv_return_dialog", self.root, "large")
        except Exception as e:
            logger.warning(f'[UI] advanced_dialogs_mixin: {e}')
        dialog.geometry("780x650")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="반품 처리", font=('맑은 고딕', 18, 'bold')).pack(pady=8)

        nb = ttk.Notebook(dialog)
        nb.pack(fill=BOTH, expand=True, padx=10, pady=5)

        # TAB 0: 단건 반품 (소량)
        self._build_return_single_tab(nb, dialog)
        # TAB 1: Excel 일괄 반품 (다량)
        self._build_return_excel_tab(nb, dialog)

        try:
            nb.select(initial_tab)
        except (tk.TclError, TypeError, IndexError) as _e:
            logger.debug(f"Notebook select: {_e}")

    def _build_return_single_tab(self, nb, dialog) -> None:
        """반품 다이얼로그 — TAB 1: 단건 입력"""
        from ..utils.constants import X, tk, ttk
        from ..utils.custom_messagebox import CustomMessageBox
        tab_single = ttk.Frame(nb)
        nb.add(tab_single, text="  📝 단건 입력  ")

        frame = ttk.LabelFrame(tab_single, text="반품 정보")
        frame.pack(fill=X, padx=15, pady=10)

        ttk.Label(frame, text="LOT 번호:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        lot_combo = ttk.Combobox(frame, width=27, state='readonly')
        lot_combo.grid(row=0, column=1, padx=5, pady=5)
        # LOT 번호 목록 DB에서 로드
        def _load_lot_list():
            if not hasattr(self, 'engine') or not self.engine:
                return
            try:
                rows = self.engine.db.fetchall(
                    "SELECT DISTINCT lot_no FROM inventory_tonbag ORDER BY lot_no")
                lots = [r['lot_no'] if isinstance(r, dict) else r[0] for r in (rows or [])]
                lot_combo['values'] = lots if lots else ["(등록된 LOT 없음)"]
                if lots:
                    lot_combo.set(lots[0])
            except (ValueError, TypeError, AttributeError) as e:
                logger.debug(f"LOT 목록 조회 오류: {e}")
        _load_lot_list()

        ttk.Label(frame, text="Tonbag No:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        tonbag_combo = ttk.Combobox(frame, width=27, state='readonly')
        tonbag_combo.grid(row=1, column=1, padx=5, pady=5)
        tonbag_combo.set("← LOT 번호 선택 후 자동 조회")

        def _on_lot_change(event=None):
            lot_no = (lot_combo.get() or '').strip()
            if not lot_no or not hasattr(self, 'engine') or lot_no == "(등록된 LOT 없음)":
                return
            try:
                rows = self.engine.db.fetchall(
                    "SELECT sub_lt, weight, status FROM inventory_tonbag "
                    "WHERE lot_no = ? ORDER BY sub_lt", (lot_no,))
                if rows:
                    values = []
                    for r in rows:
                        sub = r['sub_lt'] if isinstance(r, dict) else r[0]
                        w = r['weight'] if isinstance(r, dict) else r[1]
                        st = r['status'] if isinstance(r, dict) else r[2]
                        values.append(f"{sub} ({w:.1f}kg, {st})")
                    tonbag_combo['values'] = values
                    tonbag_combo.set(values[0])
                    first_w = rows[0]['weight'] if isinstance(rows[0], dict) else rows[0][1]
                    qty_entry.delete(0, 'end')
                    qty_entry.insert(0, f"{first_w:.1f}")
                else:
                    tonbag_combo['values'] = ["톤백 없음"]
                    tonbag_combo.set("톤백 없음")
            except (ValueError, TypeError, AttributeError) as e:
                logger.debug(f"톤백 조회 오류: {e}")

        lot_combo.bind('<<ComboboxSelected>>', _on_lot_change)
        lot_combo.bind('<FocusOut>', _on_lot_change)
        lot_combo.bind('<Return>', _on_lot_change)

        def _on_tonbag_select(event=None):
            sel = tonbag_combo.get()
            if '(' in sel and 'kg' in sel:
                try:
                    w = sel.split('(')[1].split('kg')[0]
                    qty_entry.delete(0, 'end')
                    qty_entry.insert(0, w)
                except (ValueError, TypeError, KeyError) as _e:
                    logger.debug(f"Suppressed: {_e}")

        tonbag_combo.bind('<<ComboboxSelected>>', _on_tonbag_select)

        ttk.Label(frame, text="반품 수량 (kg):").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        qty_entry = ttk.Entry(frame, width=30)
        qty_entry.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(frame, text="반품 사유:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        # v6.12.1: 표준 사유 드롭다운 + 직접 입력 가능 (Combobox)
        _reason_codes = [
            "품질 불량", "수량 오류", "고객 취소", "배송 문제",
            "파손/변질", "규격 불일치", "기타"]
        reason_combo = ttk.Combobox(frame, width=27, values=_reason_codes)
        reason_combo.grid(row=3, column=1, padx=5, pady=5)
        reason_combo.set("품질 불량")

        ttk.Label(frame, text="비고:").grid(row=4, column=0, sticky='ne', padx=5, pady=5)
        note_text = tk.Text(frame, width=30, height=3)
        note_text.grid(row=4, column=1, padx=5, pady=5)

        def _process_single_return():
            lot_no = (lot_combo.get() or '').strip()
            tonbag_sel = tonbag_combo.get().strip()
            qty_str = qty_entry.get().strip()
            reason = reason_combo.get()
            note = note_text.get("1.0", "end").strip()

            # 최소 데이터: LOT 번호, Tonbag 선택, 반품 수량(kg)
            if not lot_no or lot_no == "(등록된 LOT 없음)":
                CustomMessageBox.showwarning(dialog, "입력 필요",
                    "LOT 번호를 목록에서 선택하세요.\n\n화물 정합성을 위해 LOT 번호는 필수입니다.")
                lot_combo.focus_set()
                return
            if not tonbag_sel or tonbag_sel == "← LOT 번호 선택 후 자동 조회" or tonbag_sel == "톤백 없음":
                CustomMessageBox.showwarning(dialog, "입력 필요",
                    "LOT 번호 선택 후 반품할 톤백을 선택하세요.\n\nTonbag No는 필수입니다.")
                lot_combo.focus_set()
                return
            if not qty_str:
                CustomMessageBox.showwarning(dialog, "입력 필요",
                    "반품 수량(kg)을 입력하세요.\n\n화물 무게 정합성을 위해 반품 수량은 필수입니다.")
                qty_entry.focus_set()
                return
            try:
                qty = float(qty_str)
                if qty <= 0:
                    raise ValueError
            except ValueError:
                CustomMessageBox.showwarning(dialog, "입력 오류",
                    "반품 수량에는 0보다 큰 숫자(kg)를 입력하세요.")
                qty_entry.focus_set()
                return

            # 무게 정합성: 반품 수량이 해당 톤백 무게를 초과하면 경고
            tonbag_kg = None
            if '(' in tonbag_sel and 'kg' in tonbag_sel:
                try:
                    tonbag_kg = float(tonbag_sel.split('(')[1].split('kg')[0].strip())
                except (ValueError, TypeError, IndexError) as _e:
                    logger.debug(f"톤백 무게 파싱: {_e}")
            if tonbag_kg is not None and qty > tonbag_kg:
                CustomMessageBox.showwarning(dialog, "무게 정합성 경고",
                    f"반품 수량({qty:,.2f} kg)이 해당 톤백 무게({tonbag_kg:,.2f} kg)를 초과합니다.\n\n"
                    "수정해 주세요. (반품 수량 ≤ 톤백 무게)")
                qty_entry.focus_set()
                return

            if not CustomMessageBox.askyesno(dialog, "반품 확인",
                f"LOT: {lot_no}\n수량: {qty:,.2f} kg\n사유: {reason}\n\n반품 처리하시겠습니까?"):
                return
            sub_lt_val = 1
            if tonbag_sel and tonbag_sel[0].isdigit():
                try:
                    sub_lt_val = int(tonbag_sel.split(' ')[0])
                except (ValueError, IndexError):
                    sub_lt_val = 1
            if hasattr(self.engine, 'return_single_tonbag'):
                try:
                    result = self.engine.return_single_tonbag(
                        lot_no=lot_no, sub_lt=sub_lt_val, reason=reason, remark=note)
                except (ValueError, RuntimeError, sqlite3.OperationalError, sqlite3.IntegrityError) as e:
                    CustomMessageBox.showerror(dialog, "반품 오류", f"반품 처리 중 오류:\n{str(e)[:500]}")
                    return
            elif hasattr(self.engine, 'process_return'):
                try:
                    result = self.engine.process_return([{
                        'lot_no': lot_no, 'sub_lt': sub_lt_val,
                        'reason': reason, 'remark': note}],
                        source_type='RETURN_SINGLE')
                except (ValueError, RuntimeError, sqlite3.OperationalError, sqlite3.IntegrityError) as e:
                    CustomMessageBox.showerror(dialog, "반품 오류", f"반품 처리 중 오류:\n{str(e)[:500]}")
                    return
            else:
                CustomMessageBox.showwarning(dialog, "안내", "반품 엔진을 찾을 수 없습니다.")
                return
            if result.get('success'):
                self._log(f"✅ 반품 완료: {lot_no}-{sub_lt_val}")
                CustomMessageBox.showinfo(dialog, "완료",
                    f"반품 처리 완료\n\nLOT: {lot_no}\n톤백: {sub_lt_val}\n사유: {reason}")
                dialog.destroy()
                self._safe_refresh()
                if hasattr(self, '_refresh_dashboard'):
                    self._refresh_dashboard()
            else:
                errs = '\n'.join(result.get('errors', ['알 수 없는 오류']))
                CustomMessageBox.showerror(dialog, "오류", f"반품 실패:\n{errs}")

        s_btn = ttk.Frame(tab_single)
        s_btn.pack(pady=15)
        try:
            ttk.Button(s_btn, text="반품 처리", command=_process_single_return,
                       bootstyle="primary").pack(side='left', padx=10)
        except TypeError:
            ttk.Button(s_btn, text="반품 처리", command=_process_single_return).pack(side='left', padx=10)
        ttk.Button(s_btn, text="취소", command=dialog.destroy).pack(side='left', padx=10)

    def _build_return_excel_tab(self, nb, dialog) -> None:
        """반품 다이얼로그 — TAB 2: Excel 일괄 반품"""
        from ..utils.constants import BOTH, LEFT, RIGHT, VERTICAL, X, Y, tk, ttk
        from ..utils.custom_messagebox import CustomMessageBox
        tab_excel = ttk.Frame(nb)
        nb.add(tab_excel, text="  📂 Excel 일괄 반품  ")

        # -- top bar: 템플릿 다운로드 + 파일 라벨 --
        top_bar = ttk.Frame(tab_excel)
        top_bar.pack(fill=X, padx=10, pady=8)
        try:
            ttk.Button(top_bar, text="📥 반품 양식 다운로드",
                       command=lambda: self._bret_download_template(dialog),
                       bootstyle="info").pack(side='left', padx=5)
        except TypeError:
            ttk.Button(top_bar, text="📥 반품 양식 다운로드",
                       command=lambda: self._bret_download_template(dialog)).pack(side='left', padx=5)

        file_var = tk.StringVar(value="파일을 선택하세요...")
        ttk.Label(top_bar, textvariable=file_var, foreground=tc('text_muted')).pack(side='left', padx=10, fill=X, expand=True)

        # -- 미리보기 Treeview --
        pv_tree, summary_var = self._bret_build_preview_tree(tab_excel, tk, ttk)

        # -- 파싱된 반품 데이터 저장 --
        parsed_returns = []

        # -- 데이터 입력 버튼 --
        try:
            ttk.Button(top_bar, text="📂 데이터 입력",
                       command=lambda: self._bret_upload_return_excel(
                           dialog, file_var, pv_tree, parsed_returns, summary_var),
                       bootstyle="warning").pack(side='left', padx=5)
        except TypeError:
            ttk.Button(top_bar, text="📂 데이터 입력",
                       command=lambda: self._bret_upload_return_excel(
                           dialog, file_var, pv_tree, parsed_returns, summary_var)).pack(side='left', padx=5)

        # -- 하단 실행/취소 버튼 --
        ex_btn = ttk.Frame(tab_excel)
        ex_btn.pack(pady=8)
        try:
            ttk.Button(ex_btn, text="🔄 일괄 반품 실행",
                       command=lambda: self._bret_execute_bulk_return(dialog, parsed_returns),
                       bootstyle="danger").pack(side='left', padx=10)
        except TypeError:
            ttk.Button(ex_btn, text="🔄 일괄 반품 실행",
                       command=lambda: self._bret_execute_bulk_return(dialog, parsed_returns)).pack(side='left', padx=10)
        ttk.Button(ex_btn, text="취소", command=dialog.destroy).pack(side='left', padx=10)

    # -----------------------------------------------------------------
    # _bret_* : _build_return_excel_tab 서브 메서드  # v8.6.4 [SRP]
    # -----------------------------------------------------------------

    def _bret_download_template(self, dialog) -> None:  # v8.6.4 [SRP]
        """반품 템플릿 다운로드 — 재고 리스트 형식 + return_qty_kg, return_reason."""
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        from ..utils.constants import filedialog
        from ..utils.custom_messagebox import CustomMessageBox

        file_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="반품 양식 저장", defaultextension=".xlsx",
            initialfile="반품_양식_템플릿.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        try:
            inv_cols = getattr(self, 'INVENTORY_TEMPLATE_COLUMNS', None)
            if inv_cols is None:
                inv_cols = [
                    ('lot_no', 'LOT NO', 120, True), ('sap_no', 'SAP NO', 120, False),
                    ('bl_no', 'BL NO', 140, False), ('product', 'PRODUCT', 160, False),
                    ('net_weight', 'NET(Kg)', 100, False), ('container_no', 'CONTAINER', 130, False),
                    ('mxbg_pallet', 'MXBG', 70, False),
                ]
            return_extra = [
                ('return_qty_kg', '반품수량(갯수)', 16, True),
                ('return_reason', 'RETURN REASON', 20, True),
            ]
            headers = list(inv_cols) + return_extra
            required_return = {'lot_no', 'bl_no', 'return_qty_kg', 'return_reason'}
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "반품 데이터"
            hfont = Font(bold=True, color="FFFFFF", size=11)
            req_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
            opt_fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
            smp_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
            thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ncols = len(headers)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            ws['A1'] = "🔄 반품 템플릿 — 필수: Lot No, BL NO, 톤백중량(DB조회), 반품수량(갯수), 사유. 2행=DB필드명."
            ws['A1'].font = Font(bold=True, size=11, color="C0392B")
            ws.row_dimensions[1].height = 28
            for col, h in enumerate(headers, 1):
                db_f = h[0]
                ws.cell(row=2, column=col, value=db_f).font = Font(size=8, color="999999")
            ws.row_dimensions[2].height = 14
            for col, h in enumerate(headers, 1):
                db_f, disp, w, _ = h[0], h[1], h[2], (h[3] if len(h) > 3 else False)
                c = ws.cell(row=3, column=col, value=disp + (' *' if db_f in required_return else ''))
                c.font = hfont
                c.fill = req_fill if db_f in required_return else opt_fill
                c.alignment = Alignment(horizontal='center')
                c.border = thin
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(w, 24)
            for r in range(4, 8):
                for c in range(1, ncols + 1):
                    cell = ws.cell(row=r, column=c, value='')
                    cell.fill = smp_fill
                    cell.border = thin
            wb.save(file_path)
            CustomMessageBox.showinfo(dialog, "완료", f"반품 양식 저장 완료\n\n{file_path}\n\n재고 리스트와 동일 형식 + 반품수량·사유. 필수만 채우면 됩니다.")
            self._log(f"📥 반품 양식 다운로드: {file_path}")
        except (FileNotFoundError, OSError, PermissionError) as e:
            CustomMessageBox.showerror(dialog, "오류", f"파일 저장 실패: {e}")

    def _bret_build_preview_tree(self, tab_excel, tk, ttk):  # v8.6.4 [SRP]
        """미리보기 Treeview 구성 — 필수 4열: Lot No, BL NO, 톤백중량, 반품수량(갯수)."""
        from ..utils.constants import BOTH, LEFT, RIGHT, VERTICAL, X, Y
        pv_frame = ttk.LabelFrame(tab_excel, text="반품 미리보기 (DB 자동 조회)")
        pv_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        ttk.Label(pv_frame, text="※ 필수: Lot No, BL NO, 톤백중량, 반품수량(갯수) — 네 열 중 하나라도 비면 반품 업로드 시 오류 표시", font=('맑은 고딕', 10)).pack(anchor='w')
        cols = ('lot_no', 'bl_no', 'tonbag_no', 'product', 'weight_kg',
                'return_qty', 'reason', 'status', 'remark')
        pv_tree = ttk.Treeview(pv_frame, columns=cols, show='headings', height=10)
        for cid, txt, w in [
            ('lot_no', 'LOT NO *', 100), ('bl_no', 'BL NO *', 100),
            ('tonbag_no', 'Tonbag#', 65), ('product', 'Product', 90),
            ('weight_kg', '톤백중량(kg) *', 90), ('return_qty', '반품수량(갯수) *', 90),
            ('reason', '사유', 100), ('status', '상태', 70), ('remark', '비고', 100)]:
            pv_tree.heading(cid, text=txt, anchor='center')
            pv_tree.column(cid, width=w, anchor='center' if cid in ('tonbag_no','status') else 'w')
        pv_sb = tk.Scrollbar(pv_frame, orient=VERTICAL, command=pv_tree.yview)
        pv_tree.configure(yscrollcommand=pv_sb.set)
        pv_tree.pack(side=LEFT, fill=BOTH, expand=True)
        pv_sb.pack(side=RIGHT, fill=Y)
        try:
            from gui_app_modular.utils.tree_enhancements import TreeviewTotalFooter as _TTF
            _TTF(pv_frame, pv_tree, ['total','avail','picked'],
                 {'total':'총재고(MT)','avail':'가용(MT)','picked':'출고(MT)'},
                 {'total':',.1f','avail':',.1f','picked':',.1f'}).pack(fill='x')
        except Exception as e:
            logger.warning(f'[UI] advanced_dialogs_mixin: {e}')
        summary_var = tk.StringVar(value="데이터 붙여넣기 또는 파일 업로드하세요")
        ttk.Label(tab_excel, textvariable=summary_var, font=('맑은 고딕', 11, 'bold')).pack(pady=3)
        return pv_tree, summary_var

    def _bret_show_input_choice(self, dialog):  # v8.6.4 [SRP]
        """Excel/데이터 입력 원칙: 데이터 붙여넣기 vs 파일 업로드 선택 후 진행."""
        from ..utils.constants import tk, ttk
        from ..utils.ui_constants import (
            UPLOAD_CHOICE_BTN_PASTE,
            UPLOAD_CHOICE_BTN_UPLOAD,
            UPLOAD_CHOICE_HEADER,
            UPLOAD_CHOICE_PASTE,
            UPLOAD_CHOICE_UPLOAD,
            DialogSize,
            apply_modal_window_options,
            center_dialog,
        )
        result = [None]
        win = create_themed_toplevel(dialog)
        win.title("다량 반품 데이터 입력")
        apply_modal_window_options(win)
        win.transient(dialog)
        win.grab_set()
        win.geometry(DialogSize.get_geometry(dialog, 'small'))
        win.minsize(400, 260)
        center_dialog(win, dialog)
        f = ttk.Frame(win, padding=(20, 20, 20, 32))
        f.pack(fill=tk.BOTH, expand=True)
        ttk.Label(f, text=UPLOAD_CHOICE_HEADER, font=('맑은 고딕', 12, 'bold')).pack(anchor='w', pady=(0, 12))
        ttk.Label(f, text=UPLOAD_CHOICE_PASTE, font=('맑은 고딕', 10), wraplength=400, justify=tk.LEFT).pack(anchor='w', pady=(0, 10))
        ttk.Label(f, text=UPLOAD_CHOICE_UPLOAD, font=('맑은 고딕', 10), wraplength=400, justify=tk.LEFT).pack(anchor='w', pady=(0, 24))
        btn_f = ttk.Frame(f)
        btn_f.pack(anchor='center')
        def on_paste():
            result[0] = 'paste'
            win.destroy()
        def on_upload():
            result[0] = 'upload'
            win.destroy()
        ttk.Button(btn_f, text=UPLOAD_CHOICE_BTN_PASTE, command=on_paste, width=22).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_f, text=UPLOAD_CHOICE_BTN_UPLOAD, command=on_upload, width=22).pack(side=tk.LEFT)
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        win.wait_window(win)
        return result[0]

    def _bret_map_columns(self, df):  # v8.6.4 [SRP]
        """Excel DataFrame 컬럼 → 내부 필드 매핑 딕셔너리 반환."""
        col_map = {}
        for c in df.columns:
            cl = str(c).lower().replace(' ', '_').replace('*', '').strip()
            if 'lot' in cl and 'no' in cl:
                col_map['lot_no'] = c
            elif 'bl' in cl and 'no' in cl:
                col_map['bl_no'] = c
            elif 'tonbag' in cl or ('ton' in cl and 'bag' in cl):
                col_map['tonbag_no'] = c
            elif 'return' in cl and ('qty' in cl or 'kg' in cl):
                col_map['return_qty'] = c
            elif '반품수량' in str(c) or 'return_qty' in cl:
                col_map['return_qty'] = c
            elif 'reason' in cl or cl == 'return_reason':
                col_map['reason'] = c
            elif 'remark' in cl:
                col_map['remark'] = c
        return col_map

    def _bret_prefetch_inventory(self, df, col_map):  # v8.6.4 [SRP]
        """v8.2.0 N+1 최적화: inventory pre-fetch → {lot_no: row} 캐시."""
        _lot_col = col_map.get('lot_no', '')
        _all_lots = list(set(
            str(row.get(_lot_col, '')).strip()
            for _, row in df.iterrows()
            if str(row.get(_lot_col, '')).strip() not in ('', 'nan')
        ))
        _inv_cache = {}
        if _all_lots and hasattr(self, 'engine') and self.engine:
            try:
                _ph = ','.join('?' * len(_all_lots))
                _inv_rows = self.engine.db.fetchall(
                    f"SELECT lot_no, product, bl_no FROM inventory WHERE lot_no IN ({_ph})",
                    tuple(_all_lots)
                ) or []
                _inv_cache = {
                    (r.get('lot_no') if isinstance(r, dict) else r[0]): r
                    for r in _inv_rows
                }
            except Exception as _e:
                logger.debug(f"inventory pre-fetch 스킵: {_e}")
        return _inv_cache

    def _bret_resolve_product_bl(self, lot_no, bl_no, _inv_cache):  # v8.6.4 [SRP]
        """cache 또는 DB에서 product, bl_no 조회."""
        product = ''
        _inv_cached = _inv_cache.get(lot_no)
        if _inv_cached:
            product = (_inv_cached.get('product') if isinstance(_inv_cached, dict)
                       else _inv_cached[1]) or ''
            if not bl_no:
                bl_no = (_inv_cached.get('bl_no') if isinstance(_inv_cached, dict)
                         else _inv_cached[2]) or ''
        if hasattr(self, 'engine') and self.engine and not _inv_cached:
            try:
                inv = self.engine.db.fetchone(
                    "SELECT product, bl_no FROM inventory WHERE lot_no = ?", (lot_no,))
                if inv:
                    product = inv['product'] if isinstance(inv, dict) else inv[0]
                    if not bl_no:
                        bl_no = (inv['bl_no'] if isinstance(inv, dict) else inv[1]) or ''
            except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
                logger.warning(f"[_upload_return_excel] Suppressed: {e}")
        return product, bl_no

    def _bret_resolve_tonbags(self, lot_no, tonbag_str, return_qty_kg, returnable_statuses):  # v8.6.4 [SRP]
        """톤백 번호/수량 기반 반품 대상 톤백 목록 조회."""
        items_to_add = []
        if tonbag_str:
            sub_lt = 1
            try:
                sub_lt = int(float(tonbag_str))
            except (ValueError, TypeError):
                sub_lt = 1
            if hasattr(self, 'engine') and self.engine:
                tb = self.engine.db.fetchone(
                    "SELECT sub_lt, weight, status FROM inventory_tonbag "
                    "WHERE lot_no = ? AND sub_lt = ?", (lot_no, sub_lt))
                if tb:
                    sub_lt = tb['sub_lt'] if isinstance(tb, dict) else tb[0]
                    weight_kg = tb['weight'] if isinstance(tb, dict) else tb[1]
                    status = tb['status'] if isinstance(tb, dict) else tb[2]
                    if status in returnable_statuses:
                        items_to_add = [{'sub_lt': sub_lt, 'weight_kg': weight_kg, 'status': status}]
                if not items_to_add:
                    items_to_add = [{'sub_lt': sub_lt, 'weight_kg': 0, 'status': 'NOT FOUND'}]
            else:
                items_to_add = [{'sub_lt': sub_lt, 'weight_kg': 0, 'status': '?'}]
        elif return_qty_kg > 0 and hasattr(self, 'engine') and self.engine:
            tonbags = self.engine.db.fetchall(
                """SELECT sub_lt, weight, status FROM inventory_tonbag
                   WHERE lot_no = ? AND status IN (?,?,?,?,?)
                   ORDER BY sub_lt DESC""",
                (lot_no,) + returnable_statuses)
            remaining_kg = return_qty_kg
            for tb in (tonbags or []):
                if remaining_kg <= 0.01:
                    break
                w = float(tb['weight'] if isinstance(tb, dict) else tb[1] or 0)
                if w <= 0:
                    continue
                items_to_add.append({
                    'sub_lt': tb['sub_lt'] if isinstance(tb, dict) else tb[0],
                    'weight_kg': w,
                    'status': tb['status'] if isinstance(tb, dict) else tb[2],
                })
                remaining_kg -= w
            if not items_to_add:
                items_to_add = [{'sub_lt': 0, 'weight_kg': 0, 'status': 'NOT FOUND'}]
        else:
            if hasattr(self, 'engine') and self.engine:
                tb = self.engine.db.fetchone(
                    "SELECT sub_lt, weight, status FROM inventory_tonbag "
                    "WHERE lot_no = ? AND status IN (?,?,?,?,?) "
                    "ORDER BY sub_lt LIMIT 1", (lot_no,) + returnable_statuses)
                if tb:
                    items_to_add = [{
                        'sub_lt': tb['sub_lt'] if isinstance(tb, dict) else tb[0],
                        'weight_kg': tb['weight'] if isinstance(tb, dict) else tb[1],
                        'status': tb['status'] if isinstance(tb, dict) else tb[2],
                    }]
            if not items_to_add:
                items_to_add = [{'sub_lt': 0, 'weight_kg': 0, 'status': 'NOT FOUND'}]
        return items_to_add

    def _bret_upload_return_excel(self, dialog, file_var, pv_tree, parsed_returns, summary_var):  # v8.6.4 [SRP]
        """Excel/데이터 입력: 선택 후 파일 업로드 또는 붙여넣기 → DB 조회 → 미리보기."""
        import pandas as pd

        from ..utils.constants import filedialog
        from ..utils.custom_messagebox import CustomMessageBox

        choice = self._bret_show_input_choice(dialog)
        if not choice:
            return
        if choice == "paste":
            CustomMessageBox.showinfo(
                dialog, "다량 반품 데이터 입력",
                "다량 반품은 프로그램 내장 형식과 동일합니다.\n"
                "반품 양식 다운로드 후 엑셀에 채워 [파일 업로드]를 선택하거나,\n"
                "입고 메뉴의 [반품 입고 (Excel)]에서 데이터 붙여넣기를 사용하세요.")
            return
        fp = filedialog.askopenfilename(
            parent=self.root,
            title="반품 Excel 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not fp:
            return
        file_var.set(os.path.basename(fp))

        try:
            df = pd.read_excel(fp, header=None)
            # 헤더 행 찾기 (lot_no 또는 LOT NO 포함 행)
            header_row = None
            for idx in range(min(5, len(df))):
                row_vals = [str(v).lower().strip() for v in df.iloc[idx]]
                if any('lot_no' in v or 'lot no' in v for v in row_vals):
                    header_row = idx
                    break
            if header_row is None:
                header_row = 2  # 기본: 3행 (0-indexed 2)

            df.columns = df.iloc[header_row].astype(str).str.strip().str.lower()
            df = df.iloc[header_row+1:].reset_index(drop=True)
            df = df.dropna(how='all')

            col_map = self._bret_map_columns(df)

            if 'lot_no' not in col_map:
                CustomMessageBox.showerror(dialog, "오류", "LOT NO 컬럼을 찾을 수 없습니다.")
                return
            if 'return_qty' not in col_map and 'reason' not in col_map:
                CustomMessageBox.showwarning(dialog, "안내", "RETURN QTY (KG) 또는 RETURN REASON 컬럼을 권장합니다.")

            self._bret_populate_preview(
                dialog, df, col_map, header_row, pv_tree, parsed_returns, summary_var)

        except (FileNotFoundError, OSError, PermissionError) as e:
            CustomMessageBox.showerror(dialog, "오류", f"파일 읽기 실패: {e}")

    def _bret_populate_preview(self, dialog, df, col_map, header_row,  # v8.6.4 [SRP]
                               pv_tree, parsed_returns, summary_var):
        """미리보기 구성 + 필수 4열 검증 (Lot No, BL NO, 톤백중량, 반품수량(갯수))."""
        from ..utils.custom_messagebox import CustomMessageBox

        pv_tree.delete(*pv_tree.get_children())
        parsed_returns.clear()
        ok_count = 0
        err_count = 0
        required_missing_rows = []  # 필수 누락 행 번호(Excel 1-based)
        returnable_statuses = (STATUS_PICKED, 'CONFIRMED', 'SHIPPED', STATUS_SOLD, STATUS_RESERVED, STATUS_PARTIAL)
        data_start_row = header_row + 2  # Excel 1-based 첫 데이터 행

        _inv_cache = self._bret_prefetch_inventory(df, col_map)

        for idx, row in df.iterrows():
            lot_no = str(row.get(col_map.get('lot_no', ''), '')).strip()
            if not lot_no or lot_no == 'nan':
                continue
            excel_row = data_start_row + int(idx)

            bl_no = str(row.get(col_map.get('bl_no', ''), '')).strip()
            tonbag_str = str(row.get(col_map.get('tonbag_no', ''), '')).strip()
            qty_str = str(row.get(col_map.get('return_qty', ''), '')).strip()
            reason = str(row.get(col_map.get('reason', ''), '품질 불량')).strip()
            remark = str(row.get(col_map.get('remark', ''), '')).strip()
            if reason == 'nan': reason = '품질 불량'
            if remark == 'nan': remark = ''
            if bl_no == 'nan': bl_no = ''
            if tonbag_str == 'nan': tonbag_str = ''

            try:
                return_qty_kg = float(qty_str) if qty_str and qty_str != 'nan' else 0.0
            except (ValueError, TypeError):
                return_qty_kg = 0.0

            product, bl_no = self._bret_resolve_product_bl(lot_no, bl_no, _inv_cache)

            # 필수 4열 검증: Lot No, BL NO, 톤백중량, 반품수량(갯수)
            missing_this_row = []
            if not lot_no or lot_no == 'nan':
                missing_this_row.append('Lot No')
            if not bl_no or bl_no == 'nan':
                missing_this_row.append('BL NO')
            if not qty_str or qty_str == 'nan' or return_qty_kg <= 0:
                missing_this_row.append('반품수량(갯수)')

            items_to_add = self._bret_resolve_tonbags(
                lot_no, tonbag_str, return_qty_kg, returnable_statuses)

            has_tonbag_weight = any(float(it.get('weight_kg', 0) or 0) > 0 for it in items_to_add)
            if not has_tonbag_weight:
                missing_this_row.append('톤백중량')
            if missing_this_row:
                required_missing_rows.append((excel_row, missing_this_row))

            for it in items_to_add:
                sub_lt = it['sub_lt']
                weight_kg = it['weight_kg']
                status = it['status']
                is_ok = status in returnable_statuses and not missing_this_row
                tag = 'ok' if is_ok else 'err'
                if is_ok:
                    ok_count += 1
                else:
                    err_count += 1
                disp_qty = qty_str if qty_str and qty_str != 'nan' else (f"{weight_kg:.1f}" if weight_kg else '')
                pv_tree.insert('', 'end', values=(
                    lot_no, bl_no, sub_lt, product,
                    f"{weight_kg:.1f}" if isinstance(weight_kg, (int, float)) else weight_kg,
                    disp_qty, reason, status, remark
                ), tags=(tag,))
                parsed_returns.append({
                    'lot_no': lot_no, 'sub_lt': sub_lt,
                    'reason': reason, 'remark': remark,
                    'status': status, 'valid': is_ok,
                })

        pv_tree.tag_configure('ok', foreground=tc('success'))
        pv_tree.tag_configure('err', foreground=tc('danger'))

        if required_missing_rows:
            lines = [f"  행 {r}: {', '.join(m)}" for r, m in required_missing_rows[:20]]
            if len(required_missing_rows) > 20:
                lines.append(f"  ... 외 {len(required_missing_rows) - 20}행")
            CustomMessageBox.showerror(dialog, "필수 항목 누락",
                "반품 업로드 시 Lot No, BL NO, 톤백중량, 반품수량(갯수) 네 열은 필수입니다.\n\n"
                "다음 행에 필수 항목이 비어 있습니다:\n\n" + "\n".join(lines) + "\n\n수정 후 다시 업로드하세요.")

        summary_var.set(f"✅ 반품 가능: {ok_count}건  |  ❌ 불가: {err_count}건  |  총: {ok_count + err_count}건")

    def _bret_execute_bulk_return(self, dialog, parsed_returns):  # v8.6.4 [SRP]
        """일괄 반품 실행 — v6.12.2: 자동승인/관리자확인 워크플로우."""
        from ..utils.custom_messagebox import CustomMessageBox

        valid_items = [r for r in parsed_returns if r.get('valid')]
        if not valid_items:
            CustomMessageBox.showwarning(dialog, "안내", "반품 가능한 항목이 없습니다.")
            return

        # v6.12.2: 톤백 수 기준 자동승인 / 관리자확인 분기
        try:
            from engine_modules.constants import RETURN_AUTO_APPROVE_MAX_TONBAGS  # STATUS_* 상단 import 사용
            _max_auto = RETURN_AUTO_APPROVE_MAX_TONBAGS
        except ImportError:
            _max_auto = 5

        total_tonbags = len(valid_items)
        if total_tonbags > _max_auto:
            # 대량 반품 → 관리자 확인 강화 (사유 + 건수 명시)
            reasons = set(r.get('reason', '미기재') for r in valid_items)
            reason_summary = ', '.join(reasons) if len(reasons) <= 3 else f"{len(reasons)}종"
            confirm_msg = (
                f"⚠️ 대량 반품 — 관리자 확인 필요\n\n"
                f"반품 건수: {total_tonbags}건 (자동승인 기준 {_max_auto}건 초과)\n"
                f"반품 사유: {reason_summary}\n\n"
                f"대량 반품은 되돌릴 수 없습니다.\n"
                f"정말 진행하시겠습니까?"
            )
            if not CustomMessageBox.askyesno(dialog, "⚠️ 대량 반품 관리자 확인", confirm_msg):
                return
            # 이중 확인
            if not CustomMessageBox.askyesno(dialog, "최종 확인",
                f"총 {total_tonbags}건 반품을 최종 실행합니다.\n\n확인하셨습니까?"):
                return
        else:
            # 소량 반품 → 간단 확인
            if not CustomMessageBox.askyesno(dialog, "일괄 반품 확인",
                f"총 {total_tonbags}건을 반품 처리합니다.\n\n계속하시겠습니까?"):
                return

        if hasattr(self.engine, 'process_return'):
            try:
                result = self.engine.process_return(valid_items,
                    source_type='RETURN_EXCEL')
            except (ValueError, RuntimeError, sqlite3.OperationalError, sqlite3.IntegrityError) as e:
                CustomMessageBox.showerror(dialog, "반품 오류", f"일괄 반품 처리 중 오류:\n{str(e)[:500]}")
                return
            if result.get('success'):
                self._log(f"✅ 일괄 반품 완료: {result.get('returned', 0)}건")
                CustomMessageBox.showinfo(dialog, "완료",
                    f"일괄 반품 처리 완료\n\n"
                    f"성공: {result.get('returned', 0)}건\n"
                    f"스킵: {result.get('skipped', 0)}건")
                dialog.destroy()
                self._safe_refresh()
                if hasattr(self, '_refresh_dashboard'):
                    self._refresh_dashboard()
            else:
                errs = '\n'.join(result.get('errors', ['알 수 없는 오류'])[:5])
                CustomMessageBox.showerror(dialog, "오류", f"반품 처리 오류:\n{errs}")
        else:
            CustomMessageBox.showwarning(dialog, "안내", "반품 엔진을 찾을 수 없습니다.")

    # =========================================================================
    # v3.8.4: 수동 입고 입력 다이얼로그
    # =========================================================================

    # =========================================================================
    # v3.8.4: 문서 변환 (OCR/PDF)
    # =========================================================================

    def _show_document_convert_dialog(self) -> None:
        """v3.8.4: 문서 변환 (OCR 스캔 / PDF → Excel/Word)"""
        from ..utils.constants import X, tk, ttk
        from ..utils.custom_messagebox import CustomMessageBox

        dialog = create_themed_toplevel(self.root)
        dialog.title("📄 문서 변환 (OCR/PDF)")
        dialog.geometry("500x400")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="문서 변환", font=('맑은 고딕', 16, 'bold')).pack(pady=10)

        # 변환 모드 선택
        mode_frame = ttk.LabelFrame(dialog, text="변환 모드 선택")
        mode_frame.pack(fill=X, padx=20, pady=5)

        mode_var = tk.StringVar(value='ocr_scan')
        modes = [
            ('ocr_scan', '📷 OCR 스캔 (이미지/스캔 PDF → 텍스트 추출)'),
            ('pdf_convert', '📄 PDF → Excel/Word 변환'),
        ]
        for val, text in modes:
            ttk.Radiobutton(mode_frame, text=text, variable=mode_var,
                           value=val).pack(anchor='w', padx=10, pady=3)

        # 출력 형식
        out_frame = ttk.LabelFrame(dialog, text="출력 형식")
        out_frame.pack(fill=X, padx=20, pady=5)

        out_var = tk.StringVar(value='excel')
        ttk.Radiobutton(out_frame, text='📊 Excel (.xlsx)', variable=out_var,
                        value='excel').pack(anchor='w', padx=10, pady=3)
        ttk.Radiobutton(out_frame, text='📝 Word (.docx)', variable=out_var,
                        value='word').pack(anchor='w', padx=10, pady=3)

        # 파일 선택
        file_frame = ttk.LabelFrame(dialog, text="파일 선택")
        file_frame.pack(fill=X, padx=20, pady=5)

        file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=file_path_var, width=45).pack(side='left', padx=5, pady=5)

        def browse_file():
            from tkinter import filedialog
            filetypes = [
                ("지원 파일", "*.pdf;*.png;*.jpg;*.jpeg;*.tif;*.tiff;*.bmp"),
                ("PDF 파일", "*.pdf"),
                ("이미지 파일", "*.png;*.jpg;*.jpeg;*.tif;*.tiff;*.bmp"),
            ]
            path = filedialog.askopenfilename(parent=dialog, filetypes=filetypes)
            if path:
                file_path_var.set(path)

        ttk.Button(file_frame, text="찾아보기", command=browse_file).pack(side='left', padx=5, pady=5)

        def process_convert():
            filepath = file_path_var.get().strip()
            if not filepath or not os.path.exists(filepath):
                CustomMessageBox.showwarning(dialog, "파일 선택", "변환할 파일을 선택하세요.")
                return

            mode = mode_var.get()
            out_fmt = out_var.get()

            api_ok = getattr(self, '_api_connected', False)
            if mode == 'ocr_scan' and not api_ok:
                CustomMessageBox.showwarning(dialog, "API 필요",
                    "OCR 스캔에는 Gemini API가 필요합니다.\n\n"
                    "도구 > Gemini API 설정에서 API Key를 설정하세요.\n"
                    "https://aistudio.google.com에서 무료 발급 가능합니다.")
                return

            CustomMessageBox.showinfo(dialog, "변환 시작",
                f"변환 모드: {'OCR 스캔' if mode == 'ocr_scan' else 'PDF 변환'}\n"
                f"출력 형식: {'Excel' if out_fmt == 'excel' else 'Word'}\n"
                f"파일: {os.path.basename(filepath)}\n\n"
                f"변환 기능은 다음 업데이트에서 구현됩니다.\n"
                f"(Gemini Vision API 연동 예정)")

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=15)
        try:
            ttk.Button(btn_frame, text="🔄 변환 시작", command=process_convert,
                       bootstyle="info").pack(side='left', padx=10)
        except TypeError:
            ttk.Button(btn_frame, text="🔄 변환 시작", command=process_convert).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="닫기", command=dialog.destroy).pack(side='left', padx=10)

    # ═══════════════════════════════════════════════════════
    # v3.8.4: 출고 현황 조회
    # ═══════════════════════════════════════════════════════

    def _show_outbound_history(self) -> None:
        """출고 현황(stock_movement) 조회 팝업"""
        from datetime import datetime as _dt
        from datetime import timedelta

        from ..utils.constants import BOTH, END, LEFT, RIGHT, W, X, Y, tk, ttk
        from ..utils.custom_messagebox import CustomMessageBox

        dialog = create_themed_toplevel(self.root)
        dialog.title("📋 출고 현황 조회")
        dialog.geometry("980x560")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)

        # 입고 현황 조회와 동일한 기간 필터 UI 패턴
        top = ttk.LabelFrame(dialog, text="📅 조회 기간 (출고일 기준)")
        top.pack(fill=X, padx=10, pady=(10, 5))
        top_inner = ttk.Frame(top)
        top_inner.pack(fill=X, padx=8, pady=8)

        ttk.Label(top_inner, text="시작일").pack(side=LEFT, padx=(0, 4))
        entry_start = ttk.Entry(top_inner, width=12)
        entry_start.pack(side=LEFT, padx=(0, 8))
        ttk.Label(top_inner, text="종료일").pack(side=LEFT, padx=(0, 4))
        entry_end = ttk.Entry(top_inner, width=12)
        entry_end.pack(side=LEFT, padx=(0, 8))

        today = _dt.now()
        first = today.replace(day=1)
        entry_start.insert(0, first.strftime('%Y-%m-%d'))
        entry_end.insert(0, today.strftime('%Y-%m-%d'))

        # 기존 상세 필터(유형/LOT)는 유지
        ttk.Label(top_inner, text="유형").pack(side=LEFT, padx=(12, 4))
        type_var = tk.StringVar(value='전체')
        ttk.Combobox(
            top_inner,
            textvariable=type_var,
            state='readonly',
            width=15,
            values=['전체', 'OUTBOUND', 'CANCEL_OUTBOUND', 'INBOUND', 'RETURN']
        ).pack(side=LEFT, padx=(0, 8))
        ttk.Label(top_inner, text="LOT").pack(side=LEFT, padx=(4, 4))
        lot_var = tk.StringVar()
        ttk.Entry(top_inner, textvariable=lot_var, width=15).pack(side=LEFT, padx=(0, 8))

        # 트리뷰
        tree_frame = ttk.Frame(dialog)
        tree_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)

        cols = ('id', 'lot_no', 'type', 'qty_kg', 'customer', 'date', 'created')
        tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=15)

        for col, text, w in [
            ('id', 'ID', 50), ('lot_no', 'LOT No.', 120), ('type', 'Type', 120),
            ('qty_kg', 'Qty(Kg)', 100), ('customer', 'Customer', 120),
            ('date', 'Movement Date', 110), ('created', 'Created At', 140)
        ]:
            tree.heading(col, text=text, anchor='center')
            tree.column(col, width=w, anchor='e' if col == 'qty_kg' else 'w')

        scrollbar_y = tk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        scrollbar_x = tk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar_y.pack(side=RIGHT, fill=Y)
        try:
            from gui_app_modular.utils.tree_enhancements import TreeviewTotalFooter as _TTF
            _TTF(tree_frame, tree, ['qty_kg'],
                 {'qty_kg': '수량(kg)'}, {'qty_kg': ',.0f'}).pack(fill='x')
        except Exception as e:
            logger.warning(f'[UI] advanced_dialogs_mixin: {e}')
        scrollbar_x.pack(side='bottom', fill=X)

        # 합계
        summary_var = tk.StringVar(value="조회 버튼을 클릭하세요")
        ttk.Label(dialog, textvariable=summary_var, anchor=W).pack(fill=X, padx=10, pady=(0, 6))
        btn_export = None

        def _set_quick_range(days: int) -> None:
            entry_end.delete(0, 'end')
            entry_end.insert(0, today.strftime('%Y-%m-%d'))
            entry_start.delete(0, 'end')
            if days == 0:
                entry_start.insert(0, today.replace(day=1).strftime('%Y-%m-%d'))
            elif days == -1:
                entry_start.insert(0, '2020-01-01')
            else:
                entry_start.insert(0, (today - timedelta(days=days)).strftime('%Y-%m-%d'))
            do_search()

        def do_search():
            tree.delete(*tree.get_children())
            try:
                start = entry_start.get().strip()
                end = entry_end.get().strip()
                try:
                    _dt.strptime(start, '%Y-%m-%d')
                    _dt.strptime(end, '%Y-%m-%d')
                except (ValueError, TypeError):
                    CustomMessageBox.showwarning(dialog, "날짜 오류", "YYYY-MM-DD 형식으로 입력하세요.")
                    return

                end_plus1 = (_dt.strptime(end, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
                query = "SELECT id, lot_no, movement_type, qty_kg, '' AS customer, created_at AS movement_date, created_at FROM stock_movement WHERE 1=1"
                params = [start, end_plus1]
                query += " AND created_at >= ? AND created_at < ?"

                mv_type = type_var.get()
                if mv_type != '전체':
                    query += " AND movement_type = ?"
                    params.append(mv_type)

                lot_filter = lot_var.get().strip()
                if lot_filter:
                    query += " AND lot_no LIKE ?"
                    params.append(f"%{lot_filter}%")

                query += " ORDER BY created_at DESC LIMIT 500"

                rows = self.engine.db.fetchall(query, tuple(params))
                total_kg = 0
                for r in rows:
                    row_id = r['id'] if isinstance(r, dict) else r[0]
                    lot = r['lot_no'] if isinstance(r, dict) else r[1]
                    mtype = r['movement_type'] if isinstance(r, dict) else r[2]
                    qty = r['qty_kg'] if isinstance(r, dict) else r[3]
                    cust = r['customer'] if isinstance(r, dict) else r[4]
                    mdate = r['movement_date'] if isinstance(r, dict) else r[5]
                    created = r['created_at'] if isinstance(r, dict) else r[6]

                    qty_val = float(qty) if qty else 0
                    total_kg += qty_val

                    tree.insert('', END, values=(
                        row_id, lot, mtype, f"{qty_val:,.0f}",
                        cust or '', str(mdate or '')[:10], str(created or '')[:19]
                    ))

                summary_var.set(f"📊 {start} ~ {end} | Outbound Rows: {len(rows):,}건 | Total Weight: {total_kg:,.1f} Kg")
                if btn_export is not None:
                    btn_export.configure(state='normal' if rows else 'disabled')

            except (ValueError, TypeError, KeyError) as e:
                summary_var.set(f"오류: {e}")
                if btn_export is not None:
                    btn_export.configure(state='disabled')

        def do_export_excel():
            from ..utils.constants import filedialog
            if not tree.get_children():
                CustomMessageBox.showwarning(dialog, "경고", "내보낼 데이터가 없습니다.")
                return
            start = entry_start.get().strip().replace('-', '')
            end = entry_end.get().strip().replace('-', '')
            save_path = filedialog.asksaveasfilename(
                parent=self.root,
                title="Excel 저장",
                defaultextension=".xlsx",
                initialfile=f"출고현황_{start}_{end}.xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not save_path:
                return
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "출고 현황"
                ws.append(['ID', 'LOT No.', 'Type', 'Qty(Kg)', 'Customer', 'Movement Date', 'Created At'])
                for iid in tree.get_children():
                    vals = tree.item(iid, 'values')
                    qty_txt = str(vals[3]).replace(',', '') if len(vals) > 3 else '0'
                    try:
                        qty_val = float(qty_txt or 0)
                    except (ValueError, TypeError):
                        qty_val = 0.0
                    ws.append([
                        vals[0] if len(vals) > 0 else '',
                        vals[1] if len(vals) > 1 else '',
                        vals[2] if len(vals) > 2 else '',
                        qty_val,
                        vals[4] if len(vals) > 4 else '',
                        vals[5] if len(vals) > 5 else '',
                        vals[6] if len(vals) > 6 else '',
                    ])
                wb.save(save_path)
                CustomMessageBox.showinfo(dialog, "완료", f"저장 완료: {os.path.basename(save_path)}")
                try:
                    os.startfile(save_path)
                except (AttributeError, OSError) as e:
                    logger.warning(f"[do_export_excel] Suppressed: {e}")
            except ImportError:
                CustomMessageBox.showerror(dialog, "오류", "openpyxl이 설치되지 않았습니다.")
            except (OSError, PermissionError) as e:
                CustomMessageBox.showerror(dialog, "오류", f"파일 저장 실패:\n{e}")

        for label, days in [("이번 달", 0), ("최근 7일", 7), ("최근 30일", 30), ("최근 90일", 90), ("전체", -1)]:
            ttk.Button(top_inner, text=label, width=8, command=lambda d=days: _set_quick_range(d)).pack(side=LEFT, padx=2)
        ttk.Button(top_inner, text="🔍 조회", command=do_search).pack(side=LEFT, padx=(10, 0))

        bottom = ttk.Frame(dialog)
        bottom.pack(fill=X, padx=8, pady=(0, 8))
        btn_export = ttk.Button(bottom, text="📊 Excel 내보내기", command=do_export_excel, state='disabled')
        btn_export.pack(side=RIGHT, padx=5)
        ttk.Button(bottom, text="❌ 닫기", command=dialog.destroy).pack(side=RIGHT, padx=5)

        # 초기 로드
        do_search()

    # ═══════════════════════════════════════════════════════
    # v3.8.4 A6: 재고 추이 차트
    # ═══════════════════════════════════════════════════════

    def _show_snapshot_chart(self) -> None:
        """재고 스냅샷 추이 차트"""
        from ..utils.constants import BOTH, END, tk, ttk
        from ..utils.custom_messagebox import CustomMessageBox

        try:
            rows = self.engine.db.fetchall("""
                SELECT snapshot_date, total_lots, total_weight_kg, 
                       available_weight_kg, picked_weight_kg
                FROM inventory_snapshot 
                ORDER BY snapshot_date DESC LIMIT 30
            """)

            if not rows:
                CustomMessageBox.showinfo(self.root, "재고 추이",
                    "스냅샷 데이터가 아직 없습니다.\n\n프로그램을 매일 실행하면 자동으로 축적됩니다.")
                return

            rows = list(reversed(rows))

            dialog = create_themed_toplevel(self.root)
            dialog.title("📊 재고 추이 (최근 30일)")
            dialog.geometry("800x400")
            apply_modal_window_options(dialog)
            dialog.transient(self.root)

            # 표 형태
            tree = ttk.Treeview(dialog, columns=('date', 'lots', 'total', 'avail', 'picked'),
                               show='headings', height=15)

            for col, text, w in [('date','날짜',100), ('lots','LOT수',60),
                                 ('total','총재고(MT)',100), ('avail','판매가능(MT)',100),
                                 ('picked','출고(MT)',100)]:
                tree.heading(col, text=text, anchor='center')
                tree.column(col, width=w, anchor='e' if col != 'date' else 'w')

            for r in rows:
                tree.insert('', END, values=(
                    r['snapshot_date'],
                    r['total_lots'],
                    f"{(r['total_weight_kg'] or 0)/1000:,.1f}",
                    f"{(r['available_weight_kg'] or 0)/1000:,.1f}",
                    f"{(r['picked_weight_kg'] or 0)/1000:,.1f}",
                ))

            tree.pack(fill=BOTH, expand=True, padx=10, pady=10)
            ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=5)

        except (RuntimeError, ValueError) as e:
            CustomMessageBox.showerror(self.root, "오류", f"스냅샷 조회 오류:\n{e}")

    # ═══════════════════════════════════════════════════════
    # v3.8.4 A7: 출고 거래명세서 PDF/Excel
    # ═══════════════════════════════════════════════════════

    def _generate_outbound_invoice(self) -> None:
        """출고 거래명세서 Excel 생성"""
        from ..utils.constants import BOTH, W, filedialog, tk, ttk
        from ..utils.custom_messagebox import CustomMessageBox

        # 고객 + 기간 선택 다이얼로그
        dialog = create_themed_toplevel(self.root)
        dialog.title("📄 거래명세서 생성")
        dialog.geometry("400x250")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=15)
        frame.pack(fill=BOTH, expand=True)

        ttk.Label(frame, text="고객명:").grid(row=0, column=0, sticky=W, pady=5)
        cust_var = tk.StringVar()

        # 고객 목록 조회
        try:
            customers = self.engine.db.fetchall(
                "SELECT DISTINCT customer FROM stock_movement WHERE customer != '' ORDER BY customer")
            cust_list = [c['customer'] for c in customers if c['customer']]
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError):
            cust_list = []

        ttk.Combobox(frame, textvariable=cust_var, values=cust_list, width=30).grid(
            row=0, column=1, sticky=W, pady=5)

        ttk.Label(frame, text="시작일:").grid(row=1, column=0, sticky=W, pady=5)
        from_var = tk.StringVar(value="")
        _e_adv_from = ttk.Entry(frame, textvariable=from_var, width=15)
        _e_adv_from.grid(row=1, column=1, sticky=W, pady=5)
        try:
            from gui_app_modular.utils.tree_enhancements import attach_date_placeholder
            attach_date_placeholder(_e_adv_from, from_var)
        except Exception as e:
            logger.warning(f'[UI] advanced_dialogs_mixin: {e}')
        ttk.Label(frame, text="종료일:").grid(row=2, column=0, sticky=W, pady=5)
        from datetime import date
        to_var = tk.StringVar(value="")
        _e_adv_to = ttk.Entry(frame, textvariable=to_var, width=15)
        _e_adv_to.grid(row=2, column=1, sticky=W, pady=5)
        try:
            from gui_app_modular.utils.tree_enhancements import attach_date_placeholder
            attach_date_placeholder(_e_adv_to, to_var)
        except Exception as e:
            logger.warning(f'[UI] advanced_dialogs_mixin: {e}')
        def do_generate():
            customer = cust_var.get().strip()
            date_from = from_var.get().strip()
            date_to = to_var.get().strip()

            if not customer:
                CustomMessageBox.showwarning(dialog, "입력 필요", "고객명을 선택하세요.")
                return

            # 출고 데이터 조회 (customer 컬럼 없어도 동작)
            try:
                for q, p in [
                    ("""SELECT lot_no, movement_type, qty_kg, customer, 
                            COALESCE(movement_date, created_at) AS movement_date, created_at
                        FROM stock_movement 
                        WHERE customer = ? AND movement_type = 'OUTBOUND'
                          AND COALESCE(movement_date, created_at) >= ? AND COALESCE(movement_date, created_at) <= ?
                        ORDER BY created_at""",
                     (customer, date_from, date_to + ' 23:59:59')),
                    ("""SELECT lot_no, movement_type, qty_kg, '' AS customer, 
                            created_at AS movement_date, created_at
                        FROM stock_movement 
                        WHERE movement_type = 'OUTBOUND'
                          AND created_at >= ? AND created_at <= ?
                        ORDER BY created_at""",
                     (date_from, date_to + ' 23:59:59')),
                ]:
                    try:
                        movements = self.engine.db.fetchall(q, p)
                        break
                    except (sqlite3.OperationalError,) as e:
                        if "no such column" in str(e).lower():
                            continue
                        raise
                else:
                    movements = []

                if not movements:
                    CustomMessageBox.showinfo(dialog, "결과 없음", "해당 기간 출고 이력이 없습니다.")
                    return

                # Excel 저장
                save_path = filedialog.asksaveasfilename(
                    parent=self.root,
                    title="거래명세서 저장",
                    defaultextension=".xlsx",
                    initialfile=f"거래명세서_{customer}_{date_from}_{date_to}.xlsx",
                    filetypes=[("Excel files", "*.xlsx")]
                )

                if not save_path:
                    return
                try:
                    from ..utils.excel_file_helper import get_unique_excel_path
                    save_path = get_unique_excel_path(save_path)
                except ImportError as e:
                    logger.warning(f"[do_generate] Suppressed: {e}")

                import openpyxl
                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "거래명세서"

                # 스타일
                title_font = Font(bold=True, size=16)
                header_font = Font(bold=True, color="FFFFFF", size=10)
                header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

                # 타이틀
                ws.merge_cells('A1:F1')
                ws['A1'] = f"거래명세서 — {customer}"
                ws['A1'].font = title_font

                ws['A2'] = f"기간: {date_from} ~ {date_to}"
                ws['A2'].font = Font(size=10, color='666666')

                # 헤더
                headers = ['No', 'LOT NO', '수량(kg)', '수량(MT)', '출고일', '비고']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

                # 데이터
                total_kg = 0
                for i, mv in enumerate(movements, 1):
                    qty = mv['qty_kg'] or 0
                    total_kg += qty

                    row_data = [
                        i,
                        mv['lot_no'],
                        f"{qty:,.0f}",
                        f"{qty/1000:.3f}",
                        str(mv['movement_date'] or '')[:10],
                        ''
                    ]
                    for col, val in enumerate(row_data, 1):
                        cell = ws.cell(row=4+i, column=col, value=val)
                        cell.border = border
                        if col in (3, 4):
                            cell.alignment = Alignment(horizontal='right')

                # 합계
                sum_row = 5 + len(movements)
                ws.cell(row=sum_row, column=1, value="합계").font = Font(bold=True)
                ws.cell(row=sum_row, column=3, value=f"{total_kg:,.0f}").font = Font(bold=True)
                ws.cell(row=sum_row, column=4, value=f"{total_kg/1000:.3f}").font = Font(bold=True)

                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 16
                ws.column_dimensions['C'].width = 14
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 14
                ws.column_dimensions['F'].width = 15

                wb.save(save_path)

                dialog.destroy()
                self._log(f"✅ 거래명세서 저장: {save_path}")
                CustomMessageBox.showinfo(self.root, "완료",
                    f"거래명세서가 저장되었습니다.\n\n"
                    f"고객: {customer}\n"
                    f"건수: {len(movements)}건\n"
                    f"총량: {total_kg/1000:.3f} MT\n\n"
                    f"파일: {save_path}")

            except ImportError:
                CustomMessageBox.showerror(dialog, "오류", "openpyxl이 필요합니다.")
            except (RuntimeError, ValueError) as e:
                CustomMessageBox.showerror(dialog, "오류", f"거래명세서 생성 오류:\n{e}")

        ttk.Button(frame, text="📄 생성", command=do_generate).grid(row=3, column=1, sticky=W, pady=15)
        ttk.Button(frame, text="취소", command=dialog.destroy).grid(row=3, column=0, sticky=W, pady=15)

    def _show_return_statistics(self) -> None:
        """v6.12.1: 반품 사유 통계 리포트 다이얼로그."""
        try:
            from ..dialogs.return_statistics_dialog import ReturnStatisticsDialog
            current_theme = getattr(self, '_current_theme', 'darkly')
            ReturnStatisticsDialog(self.root, self.engine, current_theme=current_theme)
        except ImportError as e:
            logger.debug(f"반품 통계 다이얼로그 로드 실패: {e}")
            CustomMessageBox.showerror(self.root, "오류", "반품 통계 다이얼로그를 불러올 수 없습니다.")

    def _send_return_alert_email(self) -> None:
        """v6.12.2: 반품 경고 이메일 수동 발송."""
        try:
            from features.notifications.return_alert_email import (
                check_return_alerts,
                create_default_email_config,
                load_email_config,
                send_return_alert_email,
            )
        except ImportError as e:
            CustomMessageBox.showerror(self.root, "오류", f"알림 모듈 로드 실패:\n{e}")
            return

        config = load_email_config()
        if not config.get('enabled'):
            if CustomMessageBox.askyesno(self.root, "이메일 알림 미설정",
                "이메일 알림이 비활성화 상태입니다.\n\n"
                "config_email.json 템플릿을 바탕화면에 생성하시겠습니까?\n"
                "(생성 후 SMTP 정보를 입력하고 enabled=true로 변경하세요)"):
                import os
                desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
                path = create_default_email_config(desktop)
                CustomMessageBox.showinfo(self.root, "완료",
                    f"config_email.json 템플릿이 생성되었습니다.\n\n{path}\n\n"
                    f"파일을 편집한 후 앱 루트 폴더에 복사하세요.")
            return

        alerts = check_return_alerts(self.engine)
        if not alerts:
            CustomMessageBox.showinfo(self.root, "안내", "현재 반품 경고 대상 LOT이 없습니다.")
            return

        lot_summary = '\n'.join(f"  • {a['lot_no']}: {a['count']}회" for a in alerts[:5])
        if not CustomMessageBox.askyesno(self.root, "반품 경고 이메일 발송",
            f"반복 반품 LOT {len(alerts)}건:\n{lot_summary}\n\n"
            f"수신자: {', '.join(config.get('recipients', []))}\n\n발송하시겠습니까?"):
            return

        result = send_return_alert_email(self.engine, config)
        if result['sent']:
            self._log(f"📧 반품 경고 이메일 발송 완료: {result['alert_count']}건")
            CustomMessageBox.showinfo(self.root, "발송 완료",
                f"반품 경고 이메일이 발송되었습니다.\n\n대상: {result['alert_count']}건")
        else:
            CustomMessageBox.showerror(self.root, "발송 실패",
                f"이메일 발송에 실패했습니다.\n\n{result.get('error', '알 수 없는 오류')}")
    def _show_email_config(self) -> None:
        """v6.12.2: 이메일 설정 GUI 편집기."""
        try:
            from ..dialogs.email_config_dialog import EmailConfigDialog
            current_theme = getattr(self, '_current_theme', 'darkly')
            EmailConfigDialog(self.root, current_theme=current_theme)
        except ImportError as e:
            logger.debug(f"이메일 설정 다이얼로그 로드 실패: {e}")
            CustomMessageBox.showerror(self.root, "오류", "이메일 설정 다이얼로그를 불러올 수 없습니다.")

    # ═══════════════════════════════════════════════════════════════
    # v7.7.0 — 정합성 검증 리포트 다이얼로그 (시각화)
    # ═══════════════════════════════════════════════════════════════

    def _on_integrity_report_v760(self) -> None:
        """v7.7.0: 정합성 검증 리포트 (v7.5.0~v7.6.0 강화 검증 시각화)"""
        try:
            from ..dialogs.integrity_v760_dialog import IntegrityV760Dialog
            IntegrityV760Dialog(self.root, self.engine)
        except Exception as e:
            logger.error(f"[v7.7.0] 정합성 리포트 오류: {e}", exc_info=True)
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showerror(self.root, "오류", f"정합성 리포트 오류:\n{e}")
            except Exception:
                logger.debug("[SUPPRESSED] exception in advanced_dialogs_mixin.py")  # noqa

    def _on_integrity_report(self) -> None:
        """v7.0.1: 정합성 검증 리포트 생성 (PDF + Excel)"""
        from datetime import datetime

        from ..utils.constants import filedialog

        today = datetime.now().strftime('%Y%m%d')

        file_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="정합성 검증 리포트 저장",
            defaultextension=".xlsx",
            initialfile=f"SQM-IntegrityReport-{today}.xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("PDF files", "*.pdf"),
                ("All files", "*.*")
            ]
        )

        if not file_path:
            return

        try:
            if file_path.lower().endswith('.pdf'):
                from features.reports.integrity_report import (
                    generate_integrity_report_pdf,
                )
                result = generate_integrity_report_pdf(self.engine, file_path)
            else:
                from features.reports.integrity_report import (
                    generate_integrity_report_excel,
                )
                result = generate_integrity_report_excel(self.engine, file_path)

            if result:
                CustomMessageBox.showinfo(
                    self.root, "완료",
                    f"정합성 검증 리포트 생성 완료\n\n{os.path.basename(file_path)}"
                )
                # 탐색기에서 열기
                import platform
                import subprocess
                if platform.system() == 'Windows':
                    subprocess.Popen(['explorer', '/select,', file_path])
            else:
                CustomMessageBox.showerror(self.root, "오류", "리포트 생성 실패")
        except Exception as e:
            logger.error(f"정합성 리포트 생성 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(self.root, "오류", f"리포트 생성 오류:\n{e}")

    # ═══════════════════════════════════════════════════════════════
    # v6.3.3 RUBI — Allocation 양식 미리보기
    # ═══════════════════════════════════════════════════════════════

    def _show_allocation_template_preview(self) -> None:
        """
        v6.3.3 RUBI: Allocation 양식 미리보기 다이얼로그.
        Song 양식 / Woo 양식 탭 전환 + 샘플 행 주황 강조 + Excel 다운로드.
        """
        try:
            from ..dialogs.allocation_template_dialog import AllocationTemplateDialog
            AllocationTemplateDialog(self.root)
        except Exception as e:
            logger.error(f"[_show_allocation_template_preview] 오류: {e}", exc_info=True)
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showerror(self.root, "오류", f"양식 미리보기 오류:\n{e}")
            except Exception:
                logger.debug("[SUPPRESSED] exception in advanced_dialogs_mixin.py")  # noqa

    # ═══════════════════════════════════════════════════════════════
    # v6.3.3 RUBI — 반송 출고 현황 조회
    # ═══════════════════════════════════════════════════════════════

    def _show_return_export_history(self) -> None:
        """
        v6.3.3 RUBI: 반송(export_type='반송') 출고 현황 조회 다이얼로그.

        allocation_plan 테이블에서 export_type != '' 인 행을 조회.
        - 기간 필터 (created_at 기준)
        - 상태 필터 (전체 / RESERVED / STAGED / EXECUTED / CANCELLED)
        - LOT 검색
        - Excel 내보내기
        """
        import tkinter as tk
        from tkinter import ttk, filedialog
        from datetime import datetime as _dt

        try:
            from ..utils.custom_messagebox import CustomMessageBox
            _mb = CustomMessageBox
        except ImportError:
            _mb = None

        # ── 다이얼로그 생성 ──────────────────────────────────
        dialog = create_themed_toplevel(self.root)
        dialog.title("🔄 반송 출고 현황 조회")
        dialog.geometry("1150x640")
        dialog.resizable(True, True)
        try:
            from ..utils.ui_constants import apply_modal_window_options, center_dialog
            apply_modal_window_options(dialog)
            center_dialog(dialog, self.root)
        except ImportError:
            logger.debug("[SUPPRESSED] exception in advanced_dialogs_mixin.py")  # noqa
        dialog.transient(self.root)
        dialog.grab_set()

        BG      = tc('bg_secondary')
        BG2     = '#283593'
        FG      = tc('text_primary')
        FG2     = '#b0bec5'
        ACCENT  = '#00e5ff'
        STRIPE  = '#303f9f'
        SEL     = '#651fff'
        ORANGE  = '#ff9100'
        GREEN   = '#00e676'
        RED     = '#ff1744'
        dialog.configure(bg=BG)

        # ── 필터 영역 ─────────────────────────────────────────
        filter_frame = tk.LabelFrame(
            dialog, text="  🔍 조회 조건  ",
            bg=BG2, fg=ACCENT,
            font=('맑은 고딕', 10, 'bold'),
            bd=1, relief='groove', padx=10, pady=6,
        )
        filter_frame.pack(fill='x', padx=10, pady=(10, 4))

        today = _dt.now()
        first_day = today.replace(day=1)

        def _lbl(parent, text):
            return tk.Label(parent, text=text, bg=BG2, fg=FG2,
                            font=('맑은 고딕', 10))
        def _entry(parent, width=12, val=''):
            e = tk.Entry(parent, width=width, bg=tc('bg_primary'), fg=FG,
                         insertbackground=FG,
                         font=('맑은 고딕', 10), relief='flat', bd=1)
            e.insert(0, val)
            return e

        row0 = tk.Frame(filter_frame, bg=BG2)
        row0.pack(fill='x', pady=2)

        _lbl(row0, '시작일').pack(side='left', padx=(0, 4))
        ent_start = _entry(row0, val=first_day.strftime('%Y-%m-%d'))
        ent_start.pack(side='left', padx=(0, 10))

        _lbl(row0, '종료일').pack(side='left', padx=(0, 4))
        ent_end = _entry(row0, val=today.strftime('%Y-%m-%d'))
        ent_end.pack(side='left', padx=(0, 16))

        _lbl(row0, '상태').pack(side='left', padx=(0, 4))
        status_var = tk.StringVar(value='전체')
        status_cb = ttk.Combobox(
            row0, textvariable=status_var,
            values=['전체', 'RESERVED', 'STAGED', 'EXECUTED', 'CANCELLED'],
            state='readonly', width=12,
        )
        status_cb.pack(side='left', padx=(0, 16))

        _lbl(row0, 'LOT 검색').pack(side='left', padx=(0, 4))
        lot_var = tk.StringVar()
        ent_lot = tk.Entry(row0, textvariable=lot_var, width=14,
                           bg=tc('bg_primary'), fg=FG, insertbackground=FG,
                           font=('맑은 고딕', 10), relief='flat', bd=1)
        ent_lot.pack(side='left', padx=(0, 16))

        _lbl(row0, '수출유형').pack(side='left', padx=(0, 4))
        etype_var = tk.StringVar(value='전체')
        etype_cb = ttk.Combobox(
            row0, textvariable=etype_var,
            values=['전체', '반송', '일반수출'],
            state='readonly', width=10,
        )
        etype_cb.pack(side='left', padx=(0, 16))

        tk.Button(
            row0, text='🔍 조회', font=('맑은 고딕', 10, 'bold'),
            bg=tc('info'), fg=FG, activebackground=tc('info'),
            relief='flat', bd=0, padx=14, pady=3, cursor='hand2',
            command=lambda: _do_query(),
        ).pack(side='left', padx=(0, 6))

        tk.Button(
            row0, text='📊 Excel 저장', font=('맑은 고딕', 10),
            bg=tc('info'), fg=FG, activebackground=tc('info'),
            relief='flat', bd=0, padx=14, pady=3, cursor='hand2',
            command=lambda: _export_excel(),
        ).pack(side='left', padx=(0, 6))

        # ── 요약 레이블 ──────────────────────────────────────
        summary_var = tk.StringVar(value='조회 결과가 여기에 표시됩니다.')
        tk.Label(dialog, textvariable=summary_var,
                 bg=BG, fg=FG2, font=('맑은 고딕', 10),
                 anchor='w', padx=12).pack(fill='x')

        # ── Treeview ─────────────────────────────────────────
        tree_frame = tk.Frame(dialog, bg=BG)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=(0, 4))

        style = ttk.Style()
        # v9.0: theme_use('default') 제거
        style.configure('RetExport.Treeview',
                        background=BG2, fieldbackground=BG2, foreground=FG,
                        rowheight=23, font=('맑은 고딕', 10))
        style.configure('RetExport.Treeview.Heading',
                        background=BG, foreground=ACCENT,
                        font=('맑은 고딕', 10, 'bold'), relief='flat',
                        anchor='center')
        style.map('RetExport.Treeview',
                  background=[('selected', SEL)],
                  foreground=[('selected', FG)])

        COLS = [
            ('id',           'ID',         45,  'center'),
            ('lot_no',       'LOT No',     115, 'center'),
            ('sub_lt',       '톤백#',       55,  'center'),
            ('customer',     '고객',        160, 'w'),
            ('sale_ref',     'SALE REF',   80,  'center'),
            ('qty_mt',       'QTY (MT)',   80,  'center'),
            ('export_type',  '수출유형',   80,  'center'),
            ('status',       '상태',        90,  'center'),
            ('outbound_date','출고일',      95,  'center'),
            ('created_at',   '등록일시',   145, 'center'),
            ('source_file',  '소스 파일',  200, 'w'),
        ]

        tree = ttk.Treeview(
            tree_frame,
            style='RetExport.Treeview',
            columns=[c[0] for c in COLS],
            show='headings',
            selectmode='browse',
        )
        for cid, cname, cw, anchor in COLS:
            tree.heading(cid, text=cname, anchor='center')
            tree.column(cid, width=cw, minwidth=30, anchor=anchor, stretch=(cid == 'customer'))

        tree.tag_configure('normal',  background=BG2,    foreground=FG)
        tree.tag_configure('stripe',  background=STRIPE, foreground=FG)
        tree.tag_configure('bangsong',background=tc('bg_secondary'), foreground=ORANGE)
        tree.tag_configure('bangsong_stripe', background=tc('bg_secondary'), foreground=ORANGE)
        tree.tag_configure('executed',foreground=GREEN)
        tree.tag_configure('cancel',  foreground=RED)

        vsb = ttk.Scrollbar(tree_frame, orient='vertical',   command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal',  command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side='right',  fill='y')
        try:
            from gui_app_modular.utils.tree_enhancements import TreeviewTotalFooter as _TTF
            _TTF(tree_frame, tree, [],
                 {}, {}).pack(fill='x')
        except Exception as e:
            logger.warning(f'[UI] advanced_dialogs_mixin: {e}')
        hsb.pack(side='bottom', fill='x')
        tree.pack(fill='both', expand=True)

        # ── 하단 버튼 ─────────────────────────────────────────
        btn_frame = tk.Frame(dialog, bg=BG2, pady=6)
        btn_frame.pack(fill='x')
        tk.Button(
            btn_frame, text='✕  닫기',
            font=('Malgun Gothic', 10),
            bg=tc('info'), fg=FG, activebackground=tc('info'),
            relief='flat', bd=0, padx=20, pady=6, cursor='hand2',
            command=dialog.destroy,
        ).pack(side='right', padx=16)

        # ── 쿼리 함수 ─────────────────────────────────────────
        _last_rows = []

        def _do_query():
            nonlocal _last_rows
            tree.delete(*tree.get_children())

            s_date = ent_start.get().strip() or '2000-01-01'
            e_date = ent_end.get().strip() or '9999-12-31'
            status = status_var.get()
            lot_kw = lot_var.get().strip()
            etype  = etype_var.get()

            params = [s_date, e_date]
            where  = ["DATE(ap.created_at) BETWEEN ? AND ?",
                      "ap.export_type IS NOT NULL",
                      "ap.export_type != ''"]

            if status != '전체':
                where.append("ap.status = ?")
                params.append(status)
            if lot_kw:
                where.append("ap.lot_no LIKE ?")
                params.append(f'%{lot_kw}%')
            if etype != '전체':
                where.append("ap.export_type = ?")
                params.append(etype)

            sql = f"""
                SELECT
                    ap.id,
                    ap.lot_no,
                    COALESCE(ap.sub_lt, '') AS sub_lt,
                    COALESCE(ap.customer, '') AS customer,
                    COALESCE(ap.sale_ref, '') AS sale_ref,
                    COALESCE(ap.qty_mt, 0) AS qty_mt,
                    COALESCE(ap.export_type, '') AS export_type,
                    ap.status,
                    COALESCE(ap.outbound_date, '') AS outbound_date,
                    ap.created_at,
                    COALESCE(ap.source_file, '') AS source_file
                FROM allocation_plan ap
                WHERE {' AND '.join(where)}
                ORDER BY ap.created_at DESC
            """
            try:
                rows = self.engine.db.fetchall(sql, tuple(params)) or []
            except Exception as e:
                logger.error(f"[반송출고현황] 쿼리 오류: {e}")
                rows = []

            _last_rows = rows

            qty_total = 0.0
            bangsong_cnt = 0
            for i, row in enumerate(rows):
                vals = (
                    row.get('id', ''),
                    row.get('lot_no', ''),
                    row.get('sub_lt', ''),
                    row.get('customer', ''),
                    row.get('sale_ref', ''),
                    f"{float(row.get('qty_mt', 0)):.3f}",
                    row.get('export_type', ''),
                    row.get('status', ''),
                    row.get('outbound_date', ''),
                    str(row.get('created_at', ''))[:19],
                    row.get('source_file', ''),
                )
                etype_val = str(row.get('export_type', ''))
                stat_val  = str(row.get('status', ''))
                is_bangsong = ('반송' in etype_val)
                if is_bangsong:
                    bangsong_cnt += 1
                    tag = 'bangsong' if i % 2 == 0 else 'bangsong_stripe'
                elif stat_val == 'EXECUTED':
                    tag = 'executed'
                elif 'CANCEL' in stat_val:
                    tag = 'cancel'
                else:
                    tag = 'normal' if i % 2 == 0 else 'stripe'
                try:
                    qty_total += float(row.get('qty_mt', 0))
                except (ValueError, TypeError):
                    logger.debug("[SUPPRESSED] exception in advanced_dialogs_mixin.py")  # noqa
                tree.insert('', 'end', values=vals, tags=(tag,))

            summary_var.set(
                f"조회 결과: {len(rows)}건  │  반송: {bangsong_cnt}건  │  "
                f"총 QTY: {qty_total:.3f} MT  │  기간: {s_date} ~ {e_date}"
            )

        def _export_excel():
            if not _last_rows:
                if _mb:
                    _mb.showwarning(dialog, '알림', '조회 결과가 없습니다.')
                return
            save_path = filedialog.asksaveasfilename(
                parent=dialog,
                title='반송 출고 현황 저장',
                defaultextension='.xlsx',
                filetypes=[('Excel', '*.xlsx')],
                initialfile=f"반송출고현황_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx",
            )
            if not save_path:
                return
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = '반송출고현황'

                header_fill = PatternFill('solid', fgColor='1a237e')
                header_font = Font(bold=True, color='00e5ff', size=10)
                col_names = [c[1] for c in COLS]
                for ci, name in enumerate(col_names, 1):
                    cell = ws.cell(row=1, column=ci, value=name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                bangsong_fill = PatternFill('solid', fgColor='3e1a00')
                normal_fill   = PatternFill('solid', fgColor='283593')
                stripe_fill   = PatternFill('solid', fgColor='303f9f')

                for ri, row in enumerate(_last_rows, 2):
                    vals = [
                        row.get('id', ''),
                        row.get('lot_no', ''),
                        row.get('sub_lt', ''),
                        row.get('customer', ''),
                        row.get('sale_ref', ''),
                        float(row.get('qty_mt', 0)),
                        row.get('export_type', ''),
                        row.get('status', ''),
                        row.get('outbound_date', ''),
                        str(row.get('created_at', ''))[:19],
                        row.get('source_file', ''),
                    ]
                    is_bangsong = '반송' in str(row.get('export_type', ''))
                    for ci, val in enumerate(vals, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        if is_bangsong:
                            cell.fill = bangsong_fill
                            cell.font = Font(color='ff9100')
                        else:
                            cell.fill = normal_fill if ri % 2 == 0 else stripe_fill
                            cell.font = Font(color='ffffff')
                        cell.alignment = Alignment(horizontal='center')

                # 컬럼 너비
                for ci, (_, cname, cw, _) in enumerate(COLS, 1):
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(ci)
                    ].width = max(10, len(cname) + 4)
                ws.column_dimensions['D'].width = 25  # 고객
                ws.column_dimensions['K'].width = 35  # 소스파일

                wb.save(save_path)
                if _mb:
                    _mb.showinfo(dialog, '저장 완료',
                                 f'Excel 저장 완료:\n{save_path}')
            except Exception as e:
                logger.error(f"[반송출고현황] Excel 저장 오류: {e}")
                if _mb:
                    _mb.showerror(dialog, '오류', f'저장 실패:\n{e}')

        # 초기 조회
        _do_query()

    # ── v7.2.0: 입고 파싱 템플릿 관리 ───────────────────────────────────────

    def _on_inbound_template_manage(self) -> None:
        """입고 파싱 템플릿 관리 다이얼로그 열기 (메뉴 → 입고 → 📝 입고 파싱 템플릿 관리)."""
        try:
            from gui_app_modular.dialogs.inbound_template_dialog import InboundTemplateDialog
            current_theme = getattr(self, 'current_theme', 'darkly')
            InboundTemplateDialog(self.root, self.engine, current_theme=current_theme)
        except Exception as e:
            logger.error(f"[템플릿관리] 오류: {e}", exc_info=True)
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showerror(self.root, '오류', f'템플릿 관리 창을 열 수 없습니다:\n{e}')
            except Exception:
                logger.debug("[SUPPRESSED] exception in advanced_dialogs_mixin.py")  # noqa

    def _on_picking_template_manage(self) -> None:
        """출고 피킹 템플릿 관리 다이얼로그 열기 (메뉴 → 출고 → 📦 출고 피킹 템플릿 관리)."""
        try:
            from gui_app_modular.dialogs.picking_template_dialog import PickingTemplateDialog
            current_theme = getattr(self, 'current_theme', 'darkly')
            PickingTemplateDialog(self.root, self.engine, current_theme=current_theme)
        except Exception as e:
            logger.error(f"[피킹템플릿관리] 오류: {e}", exc_info=True)
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showerror(
                    self.root, '오류', f'피킹 템플릿 관리 창을 열 수 없습니다:\n{e}')
            except Exception:
                logger.debug("[SUPPRESSED] exception in advanced_dialogs_mixin.py")  # noqa

    def _on_move_approval_queue(self):
        """⑤ v6.6.0: 대량 이동 PENDING → Supervisor 승인 다이얼로그."""
        try:
            from engine_modules.inventory_modular.move_approval_dialog_helper import (
                show_move_approval_dialog
            )
        except ImportError as e:
            self._log(f"[이동승인] import 실패: {e}")
            return
        show_move_approval_dialog(self.root, self.engine)


    # ── v8.6.0: 출고 보고서 생성 (보고서 메뉴) ─────────────────────────────

    def _on_detail_of_outbound_report(self) -> None:
        """📦 Detail of Outbound — Excel + PDF 동시 생성 (보고서 메뉴)

        필터: 날짜 또는 Sales Order No 선택
        생성: Detail_of_Outbound_{date}.xlsx + .pdf
        """
        self._open_outbound_report_dialog(mode='outbound')

    def _on_sales_order_dn_report(self) -> None:
        """📋 Sales Order DN — Excel + PDF 동시 생성 (보고서 메뉴)

        필터: 날짜 또는 Sales Order No 선택
        생성: Sales_order_DN_{date}.xlsx + .pdf
        """
        self._open_outbound_report_dialog(mode='dn')

    def _open_outbound_report_dialog(self, mode: str = 'outbound') -> None:
        """출고 보고서 공통 필터 다이얼로그.

        mode='outbound' → Detail_of_Outbound
        mode='dn'       → Sales_order_DN
        """
        import os
        import sqlite3
        from datetime import date
        from ..utils.constants import BOTH, W, E, X, filedialog, tk, ttk
        from ..utils.custom_messagebox import CustomMessageBox
        from ..utils.ui_constants import create_themed_toplevel, apply_modal_window_options

        is_dn = (mode == 'dn')
        title_str  = "📋 Sales Order DN 보고서" if is_dn else "📦 Detail of Outbound 보고서"
        file_prefix = "Sales_order_DN" if is_dn else "Detail_of_Outbound"

        # ── 필터 다이얼로그 ──────────────────────────────────────────
        dialog = create_themed_toplevel(self.root)
        dialog.title(title_str)
        dialog.geometry("420x260")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding=18)
        frame.pack(fill=BOTH, expand=True)

        ttk.Label(frame, text="출고일 (YYYY-MM-DD):",
                  font=("Arial", 10)).grid(row=0, column=0, sticky=W, pady=6)
        date_var = tk.StringVar(value=date.today().strftime("%Y-%m-%d"))
        _de_frame = ttk.Frame(frame)
        _de_frame.grid(row=0, column=1, sticky=W, pady=6)
        _de = ttk.Entry(_de_frame, textvariable=date_var, width=14)
        _de.pack(side='left', padx=(0,2))
        try:
            from ..utils.tree_enhancements import show_date_calendar, attach_date_placeholder
            attach_date_placeholder(_de, date_var)
            ttk.Button(_de_frame, text="📅", width=3, command=lambda: show_date_calendar(
                dialog, date_var.get(), lambda ymd: date_var.set(ymd)
            )).pack(side='left')
        except Exception as e:
            logger.debug(f"[UI] 날짜 캘린더 연결 스킵: {e}")

        ttk.Label(frame, text="Sales Order No (선택):",
                  font=("Arial", 10)).grid(row=1, column=0, sticky=W, pady=6)

        # sold_table에서 sales_order_no 목록 조회
        try:
            so_rows = self.engine.db.fetchall(
                "SELECT DISTINCT sales_order_no FROM sold_table "
                "WHERE sales_order_no IS NOT NULL AND sales_order_no != '' "
                "ORDER BY sales_order_no DESC LIMIT 50"
            )
            so_list = [r.get('sales_order_no') if isinstance(r, dict) else r[0]
                       for r in (so_rows or [])]
        except (sqlite3.OperationalError, OSError):
            so_list = []

        so_var = tk.StringVar()
        so_cb = ttk.Combobox(frame, textvariable=so_var, values=so_list, width=30)
        so_cb.grid(row=1, column=1, sticky=W, pady=6)
        if so_list:
            so_cb.set(so_list[0])

        ttk.Label(frame, text="* Sales Order No 입력 시 날짜 무시",
                  font=("Arial", 8), foreground="gray").grid(
            row=2, column=0, columnspan=2, sticky=W)

        # 저장 경로
        ttk.Label(frame, text="저장 폴더:",
                  font=("Arial", 10)).grid(row=3, column=0, sticky=W, pady=6)
        today_str = date.today().strftime("%Y_%m_%d")
        default_dir = tk.StringVar(value=os.path.expanduser("~\\Desktop"))
        dir_entry = ttk.Entry(frame, textvariable=default_dir, width=22)
        dir_entry.grid(row=3, column=1, sticky=W+E, pady=6)

        def browse_dir():
            d = filedialog.askdirectory(parent=dialog, title="저장 폴더 선택")
            if d:
                default_dir.set(d)

        ttk.Button(frame, text="찾기", command=browse_dir, width=5).grid(
            row=3, column=2, padx=4)

        result_holder = {}

        def do_generate():
            sale_ref   = so_var.get().strip() or None
            out_date   = date_var.get().strip() or None
            save_dir   = default_dir.get().strip()

            if not sale_ref and not out_date:
                CustomMessageBox.showwarning(dialog, "입력 오류",
                    "출고일 또는 Sales Order No 중 하나는 입력해 주세요.")
                return

            # 파일명 결정
            date_tag = (out_date or today_str).replace("-", "_")
            fname = f"{file_prefix}_{date_tag}.xlsx"
            output_path = os.path.join(save_dir, fname)

            dialog.destroy()
            result_holder['sale_ref']   = sale_ref
            result_holder['out_date']   = out_date
            result_holder['output_path'] = output_path

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=12)
        ttk.Button(btn_frame, text="생성", command=do_generate,
                   width=10).pack(side='left', padx=6)
        ttk.Button(btn_frame, text="취소",
                   command=dialog.destroy, width=10).pack(side='left', padx=6)

        self.root.wait_window(dialog)

        if not result_holder:
            return

        sale_ref    = result_holder['sale_ref']
        out_date    = result_holder['out_date']
        output_path = result_holder['output_path']

        # ── 생성 실행 ────────────────────────────────────────────────
        try:
            if is_dn:
                actual = self.engine._export_sales_order_dn_report(
                    output_path, sale_ref=sale_ref, outbound_date=out_date)
                # v8.6.2: 미완료 Sales Order → 경고 팝업 후 중단
                if actual and str(actual).startswith("INCOMPLETE:"):
                    progress_info = actual.replace("INCOMPLETE:", "")
                    out_cnt, total_cnt = progress_info.split("/") if "/" in progress_info else ("?", "?")
                    warn_msg = (
                        f"Sales Order '{sale_ref}' 의 모든 화물이 출고 완료되지 않았습니다.\n\n"
                        f"  출고 완료: {out_cnt}개 LOT\n"
                        f"  배정 전체: {total_cnt}개 LOT\n\n"
                        f"모든 화물 출고 완료 후 Sales Order DN을 생성할 수 있습니다.\n"
                        f"(출고 현황은 화물 총괄 탭에서 확인하세요)"
                    )
                    CustomMessageBox.showwarning(
                        self.root, "Sales Order 미완료", warn_msg
                    )
                    return
            else:
                actual = self.engine._export_outbound_report(
                    output_path, sale_ref=sale_ref, outbound_date=out_date)

            if not actual:
                CustomMessageBox.showwarning(
                    self.root, "데이터 없음",
                    "해당 조건의 출고 데이터가 없습니다.\n"
                    "출고 처리(confirm_outbound) 완료 후 다시 시도하세요.")
                return

            # PDF 경로 안내
            pdf_path = actual.rsplit('.', 1)[0] + '.pdf'
            pdf_exists = os.path.exists(pdf_path)
            msg = (f"✅ 생성 완료\n\n"
                   f"Excel: {os.path.basename(actual)}\n"
                   f"PDF:   {'생성됨' if pdf_exists else '생성 안됨 (reportlab 필요)'}\n\n"
                   f"폴더: {os.path.dirname(actual)}\n\n"
                   f"파일을 여시겠습니까?")

            if CustomMessageBox.askyesno(self.root, title_str, msg):
                import subprocess, sys
                try:
                    if sys.platform == 'win32':
                        os.startfile(actual)
                    else:
                        subprocess.Popen(['xdg-open', actual])
                except Exception as e:
                    logger.warning(f"[UI] 파일 열기 실패: {actual}: {e}")

        except Exception as e:
            logger.error(f"[{title_str}] 생성 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(self.root, "오류",
                f"보고서 생성 중 오류가 발생했습니다:\n{e}")
