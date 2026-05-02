# -*- coding: utf-8 -*-
"""
SQM Inventory - Validation Mixin
================================

v2.9.91 - Extracted from gui_app.py

Preflight validation, data validation, and validation result dialogs
"""

from gui_app_modular.utils.ui_constants import create_themed_toplevel  # v8.0.9
from gui_app_modular.utils.ui_constants import tc
from engine_modules.constants import STATUS_AVAILABLE
import sqlite3
import logging
from ..utils.ui_constants import CustomMessageBox
from datetime import datetime

logger = logging.getLogger(__name__)


class ValidationMixin:
    """
    Validation mixin
    
    Mixed into SQMInventoryApp class
    """
    


    def _show_validation_result(self, title: str, result, validator) -> None:
        """Show validation result dialog"""
        from ..utils.constants import tk, ttk, filedialog
        from ..utils.constants import BOTH, LEFT, RIGHT, X, Y, VERTICAL, END, W
        from ..utils.ui_constants import setup_dialog_geometry_persistence
        
        dialog = create_themed_toplevel(self.root)
        dialog.title(f"Validation: {title}")
        dialog.transient(self.root)
        setup_dialog_geometry_persistence(dialog, "validation_result_dialog", self.root, "large")
        
        # Summary frame
        summary_frame = ttk.LabelFrame(dialog, text="Summary")
        summary_frame.pack(fill=X, padx=10, pady=10)
        
        status_text = "PASS" if result.is_valid else "FAIL"
        "green" if result.is_valid else "red"
        
        ttk.Label(summary_frame, text=f"Result: {status_text}",
                  font=('', 16, 'bold')).grid(row=0, column=0, sticky=W, padx=10)
        ttk.Label(summary_frame, text=f"Rows: {result.total_rows}").grid(row=0, column=1, padx=20)
        ttk.Label(summary_frame, text=f"Errors: {result.error_count}",
                  foreground=tc('danger') if result.error_count > 0 else tc('text_primary')).grid(row=0, column=2, padx=20)
        ttk.Label(summary_frame, text=f"Warnings: {result.warning_count}",
                  foreground=tc('warning') if result.warning_count > 0 else tc('text_primary')).grid(row=0, column=3, padx=20)
        
        # Detail frame
        detail_frame = ttk.LabelFrame(dialog, text="Details")
        detail_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Treeview
        columns = ("level", "row", "column", "value", "message", "suggestion")
        tree = ttk.Treeview(detail_frame, columns=columns, show="headings", height=15)
        
        tree.heading("level", text="Level", anchor='center')
        tree.heading("row", text="Row", anchor='center')
        tree.heading("column", text="Column", anchor='center')
        tree.heading("value", text="Value", anchor='center')
        tree.heading("message", text="Message", anchor='center')
        tree.heading("suggestion", text="Suggestion", anchor='center')        
        tree.column("level", width=60, anchor="center")
        tree.column("row", width=50, anchor="center")
        tree.column("column", width=100)
        tree.column("value", width=100)
        tree.column("message", width=250)
        tree.column("suggestion", width=200)
        
        scrollbar = tk.Scrollbar(detail_frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)

        # v8.1.8: 가로 스크롤바 추가
        _xsb = tk.Scrollbar(detail_frame, orient='horizontal', command=tree.xview)
        tree.configure(xscrollcommand=_xsb.set)
        _xsb.pack(side='bottom', fill='x')
        
        # Populate errors
        for err in result.errors:
            level_text = "ERROR" if err.level.value == "ERROR" else "WARNING"
            tree.insert('', END, values=(
                level_text,
                err.row if err.row > 0 else "-",
                err.column or "-",
                str(err.value)[:30] if err.value else "-",
                err.message,
                err.suggestion or "-"
            ))
        
        # Button frame
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=X, padx=10, pady=10)
        
        def export_errors():
            save_path = filedialog.asksaveasfilename(
                parent=self.root,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"validation_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not save_path:
                return
            
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Validation Errors"
                
                headers = ["Level", "Row", "Column", "Value", "Message", "Suggestion"]
                header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                
                for row_idx, err in enumerate(result.errors, 2):
                    ws.cell(row=row_idx, column=1, value=err.level.value)
                    ws.cell(row=row_idx, column=2, value=err.row if err.row > 0 else "")
                    ws.cell(row=row_idx, column=3, value=err.column or "")
                    ws.cell(row=row_idx, column=4, value=str(err.value) if err.value else "")
                    ws.cell(row=row_idx, column=5, value=err.message)
                    ws.cell(row=row_idx, column=6, value=err.suggestion or "")
                
                try:
                    from utils.sqm_excel_alignment import apply_sqm_workbook_alignment
                    apply_sqm_workbook_alignment(wb)
                except Exception:
                    pass
                wb.save(save_path)
                CustomMessageBox.showinfo(self.root, "Complete", f"Errors exported to:\n{save_path}")
                
            except (RuntimeError, ValueError) as e:
                CustomMessageBox.showerror(self.root, "Error", f"Export failed: {e}")
        
        ttk.Button(btn_frame, text="Export Errors", command=export_errors).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="Close", command=dialog.destroy).pack(side=RIGHT, padx=5)
        
        self._log(f"Validation {title}: {'PASS' if result.is_valid else f'{result.error_count} errors, {result.warning_count} warnings'}")
    


    def _check_lot_exists(self, lot_no: str) -> bool:
        """Check if LOT exists in database"""
        try:
            result = self.engine.db.fetchone(
                "SELECT 1 FROM inventory WHERE lot_no = ?",
                (lot_no,)
            )
            return result is not None
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError):
            return False
    
    def _check_tonbag_available(self, lot_no: str, sub_lt: int) -> bool:
        """Check if tonbag is available for outbound"""
        try:
            result = self.engine.db.fetchone(
                "SELECT status FROM inventory_tonbag WHERE lot_no = ? AND sub_lt = ?",
                (lot_no, sub_lt)
            )
            return result and result.get('status') == STATUS_AVAILABLE
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError):
            return False
