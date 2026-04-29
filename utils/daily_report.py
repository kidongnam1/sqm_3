# -*- coding: utf-8 -*-
"""
utils/daily_report.py — SQM v8.3.0
=====================================
일별 처리 건수·무게 자동 리포트 생성기

기능:
  1. 당일 입고·출고·반품·배정 집계
  2. 엑셀 파일로 저장 (reports/ 폴더)
  3. 이메일로 자동 전송 (error_notifier 연동)
  4. 스케줄러 연동 (매일 자정 자동 실행)

사용 예:
    from utils.daily_report import generate_daily_report

    # 오늘 리포트 생성
    result = generate_daily_report(db)
    # print(result['filepath'])  # 파일 경로 출력

    # 특정 날짜 리포트
    result = generate_daily_report(db, date_str='2026-03-23')
"""

import logging
import os
from datetime import datetime, date
from typing import Optional

logger = logging.getLogger(__name__)

REPORT_DIR = 'reports'


def _ensure_report_dir() -> str:
    """reports/ 폴더 생성."""
    os.makedirs(REPORT_DIR, exist_ok=True)
    return REPORT_DIR


def _collect_daily_data(db, date_str: str) -> dict:
    """
    특정 날짜의 처리 데이터 집계.

    Returns:
        {
          'inbound':  [{'lot_no', 'product', 'weight_kg', 'bl_no', 'created_at'}],
          'outbound': [{'lot_no', 'customer', 'weight_kg', 'created_at'}],
          'return':   [{'lot_no', 'weight_kg', 'reason', 'created_at'}],
          'reserved': [{'lot_no', 'customer', 'weight_kg', 'sale_ref', 'created_at'}],
          'summary':  {'inbound_cnt', 'outbound_cnt', 'return_cnt', 'reserved_cnt',
                       'inbound_kg', 'outbound_kg', 'return_kg'},
        }
    """
    d_start = date_str + ' 00:00:00'
    d_end   = date_str + ' 23:59:59'

    def fetch(sql, params=()):
        try:
            rows = db.fetchall(sql, params)
            return [dict(r) if isinstance(r, dict) else
                    {k: v for k, v in zip(
                        [desc[0] for desc in
                         (getattr(db, '_last_cursor_desc', None) or [])],
                        r)} for r in rows] if rows else []
        except Exception as e:
            logger.debug(f"[DailyReport] 조회 스킵: {e}")
            return []

    # ── 입고 ────────────────────────────────────────────────────
    inbound = fetch("""
        SELECT lot_no,
               COALESCE(product, '') AS product,
               COALESCE(net_weight, 0) AS weight_kg,
               COALESCE(bl_no, '') AS bl_no,
               created_at
        FROM inventory
        WHERE created_at BETWEEN ? AND ?
        ORDER BY created_at
    """, (d_start, d_end))

    # ── 출고 (stock_movement 기준) ──────────────────────────────
    outbound = fetch("""
        SELECT sm.lot_no,
               COALESCE(sm.customer, '') AS customer,
               COALESCE(sm.qty_kg, 0) AS weight_kg,
               sm.created_at
        FROM stock_movement sm
        WHERE sm.movement_type IN ('OUTBOUND', 'SOLD')
          AND sm.created_at BETWEEN ? AND ?
        ORDER BY sm.created_at
    """, (d_start, d_end))

    # ── 반품 ────────────────────────────────────────────────────
    ret = fetch("""
        SELECT lot_no,
               COALESCE(weight_kg, 0) AS weight_kg,
               COALESCE(reason, '') AS reason,
               created_at
        FROM return_history
        WHERE created_at BETWEEN ? AND ?
        ORDER BY created_at
    """, (d_start, d_end))

    # ── 배정 (Allocation RESERVED) ──────────────────────────────
    reserved = fetch("""
        SELECT lot_no,
               COALESCE(sold_to, '') AS customer,
               COALESCE(qty_mt * 1000, 0) AS weight_kg,
               COALESCE(sale_ref, '') AS sale_ref,
               created_at
        FROM allocation_plan
        WHERE status IN ('RESERVED', 'STAGED', 'PENDING_APPROVAL')
          AND created_at BETWEEN ? AND ?
        ORDER BY created_at
    """, (d_start, d_end))

    # ── 요약 집계 ────────────────────────────────────────────────
    summary = {
        'date':          date_str,
        'inbound_cnt':   len(inbound),
        'outbound_cnt':  len(outbound),
        'return_cnt':    len(ret),
        'reserved_cnt':  len(reserved),
        'inbound_kg':    sum(r.get('weight_kg', 0) for r in inbound),
        'outbound_kg':   sum(r.get('weight_kg', 0) for r in outbound),
        'return_kg':     sum(r.get('weight_kg', 0) for r in ret),
        'reserved_kg':   sum(r.get('weight_kg', 0) for r in reserved),
    }

    return {
        'inbound':  inbound,
        'outbound': outbound,
        'return':   ret,
        'reserved': reserved,
        'summary':  summary,
    }


def _write_excel(data: dict, filepath: str) -> bool:
    """집계 데이터 → 엑셀 파일 저장."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("[DailyReport] openpyxl 없음 — pip install openpyxl")
        return False

    wb = openpyxl.Workbook()
    summary = data['summary']
    date_str = summary['date']

    # ── 색상 정의 ────────────────────────────────────────────────
    HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
    HDR_FONT  = Font(color='FFFFFF', bold=True, size=10)
    SUM_FILL  = PatternFill('solid', fgColor='D6E4F0')
    SUM_FONT  = Font(bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=13, color='1F4E79')
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_header(ws, row, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

    def write_rows(ws, start_row, rows, keys):
        for r_i, row_data in enumerate(rows, start_row):
            for c_i, key in enumerate(keys, 1):
                cell = ws.cell(row=r_i, column=c_i,
                               value=row_data.get(key, ''))
                cell.border = border
                cell.alignment = Alignment(horizontal='left')

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    # ── 1. 요약 시트 ────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = '일별요약'
    ws_sum['A1'] = f'SQM 광양 창고 일별 처리 현황 — {date_str}'
    ws_sum['A1'].font = TITLE_FONT
    ws_sum.merge_cells('A1:D1')

    headers_sum = ['구분', '건수', '중량 (kg)', '중량 (MT)']
    set_header(ws_sum, 3, headers_sum)
    rows_sum = [
        ('입고',   summary['inbound_cnt'],  summary['inbound_kg'],  round(summary['inbound_kg']/1000, 3)),
        ('출고',   summary['outbound_cnt'], summary['outbound_kg'], round(summary['outbound_kg']/1000, 3)),
        ('반품',   summary['return_cnt'],   summary['return_kg'],   round(summary['return_kg']/1000, 3)),
        ('배정',   summary['reserved_cnt'], summary['reserved_kg'], round(summary['reserved_kg']/1000, 3)),
    ]
    for r_i, (label, cnt, kg, mt) in enumerate(rows_sum, 4):
        for c_i, val in enumerate([label, cnt, f'{kg:,.0f}', f'{mt:.3f}'], 1):
            cell = ws_sum.cell(row=r_i, column=c_i, value=val)
            cell.border = border
            if r_i == 4:
                cell.fill = SUM_FILL
                cell.font = SUM_FONT
    auto_width(ws_sum)

    # ── 2. 입고 시트 ────────────────────────────────────────────
    if data['inbound']:
        ws_in = wb.create_sheet('입고')
        ws_in['A1'] = f'입고 내역 — {date_str}'
        ws_in['A1'].font = TITLE_FONT
        set_header(ws_in, 3, ['LOT No', '제품명', '중량 (kg)', 'BL No', '입고시각'])
        write_rows(ws_in, 4, data['inbound'],
                   ['lot_no', 'product', 'weight_kg', 'bl_no', 'created_at'])
        auto_width(ws_in)

    # ── 3. 출고 시트 ────────────────────────────────────────────
    if data['outbound']:
        ws_out = wb.create_sheet('출고')
        ws_out['A1'] = f'출고 내역 — {date_str}'
        ws_out['A1'].font = TITLE_FONT
        set_header(ws_out, 3, ['LOT No', '고객사', '중량 (kg)', '출고시각'])
        write_rows(ws_out, 4, data['outbound'],
                   ['lot_no', 'customer', 'weight_kg', 'created_at'])
        auto_width(ws_out)

    # ── 4. 반품 시트 ────────────────────────────────────────────
    if data['return']:
        ws_ret = wb.create_sheet('반품')
        ws_ret['A1'] = f'반품 내역 — {date_str}'
        ws_ret['A1'].font = TITLE_FONT
        set_header(ws_ret, 3, ['LOT No', '중량 (kg)', '반품사유', '반품시각'])
        write_rows(ws_ret, 4, data['return'],
                   ['lot_no', 'weight_kg', 'reason', 'created_at'])
        auto_width(ws_ret)

    wb.save(filepath)
    logger.info(f"[DailyReport] 엑셀 저장: {filepath}")
    return True


def generate_daily_report(
    db,
    date_str: Optional[str] = None,
    send_email: bool = True,
) -> dict:
    """
    일별 처리 리포트 생성.

    Args:
        db        : SQMDatabase 인스턴스
        date_str  : 'YYYY-MM-DD' (기본값: 오늘)
        send_email: True이면 이메일 전송

    Returns:
        {'success', 'filepath', 'summary', 'email_sent'}
    """
    if not date_str:
        date_str = date.today().strftime('%Y-%m-%d')

    result = {
        'success':    False,
        'filepath':   '',
        'summary':    {},
        'email_sent': False,
    }

    try:
        data = _collect_daily_data(db, date_str)
        result['summary'] = data['summary']

        _ensure_report_dir()
        filename = f"SQM_일별리포트_{date_str.replace('-','')}.xlsx"
        filepath = os.path.join(REPORT_DIR, filename)

        ok = _write_excel(data, filepath)
        if ok:
            result['success']  = True
            result['filepath'] = filepath
            logger.info(f"[DailyReport] 생성 완료: {filepath}")

        # 이메일 전송
        if send_email:
            try:
                from utils.error_notifier import notify_daily_summary
                s = data['summary']
                result['email_sent'] = notify_daily_summary(
                    date_str      = date_str,
                    inbound_cnt   = s['inbound_cnt'],
                    outbound_cnt  = s['outbound_cnt'],
                    return_cnt    = s['return_cnt'],
                    total_weight_kg = s['inbound_kg'] + s['outbound_kg'],
                )
            except Exception as e:
                logger.debug(f"[DailyReport] 이메일 전송 스킵: {e}")

    except Exception as e:
        logger.error(f"[DailyReport] 생성 실패: {e}", exc_info=True)
        result['error'] = str(e)

    return result


def schedule_daily_report(db, hour: int = 23, minute: int = 55) -> None:
    """
    매일 지정 시각에 일별 리포트 자동 생성.
    별도 스레드로 실행 (tkinter GUI와 병행).

    Args:
        db     : SQMDatabase 인스턴스
        hour   : 실행 시각 (기본: 23시)
        minute : 실행 분 (기본: 55분)
    """
    import threading
    import time

    def _runner():
        logger.info(f"[DailyReport] 스케줄러 시작 — 매일 {hour:02d}:{minute:02d}")
        last_run_date = None
        while True:
            now = datetime.now()
            if (now.hour == hour and now.minute == minute
                    and now.date() != last_run_date):
                try:
                    generate_daily_report(db, send_email=True)
                    last_run_date = now.date()
                except Exception as e:
                    logger.error(f"[DailyReport] 스케줄 실행 오류: {e}")
            time.sleep(30)  # 30초마다 확인

    t = threading.Thread(target=_runner, daemon=True, name='DailyReportScheduler')
    t.start()
    logger.info("[DailyReport] 스케줄러 스레드 시작")
