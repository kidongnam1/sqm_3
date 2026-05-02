"""
SQM 프로그램 공통 엑셀 정렬 규칙.

- 숫자(int / float / Decimal, bool 제외): 가로 오른쪽, 세로 가운데
- 그 외(문자·날짜·빈 칸 등): 가로·세로 가운데

생성된 워크북 또는 저장 경로에 일괄 적용한다.
"""
from __future__ import annotations

import logging
from decimal import Decimal

logger = logging.getLogger(__name__)

try:
    from openpyxl.styles import Alignment
except ImportError:
    Alignment = None  # type: ignore

_CENTER = Alignment(horizontal="center", vertical="center") if Alignment else None
_RIGHT = Alignment(horizontal="right", vertical="center") if Alignment else None


def _is_numeric_alignment_value(val) -> bool:
    """표시용 숫자 셀 → 오른쪽 정렬 대상."""
    if val is None:
        return False
    if isinstance(val, bool):
        return False
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, Decimal):
        return True
    return False


def apply_sqm_workbook_alignment(wb) -> None:
    """
    모든 시트·사용 영역 셀에 SQM 정렬 규칙 적용 (기존 horizontal 정렬 덮어씀).
    """
    if Alignment is None:
        logger.warning("[sqm_excel_alignment] openpyxl 없음 — 건너뜀")
        return
    for ws in wb.worksheets:
        mr = ws.max_row or 1
        mc = ws.max_column or 1
        for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc):
            for cell in row:
                if _is_numeric_alignment_value(cell.value):
                    cell.alignment = _RIGHT
                else:
                    cell.alignment = _CENTER


def apply_sqm_excel_file_alignment(path: str) -> None:
    """디스크의 xlsx를 열어 정렬 적용 후 같은 경로에 저장."""
    if Alignment is None:
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("[sqm_excel_alignment] openpyxl 없음 — 파일 후처리 건너뜀")
        return
    wb = load_workbook(path)
    apply_sqm_workbook_alignment(wb)
    wb.save(path)
